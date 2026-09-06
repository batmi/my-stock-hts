# modules/account.py
import logging
from rich.table import Table
from rich.panel import Panel
from rich import box
from rich.prompt import Prompt
from rich.progress import Progress, SpinnerColumn, BarColumn, TextColumn
from rich.markup import escape
from datetime import datetime, timedelta
import os
import time
import config
from core import context # [추가]
from core import utils
import api
from core import jsonio
from core import trade_tags
from modules import db_manager
import json
import pandas as pd
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

def fetch_today_profit_summary(cano=None, acnt_prdt_cd=None, target_date=None):
    summary = {'buy_amt': 0, 'sell_amt': 0, 'total_cost': 0, 'realized_pl': 0}
    try:
        data = api.get_today_profit_summary(cano, acnt_prdt_cd, target_date=target_date)
        if data.get('rt_cd') == '0':
            out2 = data.get('output2')
            if isinstance(out2, list) and len(out2) > 0:
                summary_data = out2[0]
                summary['buy_amt'] = api.safe_int(summary_data.get('thdt_buy_amt'))
                summary['sell_amt'] = api.safe_int(summary_data.get('thdt_sll_amt'))
                summary['total_cost'] = api.safe_int(summary_data.get('thdt_tlex_amt'))
                summary['realized_pl'] = api.safe_int(summary_data.get('rlzt_pfls'))
    except Exception: pass
    return summary

def fetch_period_realized(cano=None, acnt_prdt_cd=None, start_date=None, end_date=None):
    """증권사가 집계한 기간 실현손익. 조회할 수 없으면 **None**을 돌려준다.

    [0과 '모른다'를 가른다] 이 값은 우리 DB 합계와 대조해 '우리가 실현손익을 다 알고
    있는가'를 판정하는 데 쓴다. 못 구했는데 0을 돌려주면 그 대조가 '완벽히 일치'로 읽혀
    가장 위험한 경우(모르는 매매가 있다)를 그냥 통과시킨다.

    반환: {'realized': 실현손익, 'cost': 제비용} 또는 None.
      · 두 값을 함께 주는 이유: rlzt_pfls 가 제비용을 포함하는지 여부가 계좌·응답에 따라
        다르게 관측된다. 판정하는 쪽이 두 해석(포함/미포함) 중 하나라도 맞으면 통과시킨다.
    날짜는 YYYY-MM-DD 또는 YYYYMMDD 를 받는다.
    """
    def _ymd(d):
        return d.replace("-", "") if d else None

    try:
        data = api.get_period_profit_summary(cano, acnt_prdt_cd,
                                             start_date=_ymd(start_date), end_date=_ymd(end_date))
        if not data or data.get('rt_cd') != '0':
            return None
        out2 = data.get('output2')
        if not isinstance(out2, list) or not out2:
            return None                      # 미지원(모의·토스·관찰 모드) 또는 빈 응답
        row = out2[0]
        if row.get('rlzt_pfls') is None:
            return None
        return {'realized': api.safe_int(row.get('rlzt_pfls')),
                'cost': api.safe_int(row.get('thdt_tlex_amt'))}
    except Exception as e:
        logger.debug(f"[기간 실현손익] 조회 실패: {e}")
        return None


def fetch_today_history(cano=None, acnt_prdt_cd=None, target_date=None):
    summary = {'buy_total': 0, 'sell_total': 0}
    try:
        data = api.get_today_history(cano, acnt_prdt_cd, target_date=target_date)
        if data.get('rt_cd') == '0':
            trades = data.get('output1', [])
            if trades:
                for item in trades:
                    amt = api.safe_int(item.get('tot_ccld_amt'))
                    type_cd = item.get('sll_buy_dvsn_cd')
                    if type_cd == '01': summary['sell_total'] += amt
                    elif type_cd == '02': summary['buy_total'] += amt
    except Exception: pass
    return summary

def fetch_domestic_balance(cano=None, acnt_prdt_cd=None):
    """국내 주식 잔고 데이터를 조회하여 반환. **조회 실패는 (None, None)**.

    [왜 실패를 빈 목록으로 접으면 안 되는가] api.get_domestic_balance 는 실패를
     (None, None) 으로 정확히 알려 주는데, 종전에는 여기서 그것을 [] 로 접었다.
     그 목록을 받는 곳은 손으로 손절하러 들어온 매도 화면(trading.select_stock_from_balance)
     이고, 거기서 빈 목록은 "매도 가능한 잔고가 없습니다"로 나온다 — 운영자가 **팔 것이
     없다고 믿고 나간다**. 모른다는 것을 없다고 답하지 않는다([[unknown-vs-empty]]).

    [한 줄이 깨져도 목록 전체를 잃지 않는다] 종전에는 api.safe_int(item.get('hldg_qty')) 가
     루프 밖의 except 로 튀어, 이상한 한 줄 때문에 **나머지 종목이 통째로 사라졌다**.
     읽을 수 없는 줄은 그 줄만 건너뛰고 남긴다.
    """
    holdings = []
    summary = None

    try:
        # api.py의 함수 사용 (내부에서 OPSQ2001 재시도 및 토큰 처리)
        output1, output2 = api.get_domestic_balance(cano, acnt_prdt_cd)
    except Exception as e:
        logger.error(f"국내 잔고 조회 실패: {e}")
        return None, None

    if output1 is None:
        return None, None

    for item in output1:
        try:
            qty = int(float(str(item.get('hldg_qty', '')).strip() or 0))
        except (TypeError, ValueError):
            logger.warning(f"국내 잔고 보유수량을 읽을 수 없습니다: "
                           f"{item.get('pdno')} {item.get('hldg_qty')!r} — 이 줄만 건너뜁니다")
            continue
        if qty > 0:
            holdings.append(item)
    if output2:
        summary = output2[0]

    return holdings, summary

def fetch_overseas_balance(cano=None, acnt_prdt_cd=None):
    """해외 주식 잔고 데이터를 조회하여 반환. **조회 실패는 None**(빈 목록과 구분)."""
    return api.get_overseas_balance(cano, acnt_prdt_cd)

def sync_today_trades():
    """금일 체결 내역을 API로 조회하여 DB의 단가(시장가=0) 정보를 업데이트 (모든 계좌 대상)"""
    logger.debug("[HISTORY_DEBUG] sync_today_trades() 시작")
    
    # 조회 대상 계좌 목록
    accounts = []
    if config.session.cano and config.session.acnt_prdt_cd:
        accounts.append({"cano": config.session.cano, "acnt": config.session.acnt_prdt_cd, "type": "MAIN"})
    
    if config.session.auto_cano and config.session.auto_acnt_prdt_cd:
        if config.session.auto_cano != config.session.cano or config.session.auto_acnt_prdt_cd != config.session.acnt_prdt_cd:
            accounts.append({"cano": config.session.auto_cano, "acnt": config.session.auto_acnt_prdt_cd, "type": "AUTO"})
            
    total_count = 0
    failed_accounts = []
    original_context = getattr(context.trade_context, 'use_auto_account', False)
    
    with Progress(
        SpinnerColumn(),
        TextColumn("[progress.description]{task.description}"),
        BarColumn(),
        TextColumn("[progress.percentage]{task.percentage:>3.0f}%"),
        console=config.console,
        transient=True
    ) as progress:
        task = progress.add_task("[cyan]최신 체결 내역 동기화 중...[/cyan]", total=len(accounts))
        
        try:
            for acc in accounts:
                cano = acc['cano']
                acnt = acc['acnt']
                
                try:
                    #  [계좌 라우팅 · 2026-09-06] cano 를 인자로 넘겨도 **앱키·토큰·TPS 버킷**은
                    #   thread_local(use_auto_account)이 고른다([[account-context-in-threads]]).
                    #   종전에는 그 값을 위에서 저장하고 finally 에서 되돌리기만 할 뿐 **한 번도
                    #   바꾸지 않았다** — 남아 있는 복원 코드가 그 흔적이다. 그래서 자동매매
                    #   계좌의 당일 체결을 실전 앱키로 조회했고(실측), 권한이 없으니 KIS 가
                    #   거절하거나 빈 목록을 준다. 거절이면 위 except 가 알리지만, 빈 목록이면
                    #   '오늘 체결 0건'으로 조용히 지나간다 — 그 체결에 평단·진입일·손절 기준이
                    #   붙을 근거가 통째로 사라진다.
                    with utils.AccountContext(cano):
                        data = api.get_today_history(cano, acnt)
                        ovrs_data = api.get_overseas_today_history(cano, acnt)
                    
                    all_trades = []
                    if data.get('rt_cd') == '0':
                        all_trades.extend(data.get('output1', []))
                    if ovrs_data.get('rt_cd') == '0':
                        all_trades.extend(ovrs_data.get('output', []))

                    if all_trades:
                        for item in all_trades:
                            #  [Fix 2026-09-06] 체결 **하나**의 오류가 그 계좌의 동기화를
                            #   통째로 끊던 자리(바깥 except 는 계좌 단위다). 같은 모양을
                            #   ConclusionMonitor 에서도 함께 고쳤다. 한 건이 깨지면 그 뒤
                            #   종목들의 체결까지 전부 빠진 채 화면은 건수만 말한다.
                            try:
                                odno = item.get('odno')
                                is_overseas_trade = 'ft_ccld_qty' in item
                            
                                #  증권사는 빈 문자열을 준다 — float('')·int('') 는 예외이고,
                                #  dict.get 의 기본값은 **키가 없을 때만** 쓰인다(키를 주고
                                #  값을 비우면 소용없다). api.safe_* 가 그 둘을 다 흡수한다.
                                if is_overseas_trade:
                                    avg_price = api.safe_float(item.get('ft_ccld_unpr3'), default=0.0)
                                    tot_qty = api.safe_int(item.get('ft_ccld_qty'))
                                else:
                                    avg_price = api.safe_float(item.get('avg_prvs'), default=0.0)
                                    tot_qty = api.safe_int(item.get('tot_ccld_qty'))
                            
                                if odno and avg_price > 0:
                                    # [수정] 체결 내역 분리 저장 (기존 내역 업데이트 대신 신규 추가)
                                    #  odno 는 당일 채번이라 날짜 없이는 유일하지 않다 —
                                    #  전체 이력에서 찾으면 오늘 체결이 누락된다.
                                    _scope_dt = str(item.get('ord_dt') or '')
                                    _scope = (f"{_scope_dt[:4]}-{_scope_dt[4:6]}-{_scope_dt[6:]}"
                                              if len(_scope_dt) == 8 and _scope_dt.isdigit()
                                              else datetime.now().strftime('%Y-%m-%d'))
                                    try:
                                        _already = db_manager.db.check_trade_exists(
                                            odno, "체결", on_date=_scope)
                                    except Exception as _ce:
                                        #  모르면 적지 않는다 — 중복 체결 행은 실현손익을
                                        #  이중 계상하고 되돌릴 수 없다. 다음 동기화가 다시 본다.
                                        logger.warning(f"[Account] {odno} 중복 여부를 확인하지 못해 "
                                                       f"이번 동기화에서 미룹니다: {_ce}")
                                        _already = True
                                    if not _already:
                                        if config.FILE_DEBUG_LEVEL == "DEBUG":
                                            logger.debug(f"[Account] 신규 체결 DB 저장 시도: {odno}")
                                    
                                        # 체결 시간 포맷팅
                                        ord_dt = item.get('ord_dt', '')
                                        ord_tmd = item.get('ord_tmd', '')
                                        trade_time = None
                                        if len(ord_dt) == 8 and len(ord_tmd) == 6:
                                            trade_time = f"{ord_dt[:4]}-{ord_dt[4:6]}-{ord_dt[6:]} {ord_tmd[:2]}:{ord_tmd[2:4]}:{ord_tmd[4:]}"
                                    
                                        type_cd = item.get('sll_buy_dvsn_cd')
                                        type_name = item.get('sll_buy_dvsn_cd_name')
                                    
                                        if type_name:
                                            type_str = type_name
                                        else:
                                            type_str = "매수" if type_cd == '02' else ("매도" if type_cd == '01' else "기타")
                                        
                                        # 원 주문 유형 조회 (수동/자동 태그 반영)
                                        #  위 중복 판정과 **같은 날짜**로 찾는다. 날짜가 없으면
                                        #  몇 달 전 같은 번호의 주문이 '원 주문'으로 잡혀 오늘
                                        #  체결이 그 type·손익·점수를 물려받는다(실측: 두 달 전
                                        #  '매도' 행이 오늘 매수의 원 주문으로 잡혔다).
                                        origin_trade = db_manager.db.get_trade_by_odno(odno, on_date=_scope)
                                        profit_amt = 0
                                        profit_rate = 0.0
                                        score = 0
                                        stop_loss_rate = 0.0
                                    
                                        if origin_trade:
                                            type_str = origin_trade['type'] # 기존 타입 유지
                                            profit_amt = origin_trade.get('profit_amt', 0)
                                            profit_rate = origin_trade.get('profit_rate', 0.0)
                                            score = origin_trade.get('strategy_score', 0)
                                            stop_loss_rate = api.safe_float(origin_trade.get('stop_loss_rate'), default=0.0)
                                            orig_reason = origin_trade.get('reason', '')
                                            if orig_reason and "체결 확인" not in orig_reason:
                                                reason_to_save = f"체결 확인 ({orig_reason})"
                                            else:
                                                reason_to_save = "체결 확인"
                                        else:
                                            reason_to_save = "체결 확인"
                                            # [추가] trades 테이블에 없으면 reserved_orders 테이블에서 예약 발동 주문인지 조회
                                            # [수정] DB 큐를 경유하는 전용 메서드 사용 (워커 스레드 커넥션의 교차 스레드 사용 방지)
                                            try:
                                                r_row = db_manager.db.get_reserved_order_by_odno(odno, on_date=_scope)
                                                if r_row:
                                                    t_type = "매수" if r_row['order_type'] == 'buy' else "매도"
                                                    type_str = f"{t_type}(예약)"
                                                    c_type = r_row['condition_type']
                                                    tp_val = r_row['target_price']
                                                    res_reason = f"조건: {c_type}"
                                                    if c_type == 'TIME': res_reason += f" ({r_row['target_time']})"
                                                    elif 'SCORE' in c_type: res_reason += f" (목표: {tp_val}점)"
                                                    elif 'RSI' in c_type: res_reason += f" (목표: {tp_val})"
                                                    elif 'EMA' in c_type: res_reason += f" (EMA {int(tp_val)} {'돌파' if 'UP' in c_type else '이탈'})"
                                                    elif c_type == 'TRAILING_BUY': res_reason += f" (바닥반등 {tp_val}%)"
                                                    elif c_type == 'TRAILING_SELL': res_reason += f" (고점하락 {tp_val}%)"
                                                    elif c_type == 'SMART_MONEY': res_reason += " (수급 턴어라운드)"
                                                    elif c_type.startswith('STATE_'): res_reason += f" (상태진입: { {'STATE_STRONGBUY': '강매수', 'STATE_BUY': '매수', 'STATE_MR': '역매수'}.get(c_type, c_type)})"
                                                    elif c_type == 'HOLDING_EXIT': res_reason += " (보유분석 청산)"
                                                    elif c_type == 'COMPOSITE': res_reason += " (복합조건)"
                                                    else: res_reason += f" (목표가 {tp_val})"
                                                    reason_to_save = f"체결 확인 ({res_reason})"
                                            except Exception as e:
                                                logger.debug(f"[Account] 예약 주문 조회 실패: {e}")
                                    
                                        db_manager.db.insert_trade(
                                            type_str, item.get('pdno'), item.get('prdt_name') or item.get('ovrs_item_name') or item.get('item_nm'), 
                                            tot_qty, avg_price, odno, 
                                            order_status="체결", custom_time=trade_time,
                                            reason=reason_to_save,
                                            profit_amt=profit_amt, profit_rate=profit_rate, score=score,
                                            stop_loss_rate=stop_loss_rate
                                        )
                                        # [추가] 시장가 주문 등의 경우를 위해 원 주문(접수)의 단가도 체결가로 업데이트
                                        # [수정] 원본 주문 보존을 위해 업데이트 제거
                                        # db_manager.db.update_trade(odno, price=avg_price)
                                    
                                        total_count += 1
                                    else:
                                        if config.FILE_DEBUG_LEVEL == "DEBUG":
                                            logger.debug(f"[Account] 이미 존재하는 체결 내역입니다. 저장 스킵 (ODNO: {odno})")
                            except Exception as _ie:
                                #  이 건만 건너뛰되 계좌를 실패로 표시한다 — 아래 경고가
                                #  "이 계좌의 오늘 체결이 빠져 있을 수 있다"고 밝힌다.
                                if f"{cano}-{acnt}" not in failed_accounts:
                                    failed_accounts.append(f"{cano}-{acnt}")
                                logger.warning(
                                    f"[Account] 체결 {item.get('odno')} "
                                    f"({item.get('prdt_name') or item.get('ovrs_item_name')}) "
                                    f"적재 실패 — 이 건만 건너뜁니다: "
                                    f"{type(_ie).__name__}: {_ie}", exc_info=True)
                                continue
                except Exception as e:      # noqa: BLE001
                    #  [Fix 2026-09-05] 종전에는 `except Exception: pass` 였다. 이 블록은
                    #   당일 체결을 trades 에 적재하는 전체 경로를 감싼다 — 조회 실패도,
                    #   insert_trade 실패도 여기서 통째로 사라졌다. 그러면 그 계좌의 오늘
                    #   체결이 DB 에 없는 채로 화면은 '동기화 완료'라고 말하고, 평단·진입일·
                    #   손절 기준이 붙을 자리를 잃는다(체결 기록이 그 모든 것의 근거다).
                    #   2026-09-03 에 DB 조회 실패 20곳을 드러냈는데 이 바깥 except 가 남아
                    #   있었다. 계좌별로 한 줄은 반드시 남긴다.
                    failed_accounts.append(f"{cano}-{acnt}")
                    logger.warning(f"[Account] 당일 체결 동기화 실패({cano}-{acnt}): "
                                   f"{type(e).__name__}: {e}", exc_info=True)
                progress.advance(task)
        finally:
            context.trade_context.use_auto_account = original_context

    if failed_accounts:
        #  화면에도 밝힌다 — 로그만 남기면 '동기화 완료 N건'을 그대로 믿는다.
        config.console.print(
            f"[yellow]⚠️ 당일 체결 동기화에 실패한 계좌가 있습니다: "
            f"{', '.join(failed_accounts)} — 아래 내역에 그 계좌의 오늘 체결이 빠져 있을 수 "
            f"있습니다('없음'이 아닙니다).[/yellow]")

    logger.debug(f"[HISTORY_DEBUG] sync_today_trades() 종료. 처리 건수: {total_count}")
    return total_count

def run_holding_analysis(domestic_items, overseas_items, restricted_codes=None, account=None):
    """국내/해외 보유 종목에 시스템 매도 판단을 적용한다. (읽기 전용 · 부수효과 없음)

    메뉴 2의 종목 분석은 차트만 보고 '지금 새로 살 만한가'를 판정하지만, 보유 분석은
    매입단가·수익률·최고가(트레일링)·보유일수·반익절 이력·매수 시점 ATR 손절률·개별 룰까지
    반영한다. 같은 종목이라도 두 결과가 갈릴 수 있으며, 여기서 쓰는 판정 로직은 자동매매가
    실제로 청산에 쓰는 analyze_sell과 동일하다.

    account: 'cano-acnt'. 매수 기록을 그 계좌로 거른다 — 같은 종목을 두 계좌에서 들고
     있으면 남의 매수로 손절선·진입일이 계산되기 때문이다. 자동매매 루프는 이미 계좌로
     가르고 있으므로, 여기를 비워두면 화면과 실제 판정이 갈린다.

    반환: {code: analyze_sell 결과} — 실패한 종목은 키가 없다.
    """
    #  [0% 는 '모름'이 아니다 · 2026-09-06] 종전에는 `float(... or 0)` 이었다.
    #   증권사 어댑터는 일부 필드를 0/누락으로 준다 — 그러면 0% 가 손절선(음수)보다 위라
    #   **손절 이탈 판정이 '아직 괜찮다'로 뒤집힌다**(실측: -14.3% 포지션이 0.0% 로 들어갔다).
    #   이 결과는 화면(메뉴 9-2)뿐 아니라 예약 HOLDING_EXIT 를 거쳐 실매도까지 간다.
    #   판정은 engine.holding_profit_rate 하나가 갖는다 — 모르면 None 을 올리고,
    #   analyze_holdings 가 평단·현재가로 정확히 복원한다.
    from modules.auto_trade import engine as _engine
    entries = []
    for item in domestic_items:
        try:
            qty = int(float(item.get('hldg_qty') or 0))
        except (TypeError, ValueError):
            qty = None
        entries.append({
            'code': item['pdno'],
            'name': item.get('prdt_name', ''),
            'buy_price': float(item.get('pchs_avg_pric') or 0),
            'current_price': float(item.get('prpr') or 0),
            'profit_rate': _engine.holding_profit_rate(item),
            'is_overseas': False,
            # 진입일 복원(증권사 체결 재생)에서 '조회 구간보다 오래된 포지션' 판별에 쓴다.
            'qty': qty if qty and qty > 0 else None,
        })

    for item in overseas_items:
        qty = (api.safe_float(item.get('ovrs_cblc_qty'), default=0.0)
         or api.safe_float(item.get('ord_psbl_qty'), default=0.0))
        pchs_avg = float(item.get('pchs_avg_pric') or 0)
        cur_price = float(item.get('ovrs_now_pric') or 0)
        if cur_price == 0 and qty > 0:
            # 현재가 미제공 시 평가금액에서 역산 (표시 로직과 동일 기준)
            profit = float(item.get('frcr_evlu_pfls_amt') or 0)
            cur_price = (qty * pchs_avg + profit) / qty
        entries.append({
            'code': item.get('ovrs_pdno', ''),
            'name': item.get('ovrs_item_name', ''),
            'buy_price': pchs_avg,
            'current_price': cur_price,
            #  해외 항목에는 'prpr' 이 없어 holding_profit_rate 의 복원 경로가 닿지 않는다
            #  — 위에서 이미 역산한 cur_price 가 있으므로 analyze_holdings 쪽에 맡긴다.
            'profit_rate': _engine.holding_profit_rate(item),
            'is_overseas': True,
        })

    entries = [e for e in entries if e['code']]
    if not entries:
        return {}

    try:
        from modules import auto_trade
        return auto_trade.analyze_holdings(entries, restricted_codes=restricted_codes,
                                           account=account)
    except Exception as e:
        logger.warning(f"보유 분석 실패: {e}")
        return {}

def _fmt_state_cell(res, show_auto_status=True):
    """보유분석 결과를 잔고 테이블의 '상태' 셀로 변환. (상태/점수, 청산 신호는 강조)

    시스템 자동 매도 대상이 아닌 포지션은 둘째 줄에 '수동'을 붙인다.
    청산 신호가 떠도 시스템이 팔지 않는다는 사실을 표에서 바로 알 수 있어야 한다.
    """
    if not res:
        return "[dim]-[/dim]"

    score = res.get('score')
    score_str = f" [dim]{score:.1f}[/dim]" if isinstance(score, (int, float)) else ""

    if res.get('action') == 'sell':
        cell = f"[bold blue]청산[/]{score_str}"
    else:
        state = res.get('state') or "-"
        color = (res.get('state_color') or "[white]").strip('[]') or "white"
        cell = f"[{color}]{state}[/]{score_str}"

    if show_auto_status and res.get('unmanaged'):
        cell += "\n[yellow]수동[/]"
    return cell

def _fmt_holding_days_cell(res, code=None):
    """보유일수 — 시간청산 임계에 도달하면 노란색으로 경고한다.

    매수일은 시스템 DB → 증권사 체결 내역 순으로 찾고, 둘 다 없으면 오늘 매수로 본다(0일).

    [Fix 2026-09-04] 임계는 **그 종목에 실제로 적용되는 값**을 쓴다. 개별 룰이
     time_stop_days 를 바꾸면 청산 판정(engine.analyze_sell)은 그것을 쓰는데 여기서는
     전역값만 봐서, 룰이 걸린 종목의 경고가 실제 청산 시점과 어긋났다.
    """
    if not res:
        return "[dim]-[/dim]"

    days = res.get('holding_days') or 0
    from modules.auto_trade.common import effective_time_stop_days
    limit = effective_time_stop_days(code or res.get('code'))
    if config.SELL_STRATEGY.get("TIME_STOP_USE", True) and limit and days >= limit:
        return f"[yellow]{days}일[/]"
    return f"{days}일"

def _fmt_profit_cell(amount_str, rate_str, color=""):
    """평가손익 = 금액 + 그 아래 수익률. 최고가 칸과 같은 모양으로 맞춘다.

    [표기] 두 줄 모두 우측 정렬이고 괄호는 없다.
     - 괄호: 한 칸 안에 있다는 것만으로 수익률이 금액의 부속값임은 이미 자명한데,
       괄호까지 두면 좁은 열에서 두 칸을 더 먹는다.
     - 정렬: 가운데 정렬이면 자릿수가 다른 종목마다 소수점 위치가 흔들려 세로로
       훑을 수가 없다. 금액과 같은 축(우측)에 걸어야 종목 간 비교가 눈으로 된다.
     - 색: 금액과 같은 색상룰(상승 red · 하락 blue · 보합 white)을 수익률에도 건다.
       한 칸의 두 줄이 같은 사실을 말하는데 한 줄만 무채색이면 부호를 두 번 읽게 된다.

    한 셀 안에서 줄마다 스타일이 다르므로 문자열이 아니라 Group으로 돌려준다 — 공백을
    채워 밀어내는 방식은 통하지 않는다(렌더러가 줄 끝 공백을 잘라내며, NBSP도 파이썬
    rstrip 대상이라 살아남지 못한다).
    """
    from rich.console import Group
    from rich.text import Text

    amount = f"{color}{amount_str}[/]" if color else amount_str
    rate = f"{color}{rate_str}[/]" if color else rate_str
    return Group(Text.from_markup(amount, justify="right"),
                 Text.from_markup(rate, justify="right"))


def _fmt_mfe_cell(res, is_overseas=False):
    """최고가와 그 시점의 최대 평가수익(MFE) — 수익 반납폭 확인용."""
    if not res:
        return "[dim]-[/dim]"

    highest = res.get('highest_price') or 0
    if highest <= 0:
        return "[dim]-[/dim]"

    price_str = f"${highest:,.2f}" if is_overseas else f"{int(highest):,}"
    mfe = res.get('max_profit_rate') or 0.0
    # 괄호 없이 둘째 줄로만 구분한다(평가손익 칸과 동일 규칙).
    return f"{price_str}\n[dim]{mfe:+.1f}%[/dim]"

def _fmt_ts_stop(res, is_overseas=False, buy_price=0):
    """샹들리에 TS 청산선. 고정/ATR 손절과 달리 실제 주청산선이라 별도 줄로 표시한다.

    발동선(activation)은 breakeven 모드에서 종목 변동성에 따라 20%~90%까지 벌어지므로
    무장 전후 모두 병기한다 — 고정 %일 때는 전 종목 공통 상수라 생략해도 무방했지만,
    이제는 그 값이 없으면 화면만 보고 무장 여부를 설명할 수 없다. 무장 전에는 발동
    '가격'도 함께 보여준다(같은 열의 손절선·청산선이 모두 가격이라 %만으로는 비교 불가).
    """
    from modules.auto_trade.engine import ts_activation_dynamic

    ts = (res or {}).get('ts')
    if not ts:
        return None

    def _p(v):
        return f"${v:,.2f}" if is_overseas else f"{round(v):,}"

    # 표기는 한 줄로 묶는다. 셀이 세 줄이 되면 표 전체가 종목당 3행으로 늘어난다.
    #  (열 폭이 모자라면 rich가 fold로 접어 넘긴다 — 잘라내지 않는다. add_column 주석 참조)
    #  ↑ = 이 가격에 닿으면 무장, ≥ = 무장에 필요한 MFE. 범례는 표 캡션에 한 번만 단다.
    act = ts.get('activation') or 0
    dynamic = ts_activation_dynamic()

    if not ts.get('armed'):
        if not (buy_price and buy_price > 0 and act > 0):
            return f"[dim]TS: {act:+.1f}% 도달 시[/dim]"
        arm_price = buy_price * (1 + act / 100)
        # 발동가만 보여주면 '그래서 어디서 잘리나'가 빠진다. 발동 시점의 청산선까지 같이
        #  준다 — 고점이 발동가일 때의 콜백으로 환산한다(ATR×배수는 고점에 비례해 콜백을
        #  좁히므로, 지금 콜백을 그대로 쓰면 청산선이 실제보다 낮게 나온다).
        cb = ts.get('callback') or 0
        highest = (res or {}).get('highest_price') or 0
        if cb > 0 and highest > 0:
            cb = max(config.SELL_STRATEGY.get("TRAILING_STOP_CALLBACK_RATE", 5.0),
                     cb * highest / arm_price)
            # 두 가격이 다 들어가야 하므로 %는 정수로 줄인다. 여기서 행동을 정하는 값은
            #  '얼마에 켜지나'(가격)이고 %는 맥락일 뿐이다.
            # 콜백(-N%)까지 붙이는 이유: 이 줄의 두 가격만으로는 '얼마나 밀려야 잘리나'를
            #  화면에서 알 수 없어 매번 역산해야 했다(1 - 청산선/발동가). 무장 후 행은 이미
            #  콜백을 찍고 있으므로, 병기해야 무장 전후가 같은 정보를 말한다.
            return (f"[dim]TS:[/dim]{_p(arm_price)}[dim]↑{act:.0f}%→[/dim]"
                    f"[bold magenta not dim]{_p(arm_price * (1 - cb / 100))}[/]"
                    f"[dim](-{cb:.0f}%)[/dim]")
        return f"[dim]TS:[/dim]{_p(arm_price)}[dim]↑({act:+.1f}%)[/dim]"

    # 청산선은 이 칸에서 유일하게 '실제로 포지션을 끝내는' 가격이다. ATR 손절선(파랑)과
    #  구분되게 굵은 보라로 띄운다 — 미무장 행의 → 뒤 예상 청산선도 같은 색을 쓴다.
    line = f"[dim]TS:[/dim][bold magenta not dim]{_p(ts['stop_price'])}[/][dim](-{ts['callback']:.1f}%)[/dim]"
    if dynamic and act > 0:
        # 무장 후에는 이미 넘어선 문턱이라 정수로 줄여 열 폭을 아낀다(미무장은 .1f 유지).
        line += f"[dim]≥{act:.0f}%[/dim]"
    return line

def _fmt_stop_cell(res, buy_price, is_overseas=False, code=None):
    """청산선 셀 — 손절선(ATR/고정/BEP)과 TS 청산선을 항상 두 줄로 표시한다.

    열 이름이 '손절가'였을 때는 TS 줄이 이름과 맞지 않았다. 두 줄 다 '포지션이 끝나는
    가격'이라는 공통점으로 묶고, 어느 선인지는 각 줄의 접두어(ATR/BEP/고정/TS)가 말한다.

    [중요] 표시값은 보유 분석(analyze_sell)이 실제로 적용한 손절률(applied_sl_rate)에서
    유도한다. 예전에는 이 셀이 DB의 매수 기록을 따로 읽어 손절선을 재구성했는데, 기록이
    없는 HTS 직접 매수분은 '미사용'으로 비워두면서도 엔진은 전역 고정 손절로 판정하고
    있어 화면과 판정이 어긋났다(잔고에 '미사용'인데 청산 사유는 '손절').

    접두어는 실제로 지배하는 선을 가리킨다 — BEP(본전 청산이 손절선을 끌어올린 상태)
    > ATR(변동성 기반) > 고정(USE_ATR_STOP이 꺼졌거나 ATR 산출 불가).
    """
    def _p(v):
        # int() 절삭이 아니라 반올림 — buy_price*(1+rate/100)은 부동소수점 오차로 정수 바로
        #  아래에 떨어지는 일이 흔해(10000*1.005 = 10049.999…) 손절가가 1원씩 낮게 찍혔다.
        return f"${v:,.2f}" if is_overseas else f"{round(v):,}"

    parts = []
    sl_rate = (res or {}).get('applied_sl_rate')
    if sl_rate is not None and sl_rate != 0 and buy_price > 0:
        if (res or {}).get('is_bep_applied'):
            label = "BEP"
        elif (res or {}).get('is_atr_stop'):
            label = "ATR"
        else:
            label = "고정"
        stop_price = buy_price * (1 + sl_rate / 100)
        # 손절가도 TS 청산선과 같이 굵게 띄운다 — 열 스타일이 dim이라 not dim으로 끊어야
        #  실제로 강조된다(bold만 주면 dim과 겹쳐 오히려 흐려진다).
        parts.append(f"[dim]{label}:[/dim][bold blue not dim]{_p(stop_price)}[/]"
                     f"[dim]({sl_rate:+.1f}%)[/dim]")

    ts_line = _fmt_ts_stop(res, is_overseas=is_overseas, buy_price=buy_price)
    if ts_line:
        parts.append(ts_line)

    return "\n".join(parts) if parts else "[dim]미사용[/dim]"

def _decorate_name(name, code, marks_ctx):
    """종목명 뒤에 제한(-)/개별룰(+)/메모(=)/예약(*) 마크를 붙인다."""
    if not marks_ctx:
        return name

    marks = []
    if code in marks_ctx.get('restricted', ()): marks.append("-")
    if code in marks_ctx.get('rules', ()): marks.append("+")
    if code in marks_ctx.get('memo', ()): marks.append("=")
    if code in marks_ctx.get('reserved', ()): marks.append("[magenta]*[/magenta]")

    mark_str = "".join(marks)
    return f"{name}[dim]{mark_str}[/dim]".strip() if mark_str else name

def _print_ts_legend():
    """TS 발동 표기 범례. 발동선이 종목마다 다른 체제에서만 필요하다.

    표 캡션(표 바로 아래)이 아니라 합계 줄 다음에 한 줄 띄우고 낸다 — 표에 바로 붙으면
    마지막 종목의 행처럼 읽힌다.
    """
    from modules.auto_trade.engine import ts_activation_dynamic
    if not ts_activation_dynamic():
        return
    config.console.print()
    config.console.print(
        "[dim]  ※ TS 발동: 손익분기 연동(종목 변동성이 결정) · "
        "↑=TS가 켜지는 가격(매수가 대비 %) · →=그때 생기는 청산선 · ≥=TS가 켜진 MFE[/dim]")


def build_domestic_holdings_table(items, holding_analysis, marks_ctx=None, title="\n[국내] 계좌 잔고 현황", show_auto_status=True):
    """국내 보유 종목 표를 만든다. ([9]-2 잔고와 [9]-5 포지션 분석이 공유)

    items: KIS 국내 잔고 output1 형식 (pdno/prdt_name/hldg_qty/pchs_avg_pric/prpr/...)
    반환: (table, {'pchs','eval','profit','count'}, [(종목명, 코드, 청산사유, 미관리사유)])
    """
    # show_lines: 상태 칸의 '수동'·손절가의 TS 등 셀이 여러 줄이라 종목 경계가 흐려진다.
    #  행 사이 흐린 실선으로 구분한다.
    table = Table(title=title, box=box.HORIZONTALS, show_header=True, header_style="dim",
                  border_style="dim", show_lines=True)
    table.add_column("종목명", justify="left")
    table.add_column("코드", justify="center", style="dim")
    table.add_column("상태", justify="center")            # [추가] 보유 분석 결과
    table.add_column("보유수량", justify="right")
    table.add_column("매입단가", justify="right")
    table.add_column("현재가", justify="right")
    table.add_column("매입금액", justify="right")
    table.add_column("평가금액", justify="right")
    # 평가손익과 수익률은 한 칸에 두 줄로 묶는다. 어차피 행이 2줄이라 세로 공간은 공짜인
    #  반면, 열을 하나 줄이면 그만큼이 손절가 칸으로 간다(TS 표기가 잘리던 자리).
    table.add_column("평가손익(원/%)", justify="right")
    table.add_column("보유일", justify="right", style="dim")
    table.add_column("최고가", justify="right", style="dim")
    # [overflow] 청산선 셀은 'TS:57,470↑37%→44,245(-23.4%)≥30%'처럼 공백이 없는 한 덩어리라
    #  rich가 줄바꿈할 지점을 찾지 못하고 기본값(ellipsis)으로 뒤를 잘라낸다. 잘린 자리에
    #  들어가는 것이 하필 '실제 청산 가격'이라 화면만 보고는 어디서 잘리는지 알 수 없었다.
    #  fold로 바꿔 폭이 모자라면 다음 줄로 접어 넘긴다 — 좁은 터미널에서 행이 한 줄 늘 수는
    #  있어도 값이 사라지지는 않는다.
    table.add_column("청산선", justify="right", style="dim", overflow="fold")

    totals = {'pchs': 0, 'eval': 0, 'profit': 0, 'count': 0}
    sell_signals = []

    for item in items:
        code = item['pdno']
        raw_name = item['prdt_name']
        res = holding_analysis.get(code)

        qty = api.safe_int(item.get('hldg_qty'))
        buy_price = api.safe_float(item.get('pchs_avg_pric'), default=0.0)
        cur_price = api.safe_int(item.get('prpr'))
        eval_amt = api.safe_int(item.get('evlu_amt'))
        profit = api.safe_int(item.get('evlu_pfls_amt'))
        rate = api.safe_float(item.get('evlu_pfls_rt'), default=0.0)
        pchs_amt = int(qty * buy_price)

        totals['pchs'] += pchs_amt
        totals['eval'] += eval_amt
        totals['profit'] += profit
        totals['count'] += 1

        if res and res.get('action') == 'sell':
            sell_signals.append((raw_name, code, res.get('reason', '')))

        p_color = "[red]" if rate > 0 else ("[blue]" if rate < 0 else "[white]")
        table.add_row(
            _decorate_name(raw_name, code, marks_ctx),
            code,
            _fmt_state_cell(res, show_auto_status=show_auto_status),
            f"{qty:,}주",
            f"{buy_price:,.0f}원",
            f"{cur_price:,}원",
            f"{pchs_amt:,}원",
            f"{eval_amt:,}원",
            _fmt_profit_cell(f"{profit:+,}원", f"{rate:+.2f}%", p_color),
            _fmt_holding_days_cell(res, code=code),
            _fmt_mfe_cell(res, is_overseas=False),
            _fmt_stop_cell(res, buy_price, is_overseas=False, code=code)
        )

    return table, totals, sell_signals

def build_overseas_holdings_table(items, holding_analysis, marks_ctx=None, title="\n[해외] 계좌 잔고 현황", show_auto_status=True):
    """해외 보유 종목 표를 만든다. ([9]-2 잔고와 [9]-5 포지션 분석이 공유)

    items: KIS 해외 잔고 형식 (ovrs_pdno/ovrs_item_name/ovrs_cblc_qty/pchs_avg_pric/...)
    반환: (table, {'pchs','eval','profit','count'}, [(종목명, 코드, 청산사유, 미관리사유)])
    """
    # show_lines: 상태 칸의 '수동'·손절가의 TS 등 셀이 여러 줄이라 종목 경계가 흐려진다.
    #  행 사이 흐린 실선으로 구분한다.
    table = Table(title=title, box=box.HORIZONTALS, show_header=True, header_style="dim",
                  border_style="dim", show_lines=True)
    table.add_column("종목명", justify="left")
    table.add_column("코드", justify="center", style="dim")
    table.add_column("상태", justify="center")            # [추가] 보유 분석 결과
    table.add_column("거래소", justify="center")
    table.add_column("보유수량", justify="right")
    table.add_column("매입단가($)", justify="right")
    table.add_column("현재가($)", justify="right")
    table.add_column("매입금액($)", justify="right")
    table.add_column("평가금액($)", justify="right")
    table.add_column("평가손익($/%)", justify="right")   # 국내 표와 같은 이유로 두 줄 묶음
    table.add_column("보유일", justify="right", style="dim")
    table.add_column("최고가", justify="right", style="dim")
    table.add_column("청산선", justify="right", style="dim", overflow="fold")  # 잘림 방지(국내 표 주석 참조)

    totals = {'pchs': 0.0, 'eval': 0.0, 'profit': 0.0, 'count': 0}
    sell_signals = []

    for item in items:
        qty = (api.safe_float(item.get('ovrs_cblc_qty'), default=0.0)
         or api.safe_float(item.get('ord_psbl_qty'), default=0.0))
        if qty <= 0:
            continue

        code = item.get('ovrs_pdno', '-')
        raw_name = item.get('ovrs_item_name', '-')
        res = holding_analysis.get(code)

        pchs_avg = api.safe_float(item.get('pchs_avg_pric'), default=0.0)
        profit = api.safe_float(item.get('frcr_evlu_pfls_amt'), default=0.0)
        rate = api.safe_float(item.get('evlu_pfls_rt'), default=0.0)
        exc_name = item.get('_exchange', '')
        cur_price = api.safe_float(item.get('ovrs_now_pric'), default=0.0)
        item_pchs = qty * pchs_avg
        item_eval = item_pchs + profit
        if cur_price == 0 and qty > 0: cur_price = item_eval / qty

        totals['pchs'] += item_pchs
        totals['eval'] += item_eval
        totals['profit'] += profit
        totals['count'] += 1

        if res and res.get('action') == 'sell':
            sell_signals.append((raw_name, code, res.get('reason', '')))

        color = "[red]" if profit > 0 else ("[blue]" if profit < 0 else "[white]")
        table.add_row(
            _decorate_name(raw_name, code, marks_ctx),
            code,
            _fmt_state_cell(res, show_auto_status=show_auto_status),
            exc_name,
            f"{qty:,.0f}",
            f"{pchs_avg:,.2f}",
            f"{cur_price:,.2f}",
            f"{item_pchs:,.2f}",
            f"{item_eval:,.2f}",
            _fmt_profit_cell(f"{profit:+,.2f}", f"{rate:+.2f}%", color),
            _fmt_holding_days_cell(res, code=code),
            _fmt_mfe_cell(res, is_overseas=True),
            _fmt_stop_cell(res, pchs_avg, is_overseas=True, code=code)
        )

    return table, totals, sell_signals

def _print_sell_signals(signals):
    """청산 신호가 걸린 종목의 사유를 표 아래 각주로 출력한다.

    signals: [(종목명, 코드, 사유)]
    (시스템 미관리 여부는 표의 상태 칸 '수동' 표기로 드러나므로 여기서 중복하지 않는다)
    """
    if not signals:
        return

    config.console.print()
    config.console.print(f"[bold blue]  ⚠ 청산 신호 {len(signals)}건[/bold blue] [dim](시스템 트레이딩 매도 기준)[/dim]")
    for name, code, reason in signals:
        # 사유에 '[점수:3.1, RSI:...]' 형태의 대괄호가 포함되므로 rich 마크업으로 해석되지 않게 이스케이프
        config.console.print(f"    [dim]·[/dim] [bold]{name}[/bold]([dim]{code}[/dim]): {escape(reason)}")

def _ask_positive_number(prompt, cast=float, default=None, allow_cancel=True):
    """양수 입력 헬퍼. 취소(b/q)면 None을 반환한다."""
    while True:
        raw = Prompt.ask(prompt, default=default) if default is not None else Prompt.ask(prompt)
        if raw is None:
            return None
        raw = str(raw).strip().replace(",", "")
        if allow_cancel and raw.lower() in ('b', 'q', ''):
            return None
        try:
            val = cast(raw)
        except ValueError:
            config.console.print("[red]숫자를 입력하세요.[/red]")
            continue
        if val <= 0:
            config.console.print("[red]0보다 큰 값을 입력하세요.[/red]")
            continue
        return val

def _ask_buy_date(prompt, default="", empty_as_today=False):
    """매수일 입력 헬퍼. 반환 (date, 취소여부).

    빈 입력이면 (None, False) — 매수일 미상 처리 (empty_as_today=True인 경우 오늘 날짜).
    수정 시에는 현재값을 default로 넘기며, '-'를 입력하면 기존 매수일을 지운다.
    """
    while True:
        raw = Prompt.ask(prompt, default=default)
        raw = (raw or "").strip()
        if raw.lower() in ('b', 'q'):
            return None, True
        if raw == '-':
            return None, False
        if not raw:
            if empty_as_today:
                return datetime.now().date(), False
            return None, False

        digits = raw.replace("-", "").replace("/", "").replace(".", "")
        try:
            parsed = datetime.strptime(digits, "%Y%m%d").date()
        except ValueError:
            config.console.print("[red]YYYY-MM-DD 형식으로 입력하세요. (예: 2026-07-01)[/red]")
            continue

        if parsed > datetime.now().date():
            config.console.print("[red]미래 날짜는 입력할 수 없습니다.[/red]")
            continue
        return parsed, False

MANUAL_POSITIONS_FILE = os.path.join(config.JSON_DIR, "manual_positions.json")

def load_manual_positions():
    """저장된 수동 보유 포지션을 불러온다. (매수일은 date로 복원)"""
    rows = jsonio.load_json(MANUAL_POSITIONS_FILE, default=[]) or []
    positions = []
    for row in rows:
        try:
            buy_date = None
            if row.get('buy_date'):
                buy_date = datetime.strptime(row['buy_date'], "%Y-%m-%d").date()
            positions.append({
                'code': str(row['code']), 'name': row.get('name') or str(row['code']),
                'is_overseas': bool(row.get('is_overseas')),
                'buy_price': float(row['buy_price']), 'qty': int(row['qty']),
                'buy_date': buy_date,
            })
        except (KeyError, TypeError, ValueError) as e:
            logger.debug(f"수동 포지션 항목 파싱 실패({row}): {e}")
    return positions

def save_manual_positions(positions):
    """수동 보유 포지션을 저장한다. (현재가/수익률 등 조회값은 저장하지 않음)"""
    rows = [{
        'code': p['code'], 'name': p['name'], 'is_overseas': bool(p['is_overseas']),
        'buy_price': p['buy_price'], 'qty': p['qty'],
        'buy_date': p['buy_date'].strftime("%Y-%m-%d") if p.get('buy_date') else None,
    } for p in positions]
    return jsonio.save_json(MANUAL_POSITIONS_FILE, rows)

def _qty_unit(code):
    """보유수량 단위 — KRX 금현물처럼 주식이 아닌 상품은 'g'로 센다."""
    from modules import market
    return "g" if market.resolve_index_product(code) else "주"


def _collect_manual_position():
    """포지션 분석용 포지션 정보를 입력받는다. 취소 시 None.

    직접 입력에서는 종목 코드·티커뿐 아니라 지수 목록 상품(KRX 금현물 = KRXGOLD)도 받는다.
    """
    from modules import auto_trade

    code, name, is_overseas = auto_trade._select_stock_for_rules(allow_index_products=True)
    if not code:
        return None

    unit = "$" if is_overseas else "원"
    qty_unit = _qty_unit(code)
    config.console.print()
    config.console.print(f"[bold cyan]{name}({code})[/bold cyan] 보유 정보를 입력하세요. [dim](취소: b)[/dim]")

    buy_price = _ask_positive_number(f"  매수단가 ({unit})", cast=float)
    if buy_price is None:
        return None

    qty = _ask_positive_number(f"  보유수량 ({qty_unit})", cast=int)
    if qty is None:
        return None

    buy_date, cancelled = _ask_buy_date("  매수일 (YYYY-MM-DD, 미입력 시 오늘날짜 적용)", empty_as_today=True)
    if cancelled:
        return None

    return {
        'code': code, 'name': name, 'is_overseas': is_overseas,
        'buy_price': buy_price, 'qty': qty, 'buy_date': buy_date,
    }

def _print_saved_positions(positions):
    """저장된 수동 포지션 목록을 요약 출력한다. (분석 전 확인용)"""
    table = Table(title="\n저장된 수동 보유 포지션", box=box.HORIZONTALS,
                  show_header=True, header_style="dim", border_style="dim")
    table.add_column("#", justify="right", style="dim")
    table.add_column("종목명", justify="left")
    table.add_column("코드", justify="center", style="dim")
    table.add_column("구분", justify="center", style="dim")
    table.add_column("매수단가", justify="right")
    table.add_column("보유수량", justify="right")
    table.add_column("매수일", justify="center")

    for i, p in enumerate(positions, 1):
        unit = "$" if p['is_overseas'] else "원"
        price = f"{p['buy_price']:,.2f}{unit}" if p['is_overseas'] else f"{p['buy_price']:,.0f}{unit}"
        
        qty_unit = _qty_unit(p['code'])
        # '주'(전각, 2칸)와 'g'(반각, 1칸)의 우측 정렬 시 숫자 위치를 맞추기 위해 뒤에 공백(1칸) 추가
        # rich 라이브러리가 우측 공백을 지우는 것을 방지하기 위해 Zero-width space(\u200b) 사용
        qty_str = f"{p['qty']:,}{qty_unit} \u200b" if qty_unit == "g" else f"{p['qty']:,}{qty_unit}"
        
        table.add_row(
            str(i), p['name'], p['code'], "해외" if p['is_overseas'] else "국내",
            price, qty_str,
            p['buy_date'].strftime("%Y-%m-%d") if p.get('buy_date') else "[dim]-[/dim]",
        )

    config.console.print(table)

def _fmt_position_summary(pos):
    """포지션 한 줄 요약 (수정/삭제 대상 확인용)."""
    price = f"${pos['buy_price']:,.2f}" if pos['is_overseas'] else f"{pos['buy_price']:,.0f}원"
    date_str = pos['buy_date'].strftime("%Y-%m-%d") if pos.get('buy_date') else "매수일 없음"
    return (f"[bold]{pos['name']}[/bold]([dim]{pos['code']}[/dim]) {price} · "
            f"{pos['qty']:,}{_qty_unit(pos['code'])} · {date_str}")

def _edit_position(pos):
    """저장된 포지션의 매수 정보를 수정한다. 취소하면 None (원본 유지).

    종목 자체는 바꾸지 않는다 — 다른 종목은 신규 입력으로 추가한다.
    """
    unit = "$" if pos['is_overseas'] else "원"
    cur_price = f"{pos['buy_price']:.2f}" if pos['is_overseas'] else f"{pos['buy_price']:.0f}"
    cur_date = pos['buy_date'].strftime("%Y-%m-%d") if pos.get('buy_date') else ""

    config.console.print()
    config.console.print(f"[bold cyan]{pos['name']}({pos['code']})[/bold cyan] 수정 "
                         f"[dim](Enter=현재값 유지, 취소: b)[/dim]")

    buy_price = _ask_positive_number(f"  매수단가 ({unit})", cast=float, default=cur_price)
    if buy_price is None:
        return None

    qty = _ask_positive_number(f"  보유수량 ({_qty_unit(pos['code'])})", cast=int, default=str(pos['qty']))
    if qty is None:
        return None

    buy_date, cancelled = _ask_buy_date(
        "  매수일 (YYYY-MM-DD, 지우려면 '-')", default=cur_date)
    if cancelled:
        return None

    updated = dict(pos)
    updated.update({'buy_price': buy_price, 'qty': qty, 'buy_date': buy_date})
    return updated

def _select_position(positions, action_label):
    """목록에서 번호로 포지션을 고른다. 취소하거나 잘못된 번호면 None."""
    raw = Prompt.ask(f"{action_label}할 번호 (취소: Enter)", default="")
    raw = (raw or "").strip()
    if not raw or raw.lower() in ('b', 'q'):
        return None

    if not raw.isdigit() or not (1 <= int(raw) <= len(positions)):
        config.console.print("[red]목록에 있는 번호를 입력하세요.[/red]")
        return None

    return int(raw) - 1

def _modify_saved_position(positions):
    """저장 목록의 한 항목을 수정한다. 반환: (포지션 리스트, 변경 여부)."""
    idx = _select_position(positions, "수정")
    if idx is None:
        return positions, False

    updated = _edit_position(positions[idx])
    if not updated:
        return positions, False

    positions[idx] = updated
    config.console.print(f"[green]✓ 수정됨: {_fmt_position_summary(updated)}[/green]")
    return positions, True

def _delete_saved_position(positions):
    """저장 목록의 한 항목을 삭제한다. 반환: (포지션 리스트, 변경 여부)."""
    idx = _select_position(positions, "삭제")
    if idx is None:
        return positions, False

    target = positions[idx]
    config.console.print(f"  → {_fmt_position_summary(target)}")
    if Prompt.ask("  삭제하시겠습니까?", choices=["y", "n"], default="n") != "y":
        return positions, False

    positions.pop(idx)
    config.console.print(f"[yellow]✗ 삭제됨: {_fmt_position_summary(target)}[/yellow]")
    return positions, True

def _add_manual_positions(positions, base_breadcrumb_len):
    """신규 종목을 연속 입력받아 목록에 추가한다. 반환: (목록, 추가 여부)."""
    added = False
    while True:
        context.USER_ACTION_BREADCRUMB = context.USER_ACTION_BREADCRUMB[:base_breadcrumb_len]
        pos = _collect_manual_position()
        if pos:
            positions.append(pos)
            added = True
            config.console.print(f"[green]✓ 추가됨: {_fmt_position_summary(pos)}[/green]")

        config.console.print()
        if Prompt.ask("종목을 더 추가하시겠습니까?", choices=["y", "n"], default="n") != "y":
            break

    return positions, added

def manual_holding_analysis():
    """[9]-5 포지션 분석 — 계좌 잔고에 없는 포지션을 직접 입력해 [9]-2와 같은 판정을 본다.

    타 계좌 보유분, 매수 검토 중인 시나리오, 시스템 도입 전 매수분 등 잔고 API로는
    잡히지 않는 포지션을 같은 기준(analyze_sell)으로 확인하기 위한 메뉴다.
    입력한 포지션은 저장해 두고 다음 실행 때 바로 재분석할 수 있다.
    """
    base_breadcrumb_len = len(context.USER_ACTION_BREADCRUMB)

    positions = load_manual_positions()
    dirty = False

    if positions:
        _print_saved_positions(positions)
    else:
        # 저장분이 없어도 곧바로 입력으로 밀어넣지 않는다. 메뉴를 잘못 눌렀을 때
        # 빠져나갈 곳 없이 종목 입력이 시작되던 문제.
        config.console.print("\n[dim]저장된 수동 보유 포지션이 없습니다. [1] 추가로 종목을 입력하세요.[/dim]")

    while True:
        config.console.print()
        choice = Prompt.ask("작업 선택 (0: 분석, 1: 추가, 2: 수정, 3: 삭제) [dim](이전: b, 메인: q)[/dim]",
                            choices=["0", "1", "2", "3", "b", "q"], default="0")
        if choice.lower() in ('b', 'q'):
            # 분석 없이 나가더라도 편집분은 잃지 않도록 저장 여부는 물어본다.
            if dirty:
                _prompt_save_positions(positions)
            return False

        if choice == "0":
            if positions:
                break
            config.console.print("[yellow]분석할 종목이 없습니다. [1] 추가로 종목을 먼저 입력하세요.[/yellow]")
            continue

        if choice == "1":
            positions, changed = _add_manual_positions(positions, base_breadcrumb_len)
        elif not positions:
            config.console.print("[yellow]수정·삭제할 종목이 없습니다.[/yellow]")
            continue
        elif choice == "2":
            positions, changed = _modify_saved_position(positions)
        else:
            positions, changed = _delete_saved_position(positions)
        dirty = dirty or changed

        if positions:
            _print_saved_positions(positions)

    # 현재가 조회 → 평가금액/수익률 산출 (사용자 입력은 매수 정보뿐)
    with Progress(
        SpinnerColumn(),
        TextColumn("[progress.description]{task.description}"),
        BarColumn(),
        console=config.console,
        transient=True
    ) as progress:
        task = progress.add_task("[cyan]현재가 조회 중...[/cyan]", total=None)

        priced = []
        for pos in positions:
            cur_price = api.get_current_price(pos['code'], pos['is_overseas'])
            if not cur_price or cur_price <= 0:
                config.console.print(f"[red]현재가 조회 실패: {pos['name']}({pos['code']}) — 제외합니다.[/red]")
                continue
            pos['current_price'] = float(cur_price)
            pos['profit_rate'] = ((pos['current_price'] - pos['buy_price']) / pos['buy_price']) * 100
            priced.append(pos)

        if not priced:
            config.console.print("[red]현재가를 조회할 수 있는 종목이 없습니다.[/red]")
            if dirty:
                save_manual_positions(positions)
            return True

        progress.update(task, description="[cyan]보유 종목 분석 중 (시스템 매도 기준)...[/cyan]")
        holding_analysis = _analyze_manual_positions(priced)

    _print_manual_positions(priced, holding_analysis)

    # 추가·수정·삭제가 있었을 때만 저장 여부를 묻는다. (변경 없이 재분석만 한 경우는 묻지 않음)
    if dirty:
        _prompt_save_positions(positions)

    return True

def _prompt_save_positions(positions):
    """변경된 수동 포지션의 저장 여부를 묻고 저장한다."""
    config.console.print()
    if Prompt.ask("변경한 보유 정보를 저장하시겠습니까? [dim](다음 실행 시 자동 표시)[/dim]",
                  choices=["y", "n"], default="y") != "y":
        return

    if save_manual_positions(positions):
        config.console.print(f"[green]✓ 저장 완료 ({len(positions)}종목)[/green]")
    else:
        config.console.print("[red]저장에 실패했습니다. 로그를 확인하세요.[/red]")

def _analyze_manual_positions(positions):
    """수동 입력 포지션을 [9]-2와 같은 기준으로 분석한다. (읽기 전용)

    잔고 경로와 달리 보유일수는 입력한 매수일에서, 트레일링 스탑 앵커(최고가)는
    매수일 이후 실제 일봉 고가에서 유도한다. (DB에 매수·최고가 기록이 없는 포지션)
    """
    from modules import auto_trade

    restricted_codes = {}
    try:
        restricted_codes = auto_trade.get_restricted_stocks()
    except Exception as e:
        logger.debug(f"제한 종목 조회 실패: {e}")

    entries = []
    for pos in positions:
        entry = {
            'code': pos['code'], 'name': pos['name'],
            'buy_price': pos['buy_price'], 'current_price': pos['current_price'],
            'profit_rate': pos['profit_rate'], 'is_overseas': pos['is_overseas'],
        }
        if pos.get('buy_date'):
            entry['holding_days'] = (datetime.now().date() - pos['buy_date']).days
            entry['highest_since'] = pos['buy_date']
        entries.append(entry)

    try:
        return auto_trade.analyze_holdings(entries, restricted_codes=restricted_codes)
    except Exception as e:
        logger.warning(f"포지션 분석 실패: {e}")
        return {}

def _print_manual_positions(positions, holding_analysis):
    """수동 입력 포지션을 [9]-2 잔고와 같은 표로 출력한다."""
    domestic_items = []
    overseas_items = []

    for pos in positions:
        qty = pos['qty']
        buy_price = pos['buy_price']
        cur_price = pos['current_price']
        eval_amt = qty * cur_price
        profit = eval_amt - (qty * buy_price)

        if pos['is_overseas']:
            overseas_items.append({
                'ovrs_pdno': pos['code'], 'ovrs_item_name': pos['name'],
                'ovrs_cblc_qty': qty, 'pchs_avg_pric': buy_price,
                'ovrs_now_pric': cur_price, 'frcr_evlu_pfls_amt': profit,
                'evlu_pfls_rt': pos['profit_rate'], '_exchange': '-',
            })
        else:
            domestic_items.append({
                'pdno': pos['code'], 'prdt_name': pos['name'],
                'hldg_qty': qty, 'pchs_avg_pric': buy_price,
                'prpr': int(cur_price), 'evlu_amt': int(eval_amt),
                'evlu_pfls_amt': int(profit), 'evlu_pfls_rt': pos['profit_rate'],
            })

    if domestic_items:
        table, totals, signals = build_domestic_holdings_table(
            domestic_items, holding_analysis, title="\n[국내] 포지션 분석", show_auto_status=False)
        config.console.print(table)

        total_rate = (totals['profit'] / totals['pchs'] * 100) if totals['pchs'] > 0 else 0.0
        p_color = "[red]" if totals['profit'] > 0 else ("[blue]" if totals['profit'] < 0 else "[white]")
        config.console.print(f"[bold dim]  국내 총 매입금액:[/bold dim] {totals['pchs']:,}원  |  [bold dim]총 평가금액:[/bold dim] {totals['eval']:,}원  |  [bold dim]총 평가손익:[/bold dim] {p_color}{totals['profit']:+,}원 ({total_rate:+.2f}%)[/]")
        _print_ts_legend()
        _print_sell_signals(signals)

    if overseas_items:
        if domestic_items:
            config.console.print()
        table, totals, signals = build_overseas_holdings_table(
            overseas_items, holding_analysis, title="\n[해외] 포지션 분석", show_auto_status=False)
        config.console.print(table)

        total_rate = (totals['profit'] / totals['pchs'] * 100) if totals['pchs'] > 0 else 0.0
        p_color = "[red]" if totals['profit'] > 0 else ("[blue]" if totals['profit'] < 0 else "[white]")
        config.console.print(f"[bold dim]  해외 총 매입금액:[/bold dim] ${totals['pchs']:,.2f}  |  [bold dim]총 평가금액:[/bold dim] ${totals['eval']:,.2f}  |  [bold dim]총 평가손익:[/bold dim] {p_color}${totals['profit']:+,.2f} ({total_rate:+.2f}%)[/]")
        _print_ts_legend()
        _print_sell_signals(signals)

def _display_balance_details(cano, acnt_prdt_cd):
    """특정 계좌의 잔고 상세 출력"""

    reserved_codes = set()
    try:
        pending_reserves = db_manager.db.get_pending_reserved_orders()
        reserved_codes = set(o['code'] for o in pending_reserves if o.get('cano') == cano and o.get('acnt') == acnt_prdt_cd)
    except Exception as e:
        logger.debug(f"reserved_codes fetch error: {e}")
        
    # [추가] 제한 종목 및 개별 룰 로드
    from modules import auto_trade
    restricted_stocks = auto_trade.get_restricted_stocks(cano, acnt_prdt_cd)
    custom_rules = db_manager.db.get_all_stock_strategies()
    rules_map = {r['code']: True for r in custom_rules}

    marks_ctx = {
        'restricted': restricted_stocks,
        'rules': rules_map,
        'memo': utils.get_memo_codes(),
        'reserved': reserved_codes,
    }

    # ---------------------------
    # [잔고 조회 + 보유 분석]
    # 국내·해외를 먼저 모두 조회한 뒤 보유 분석을 1회만 수행한다.
    # (DB 배치 로드와 시장 국면 조회가 종목 수와 무관하게 1회로 끝남)
    # ---------------------------
    with Progress(
        SpinnerColumn(),
        TextColumn("[progress.description]{task.description}"),
        BarColumn(),
        console=config.console,
        transient=True
    ) as progress:
        task = progress.add_task("[cyan]국내 잔고 조회 중...[/cyan]", total=None)
        # [수정] api.get_domestic_balance 직접 호출
        #  [계좌 라우팅] 앱키·토큰·TPS 버킷은 thread_local 이 고른다 — cano 인자로는
        #   바뀌지 않는다([[account-context-in-threads]]).
        with utils.AccountContext(cano):
            raw_holdings, raw_summary = api.get_domestic_balance(cano, acnt_prdt_cd)

        if raw_holdings is None:
            config.console.print("[red]잔고 조회 실패 (API 오류)[/red]")
            return

        # 보유수량 0 이상인 종목만 필터링
        output1 = [item for item in raw_holdings if api.safe_int(item.get('hldg_qty')) > 0]
        summary = raw_summary[0] if raw_summary else None

        progress.update(task, description="[cyan]해외 잔고 조회 중...[/cyan]")
        # [수정] api.get_overseas_balance 직접 호출
        with utils.AccountContext(cano):
            all_overseas_holdings = api.get_overseas_balance(cano, acnt_prdt_cd)
        #  None = 조회 실패. 종전의 `or []` 는 실패를 '해외 보유 없음'으로 접어, 화면에서
        #  해외분이 조용히 사라졌다(합계도 그만큼 적게 나온다). 사실을 밝히고 비운다.
        overseas_failed = all_overseas_holdings is None
        ovrs_output = [item for item in (all_overseas_holdings or [])
                       if (api.safe_float(item.get('ovrs_cblc_qty'), default=0.0)
                        or api.safe_float(item.get('ord_psbl_qty'), default=0.0)) > 0]
        if overseas_failed:
            config.console.print("[yellow]해외 잔고를 조회하지 못했습니다 — 아래 표에 해외분이 빠져 있습니다.[/yellow]")

        # [추가] 보유 분석 — 자동매매가 실제로 쓰는 매도 판단(analyze_sell)을 그대로 적용
        progress.update(task, description="[cyan]보유 종목 분석 중 (시스템 매도 기준)...[/cyan]")
        holding_analysis = run_holding_analysis(output1, ovrs_output, restricted_codes=restricted_stocks,
                                                account=f"{cano}-{acnt_prdt_cd}")

    # ---------------------------
    # [국내 주식 잔고]
    # ---------------------------
    if output1:
        table, totals, sell_signals = build_domestic_holdings_table(
            output1, holding_analysis, marks_ctx=marks_ctx)
        config.console.print(table)

        # 요약 정보 출력
        if summary:
            total_rate = 0.0
            if totals['pchs'] > 0:
                total_rate = (totals['profit'] / totals['pchs']) * 100

            profit_color = "[red]" if totals['profit'] > 0 else ("[blue]" if totals['profit'] < 0 else "[white]")
            config.console.print(f"[bold dim]  국내 총 매입금액:[/bold dim] {totals['pchs']:,}원  |  [bold dim]총 평가금액:[/bold dim] {totals['eval']:,}원  |  [bold dim]총 평가손익:[/bold dim] {profit_color}{totals['profit']:+,}원 ({total_rate:+.2f}%)[/]")
            _print_ts_legend()

        _print_sell_signals(sell_signals)
    else:
        config.console.print("\n[yellow]국내 보유 종목이 없습니다.[/yellow]")

    config.console.print()

    # ---------------------------
    # [해외 주식 잔고]
    # ---------------------------
    if not all_overseas_holdings:
        config.console.print("\n[yellow]해외 보유 종목이 없습니다.[/yellow]\n")
    else:
        table_ovrs, totals_ovrs, ovrs_sell_signals = build_overseas_holdings_table(
            all_overseas_holdings, holding_analysis, marks_ctx=marks_ctx)

        if totals_ovrs['count'] > 0:
            config.console.print(table_ovrs)
            total_ovrs_rate = 0.0
            if totals_ovrs['pchs'] > 0:
                total_ovrs_rate = (totals_ovrs['profit'] / totals_ovrs['pchs']) * 100

            profit_color = "[red]" if totals_ovrs['profit'] > 0 else ("[blue]" if totals_ovrs['profit'] < 0 else "[white]")
            config.console.print(f"[bold dim]  해외 총 매입금액:[/bold dim] ${totals_ovrs['pchs']:,.2f}  |  [bold dim]총 평가금액:[/bold dim] ${totals_ovrs['eval']:,.2f}  |  [bold dim]총 평가손익:[/bold dim] {profit_color}${totals_ovrs['profit']:+,.2f} ({total_ovrs_rate:+.2f}%)[/]")
            _print_ts_legend()

            _print_sell_signals(ovrs_sell_signals)

        else:
            config.console.print("\n[yellow]해외 보유 종목이 없습니다 (수량 0).[/yellow]")

def _display_account_targets():
    """화면에 나눠 찍을 계좌 목록 — [(계좌번호, 상품코드, 표시명)].

    [왜 모드로 가르나] 계좌가 둘로 갈리는 것은 실전(mode 2)에서 시스템 트레이딩 계좌를
    따로 잡았을 때뿐이다. 토스는 주식계좌가 하나뿐이고(session이 auto_cano를 cano로
    동기화한다), 가상투자도 가상 계좌 하나로 돈다. 종전에는 '한투증권 (수동)' 줄을
    모드와 무관하게 무조건 넣어, 토스 모드에서 같은 계좌가 '토스증권'과
    '한투증권 (수동)'으로 두 번 찍혔다(실측 2026-08-26).
    """
    s = config.session
    if s.is_toss:
        return [(s.cano, s.acnt_prdt_cd, "토스증권")]
    if getattr(s, 'is_paper', False):
        return [(s.cano, s.acnt_prdt_cd, "가상투자")]

    # 한투 실전 — 자동매매 계좌가 따로 있을 때만 수동/자동으로 가른다.
    has_auto = bool(s.auto_cano and s.auto_acnt_prdt_cd and
                    (s.auto_cano != s.cano or s.auto_acnt_prdt_cd != s.acnt_prdt_cd))
    if not has_auto:
        return [(s.cano, s.acnt_prdt_cd, "한투증권")]
    return [(s.cano, s.acnt_prdt_cd, "한투증권 (수동)"),
            (s.auto_cano, s.auto_acnt_prdt_cd, "한투증권 (자동)")]


def get_account_balance():
    """보유 잔고 조회 (메인/자동 계좌 순차 조회)"""
    time.sleep(0.5)
    
    accounts = _display_account_targets()

    for i, (cano, acnt, label) in enumerate(accounts):
        if i > 0: config.console.print("\n")
        config.console.print(f"\n[bold cyan]{label} 계좌 잔고 ({cano}{'-' + acnt if acnt else ''})[/]")
        _display_balance_details(cano, acnt)

def get_asset_status_data(cano, acnt_prdt_cd, progress=None, task=None):
    """자산 현황 데이터 조회 및 계산 (UI 로직 없음)"""
    summary_data = {
        "withdraw": 0,      "tot_asset": 0,
        "dep_dom": 0,       "dep_ovs": 0,
        "d1_dep": 0,        "d2_dep": 0,
        "sec_buy": 0,       "sec_eval": 0,      "sec_pl": 0,
        "realized_pl": 0,   "total_cost": 0,
        "buy_today": 0,     "sell_today": 0,
        "ovrs_eval_krw": 0, "ovrs_pl_krw": 0,
        "order_possible": 0, # [추가] 주문가능금액
        "d2_real": 0,        # [추가] 실제 D+2 예수금
        "next_day_plus": 0,  # [추가] 익일결재(+)
        "next_day_minus": 0, # [추가] 익일결재(-)
        "api_tot_asset": 0,  # [추가] API 제공 총 평가금액 (검증용)
        #  [추가 2026-09-06] **이 값이 온전한가.** 아래 네 구간은 각자 예외를 삼키고
        #   넘어가므로, 한 구간이 실패해도 tot_asset 은 **숫자로** 나온다 — 그저 그만큼
        #   작을 뿐이다. 실측(주식비중 36% 계좌, 국내 잔고 조회만 실패):
        #       정상          : 총자산 10,000,000 (주식 3,600,000 + 현금 6,400,000)
        #       국내잔고 실패 : 총자산  6,400,000 (주식 0 + 현금 6,400,000)
        #   호출부는 이것을 '자산이 36% 줄었다'로 읽는다. 기존 방어 둘(차단기의 '비정상
        #   급감' 문턱, is_plausible_baseline)은 **직전 대비 0.5배**를 본다 — 이 시스템의
        #   노출 상한이 40%라(4슬롯·균등배분) 주식 평가액이 통째로 빠져도 그 문턱에
        #   영영 닿지 않는다. 문턱을 낮추는 것은 답이 아니다. **못 읽었다는 사실 자체**를
        #   전한다 — 기준선처럼 되돌릴 수 없는 결정은 그때 판단을 미룬다.
        "degraded": []       # 값을 채우지 못한 구간 이름들. 비어 있어야 온전한 값이다.
    }
    
    # 1. 금일 데이터 조회
    if progress: progress.update(task, description="[cyan]금일 매매 손익 조회 중...[/cyan]")
    try:
        # [원복] 항상 현재 날짜 기준 조회 (새벽 로직 제거)
        #  [계좌 라우팅] cano 인자만으로는 앱키·토큰·TPS 버킷이 바뀌지 않는다.
        #   같은 함수의 예수금 조회(3번)에는 이미 걸려 있었는데 1·2번에는 없었다.
        with utils.AccountContext(cano):
            profit_data = fetch_today_profit_summary(cano, acnt_prdt_cd)
            backup_data = fetch_today_history(cano, acnt_prdt_cd)
        summary_data['buy_today'] = profit_data['buy_amt']
        summary_data['sell_today'] = profit_data['sell_amt']
        summary_data['total_cost'] = profit_data['total_cost']
        summary_data['realized_pl'] = profit_data['realized_pl']

        # [수정] 기간별 손익 API가 매매금액을 0으로 반환하는 경우가 많으므로
        # 체결 내역(fetch_today_history)을 조회하여 값이 더 크다면(누락된 경우) 덮어쓰기 수행
        #  (조회는 위 AccountContext 안에서 이미 끝냈다)
        if backup_data['buy_total'] > summary_data['buy_today']:
            summary_data['buy_today'] = backup_data['buy_total']
        if backup_data['sell_total'] > summary_data['sell_today']:
            summary_data['sell_today'] = backup_data['sell_total']
            
        # [추가] 모의투자이거나 실현손익이 0인 경우 DB에서 금일 손익 및 매매금액 합산 (Fallback)
        # 모의투자는 기간별 손익 API를 지원하지 않으므로 DB 활용 필수
        if summary_data['realized_pl'] == 0:
            try:
                today_str = datetime.now().strftime("%Y-%m-%d")
                target_acc = f"{cano}-{acnt_prdt_cd}"
                
                # DB 조회
                db_trades = db_manager.db.get_trades(
                    start_date=today_str, end_date=today_str,
                    is_sim=False, account=target_acc
                )
                
                db_pl = 0
                db_buy = 0
                db_sell = 0
                
                for t in db_trades:
                    # [수정] 중복 합산 방지: '체결 확인' 사유가 있는 확정 내역만 집계
                    # (DB에는 원주문(접수)과 체결 내역이 모두 존재할 수 있어 단순 합산 시 이중 계산됨)
                    if "체결 확인" not in t.get('reason', ''):
                        continue

                    type_str = t.get('type', '').lower()
                    price = api.safe_float(t.get('price'), default=0.0)
                    qty = api.safe_int(t.get('qty'))
                    amt = int(price * qty)
                    
                    if "sell" in type_str or "매도" in type_str:
                        db_pl += int(t.get('profit_amt') or 0)
                        db_sell += amt
                    elif "buy" in type_str or "매수" in type_str:
                        db_buy += amt
                
                if summary_data['realized_pl'] == 0: summary_data['realized_pl'] = db_pl
                if db_buy > summary_data['buy_today']: summary_data['buy_today'] = db_buy
                if db_sell > summary_data['sell_today']: summary_data['sell_today'] = db_sell
                    
            except Exception as e:
                logger.debug(f"DB 금일 데이터 조회 실패: {e}")

    except Exception as e:
        #  실현손익이 0으로 남으면 그 차액이 **가짜 입금**으로 둔갑해 자산 기준선이
        #  밀린다([[daily-asset-baseline-transfers]]). 조용히 넘길 수 없다.
        summary_data['degraded'].append("금일손익")
        logger.error(f"자산 현황 — 금일 매매 손익 조회 실패: {type(e).__name__}: {e}")

    # 2. 국내 주식 잔고 및 자산
    if progress: progress.update(task, description="[cyan]국내 주식 잔고 및 평가금 조회 중...[/cyan]")
    try:
        # api.get_domestic_balance 사용 (내부에서 OPSQ2001 처리)
        with utils.AccountContext(cano):
            output1, output2 = api.get_domestic_balance(cano, acnt_prdt_cd)

        if output1 is None:
            #  None = 조회 실패다(빈 리스트는 '보유 없음'이라 정상). 예외가 안 났을 뿐
            #  주식 평가액을 모르는 것은 같다.
            summary_data['degraded'].append("국내잔고")
            logger.warning("자산 현황 — 국내 잔고를 읽지 못했습니다(주식 평가액이 빠집니다)")
        else:
            # [수정] 보유 중인 종목만 필터링
            holdings = [h for h in output1 if api.safe_int(h.get('hldg_qty')) > 0]
            calc_buy = 0; calc_eval = 0; calc_pl = 0
            degraded = []
            for item in holdings:
                #  [Fix 2026-09-06] 바로 아래 두 줄은 이미 safe_int 를 쓰는데 이 둘만
                #   하드 서브스크립트였다. KIS 는 값이 없을 때 키를 주고 **빈 문자열**을
                #   담으므로 int('')/float('') 가 ValueError 를 낸다 — 그러면 이 블록의
                #   바깥 except 로 튀어 **주식 평가액이 통째로 빠진 총자산**이 만들어진다
                #   (트레이더가 '통합 자산 조회 이상'으로 읽는 바로 그 상태다).
                #   한 종목이 이상해도 나머지 집계는 살린다 — 대신 빠진 사실을 남긴다.
                qty = api.safe_int(item.get('hldg_qty'))
                avg_pric = api.safe_float(item.get('pchs_avg_pric'), default=0.0)
                if qty <= 0 or avg_pric <= 0:
                    degraded.append(str(item.get('pdno') or '?'))
                calc_buy += int(qty * avg_pric) # 매입금액 직접 계산 (API 누락 방지)
                calc_eval += api.safe_int(item.get('evlu_amt'))
                calc_pl += api.safe_int(item.get('evlu_pfls_amt'))
            if degraded:
                logger.warning(f"자산 집계 — 수량·평단을 읽을 수 없는 종목 {len(degraded)}건이 "
                               f"매입금액에서 빠졌습니다: {', '.join(degraded[:5])}")
            summary_data['sec_buy'] = calc_buy
            summary_data['sec_eval'] = calc_eval
            summary_data['sec_pl'] = calc_pl

            if output2:
                summary = output2[0]
                # [수정] 실전/모의 공통으로 D+1, D+2, 예수금 데이터 파싱
                summary_data['api_tot_asset'] = api.safe_int(summary.get('tot_evlu_amt')) # API 제공 총평가금
                summary_data['d1_dep'] = api.safe_int(summary.get('nxdy_excc_amt'))
                summary_data['d2_dep'] = api.safe_int(summary.get('prvs_rcdl_excc_amt'))
                summary_data['dep_dom'] = api.safe_int(summary.get('dnca_tot_amt'))

                if config.FILE_DEBUG_LEVEL == "DEBUG":
                    logger.debug(f"[ACCOUNT_DEBUG] Balance Summary (Output2): {summary}")
                
                # [추가] 익일결재 금액 계산 (전일 매도/매수 기준)
                bfdy_sll = api.safe_int(summary.get('bfdy_sll_amt'))
                bfdy_buy = api.safe_int(summary.get('bfdy_buy_amt'))
                bfdy_tlex = api.safe_int(summary.get('bfdy_tlex_amt'))
                summary_data['next_day_plus'] = bfdy_sll - bfdy_tlex
                summary_data['next_day_minus'] = bfdy_buy
                
                # [추가] 금일 제비용 보정 (기간별 손익 API 누락 시 잔고 요약 데이터 활용)
                tlex_amt = api.safe_int(summary.get('thdt_tlex_amt'))
                if tlex_amt > summary_data['total_cost']:
                    summary_data['total_cost'] = tlex_amt
                    
                summary_data['withdraw'] = summary_data['d2_dep'] 

    except Exception as e:
        #  이 구간이 실패하면 **주식 평가액이 통째로 빠진** 총자산이 만들어진다.
        summary_data['degraded'].append("국내잔고")
        logger.error(f"자산 현황 조회 오류: {str(e)}")
        
    # [추가] 해외 주식 잔고 합산 (원화 환산)
    if progress: progress.update(task, description="[cyan]해외 주식 잔고 및 환산액 계산 중...[/cyan]")
    try:
        with utils.AccountContext(cano):
            ovrs_holdings = fetch_overseas_balance(cano, acnt_prdt_cd)
        #  None = 조회 실패. 그대로 순회하면 TypeError 가 바깥 except 로 튀어 **국내분
        #  집계까지 함께 사라진다**. 해외만 비우고 실패 사실을 남긴다.
        if ovrs_holdings is None:
            logger.warning("자산 현황 — 해외 잔고 조회 실패(해외분 제외하고 계산합니다)")
            summary_data['degraded'].append("해외잔고")
            ovrs_holdings = []
        ovrs_buy_usd = 0.0
        ovrs_eval_usd = 0.0
        ovrs_pl_usd = 0.0
        
        for item in ovrs_holdings:
            #  맨 float() 였다 — 증권사는 값이 없을 때 빈 문자열을 준다. `A or B` 안에
            #  있어 형태만 다를 뿐 같은 사고다(빈 문자열은 거짓이라 B 로 넘어가지만,
            #  B 도 빈 문자열이면 float('') 가 터져 **해외분 전체**가 사라진다).
            qty = api.safe_float(item.get('ovrs_cblc_qty'), default=0.0) \
                or api.safe_float(item.get('ord_psbl_qty'), default=0.0)
            if qty > 0:
                pchs = api.safe_float(item.get('pchs_avg_pric'), default=0.0)
                profit = api.safe_float(item.get('frcr_evlu_pfls_amt'), default=0.0)
                buy_amt = qty * pchs
                eval_amt = buy_amt + profit
                
                ovrs_buy_usd += buy_amt
                ovrs_eval_usd += eval_amt
                ovrs_pl_usd += profit
        
        exchange_rate = utils.get_exchange_rate()

        ovrs_eval_krw = int(ovrs_eval_usd * exchange_rate)
        summary_data['ovrs_eval_krw'] = ovrs_eval_krw
        ovrs_pl_krw = int(ovrs_pl_usd * exchange_rate)
        summary_data['ovrs_pl_krw'] = ovrs_pl_krw
        
        if config.FILE_DEBUG_LEVEL == "DEBUG":
            logger.debug(f"CALC (Ovrs->KRW) | USD: Buy={ovrs_buy_usd:.2f}, Eval={ovrs_eval_usd:.2f}, PL={ovrs_pl_usd:.2f} | Rate: {exchange_rate} | KRW: Eval={ovrs_eval_krw}, PL={ovrs_pl_krw}")
        
        summary_data['sec_buy'] += int(ovrs_buy_usd * exchange_rate)
        summary_data['sec_eval'] += ovrs_eval_krw
        summary_data['sec_pl'] += ovrs_pl_krw
    except Exception as e:
        #  `except Exception: pass` 였다. 바로 위 None 처리는 실패를 남기는데, 예외
        #  경로만 조용했다 — 해외 보유가 있는 계좌에서는 그만큼 총자산이 작아진다.
        summary_data['degraded'].append("해외잔고")
        logger.error(f"자산 현황 — 해외 잔고 합산 실패: {type(e).__name__}: {e}")

    # 3. 예수금 조회
    if progress: progress.update(task, description="[cyan]예수금 조회 및 최종 집계 중...[/cyan]")
    try:
        with utils.AccountContext(cano):
            dep_data = api.get_deposit_balance(cano, acnt_prdt_cd)
            
            if dep_data:
                # [수정] 실전/모의 모두 상세 예수금 정보로 업데이트 (잔고 조회 API보다 정확함)
                summary_data['dep_dom'] = dep_data['deposit']
                summary_data['d2_dep'] = dep_data['d2_deposit']
                summary_data['withdraw'] = dep_data['withdraw']
                summary_data['order_possible'] = dep_data.get('order_possible', 0)
                summary_data['d2_real'] = dep_data.get('d2_real', 0)
                
                # [수정] 실전투자일 경우 UI 표시용 D+2 값을 실제 D+2(가수도) 값으로 덮어쓰기
                if summary_data['d2_real'] > 0:
                    summary_data['d2_dep'] = summary_data['d2_real']
                
                summary_data['dep_ovs'] = dep_data['foreign_deposit']
            
            if config.FILE_DEBUG_LEVEL == "DEBUG":
                logger.debug(f"[ACCOUNT_DEBUG] Deposit Detail: {dep_data}")
            if not dep_data:
                #  예수금을 못 읽으면 총자산에서 **현금이 통째로** 빠진다(구간 2의
                #  output2 값이 있으면 그것으로 버티지만, 없으면 0원이다).
                summary_data['degraded'].append("예수금")
                logger.warning("자산 현황 — 예수금을 읽지 못했습니다")
    except Exception as e:
        summary_data['degraded'].append("예수금")
        logger.error(f"자산 현황 — 예수금 조회 실패: {type(e).__name__}: {e}")
    
    # 4. 최종 계산
    # API 지연(Lag)에 의한 총 자산 금액의 왜곡을 방지하기 위해
    # API가 제공하는 tot_evlu_amt 대신 개별 종목 합산 기반으로 직접 계산하여 일관성 유지
    # [Fix] sec_eval에는 이미 해외 평가금(ovrs_eval_krw)이 합산되어 있으므로(위 2번 단계)
    #       여기서 ovrs_eval_krw를 다시 더하면 해외 평가금이 이중 계산되어 총자산이 부풀려진다.
    real_cash = summary_data['d2_dep']
    summary_data['tot_asset'] = real_cash + summary_data['dep_ovs'] + summary_data['sec_eval']

    if config.FILE_DEBUG_LEVEL == "DEBUG":
        logger.debug(f"[ACCOUNT_DEBUG] Calculated Total Asset: {summary_data['tot_asset']:,} (D2 + Ovs예수금 + Sec(국내+해외))")
    
    return summary_data

def _display_asset_status(cano, acnt_prdt_cd):
    """특정 계좌의 자산 현황 출력 (UI)"""
    
    summary_data = None
    with Progress(
        SpinnerColumn(),
        TextColumn("[progress.description]{task.description}"),
        BarColumn(),
        console=config.console,
        transient=True
    ) as progress:
        task = progress.add_task("[cyan]자산 현황 조회 시작...[/cyan]", total=None)
        summary_data = get_asset_status_data(cano, acnt_prdt_cd, progress, task)

    display_tot_deposit = summary_data['dep_dom'] + summary_data['dep_ovs']
    roi = 0.0
    if summary_data['sec_buy'] > 0: roi = (summary_data['sec_pl'] / summary_data['sec_buy']) * 100
    
    def get_color(val): return "red" if val > 0 else ("blue" if val < 0 else "white")

    summary_table = Table(box=box.HORIZONTALS, show_header=False, padding=(0, 2), expand=True, border_style="dim", show_edge=False)
    summary_table.add_column("Item", justify="left", style="white", ratio=6) 
    summary_table.add_column("Value", justify="right", style="white", ratio=4)

    summary_table.add_row("총 평가금액", f"{summary_data['tot_asset']:,}원")
    summary_table.add_row("총 예수금(D+0)", f"{display_tot_deposit:,}원")
    summary_table.add_row("    원화 예수금", f"{summary_data['dep_dom']:,}원", style="dim")
    
    # [수정] 모의투자도 D+1, D+2 정보 표시 (사용자 요청)
    d2_val = summary_data.get('d2_real', 0)
    if d2_val == 0: d2_val = summary_data['d2_dep']
    
    summary_table.add_row("      └ D+1 (익일)", f"{summary_data['d1_dep']:,}원", style="dim")
    summary_table.add_row("      └ D+2 (가수도)", f"{d2_val:,}원", style="dim")
    
    # [추가] 익일결재 정보 표시 (값이 있을 때만)
    if summary_data['next_day_plus'] > 0:
        summary_table.add_row("      └ 익일결재(+)", f"{summary_data['next_day_plus']:,}원", style="dim")
    if summary_data['next_day_minus'] > 0:
        summary_table.add_row("      └ 익일결재(-)", f"{summary_data['next_day_minus']:,}원", style="dim")
    
    summary_table.add_row("    외화예수금", f"{summary_data['dep_ovs']:,}원", style="dim")
    summary_table.add_row("주문가능금액", f"[bold green]{summary_data['order_possible']:,}원[/]")
    summary_table.add_row("출금가능금액", f"{summary_data['withdraw']:,}원")
    summary_table.add_section()
    summary_table.add_row("유가증권매입금액", f"{summary_data['sec_buy']:,}원")
    summary_table.add_row("유가증권평가금액", f"{summary_data['sec_eval']:,}원")
    if summary_data['ovrs_eval_krw'] > 0:
        summary_table.add_row("  └ 해외주식(원화)", f"{summary_data['ovrs_eval_krw']:,}원", style="dim")
    
    pl_str = f"{summary_data['sec_pl']:,}원  ({roi:.2f}%)"
    summary_table.add_row("평가손익금액(보유)", f"[{get_color(summary_data['sec_pl'])}]{pl_str}[/]")
    if summary_data['ovrs_eval_krw'] > 0:
        ovrs_pl_val = summary_data['ovrs_pl_krw']
        summary_table.add_row("  └ 해외손익(원화)", f"[{get_color(ovrs_pl_val)}]{ovrs_pl_val:+,}원[/]", style="dim")

    summary_table.add_section()
    
    # [원복] 라벨 고정
    summary_table.add_row("금일 매수 체결합계", f"{summary_data['buy_today']:,}원")
    summary_table.add_row("금일 매도 체결합계", f"{summary_data['sell_today']:,}원")
    summary_table.add_row("금일 제비용", f"{summary_data['total_cost']:,}원")
    summary_table.add_row("금일 실현 손익 (확정)", f"[{get_color(summary_data['realized_pl'])}]{summary_data['realized_pl']:,}원[/]")

    panel = Panel(
        summary_table,
        title="계좌 자산 현황 요약",
        subtitle=f"[dim]업데이트: {datetime.now().strftime('%H:%M:%S')}[/]",
        subtitle_align="right",
        width=70,
        border_style="green"
    )

    config.console.print()
    config.console.print(panel)
    config.console.print("\n")

def get_deposit_balance():
    """자산 현황 조회 (메인/자동 계좌 순차 조회)"""
    time.sleep(0.5)
    
    accounts = _display_account_targets()

    for cano, acnt, label in accounts:
        config.console.print(f"\n[bold cyan]{label} 자산 현황 ({cano}{'-' + acnt if acnt else ''})[/]")
        _display_asset_status(cano, acnt)
        
def export_trade_history_to_excel():
    """전체 거래 내역을 엑셀 파일로 저장"""
    try:
        trades = db_manager.db.get_trades(is_sim=False, limit=None)
        if not trades:
            config.console.print("\n[yellow]저장할 거래 내역이 없습니다.[/yellow]")
            return

        # DataFrame 생성
        df = pd.DataFrame(trades)
        
        # [추가] 단가 포맷팅 (국내 주식 소수점 제거)
        if 'price' in df.columns and 'code' in df.columns:
            def _format_price(row):
                try:
                    val = float(row['price'])
                    code = str(row['code'])
                    # 국내 주식 (6자리, 숫자로 시작)
                    if len(code) == 6 and code[0].isdigit() and code.isalnum():
                        return int(val)
                    return val
                except Exception: return row['price']
            df['price'] = df.apply(_format_price, axis=1)

        # [추가] 수익률 및 손익금 포맷팅 (+/- 기호 추가)
        if 'profit_rate' in df.columns:
            def _format_rate(val):
                try:
                    if val is None or val == '': return "0.00"
                    f = float(val)
                    return f"{f:+.2f}"
                except Exception: return val
            df['profit_rate'] = df['profit_rate'].apply(_format_rate)

        if 'profit_amt' in df.columns and 'code' in df.columns:
            def _format_amt(row):
                val = row.get('profit_amt')
                code = str(row.get('code', ''))
                try:
                    if val is None or val == '': return "0"
                    f = float(val)
                    # 국내 주식 (6자리, 숫자로 시작)
                    if len(code) == 6 and code[0].isdigit() and code.isalnum():
                        return f"{int(f):+,}"
                    # 해외 주식
                    return f"{f:+,.2f}"
                except Exception: return val
            df['profit_amt'] = df.apply(_format_amt, axis=1)

        if 'snapshot' in df.columns:
            def _process_snapshot(row):
                val = row.get('snapshot')
                score = row.get('strategy_score')
                
                data = {}
                # 점수 정보 병합 (가장 앞에 추가)
                if score is not None and score != '':
                    try:
                        data['score'] = float(score)
                    except Exception: pass

                try:
                    if val:
                        loaded = json.loads(val)
                        if isinstance(loaded, dict):
                            data.update(loaded)
                except Exception: pass

                if not data: return val
            
                def _recursive_round(obj):
                    if isinstance(obj, float): return round(obj, 2)
                    if isinstance(obj, dict): return {k: _recursive_round(v) for k, v in obj.items()}
                    if isinstance(obj, list): return [_recursive_round(v) for v in obj]
                    return obj
                return json.dumps(_recursive_round(data), ensure_ascii=False)
            
            # 행 단위(axis=1)로 처리하여 점수 컬럼 접근
            df['snapshot'] = df.apply(_process_snapshot, axis=1)

        # 컬럼 순서 및 이름 변경 (사용자 친화적)
        columns_map = {
            'time': '일시',
            'account': '계좌번호',
            'is_sim': '종류',
            'odno': '주문번호',
            'org_odno': '원주문',
            'type': '유형',
            'order_status': '상태',
            'name': '종목명',
            'code': '종목코드',
            'qty': '수량',
            'price': '단가',
            'profit_amt': '손익금',
            'profit_rate': '수익률',
            'reason': '매매사유',
            'snapshot': '스냅샷'
        }
        
        # 존재하는 컬럼만 선택하여 순서대로 정렬 (없는 컬럼은 제외)
        target_cols = [c for c in columns_map.keys() if c in df.columns]
        df = df[target_cols]
        df.rename(columns=columns_map, inplace=True)
        
        # 모의투자여부 가독성 좋게 변경
        if '종류' in df.columns:
            if '유형' in df.columns:
                # 실전(0)이면서 유형에 'AUTO'나 '자동'이 포함되면 '자동'으로 표시
                df['종류'] = df.apply(lambda row: '모의' if row['종류'] == 1 else ('자동' if 'AUTO' in str(row['유형']) or '자동' in str(row['유형']) else '실전'), axis=1)
            else:
                df['종류'] = df['종류'].apply(lambda x: '모의' if x == 1 else '실전')

        # 파일명 생성
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename_xlsx = os.path.join(config.DATA_DIR, f"trade_history_{timestamp}.xlsx")
        
        # 엑셀 저장
        try:
            with Progress(
                SpinnerColumn(),
                TextColumn("[progress.description]{task.description}"),
                BarColumn(),
                TextColumn("[progress.percentage]{task.percentage:>3.0f}%"),
                console=config.console,
                transient=True
            ) as progress:
                task = progress.add_task(f"[cyan]'{os.path.basename(filename_xlsx)}' 파일로 저장 중...[/cyan]", total=None)
                
                with pd.ExcelWriter(filename_xlsx, engine='openpyxl') as writer:
                    # [추가] 컬럼 너비 자동 조절 헬퍼 함수
                    def _auto_adjust_width(worksheet):
                        for i, col in enumerate(worksheet.columns):
                            col_idx = i + 1
                            col_letter = get_column_letter(col_idx)
                            
                            # 헤더 길이 계산
                            header_val = worksheet.cell(row=1, column=col_idx).value
                            s_header = str(header_val) if header_val else ""
                            max_width = len(s_header) + sum(0.7 for c in s_header if ord(c) > 127)
                            
                            # 데이터 길이 계산
                            for cell in col[1:]:
                                val = cell.value
                                if val:
                                    s_val = str(val)
                                    length = len(s_val) + sum(0.7 for c in s_val if ord(c) > 127)
                                    if length > max_width: max_width = length
                            
                            # 최대 너비 제한 (스냅샷 등 긴 컬럼 고려)
                            limit = 100 if s_header in ["매매사유", "스냅샷", "비고"] else 60
                            worksheet.column_dimensions[col_letter].width = min(max_width * 1.2, limit)

                    if '계좌번호' in df.columns:
                        # 계좌번호가 없는 데이터 처리
                        df['계좌번호'] = df['계좌번호'].fillna('기타')
                        
                        # 계좌번호별로 시트 분리 저장
                        accounts = df['계좌번호'].unique()
                        progress.update(task, total=len(accounts))
                        
                        for acc in accounts:
                            # 시트 이름 정제 (특수문자 제거 및 길이 제한 31자)
                            sheet_name = str(acc).replace(':', '').replace('\\', '').replace('/', '').replace('?', '').replace('*', '').replace('[', '').replace(']', '')[:31]
                            if not sheet_name: sheet_name = "Unknown"
                            df[df['계좌번호'] == acc].to_excel(writer, sheet_name=sheet_name, index=False)
                            
                            # [추가] 너비 조절 적용
                            _auto_adjust_width(writer.sheets[sheet_name])
                            
                            progress.advance(task)
                    else:
                        progress.update(task, total=1)
                        sheet_name = '전체내역'
                        df.to_excel(writer, sheet_name=sheet_name, index=False)
                        
                        # [추가] 너비 조절 적용
                        _auto_adjust_width(writer.sheets[sheet_name])
                        
                        progress.advance(task)

            config.console.print(f"\n[bold green]성공적으로 저장되었습니다: {os.path.basename(filename_xlsx)}[/bold green]")
            config.console.print("[dim]  - 탭 구분: 계좌번호[/dim]")
        except ImportError:
            config.console.print("\n[yellow]openpyxl 라이브러리가 설치되지 않아 엑셀(.xlsx) 저장이 불가능합니다.[/yellow]")
            config.console.print()
            if Prompt.ask("대신 CSV 파일로 저장하시겠습니까?", choices=["y", "n"], default="y") == "y":
                config.console.print()
                filename_csv = os.path.join(config.DATA_DIR, f"trade_history_{timestamp}.csv")
                df.to_csv(filename_csv, index=False, encoding='utf-8-sig')
                config.console.print(f"\n[bold green]성공적으로 저장되었습니다: {os.path.basename(filename_csv)}[/bold green]")
            else:
                config.console.print("[dim]저장을 취소했습니다. (터미널에서 'pip install openpyxl'을 실행하세요)[/dim]")

    except Exception as e:
        config.console.print(f"\n[bold red]저장 실패: {e}[/bold red]")

def offer_holdings_backfill():
    """매수 기록이 없는 보유 종목을 찾아 복원을 제안한다(거래 내역 조회 진입 시).

    HTS·MTS 직접 매수분, 시스템 도입 이전 포지션, DB 이관 중 잃은 기록은 매수 이력이
    비어 있다. 그러면 실현손익·평단·보유일수가 전부 빈 채로 남는데, 정작 그 사실이
    드러나는 곳이 이 화면이다. 여기서 바로 복원할 수 있게 한다.

    조회만으로 API를 더 쓰지 않도록, 먼저 DB로 '기록 없는 보유 종목'이 있는지부터 본다.
    """
    from modules import holdings_backfill as hb

    # 가상투자·토스는 증권사 체결 이력이 없다. 제안 자체를 띄우지 않는다.
    if not hb.supports_broker_history():
        return

    cano, acnt = config.session.cano, config.session.acnt_prdt_cd
    holdings, _ = fetch_domestic_balance(cano, acnt)
    if holdings is None:
        logger.debug("[백필제안] 국내 잔고 조회 실패 — 제안을 건너뜁니다('보유 없음'이 아닙니다)")
        return
    if not holdings:
        return

    missing = []
    for h in holdings:
        try:
            if int(float(h.get('hldg_qty') or 0)) <= 0:
                continue
        except (TypeError, ValueError):
            continue
        code = str(h.get('pdno') or '').strip()
        try:
            rows = db_manager.db.get_trades(code=code, is_sim=False) or []
        except Exception:
            continue
        if not any('매수' in str(r.get('type') or '') for r in rows):
            missing.append((code, str(h.get('prdt_name') or '').strip()))

    if not missing:
        return

    config.console.print()
    config.console.print(
        f"[yellow]ℹ 매수 기록이 없는 보유 종목이 {len(missing)}개 있습니다[/yellow] "
        f"[dim]— {', '.join(n or c for c, n in missing[:3])}"
        f"{' 외 ' + str(len(missing) - 3) + '종목' if len(missing) > 3 else ''}[/dim]")
    config.console.print(
        "[dim]  증권사 체결 내역에서 복원하면 실현손익·평단·보유일수가 채워집니다.[/dim]")

    utils.print_breadcrumb()
    if Prompt.ask("지금 복원하시겠습니까?", choices=["y", "n"], default="n") != "y":
        return

    plans = hb.plan(holdings, cano=cano, acnt_prdt_cd=acnt)
    if not plans:
        config.console.print("[yellow]증권사 체결 내역에서 복원할 것을 찾지 못했습니다.[/yellow]")
        return

    for p in plans:
        new_cnt = len(p['records']) - p['already']
        if new_cnt <= 0 and p['missing'] <= 0:
            continue
        line = f"  {p['name']}({p['code']}) 보유 {p['qty']}주 → 신규 {new_cnt}건"
        if p['missing'] > 0:
            line += f" [yellow](부분 복원: {p['missing']}주가 조회 구간보다 과거)[/yellow]"
        config.console.print(line)

    written, skipped = hb.apply(plans, cano=cano, acnt_prdt_cd=acnt)
    config.console.print(f"[green]복원 완료: {written}건[/green] [dim]/ 건너뜀 {skipped}건[/dim]")


def view_trade_history():
    """DB에 저장된 거래 내역 조회"""
    logger.debug("[HISTORY_DEBUG] view_trade_history() 진입")
    
    menu_items = [
        ("1", "전체 내역 (최신순 50건)", "All - Latest 50"), ("2", "최근 30일 내역", "Last 30 Days"),
        ("3", "종목코드(티커) 검색", "Search by Ticker"), ("4", "전체 거래 내역 저장", "Save to Excel")
    ]
    choice = utils.show_menu("거래 내역 조회 옵션 (Trade History Options)", menu_items, default_choice="1")
    logger.debug(f"[HISTORY_DEBUG] 사용자 선택: {choice}")
    
    menu_map = {"1": "전체 내역", "2": "최근 30일", "3": "종목 검색", "4": "엑셀 저장"}
    if choice in menu_map:
        context.USER_ACTION_BREADCRUMB.append(f"[{choice}] {menu_map[choice]}")

    if choice.lower() in ['b', 'q']:
        logger.debug("[HISTORY_DEBUG] 사용자 취소(q)로 종료")
        return False

    # [추가] 조회 전 금일 체결 내역 동기화 (시장가 주문 단가 업데이트)
    try:
        logger.debug("[HISTORY_DEBUG] 체결 내역 동기화 시도")
        sync_today_trades()
    except Exception as e:
        config.console.print(f"[dim red]⚠️ 체결 내역 동기화 중 오류 발생: {e}[/dim red]")
        logger.error(f"[HISTORY_DEBUG] sync_today_trades error: {e}")

    # [추가] 매수 기록이 없는 보유 종목을 이 자리에서 알린다.
    #  별도 메뉴 항목으로 두면 '존재를 기억해야 쓰는' 기능이 된다. 거래 내역이 비어 있는
    #  것을 확인하는 화면이 곧 복원을 제안할 자리다.
    try:
        offer_holdings_backfill()
    except Exception as e:
        logger.debug(f"[HISTORY_DEBUG] 보유분 복원 제안 실패: {e}")

    trades = []
    reserved_trades = []
    keyword = ""
    start_dt = ""
    
    if choice == "1":
        logger.info("운영자 실행: " + " - ".join(context.USER_ACTION_BREADCRUMB))
        try:
            logger.debug("[HISTORY_DEBUG] DB 조회 요청 (limit=50)")
            trades = db_manager.db.get_trades(is_sim=False, limit=50)
            logger.debug(f"[HISTORY_DEBUG] DB 조회 완료. 건수: {len(trades)}")
        except Exception as e:
            logger.error(f"[HISTORY_DEBUG] DB 조회 실패: {e}")
            config.console.print(f"[bold red]❌ 거래 내역 조회 실패: {e}[/bold red]")
            return
    elif choice == "2":
        logger.info("운영자 실행: " + " - ".join(context.USER_ACTION_BREADCRUMB))
        start_dt = (datetime.now() - timedelta(days=30)).strftime("%Y-%m-%d")
        try:
            logger.debug(f"[HISTORY_DEBUG] DB 조회 요청 (start_date={start_dt})")
            trades = db_manager.db.get_trades(is_sim=False, start_date=start_dt)
            logger.debug(f"[HISTORY_DEBUG] DB 조회 완료. 건수: {len(trades)}")
        except Exception as e:
            logger.error(f"[HISTORY_DEBUG] DB 조회 실패: {e}")
            config.console.print(f"[bold red]❌ 거래 내역 조회 실패: {e}[/bold red]")
            return
    elif choice == "3":
        config.console.print()
        keyword = Prompt.ask("검색할 종목코드(티커) 입력 [dim](이전: b, 메인: q)[/dim]")
        config.console.print()
        if keyword.lower() in ['b', 'q']: return False
        context.USER_ACTION_BREADCRUMB.append(f"[검색] {keyword}")
        logger.info("운영자 실행: " + " - ".join(context.USER_ACTION_BREADCRUMB))
        try:
            logger.debug(f"[HISTORY_DEBUG] DB 조회 요청 (code={keyword})")
            trades = db_manager.db.get_trades(is_sim=False, code=keyword)
            logger.debug(f"[HISTORY_DEBUG] DB 조회 완료. 건수: {len(trades)}")
        except Exception as e:
            logger.error(f"[HISTORY_DEBUG] DB 조회 실패: {e}")
            config.console.print(f"[bold red]❌ 거래 내역 조회 실패: {e}[/bold red]")
            return
    elif choice == "4":
        logger.info("운영자 실행: " + " - ".join(context.USER_ACTION_BREADCRUMB))
        export_trade_history_to_excel()
        return

    # [추가] 예약 주문 내역 병합
    if choice in ["1", "2", "3"]:
        try:
            res_rows = []
            if choice == "1":
                res_rows = db_manager.db.get_completed_reserved_orders()
            elif choice == "2":
                res_rows = db_manager.db.get_completed_reserved_orders(start_date=start_dt)
            elif choice == "3":
                res_rows = db_manager.db.get_completed_reserved_orders(keyword=keyword)
                
            for r in res_rows:
                cano = r['cano']
                acnt = r['acnt']
                acc_str = f"{cano}-{acnt}"
                
                # 모의/실전 필터링
                is_sim_account = False
                if cano == config.session.cano and acnt == config.session.acnt_prdt_cd:
                    pass 
                elif config.session.auto_cano and cano == config.session.auto_cano and acnt == config.session.auto_acnt_prdt_cd:
                    pass 
                else:
                    continue

                t_type = "매수" if r['order_type'] == 'buy' else "매도"
                reason = f"조건: {r['condition_type']}"
                
                c_type = r['condition_type']
                if c_type == 'TIME': reason += f" ({r['target_time']})"
                elif 'SCORE' in c_type: reason += f" (목표점수: {r['target_price']}점)"
                elif 'RSI' in c_type: reason += f" (목표RSI: {r['target_price']})"
                elif c_type == 'TRAILING_BUY': reason += f" (바닥 반등: {r['target_price']}%)"
                elif c_type == 'TRAILING_SELL': reason += f" (고점 하락: {r['target_price']}%)"
                elif c_type == 'SMART_MONEY': reason += " (수급 턴어라운드)"
                elif c_type.startswith('STATE_'): reason += f" (상태진입: { {'STATE_STRONGBUY': '강매수', 'STATE_BUY': '매수', 'STATE_MR': '역매수'}.get(c_type, c_type)})"
                elif c_type == 'HOLDING_EXIT': reason += " (보유분석 청산)"
                elif c_type == 'COMPOSITE': reason += " (복합조건)"
                else: reason += f" ({r['target_price']})"
                    
                status_str = r['status']
                fail_reason = r['fail_reason'] if 'fail_reason' in r.keys() else None
                if status_str == 'CANCELED': status_str = '예약취소'
                elif status_str == 'FAILED': 
                    status_str = '발동실패'
                    if fail_reason: reason += f" [실패 사유: {fail_reason}]"
                elif status_str == 'TRIGGERED': status_str = '예약발동'
                elif status_str == 'EXPIRED': status_str = '기간만료'
                
                time_val = r['created_at']
                if time_val:
                    try:
                        dt = datetime.strptime(time_val, "%Y-%m-%d %H:%M:%S") + timedelta(hours=9)
                        time_val = dt.strftime("%Y-%m-%d %H:%M:%S")
                    except Exception:
                        pass

                reserved_trades.append({
                    'id': f"R{r['id']}", 'time': time_val, 'type': f"{t_type}(예약)",
                    'code': r['code'], 'name': r['name'], 'qty': r['qty'], 'price': r['order_price'],
                    'odno': r['odno'] or '-', 'org_odno': None, 'account': acc_str,
                    'is_sim': 1 if is_sim_account else 0, 'snapshot': None, 'profit_amt': 0,
                    'profit_rate': 0.0, 'reason': reason, 'strategy_score': 0,
                    'order_status': status_str, 'stop_loss_rate': 0.0
                })
        except Exception as e:
            logger.error(f"[HISTORY_DEBUG] 예약 주문 내역 조회 실패: {e}")
            
    if reserved_trades:
        trades.extend(reserved_trades)
        trades.sort(key=lambda x: x['time'], reverse=True)
        if choice == "1":
            trades = trades[:50]

    if not trades:
        logger.debug("[HISTORY_DEBUG] 조회된 내역 없음. 리턴.")
        config.console.print("\n[yellow]검색된 거래 내역이 없습니다.[/yellow]")
        return

    # [추가] 현재 설정된 계좌 정보 확인 (그룹핑용)
    current_main_acc = f"{config.session.cano}-{config.session.acnt_prdt_cd}"
    current_auto_acc = ""
    if config.session.auto_cano:
        current_auto_acc = f"{config.session.auto_cano}-{config.session.auto_acnt_prdt_cd}"

    # [추가] 현재 실행 중인 세션의 계좌만 표시하기 위한 유효 계좌 집합.
    #        토스/모의/한투는 모두 세션 계좌가 다르므로, 실제 계좌번호로 격리한다.
    #        (예: 모드2 한투 실행 시 토스 내역이 섞여 보이던 문제 해결)
    #        계좌번호 끝의 '-'(토스 등 상품코드 없는 계좌)는 정규화해 비교한다.
    valid_accs_norm = {current_main_acc.rstrip('-')}
    if current_auto_acc:
        valid_accs_norm.add(current_auto_acc.rstrip('-'))

    # [수정] 데이터 분류 및 그룹핑 (계좌번호 정규화 단위)
    grouped_trades = {} # acc_norm -> list

    for t in trades:
        # 1. 모드 필터링 (모의투자 모드면 모의내역만, 실전이면 실전/자동 내역만)
        is_sim_data = bool(t['is_sim'])
        if is_sim_data: continue

        acc_no = t.get('account', '')
        acc_norm = acc_no.rstrip('-')

        # 1-2. 현재 세션 계좌 격리: 실행 중인 계좌의 내역만 표시.
        #      계좌번호가 없는 오염 데이터('')나 다른 계좌(토스↔한투) 내역은 제외.
        if acc_norm not in valid_accs_norm:
            continue

        grouped_trades.setdefault(acc_norm, []).append(t)

    # [추가] 현재 세션에서 표시할 계좌 섹션(수동+자동)을 정의한다. 거래가 0건이어도
    #        섹션은 항상 노출하고, 내역이 없으면 노란색 안내를 출력한다.
    #        (예: 시스템 트레이딩(자동) 계좌를 동시 운용하지만 아직 자동매매 체결이
    #         없는 경우에도 해당 계좌 섹션을 확인할 수 있도록 함)
    #        [용어 통일] 계좌종류 라벨을 제한종목 화면과 동일하게
    #        (모의 / 토스 / 한투-수동 / 한투-자동)으로 표기한다.
    expected_sections = []  # (category, display_acc, acc_norm)
    if getattr(config.session, 'is_toss', False):
        expected_sections.append(("토스", current_main_acc, current_main_acc.rstrip('-')))
    else:
        expected_sections.append(("한투-수동", current_main_acc, current_main_acc.rstrip('-')))
        if current_auto_acc and current_auto_acc.rstrip('-') != current_main_acc.rstrip('-'):
            expected_sections.append(("한투-자동", current_auto_acc, current_auto_acc.rstrip('-')))

    for cat, disp_acc, acc_norm in expected_sections:
        t_list = grouped_trades.get(acc_norm, [])

        # 거래 내역이 없는 계좌(예: 아직 체결이 없는 자동매매 계좌)는 노란색 안내 출력
        if not t_list:
            logger.debug(f"[HISTORY_DEBUG] 내역 없음: {cat} {acc_norm}")
            config.console.print(
                f"\n[yellow][{cat}] 거래 히스토리 (계좌: {disp_acc.rstrip('-')}) - 거래 내역이 없습니다.[/yellow]"
            )
            continue

        acc = disp_acc  # 제목 표시용 (아래 rstrip 처리)
        logger.debug(f"[HISTORY_DEBUG] 테이블 생성 및 출력: {cat} {acc} ({len(t_list)}건)")

        # 테이블 생성 (제목에 계좌번호 포함, 끝의 '-'는 제거하여 표시)
        table_title = f"\n[{cat}] 거래 히스토리 (계좌: {acc.rstrip('-')}) - {len(t_list)}건"
        table = Table(title=table_title, box=box.HORIZONTALS, header_style="dim", border_style="dim", show_lines=True)
        table.add_column("시간", justify="center", style="dim", width=15, overflow="fold")
        table.add_column("주문번호", justify="center", style="dim", width=10, overflow="fold")
        # 계좌 컬럼 제거됨
        table.add_column("유형", justify="center", width=14, no_wrap=True)
        table.add_column("상태", justify="center", width=10, overflow="fold")
        table.add_column("종목명(코드)", justify="left", overflow="fold")
        table.add_column("수량", justify="right", width=6, overflow="fold")
        table.add_column("단가", justify="right", width=9, overflow="fold")
        table.add_column("금액", justify="right", width=10, overflow="fold")
        table.add_column("손익(수익률)", justify="right", overflow="fold")
        table.add_column("사유", justify="left", overflow="fold")

        for i, t in enumerate(t_list):
            # [수정] 유형 표기 개선 (줄바꿈 및 색상 적용)
            raw_type = t['type']
            clean_type = raw_type.replace("buy", "매수").replace("BUY", "매수").replace("sell", "매도").replace("SELL", "매도").replace("AUTO", "자동")
            
            base_type = "기타"
            is_buy = "매수" in clean_type
            is_sell = "매도" in clean_type
            is_mod = "정정" in clean_type
            is_cancel = "취소" in clean_type
            
            if is_mod:
                if is_buy: base_type = "매수정정"
                elif is_sell: base_type = "매도정정"
                else: base_type = "정정"
            elif is_cancel:
                if is_buy: base_type = "매수취소"
                elif is_sell: base_type = "매도취소"
                else: base_type = "취소"
            elif is_buy:
                base_type = "매수"
            elif is_sell:
                base_type = "매도"
            
            type_disp = base_type
            if base_type == "매수": type_disp = "[red]매수[/]"
            elif base_type == "매도": type_disp = "[blue]매도[/]"
            elif base_type == "매수정정": type_disp = "[red]매수[/][magenta]정정[/]"
            elif base_type == "매도정정": type_disp = "[blue]매도[/][magenta]정정[/]"
            elif base_type == "매수취소": type_disp = "[red]매수[/][yellow]취소[/]"
            elif base_type == "매도취소": type_disp = "[blue]매도[/][yellow]취소[/]"
            elif base_type == "정정": type_disp = "[magenta]정정[/]"
            elif base_type == "취소": type_disp = "[yellow]취소[/]"
            
            tag_disp = ""
            if "자동" in clean_type: tag_disp = "([yellow]자동[/])"
            elif "수동" in clean_type: tag_disp = "([green]수동[/])"
            elif "예약" in clean_type: tag_disp = "([magenta]예약[/])"
            else: tag_disp = "([dim]외부[/])"
            
            type_str = f"{type_disp}{tag_disp}"

            # 상태 표시
            raw_status_str = t.get('order_status', '접수')
            status_str = raw_status_str
            if "부분체결" in status_str: status_str = "[bold cyan]부분체결[/]"
            elif status_str == "체결": status_str = "[green]체결[/]"
            elif "체결(추정)" in status_str: status_str = "[green]체결 추정[/]" # [수정] 괄호 제거 및 색상 적용
            elif "취소(추정)" in status_str: status_str = "[bold yellow]취소 추정[/]"
            elif "거부" in status_str or "에러" in status_str or "REJECTED" in status_str.upper(): status_str = "[bold red]주문거부[/]"
            elif "취소" in status_str: status_str = f"[yellow]{status_str}[/]"
            elif "정정" in status_str: status_str = f"[magenta]{status_str}[/]"
            elif "예약발동" in status_str: status_str = "[bold green]예약발동[/]"
            elif "발동실패" in status_str: status_str = "[bold red]발동실패[/]"
            elif "기간만료" in status_str: status_str = "[dim]기간만료[/]"
            else: status_str = f"[dim]{status_str}[/]"

            # 가격 포맷팅
            price_display = t['price']
            try:
                p_val = float(t['price'])
                code = str(t.get('code', ''))
                is_domestic = (len(code) == 6 and code[0].isdigit() and code.isalnum())

                if p_val > 0:
                    if is_domestic:
                        price_display = f"{int(p_val):,}"
                    else:
                        price_display = f"{p_val:,.2f}"
                elif p_val == 0:
                    if "취소" in t['type'] or "cancel" in t['type'].lower():
                        price_display = "-"
                    else:
                        price_display = "시장가"
            except Exception: pass
            
            # [추가] 체결금액 계산 (단가 * 수량)
            total_amt_display = "-"
            try:
                p_val = float(t['price'])
                q_val = float(t['qty'])
                if p_val > 0 and q_val > 0:
                    tot = p_val * q_val
                    code = str(t.get('code', ''))
                    is_domestic = (len(code) == 6 and code[0].isdigit() and code.isalnum())
                    if is_domestic:
                        total_amt_display = f"{int(tot):,}"
                    else:
                        total_amt_display = f"{tot:,.2f}"
            except Exception: pass
            
            # 손익 정보
            profit_display = "-"
            # [수정] 정정/취소이더라도 원본이 매도라면 상속받은 손익을 화면에 정상 출력
            if base_type == "매도" or "매도" in raw_type:
                amt = t.get('profit_amt', 0)
                rate = t.get('profit_rate', 0.0)
                if amt is not None and rate is not None:
                    try:
                        if int(amt) != 0 or float(rate) != 0.0:
                            color = "red" if int(amt) > 0 else ("blue" if int(amt) < 0 else "white")
                            profit_display = f"[{color}]{int(amt):+,}원 ({float(rate):+.2f}%)[/]"
                    except Exception: pass

            # [추가] 사유 상세화: 스냅샷 정보를 활용하여 지표 정보 보강
            reason_display = t.get('reason') or "-"
            
            if t.get('snapshot'):
                try:
                    snap_data = json.loads(t['snapshot'])
                    
                    # 자동매매 상태(강매수, 역매수 등)가 스냅샷에 존재하면 사유에 명시적 추가
                    state_val = snap_data.get('state')
                    if state_val and is_buy and state_val not in reason_display:
                        reason_display = f"[{state_val}] {reason_display}"
                    
                    # 사용자 수동 주문인 경우 스냅샷에서 지표 정보 추출하여 표시
                    if "수동" in reason_display and 'indicators' in snap_data:
                        ind = snap_data['indicators']
                        add_info = []
                        if ind.get('rsi') is not None: add_info.append(f"RSI:{ind['rsi']:.1f}")
                        if ind.get('adx') is not None: add_info.append(f"ADX:{ind['adx']:.1f}")
                        if ind.get('cci') is not None: add_info.append(f"CCI:{ind['cci']:.1f}")
                        if add_info:
                            reason_display += f" [{', '.join(add_info)}]"
                except Exception: pass

            # [사유 태그] 분류 어휘는 core.trade_tags 단일 소스 — 텔레그램 /history 와 공유한다.
            if is_buy and reason_display != "-":
                reason_display = trade_tags.apply_buy_tag(reason_display)
            if is_sell and reason_display != "-":
                reason_display = trade_tags.apply_sell_tag(reason_display)

            # [추가] 기간만료/발동실패 상태 사유 태그 적용
            if reason_display != "-" and ("기간만료" in raw_status_str or "발동실패" in raw_status_str):
                if not reason_display.startswith("["):
                    fail_tag = "기간만료" if "기간만료" in raw_status_str else "발동실패"
                    reason_display = f"[{fail_tag}] {reason_display}"

            # [추가] 예약취소 상태 사유 태그 적용
            if reason_display != "-" and ("예약취소" in raw_status_str or "RES_CAN" in str(t.get('odno', ''))):
                if not reason_display.startswith("["):
                    reason_display = f"[예약취소] {reason_display}"

            # [추가] 예약 주문 발동 사유 태그 적용
            if reason_display != "-" and ("예약" in clean_type or "예약발동" in status_str):
                if not reason_display.startswith("["):
                    reserve_tag = "예약매수" if is_buy else ("예약매도" if is_sell else "예약발동")
                    reason_display = f"[{reserve_tag}] {reason_display}"

            # [추가] 정정/취소 주문 사유 태그 적용
            if reason_display != "-" and (is_mod or is_cancel):
                if not reason_display.startswith("["):
                    mod_tag = "정정" if is_mod else "취소"
                    reason_display = f"[{mod_tag}] {reason_display}"

            # [추가] 부분체결 상태 사유 태그 적용
            if reason_display != "-" and "부분체결" in raw_status_str:
                if not reason_display.startswith("["):
                    reason_display = f"[부분체결] {reason_display}"

            # [추가] 체결(추정) 상태 사유 태그 적용
            if reason_display != "-" and "체결(추정)" in raw_status_str:
                if not reason_display.startswith("["):
                    reason_display = f"[체결추정] {reason_display}"

            # [추가] 취소(추정) 상태 사유 태그 적용
            if reason_display != "-" and "취소(추정)" in raw_status_str:
                if not reason_display.startswith("["):
                    reason_display = f"[취소추정] {reason_display}"

            # [추가] 주문거부 상태 사유 태그 적용
            if reason_display != "-" and ("거부" in raw_status_str or "에러" in raw_status_str or "REJECTED" in raw_status_str.upper()):
                if not reason_display.startswith("["):
                    reason_display = f"[bold red]\\[주문거부][/] {reason_display}"

            # [추가] 접수(미체결) 상태 사유 태그 적용
            if reason_display != "-" and raw_status_str == "접수":
                if not reason_display.startswith("["):
                    reason_display = f"[미체결] {reason_display}"

            # [추가] 자동/수동/외부 사유 태그 적용
            if reason_display != "-":
                if "자동" in clean_type and "[자동]" not in reason_display:
                    reason_display = f"[자동] {reason_display}"
                elif "수동" in clean_type and "[수동]" not in reason_display:
                    reason_display = f"[수동] {reason_display}"
                elif "자동" not in clean_type and "수동" not in clean_type and "예약" not in clean_type and "[외부]" not in reason_display:
                    reason_display = f"[외부] {reason_display}"

            # [수정] 사유 내 강제 줄바꿈 제거 (2줄 내 유동적 출력 지원)
            reason_display = reason_display.replace('\n', ' ')

            table.add_row(
                t['time'][5:19], # MM-DD HH:MM:SS
                utils.format_order_no(t['odno']),  # 표시만 토스 뒤 10자리(DB엔 원본 저장)
                type_str,
                status_str,
                f"{t['name']}\n({t['code']})",
                f"{int(float(t['qty'])):,}",
                price_display,
                total_amt_display,
                profit_display,
                reason_display
            )
            
            # [추가] 5개마다 실선 추가
            if (i + 1) % 5 == 0 and (i + 1) < len(t_list):
                table.add_section()

        config.console.print(table)
        logger.debug("[HISTORY_DEBUG] 테이블 출력 완료")

def asset_management_menu():
    """자산 관리 메인 메뉴"""
    base_breadcrumb_len = len(context.USER_ACTION_BREADCRUMB)
    last_choice = "2"
    while True:
        context.USER_ACTION_BREADCRUMB = context.USER_ACTION_BREADCRUMB[:base_breadcrumb_len]
        
        menu_items = [("1", "자산 조회", "Asset Inquiry"), ("2", "보유 잔고", "Holdings"), ("3", "거래 내역", "Trade History"), ("4", "거래 평가", "Trading Report"), ("5", "포지션 분석", "Position Analysis"), ("6", "가상투자 관리", "Paper Account")]
        choice = utils.show_menu("자산 관리 (Asset Management)", menu_items, default_choice=last_choice)
        
        if choice.lower() in ['b', 'q']: return False
        if choice.lower() == 'h':
            if getattr(utils, 'show_help', None):
                utils.show_help()
                utils.pause()
            continue
        
        last_choice = choice
        menu_map = {"1": "자산 조회", "2": "보유 잔고", "3": "거래 내역", "4": "거래 평가", "5": "포지션 분석", "6": "가상투자 관리"}
        if choice in menu_map:
            context.USER_ACTION_BREADCRUMB.append(f"[{choice}] {menu_map[choice]}")

        if choice == "1":
            logger.info("운영자 실행: " + " - ".join(context.USER_ACTION_BREADCRUMB))
            get_deposit_balance()
            utils.pause()
        elif choice == "2":
            logger.info("운영자 실행: " + " - ".join(context.USER_ACTION_BREADCRUMB))
            get_account_balance()
            utils.pause()
        elif choice == "3":
            if view_trade_history() is not False: utils.pause()
        elif choice == "4":
            logger.info("운영자 실행: " + " - ".join(context.USER_ACTION_BREADCRUMB))
            from modules.trading import select_account
            from modules.auto_trade import AutoTrader
            target_cano, target_acnt, acc_label = select_account(title="평가할 계좌를 선택하세요")
            if target_cano:
                trader = AutoTrader()
                target_acc = f"{target_cano}-{target_acnt}"
                if trader.print_report(target_account=target_acc) is not False:
                    utils.pause()
        elif choice == "5":
            logger.info("운영자 실행: " + " - ".join(context.USER_ACTION_BREADCRUMB))
            if manual_holding_analysis() is not False:
                utils.pause()
        elif choice == "6":
            # 가상투자 계좌 관리(입출금·초기화·자산곡선). 계좌 성격의 기능이라 자산 관리에 둔다.
            logger.info("운영자 실행: " + " - ".join(context.USER_ACTION_BREADCRUMB))
            from modules import paper_report
            # 서브메뉴에서 q를 누르면 메인 메뉴로 빠져나가야 한다(다른 메뉴와 동일).
            if paper_report.show_paper_menu() is False:
                return False
