# modules/analysis.py
from rich.table import Table
from rich.prompt import Prompt
from rich.progress import Progress, SpinnerColumn, BarColumn, TextColumn, TimeRemainingColumn, DownloadColumn, TransferSpeedColumn
from rich import box
from rich.cells import cell_len

# [표 정렬] '분류 (TQ)' 컬럼은 한 셀에 두 값을 담는다. 분류 문구 길이가 4~6폭으로
#  들쭉날쭉해 그냥 이어 붙이면 분류가 세로로 안 읽힌다. **분류만 고정폭(6 = 최대 문구
#  '강매수·역매수')으로 채워 좌측 정렬**하고, TQ는 그 바로 뒤에 붙인다.
#  TQ까지 우측 정렬하면 자릿수는 맞지만 짧은 값(예: 상승 (4))에서 사이가 5칸까지 벌어져
#  두 값이 한 덩어리로 안 보인다 — 붙여 두는 쪽을 택했다.
#  구분 공백은 두지 않는다. 최대 문구('강매수(124)')는 붙지만 그 편이 한 덩어리로 읽힌다.
_CLASS_W = 6


def _class_tq_cell(left):
    """'분류 (TQ)' 셀에서 분류 뒤에 붙일 여백 — 분류를 고정폭으로 채워 좌측 정렬한다.

    헤더와 데이터 행이 **같은 함수**를 써야 폭 상수를 바꿔도 둘이 어긋나지 않는다.
    left는 색 마크업이 없는 순수 문자열이어야 한다(폭 계산이 틀어진다).
    """
    return " " * max(0, _CLASS_W - cell_len(left))
import config
from core import context
import api
import json
import logging
import contextlib
from core import indicators
from core import utils
from core import caching
import time
from datetime import datetime, timedelta
import urllib.request
import sys
import zipfile
import threading # [추가]
import os
import pandas as pd
import numpy as np
import requests
import concurrent.futures
import shutil
import sqlite3
import json
import math
import re
from modules import db_manager

logger = logging.getLogger(__name__)

# [Pylance 에러 방지용] 잔여 코드 및 중복 함수로 인한 미정의 변수 참조 경고 차단
reserved_codes = set()
m_codes = set()
restricted_stocks = set()
rules_map = {}

# [리팩토링] 스마트머니 캐시 (종목코드 -> (flag, reason)) — 공용 TTLCache 사용
_SMART_MONEY_CACHE = caching.TTLCache()

# [최적화] 시장 국면(get_market_regime) 단기 캐시 — 자동매매 주기당 반복 호출 제거
_MARKET_REGIME_CACHE = caching.TTLCache(max_size=8)
_MARKET_REGIME_TTL_SEC = 60.0

# [추가] tvDatafeed 싱글턴 — 토스 모드에서 KIS 미제공 지수(코스피·코스닥·코스피200·코스닥150)를
#  TradingView 시세로 보강한다. 익명 웹소켓은 연결 드롭이 잦아 단일 인스턴스를 재사용하고
#  호출을 전역 락으로 직렬화(페이싱)해 안정성을 높인다. get_domestic_index_data의 TTL 캐시가
#  실호출 빈도를 크게 낮춘다.
_TVDATAFEED_INSTANCE = None
_TVDATAFEED_LOCK = threading.Lock()
# 현재 인스턴스가 로그인 상태인지(True) — 로그인 상태에서는 실패 버스트 시 인스턴스를
#  재생성하지 않는다(재생성=재-signin이라 캡차·차단 위험).
_TVDATAFEED_LOGGED_IN = False
# 생성 전용 락(조회 직렬화용 _TVDATAFEED_LOCK과 분리) — 스레드풀에서 동시에 첫 호출이
#  들어와도 signin이 중복 발생하지 않도록 한다.
_TVDATAFEED_INIT_LOCK = threading.Lock()
# market_type -> (TradingView 심볼, 거래소)
#  코스피/코스닥은 yfinance(^KS11/^KQ11)가 최신 거래일 종가를 NaN으로 주는 일이 잦아(fast_info도
#  None) 지수·등락률이 '-'로 표시된다 → tvDatafeed를 1순위로 쓰고 실패 시 yfinance로 폴백한다.
# 토스 시장지표(/api/v1/market-indicators, API 1.2.4)로 조회 가능한 지수.
#  코스피200·코스닥150은 토스 심볼 카탈로그에 없어 tvDatafeed가 1순위로 남는다.
_TOSS_INDEX_MARKET_TYPES = ("KOSPI", "KOSDAQ")

_TVDATAFEED_INDEX_SYMBOLS = {
    "KOSPI": ("KOSPI", "KRX"),
    "KOSDAQ": ("KOSDAQ", "KRX"),
    "KOSPI200": ("KOSPI200", "KRX"),
    "KOSDAQ150": ("KOSDAQ150", "KRX"),
}

def _log_startup_info(msg):
    """기동 시 1회성 INFO를 FILE_DEBUG_LEVEL(기본 WARNING)과 무관하게 파일 로그에 남긴다.

    로거 레벨에서 걸리는 경우 레코드를 직접 만들어 핸들러로 넘긴다(핸들러 레벨은 NOTSET).
    남용하면 로그가 불어나므로 '운영 중 상태 확인에 반드시 필요한 1회성 기록'에만 쓴다.
    """
    if logger.isEnabledFor(logging.INFO):
        logger.info(msg)
        return
    try:
        logger.handle(logger.makeRecord(logger.name, logging.INFO, __file__, 0, msg, (), None))
    except Exception:
        pass


# [추가] TradingView 로그인. 라이브러리(TvDatafeed(username, password))의 내장 signin은
#  Referer 헤더만 보내는 구식 요청이라 현재 TradingView가 거부한다(→ 'error while signin' 후
#  익명 폴백). 동일 엔드포인트에 브라우저 수준 헤더를 붙여 우리가 직접 토큰을 받고,
#  익명으로 만든 인스턴스에 주입한다(get_hist는 호출마다 웹소켓에 토큰을 재전송하므로 유효).
_TV_SIGNIN_URL = "https://www.tradingview.com/accounts/signin/"
_TV_SIGNIN_HEADERS = {
    "Referer": "https://www.tradingview.com/",
    "Origin": "https://www.tradingview.com",
    "User-Agent": ("Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 "
                   "(KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36"),
    "Content-Type": "application/x-www-form-urlencoded",
    "Accept": "application/json, text/plain, */*",
}


# [추가] 발급받은 토큰은 파일에 캐시해 재사용한다. TradingView는 짧은 간격의 반복 로그인에
#  캡차("confirm that you are not a robot")를 요구하므로, 프로그램을 재시작할 때마다
#  signin을 하면 결국 로그인이 막힌다. 캐시가 유효하면 signin 없이 토큰만 주입한다.
_TV_TOKEN_CACHE_PATH = os.path.join(config.DATA_DIR, "tv_token.json")
_TV_TOKEN_TTL_SEC = 7 * 24 * 3600   # 7일


def _load_tv_token(user):
    """캐시된 TradingView 토큰을 읽는다(계정 불일치·만료·손상 시 None)."""
    try:
        with open(_TV_TOKEN_CACHE_PATH, 'r', encoding='utf-8') as f:
            data = json.load(f)
        if data.get('user') != user:
            return None
        if (time.time() - float(data.get('ts', 0))) > _TV_TOKEN_TTL_SEC:
            return None
        token = data.get('token')
        return token if token and token != "unauthorized_user_token" else None
    except Exception:
        return None


def _save_tv_token(user, token):
    """발급 토큰을 캐시한다(자격증명 성격이라 소유자 전용 권한 0600)."""
    try:
        os.makedirs(config.DATA_DIR, exist_ok=True)
        with open(_TV_TOKEN_CACHE_PATH, 'w', encoding='utf-8') as f:
            json.dump({'user': user, 'token': token, 'ts': time.time()}, f)
        os.chmod(_TV_TOKEN_CACHE_PATH, 0o600)
    except Exception as e:
        logger.debug(f"[TVDATAFEED] 토큰 캐시 저장 실패: {e}")


def _tv_signin(user, pw):
    """TradingView 인증 토큰을 발급받는다. 성공 시 (token, None), 실패 시 (None, 사유).

    사유 문자열에는 비밀번호를 포함하지 않는다(파일 로그에 남기 때문).
    """
    try:
        import requests
        resp = requests.post(_TV_SIGNIN_URL,
                             data={"username": user, "password": pw, "remember": "on"},
                             headers=_TV_SIGNIN_HEADERS, timeout=15)
    except Exception as e:
        return None, f"요청 실패({type(e).__name__}: {e})"
    try:
        body = resp.json()
    except Exception:
        return None, f"응답 파싱 실패(HTTP {resp.status_code})"
    token = (body.get("user") or {}).get("auth_token")
    if token:
        return token, None
    # TradingView는 실패 사유를 error/code(예: 캡차 요구, 2FA, 자격증명 오류)로 준다.
    err = body.get("error") or body.get("code") or body.get("detail") or ""
    if not err and body.get("two_factor_types"):
        err = "2단계 인증(2FA) 필요 — 계정에 2FA가 켜져 있으면 비밀번호 로그인 불가"
    return None, f"HTTP {resp.status_code}, 사유: {err or '토큰 없음'}"


# tvdatafeed는 PyPI 미배포(git 전용)라 배포 환경(예: 라즈베리파이)에 누락될 수 있다.
# 설치는 run.sh의 '필수 라이브러리 설치 상태 스캔' 단계가 git URL로 담당한다(여기서는 조회만).
def _get_tvdatafeed():
    """tvDatafeed 인스턴스를 지연 생성해 재사용한다. 미설치/초기화 실패 시 None.

    TV_USERNAME/TV_PASSWORD(config, ~/.htsrc의 export)가 모두 있으면 로그인 모드로 생성하고,
    없으면 종전처럼 익명(nologin)으로 생성한다. 로그인 결과는 파일 로그에 남긴다
    (성공=INFO / 미로그인·실패=WARNING). 로그는 인스턴스 생성 시점에만 찍힌다(싱글턴).
    """
    global _TVDATAFEED_INSTANCE, _TVDATAFEED_LOGGED_IN
    if _TVDATAFEED_INSTANCE is not None:
        return _TVDATAFEED_INSTANCE
    with _TVDATAFEED_INIT_LOCK:
        if _TVDATAFEED_INSTANCE is not None:  # 락 대기 중 다른 스레드가 생성했으면 재사용
            return _TVDATAFEED_INSTANCE
        return _init_tvdatafeed()


def _init_tvdatafeed():
    """_get_tvdatafeed 전용 생성 루틴(락 보유 상태에서만 호출)."""
    global _TVDATAFEED_INSTANCE, _TVDATAFEED_LOGGED_IN
    user = (getattr(config, 'TV_USERNAME', '') or '').strip()
    pw = (getattr(config, 'TV_PASSWORD', '') or '').strip()
    try:
        from tvDatafeed import TvDatafeed
        if user and pw:
            token, reason, cached = _load_tv_token(user), None, True
            if not token:  # 캐시 미스·만료 시에만 실제 signin(캡차 유발 최소화)
                cached = False
                token, reason = _tv_signin(user, pw)
                if token:
                    _save_tv_token(user, token)
            # 익명 생성 후 토큰 주입(내장 signin 우회). 토큰이 있으면 라이브러리의
            #  'you are using nologin method' 경고는 사실과 달라 생성 중에만 억제한다.
            _tvlog = logging.getLogger("tvDatafeed.main")
            _prev_level = _tvlog.level
            if token:
                _tvlog.setLevel(logging.ERROR)
            try:
                _TVDATAFEED_INSTANCE = TvDatafeed()
            finally:
                _tvlog.setLevel(_prev_level)
            if token:
                _TVDATAFEED_INSTANCE.token = token
                _TVDATAFEED_LOGGED_IN = True
                _log_startup_info(f"[TVDATAFEED] TradingView 로그인 성공 (계정: {user}"
                                  f"{', 캐시 토큰 재사용' if cached else ''})")
            else:
                _TVDATAFEED_LOGGED_IN = False
                logger.warning(
                    f"[TVDATAFEED] TradingView 로그인 실패 — 익명(nologin)으로 동작합니다 "
                    f"(계정: {user}, {reason})")
        else:
            _TVDATAFEED_INSTANCE = TvDatafeed()  # 익명(nologin)
            _TVDATAFEED_LOGGED_IN = False
            logger.warning(
                "[TVDATAFEED] TV_USERNAME/TV_PASSWORD 환경변수가 없어 로그인하지 않았습니다 "
                "— 익명(nologin)으로 동작합니다(데이터 한도·안정성 제한).")
    except Exception as e:
        logger.debug(f"[TVDATAFEED] 초기화 실패(라이브러리 미설치 가능): {e}")
        _TVDATAFEED_INSTANCE = None
        _TVDATAFEED_LOGGED_IN = False
    return _TVDATAFEED_INSTANCE

# KRX 공식 확정 봉을 '이력'으로 덮어쓸 지수 — 실시간 소스를 대체하지 않고 뼈대만 바꾼다.
#  V코스피200은 뺀다: KRX가 종가만 주므로(선물 응답의 SPOT_PRC) KIS의 OHLC를 덮으면
#  오히려 지표가 나빠진다. 그쪽은 KIS를 못 쓰는 모드에서만 폴백으로 쓴다.
_KRX_INDEX_MERGE_TYPES = ("KOSPI", "KOSDAQ", "KOSPI200", "KOSDAQ150")

# 지수 조회 창(달력일) — 250거래일(EMA120·52주)을 덮도록 여유를 둔다.
_KRX_INDEX_DAYS = 400


def _fetch_index_via_krx(market_type):
    """data.krx.co.kr 공식 지수·파생 일봉. 자격증명 없음·실패 시 None.

    지수(코스피 계열)·V코스피200·코스피200 선물(주간/야간)을 한 입구로 모은다 —
    호출부가 소스 종류를 몰라도 되게 하려는 것이다. (V코스피200·선물은 확정 봉뿐이라
    화면 경로에서는 쓰지 않는다 — `_fetch_domestic_index_data` 주석 참조)
    """
    try:
        from modules import krx_data
        if market_type == "VKOSPI":
            return krx_data.get_vkospi_daily(_KRX_INDEX_DAYS)
        if market_type in ("K200FUT_F", "K200FUT_CM"):
            return krx_data.get_k200_futures_daily(
                "F" if market_type == "K200FUT_F" else "CM", _KRX_INDEX_DAYS)
        return krx_data.get_index_daily(market_type, _KRX_INDEX_DAYS)
    except Exception as e:      # noqa: BLE001 - 어떤 실패든 종전 소스로 폴백한다
        logger.debug(f"[KRX] {market_type} 공식 조회 실패: {e}")
        return None


def _date_key(value):
    """비교용 'YYYYMMDD' 문자열. 소스마다 date 타입이 다르다(KIS/토스=문자열, tvDatafeed=datetime)."""
    try:
        return pd.to_datetime(value).strftime("%Y%m%d")
    except Exception:       # noqa: BLE001
        return str(value)


def _merge_index_history(hist, live):
    """KRX 확정 봉(hist) 위에 실시간 소스(live)의 '더 최신 날짜'만 얹는다.

    [왜 병합인가] KRX는 마감 후 확정 봉만 주므로 장중 당일 값이 없다. 반대로 실시간
     소스는 당일 값을 주지만 tvDatafeed는 간헐 빈 응답이고 지수 거래량을 0으로 준다.
     그래서 **판단(지표)의 뼈대는 KRX 확정 봉**으로 두고 KRX가 아직 갖지 않은 날짜만
     실시간 소스에서 가져온다 — 국내 일봉(krx_daily + 당일 오버레이)과 같은 구조다.

    한쪽이 없으면 있는 쪽을 그대로 돌려준다(그래서 tvDatafeed가 통째로 죽어도
    코스닥150이 살아남는다 — 야후·FDR에는 티커 자체가 없어 종전엔 복구 불가였다).
    반환 date는 'YYYYMMDD' 문자열로 통일한다(KIS·토스 경로와 같은 표기).
    """
    if hist is None or getattr(hist, "empty", True):
        return live
    if live is None or getattr(live, "empty", True):
        return hist

    out = hist.copy()
    out["date"] = out["date"].map(_date_key)
    tail = live.copy()
    tail["date"] = tail["date"].map(_date_key)
    tail = tail[tail["date"] > out["date"].max()]
    if not tail.empty:
        out = pd.concat([out, tail[out.columns]], ignore_index=True)
    out = out.sort_values("date").reset_index(drop=True)
    out.attrs["source"] = hist.attrs.get("source", "KRX")
    return out


def _fetch_index_via_tvdatafeed(market_type, n_bars=260):
    """토스 모드 국내 지수(코스피·코스닥·코스피200·코스닥150)를 TradingView(tvDatafeed)로 조회한다.

    반환 스키마는 KIS/yfinance 경로와 동일: ['date','open','high','low','close','volume']
    (date=datetime, RangeIndex, 오름차순, attrs['source']='TVDATAFEED'). 실패 시 None.
    """
    sym = _TVDATAFEED_INDEX_SYMBOLS.get(market_type)
    if not sym:
        return None
    symbol, exchange = sym
    tv = _get_tvdatafeed()
    if tv is None:
        return None
    try:
        from tvDatafeed import Interval
    except Exception:
        return None

    # 익명 웹소켓은 웜 인스턴스에서도 ~1/3 확률로 빈 응답이 오고 버스트로 연속 실패도 있어
    # 최대 4회 재시도한다. 호출은 전역 락으로 직렬화(페이싱)한다. get_hist는 호출마다 웹소켓을
    # 새로 맺으므로(인증 토큰 재전송 포함) 인스턴스를 재생성할 필요가 없다 — 로그인 상태에서
    # 반복 재생성 시 매번 재-signin(캡차/차단 위험)이 발생하므로 인스턴스는 재사용한다.
    #  [회로차단 2026-08-19] 직전에 tvDatafeed가 전 재시도 실패했다면 1회만 시도한다.
    #   토스 모드 지수 화면은 7개가 이 소스를 전역 락으로 직렬 사용하므로, 각자 4회씩
    #   재시도하면 그 대기가 그대로 화면 정지로 나타난다.
    max_attempts = 1 if _tv_circuit_open() else 4
    df = None
    _t0 = time.time()
    for attempt in range(max_attempts):
        try:
            with _TVDATAFEED_LOCK:
                df = tv.get_hist(symbol=symbol, exchange=exchange,
                                 interval=Interval.in_daily, n_bars=n_bars)
            if df is not None and not df.empty:
                break
        except Exception as e:
            logger.debug(f"[TVDATAFEED] {market_type} 조회 오류(attempt={attempt}): {e}")
            df = None
        if attempt < max_attempts - 1:
            time.sleep(0.8 * (attempt + 1))  # 페이싱 후 재시도(UI 지연 최소화 위해 점증 백오프)

    if df is None or df.empty:
        logger.warning(f"[TVDATAFEED] {market_type} 데이터 없음 — {max_attempts}회 시도 실패 "
                       f"({time.time() - _t0:.1f}s)")
        _tv_note_failure()
        return None
    _tv_note_success()

    try:
        out = df.reset_index().rename(columns={'datetime': 'date'})
        for col in ['open', 'high', 'low', 'close', 'volume']:
            if col not in out.columns:
                out[col] = 0.0
        out = out[['date', 'open', 'high', 'low', 'close', 'volume']].copy()
        out = out.sort_values('date', ascending=True).reset_index(drop=True)
        # [추가] tvDatafeed는 지수 거래량을 0으로 준다 → OBV 계산 불가('-'). yfinance 거래량을
        #  날짜 매칭으로 채워 OBV를 살린다(코스피/코스닥/코스피200; ^KQ150은 실측 없음 → 0 유지).
        out = _merge_index_volume_from_yfinance(out, _INDEX_YF_TICKERS.get(market_type))
        out.attrs['source'] = 'TVDATAFEED'
        return out
    except Exception as e:
        logger.debug(f"[TVDATAFEED] {market_type} 스키마 변환 실패: {e}")
        return None


# market_type -> yfinance 티커 (거래량 보강·최후 폴백 공용)
_INDEX_YF_TICKERS = {
    "KOSPI": "^KS11",
    "KOSDAQ": "^KQ11",
    "KOSPI200": "^KS200",
    "KOSDAQ150": "^KQ150",
}


def _merge_index_volume_from_yfinance(df, yf_ticker):
    """tvDatafeed 지수 df(거래량=0)에 yfinance 일봉 거래량을 '날짜 매칭'으로 채운다(OBV 계산용).

    가격(OHLC/close)은 tvDatafeed 값을 그대로 유지하고 volume만 교체한다. yfinance 조회 실패·
    빈 응답(^KQ150 등)·매칭 실패 시 원본(volume=0)을 그대로 반환한다. 실 호출은 tvDatafeed
    캐시 미스 시점에만 일어난다(get_domestic_index_data TTL 캐시가 빈도를 낮춤).
    """
    if not yf_ticker or df is None or df.empty or 'date' not in df.columns:
        return df
    # [교착 수정 2026-08-19] 종전에는 api.get_chart_data(yf_ticker)를 불렀다. 그런데 지수
    #  '단일 소스' 규칙(2026-07)이 들어오면서 ^KS200·^KQ150은 그 안에서 다시 지수 소스
    #  체인으로 되돌려진다 → get_domestic_index_data → _fetch_index_via_tvdatafeed →
    #  여기 → get_chart_data … 사이클이 생겼고, get_domestic_index_data의 market_type별
    #  single-flight 락을 **같은 스레드가 다시 잡으며 영구 교착**했다(토스 모드에서
    #  코스피200·코스닥150이 조회 중 멈춤 — 2026-08-19 신고).
    #  거래량 보강에 필요한 것은 야후 원본 거래량뿐이므로 소스 체인을 타지 않고 직접 받는다.
    try:
        raw = api.fetch_yfinance_data(yf_ticker, period="1y")
    except Exception as e:
        logger.debug(f"[TVDATAFEED] 거래량 보강 yfinance 조회 실패({yf_ticker}): {e}")
        return df
    if raw is None or getattr(raw, 'empty', True):
        return df

    vol_map = {}
    try:
        cols = raw.columns
        if hasattr(cols, 'nlevels') and cols.nlevels > 1:
            # 단일 티커라도 group_by에 따라 MultiIndex가 올 수 있다.
            raw = raw.xs(yf_ticker, axis=1, level=-1) if yf_ticker in cols.get_level_values(-1) \
                else raw.droplevel(-1, axis=1)
        if 'Volume' not in raw.columns:
            return df
        for idx, v in raw['Volume'].items():
            try:
                v = float(v)
            except (TypeError, ValueError):
                continue
            if v == v and v > 0:  # NaN 제외 & 양수만
                key = idx.strftime('%Y%m%d') if hasattr(idx, 'strftime') else str(idx).replace('-', '')[:8]
                vol_map[key] = v
    except Exception as e:
        logger.debug(f"[TVDATAFEED] 거래량 보강 파싱 실패({yf_ticker}): {e}")
        return df
    if not vol_map:
        return df

    def _vol_for(dt):
        key = dt.strftime('%Y%m%d') if hasattr(dt, 'strftime') else str(dt).replace('-', '')[:8]
        return vol_map.get(key, 0.0)

    out = df.copy()
    out['volume'] = out['date'].map(_vol_for)
    return out

# [추가] 미국채 현물 금리(TVC:US02Y/US05Y/US10Y/US30Y) 캐시 — 현물 금리는 장외(아시아장)에도
#  거의 24시간 갱신되어 선물 프록시 추정 없이 실제 호가를 표시할 수 있다. 2년물은 야후에
#  현물 지수가 없고 CBOT 금리선물(2YY=F)마저 유동성 고갈로 시세가 수일 지연되므로(현물
#  4.17% vs 선물 3.99%) tvDatafeed가 유일한 현물 소스. TV 일봉 마지막 봉은 장중 실시간
#  갱신되므로 짧은 TTL 캐시로 준실시간을 유지한다.
_US_TREASURY_SPOT_CACHE = {}   # symbol -> {"df": DataFrame, "time": datetime, "fail": datetime|None}
_US_TREASURY_TTL_SEC = 120     # 미국 장중 금리 변동 대응 (다른 지수 fast_info 60s와 유사 수준)
_US_TREASURY_NEG_TTL_SEC = 180  # 실패 음성 캐시(익명 웹소켓 다운 시 매 조회 재시도로 UI 지연 방지)
# [회로차단] 한 심볼이 전 재시도 실패(TV 전면 장애 가능성)하면 직후 다른 심볼은 재시도
#  1회만 수행해 지수 화면 콜드스타트가 수십 초 지연되는 것을 막는다.
#
#  [확대 2026-08-19] 종전에는 국채 경로만 이 차단을 썼다. 그런데 토스 모드 지수 화면은
#  코스피200·코스닥150·국채 4개 테너·HY OAS까지 **7개가 tvDatafeed**를 쓰고, 그 호출은
#  _TVDATAFEED_LOCK으로 한 번에 하나씩만 돈다. TV가 흔들릴 때 국내 지수(4회)·FRED(4회)가
#  각자 재시도 예산을 다 쓰면 직렬 대기가 분 단위로 쌓여 화면이 멈춘 것처럼 보인다
#  (실측 신고: 전체 지수 96%에서 정지). 그래서 신호를 **세 경로가 공유**한다.
_TV_ANY_FAIL_TIME = None
_TV_CIRCUIT_SEC = 120


def _tv_circuit_open(now=None):
    """직전에 tvDatafeed가 전 재시도 실패했는가 — True면 이번 호출은 1회만 시도한다."""
    if _TV_ANY_FAIL_TIME is None:
        return False
    now = now or datetime.now()
    return (now - _TV_ANY_FAIL_TIME).total_seconds() < _TV_CIRCUIT_SEC


def _tv_note_failure(now=None):
    global _TV_ANY_FAIL_TIME
    _TV_ANY_FAIL_TIME = now or datetime.now()


def _tv_note_success():
    global _TV_ANY_FAIL_TIME
    _TV_ANY_FAIL_TIME = None


def reset_tvdatafeed_circuit():
    """회로차단만 해제한다(음성 캐시는 각 경로의 reset_*가 담당)."""
    _tv_note_success()


def reset_us_treasury_spot_failures():
    """국채 현물(TVC) 음성 캐시·회로차단을 해제한다.

    사용자가 지수 화면에서 명시적으로 재시도(y)할 때 호출 — 해제하지 않으면
    음성 캐시(600s) 동안 재시도가 즉시 실패를 반환해 무의미해진다.
    """
    reset_tvdatafeed_circuit()
    for ent in _US_TREASURY_SPOT_CACHE.values():
        ent["fail"] = None

def get_us_treasury_spot_data(symbol, n_bars=300):
    """미국채 현물 금리(TVC:USxxY) 일봉을 tvDatafeed로 조회한다(5분 TTL 캐시).

    symbol: "US02Y"|"US05Y"|"US10Y"|"US30Y".
    반환 스키마는 지수 경로와 동일: ['date','open','high','low','close','volume']
    (volume=0 — 금리 지수라 OBV 불가, attrs['source']='TVDATAFEED'). 실패 시 None.
    """
    now = datetime.now()
    ent = _US_TREASURY_SPOT_CACHE.setdefault(symbol, {"df": None, "time": None, "fail": None})
    cached = ent["df"]
    if cached is not None and ent["time"] and (now - ent["time"]).total_seconds() < _US_TREASURY_TTL_SEC:
        return cached
    if ent["fail"] and (now - ent["fail"]).total_seconds() < _US_TREASURY_NEG_TTL_SEC:
        return cached  # 음성 캐시 구간엔 만료된 성공 캐시라도 재사용(없으면 None)

    tv = _get_tvdatafeed()
    if tv is None:
        return cached
    try:
        from tvDatafeed import Interval
    except Exception:
        return cached

    # 회로차단: 직전 다른 심볼이 전 재시도 실패였다면 이번 심볼은 1회만 시도.
    # 콜드스타트(성공 캐시 없음)는 실패 시 표시할 값 자체가 없으므로(특히 2년물은 폴백도 없음)
    # 재시도를 6회로 늘려 간헐 실패 버스트를 견딘다.
    max_attempts = 1 if _tv_circuit_open(now) else (6 if cached is None else 4)

    df = None
    for attempt in range(max_attempts):  # 익명 웹소켓 간헐 실패 대응(국내 지수 경로와 동일 재시도 정책)
        try:
            with _TVDATAFEED_LOCK:
                df = tv.get_hist(symbol=symbol, exchange="TVC",
                                 interval=Interval.in_daily, n_bars=n_bars)
            if df is not None and not df.empty:
                break
        except Exception as e:
            logger.debug(f"[TVDATAFEED] {symbol} 조회 오류(attempt={attempt}): {e}")
            df = None
        if attempt < max_attempts - 1:
            # 실패 버스트가 이어지면 마지막 시도 전 익명 인스턴스를 재생성해 웹소켓
            # 불량 상태를 리셋한다(익명이라 재-signin/캡차 위험 없음).
            # 로그인 인스턴스는 재생성 시 매번 재-signin이 발생하므로 재사용만 한다.
            if attempt == max_attempts - 2 and not _TVDATAFEED_LOGGED_IN:
                global _TVDATAFEED_INSTANCE
                with _TVDATAFEED_LOCK:
                    _TVDATAFEED_INSTANCE = None
                tv = _get_tvdatafeed()
                if tv is None:
                    break
            time.sleep(0.8 * (attempt + 1))

    if df is None or df.empty:
        # 실패는 warning으로 남긴다 — debug로 두면 FILE_DEBUG_LEVEL=DEBUG가 아닌 운영에서
        #  '화면이 왜 느렸는지'를 사후에 확인할 방법이 없다(2026-08-19 신고 건이 그랬다).
        logger.warning(f"[TVDATAFEED] {symbol} 데이터 없음 — {max_attempts}회 시도 실패")
        ent["fail"] = now
        _tv_note_failure(now)
        return cached

    try:
        out = df.reset_index().rename(columns={'datetime': 'date'})
        for col in ['open', 'high', 'low', 'close', 'volume']:
            if col not in out.columns:
                out[col] = 0.0
        out = out[['date', 'open', 'high', 'low', 'close', 'volume']].copy()
        out = out.sort_values('date', ascending=True).reset_index(drop=True)
        out.attrs['source'] = 'TVDATAFEED'
        ent["df"] = out
        ent["time"] = now
        ent["fail"] = None
        _tv_note_success()  # 성공 → 회로차단 해제
        return out
    except Exception as e:
        logger.debug(f"[TVDATAFEED] {symbol} 스키마 변환 실패: {e}")
        ent["fail"] = now
        return cached

# [추가] FRED 계열(HY OAS 등) 캐시 — 국채 현물(_US_TREASURY_SPOT_CACHE)과 같은 구조.
#  FRED는 일 1회 갱신되는 거시 시계열이라 장중 재조회 가치가 낮다 → TTL을 국채(120s)보다
#  길게 잡아 tvDatafeed 전역 락 경합 자체를 줄인다.
_FRED_CACHE = {}            # symbol -> {"df": DataFrame, "time": datetime, "fail": datetime|None}
_FRED_TTL_SEC = 1800        # 30분 (일 1회 갱신 시계열)
_FRED_NEG_TTL_SEC = 180     # 실패 음성 캐시 — 익명 웹소켓 다운 시 매 렌더 재시도로 UI가 지연되는 것 방지


def reset_fred_failures():
    """FRED 음성 캐시를 해제한다(사용자가 지수 화면에서 명시적으로 재시도할 때)."""
    reset_tvdatafeed_circuit()   # 재시도인데 회로가 닫혀 있으면 1회 시도로 끝나 무의미하다
    for ent in _FRED_CACHE.values():
        ent["fail"] = None


def get_fred_data(symbol, n_bars=300):
    """FRED 일봉을 tvDatafeed로 조회한다(TTL 캐시 + 재시도).

    [종전] 단 1회만 시도하고 캐시도 폴백도 없었다. 그런데 익명 웹소켓은 웜 인스턴스에서도
     ~1/3 확률로 빈 응답을 주고 실패가 버스트로 몰린다(_fetch_index_via_tvdatafeed 주석).
     같은 소스를 쓰는 국내 지수는 4회, 국채 현물은 최대 6회 재시도 + 성공값 폴백을 두었는데
     이 경로만 무방비여서, tvDatafeed 호출이 많은 토스/가상투자 모드(코스피200·코스닥150이
     tvDatafeed 1순위)에서 HY OAS만 상시 실패로 보였다. 재시도 정책을 지수 경로와 맞춘다.
    """
    now = datetime.now()
    ent = _FRED_CACHE.setdefault(symbol, {"df": None, "time": None, "fail": None})
    cached = ent["df"]
    if cached is not None and ent["time"] and (now - ent["time"]).total_seconds() < _FRED_TTL_SEC:
        return cached
    if ent["fail"] and (now - ent["fail"]).total_seconds() < _FRED_NEG_TTL_SEC:
        return cached  # 음성 캐시 구간엔 만료된 성공 캐시라도 재사용(없으면 None)

    tv = _get_tvdatafeed()
    if tv is None:
        return cached
    try:
        from tvDatafeed import Interval
    except Exception:
        return cached

    df = None
    # 익명 웹소켓 간헐 실패 대응(국내 지수 경로와 동일 재시도 정책).
    #  [회로차단 2026-08-19] TV가 전면으로 흔들리는 중이면 1회만 — 같은 락을 기다리는
    #  다른 지수까지 이 재시도 시간을 그대로 물려받는다.
    max_attempts = 1 if _tv_circuit_open(now) else 4
    _t0 = time.time()
    for attempt in range(max_attempts):
        try:
            with _TVDATAFEED_LOCK:
                df = tv.get_hist(symbol=symbol, exchange="FRED",
                                 interval=Interval.in_daily, n_bars=n_bars)
            if df is not None and not df.empty:
                break
        except Exception as e:
            logger.debug(f"[TVDATAFEED] FRED:{symbol} 조회 오류(attempt={attempt}): {e}")
            df = None
        if attempt < max_attempts - 1:
            time.sleep(0.8 * (attempt + 1))   # 페이싱 후 재시도(점증 백오프)

    if df is None or df.empty:
        logger.warning(f"[TVDATAFEED] FRED:{symbol} 데이터 없음 — {max_attempts}회 시도 실패 "
                       f"({time.time() - _t0:.1f}s)")
        ent["fail"] = now
        _tv_note_failure(now)
        return cached
    _tv_note_success()

    try:
        out = df.reset_index().rename(columns={'datetime': 'date'})
        for col in ['open', 'high', 'low', 'close', 'volume']:
            if col not in out.columns:
                out[col] = 0.0
        out = out[['date', 'open', 'high', 'low', 'close', 'volume']].copy()
        out = out.sort_values('date', ascending=True).reset_index(drop=True)
        out.attrs['source'] = 'TVDATAFEED'
        ent["df"] = out
        ent["time"] = now
        ent["fail"] = None
        return out
    except Exception as e:
        logger.debug(f"[TVDATAFEED] FRED:{symbol} 스키마 변환 실패: {e}")
        ent["fail"] = now
        return cached

# ==========================================================
# [추가] KRX 금현물 (금 99.99_1Kg, 원/g) — 네이버 원자재 시세
# ==========================================================
#  국제 금(COMEX GC=F, USD/온스)과 달리 KRX 금시장 시세는 KIS·토스·yfinance·pykrx 어디에도
#  없다(ETF 411060은 ETF 주가라 대용 불가). KRX 정보데이터시스템 JSON은 세션 쿠키를 심어도
#  LOGOUT을 돌려준다 → 네이버 원자재 API가 유일한 실시간 소스다(delayTime=0, 거래소 KRX).
#
#  [캐시를 둘로 나누는 이유] 현재가는 장중 계속 바뀌지만 과거 종가는 불변이다. 한 덩어리로
#  묶으면 60초마다 5페이지(300거래일)를 다시 받게 된다 → 현재가만 짧은 TTL로 갱신하고
#  시계열은 6시간(날짜가 바뀌면 즉시) 캐시한다. 정상 구간의 반복 조회는 1콜이면 끝난다.
#
#  [종가만 있는 시계열] 네이버 일별 시세는 종가만 유효하고 시·고·저는 0으로 내려온다 →
#  모든 봉을 종가로 평탄화한 OHLC로 만든다. EMA·RSI·CCI·MACD·ADX·SAR은 종가 기준으로
#  정상 산출되며(고·저 대신 종가 차분이 True Range가 된다), 거래량 이력이 없어 OBV만
#  '-'로 남는다(지수 화면이 vol_sum==0을 이미 그렇게 처리한다).
_KRX_GOLD_URL = "https://api.stock.naver.com/marketindex/metals"
_KRX_GOLD_HEADERS = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 '
                  '(KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
    'Referer': 'https://m.stock.naver.com/',
}
_KRX_GOLD_PAGE_SIZE = 60        # 네이버 상한 (초과하면 400)
_KRX_GOLD_PAGES = 5             # 60 x 5 = 300거래일 — EMA120·52주(250봉)를 덮는다
_KRX_GOLD_TIMEOUT = 5
_KRX_GOLD_QUOTE_TTL_SEC = 60    # 다른 지수의 fast_info 캐시(60초)와 같은 신선도
_KRX_GOLD_HIST_TTL_SEC = 21600  # 6시간 (과거 종가는 불변, 날짜가 바뀌면 아래에서 별도 무효화)
_KRX_GOLD_NEG_TTL_SEC = 180     # 실패 음성 캐시 — 네이버 장애 시 매 렌더 재시도로 UI가 멈추는 것 방지
_KRX_GOLD_CACHE = {}            # symbol -> {"hist","hist_time","hist_day","quote","quote_time","fail"}
_KRX_GOLD_LOCK = threading.RLock()


def reset_krx_gold_failures():
    """KRX 금 음성 캐시를 해제한다(사용자가 지수 화면에서 명시적으로 재시도할 때)."""
    with _KRX_GOLD_LOCK:
        for ent in _KRX_GOLD_CACHE.values():
            ent["fail"] = None


def _krx_gold_entry(symbol):
    with _KRX_GOLD_LOCK:
        return _KRX_GOLD_CACHE.setdefault(symbol, {
            "hist": None, "hist_time": None, "hist_day": None,
            "quote": None, "quote_time": None, "fail": None,
            # quote_fail: KRX 경로에서 현재가만 실패했을 때의 음성 캐시.
            #  시계열 실패(fail)와 분리해야 현재가 장애가 시계열 재시도를 막지 않는다.
            "quote_fail": None,
        })


def _krx_gold_num(text):
    """네이버 표기('203,410')를 수치로. 값이 없거나 0이면 None."""
    try:
        val = float(str(text).replace(",", "").strip())
    except (TypeError, ValueError):
        return None
    return val if val > 0 else None


def _fetch_krx_gold_history(symbol):
    """일별 종가 시계열 [(date, close)] — 최근 것부터 페이지 단위로 받아 온다."""
    rows = []
    for page in range(1, _KRX_GOLD_PAGES + 1):
        res = requests.get(f"{_KRX_GOLD_URL}/{symbol}/prices",
                           params={'page': page, 'pageSize': _KRX_GOLD_PAGE_SIZE},
                           headers=_KRX_GOLD_HEADERS, timeout=_KRX_GOLD_TIMEOUT)
        page_rows = res.json()
        if not isinstance(page_rows, list) or not page_rows:
            break
        for row in page_rows:
            close = _krx_gold_num(row.get('closePrice'))
            traded_at = row.get('localTradedAt')
            if close is None or not traded_at:
                continue
            rows.append((pd.to_datetime(str(traded_at)[:10]), close))
        if len(page_rows) < _KRX_GOLD_PAGE_SIZE:
            break   # 마지막 페이지 — 더 받아도 빈 응답이다
    return rows


def _fetch_krx_gold_quote(symbol):
    """현재가와 전일 종가 (current, prev). 장 종료 뒤에는 그날의 최종 종가가 내려온다."""
    res = requests.get(f"{_KRX_GOLD_URL}/{symbol}", headers=_KRX_GOLD_HEADERS,
                       timeout=_KRX_GOLD_TIMEOUT)
    data = res.json()
    current = _krx_gold_num(data.get('closePrice'))
    if current is None:
        return None
    traded_at = data.get('localTradedAt')
    day = pd.to_datetime(str(traded_at)[:10]) if traded_at else None
    return {'date': day, 'close': current}


# KRX 공식 금현물 조회 창(달력일). 네이버 경로(_KRX_GOLD_PAGES × 60 = 300거래일)와 같은
#  분량을 목표로 여유 있게 잡는다 — EMA120·52주(250봉)를 덮어야 한다.
_KRX_GOLD_OFFICIAL_DAYS = 400


def _krx_gold_official():
    """KRX 공식 금현물 일봉(확정 봉). 자격증명 없음·실패 시 None → 네이버로 폴백한다."""
    try:
        from modules import krx_data
        return krx_data.get_gold_daily(_KRX_GOLD_OFFICIAL_DAYS)
    except Exception as e:      # noqa: BLE001 - 어떤 실패든 네이버 폴백으로 넘긴다
        logger.debug(f"[KRX] 금현물 공식 조회 실패: {e}")
        return None


def _krx_gold_quote_cached(symbol, ent, now):
    """네이버 현재가(60초 캐시). KRX 경로 전용 — 실패는 quote_fail로 따로 묶는다.

    시계열 실패(ent['fail'])와 섞지 않는다. 섞으면 현재가 한 번 실패가 시계열 재시도까지
    막아 버린다(실제로 그렇게 짰다가 음성 캐시 테스트가 잡아냈다).
    """
    quote = ent.get("quote")
    if quote and ent.get("quote_time") and \
            (now - ent["quote_time"]).total_seconds() < _KRX_GOLD_QUOTE_TTL_SEC:
        return quote
    last_fail = ent.get("quote_fail")
    if last_fail and (now - last_fail).total_seconds() < _KRX_GOLD_NEG_TTL_SEC:
        return quote                      # 장애 구간엔 재요청하지 않는다(만료된 값이라도 그대로)
    try:
        fetched = _fetch_krx_gold_quote(symbol)
        if fetched:
            with _KRX_GOLD_LOCK:
                ent["quote"], ent["quote_time"], ent["quote_fail"] = fetched, now, None
            return fetched
    except Exception as e:      # noqa: BLE001 - 현재가는 없어도 확정 봉만으로 표를 채울 수 있다
        logger.debug(f"[NAVER] KRX 금 현재가 조회 실패: {e}")
    with _KRX_GOLD_LOCK:
        ent["quote_fail"] = now
    return quote


def _overlay_gold_quote(df, quote):
    """확정 일봉에 장중 현재가를 덮는다(KRX는 마감 후 확정 봉만 주기 때문).

    같은 날짜 봉이 있으면 종가를 갈아끼우면서 **고·저를 그 값까지 넓힌다** — 종가가 봉
    밖으로 벗어나면 True Range 가 음수가 되어 ATR·SAR 이 망가진다. 봉이 없으면 새로 덧붙인다.
    """
    if df is None or df.empty:
        return df
    out = df.copy()
    # 네이버 경로와 날짜 타입을 맞춘다(그쪽은 Timestamp를 쓴다) — 호출부가 둘을 구분하지 않는다.
    out['date'] = pd.to_datetime(out['date'], format='%Y%m%d', errors='coerce')
    out = out.dropna(subset=['date'])
    if not quote or quote.get('date') is None or not quote.get('close'):
        out.attrs['source'] = 'KRX'
        return out.reset_index(drop=True)

    day, price = pd.to_datetime(quote['date']), float(quote['close'])
    hit = out.index[out['date'] == day]
    if len(hit):
        i = hit[-1]
        out.loc[i, 'high'] = max(float(out.loc[i, 'high']), price)
        out.loc[i, 'low'] = min(float(out.loc[i, 'low']), price)
        out.loc[i, 'close'] = price
    else:
        out = pd.concat([out, pd.DataFrame([{
            'date': day, 'open': price, 'high': price, 'low': price,
            'close': price, 'volume': 0.0}])], ignore_index=True)
    out = out.sort_values('date').reset_index(drop=True)
    out.attrs['source'] = 'KRX'
    return out


def get_krx_gold_data(symbol=None):
    """KRX 금현물(원/g) 일봉 — **KRX 공식(1순위) / 네이버(폴백)**.

    반환 스키마는 다른 지수 전용 소스와 동일: ['date','open','high','low','close','volume'].

    [왜 KRX가 1순위인가] 네이버 일별 시세는 **종가만** 유효하고 시·고·저가 0으로 내려와
     모든 봉을 종가로 평탄화해야 했다 — True Range 가 종가 차분이 되어 ATR·ADX 가 왜곡되고,
     거래량 이력이 없어 OBV 는 '-' 로 남았다. data.krx.co.kr 로그인이 생기면서 실제 OHLC 와
     거래량을 받게 됐다(네이버 종가와 겹치는 60일 불일치 0으로 드롭인 확인).
     KRX_ID/KRX_PW 가 없으면 종전 네이버 경로로 그대로 폴백하므로 동작은 종전과 같다.

    KRX는 마감 후 확정 봉만 주므로 **장중 현재가는 네이버 현재가로 덮는다**(60초 캐시).
    attrs['source'] 는 'KRX' 또는 'NAVER'. 실패 시 None(성공 캐시가 있으면 그것을 돌려준다).
    """
    symbol = symbol or config.KRX_GOLD_SYMBOL
    now = datetime.now()
    ent = _krx_gold_entry(symbol)
    today = now.strftime("%Y%m%d")

    official = _krx_gold_official()
    if official is not None and not official.empty:
        # KRX가 확정 봉을 줬다 — 장중 현재가만 네이버에서 덧댄다(네이버 장애는 무해하다).
        return _overlay_gold_quote(official, _krx_gold_quote_cached(symbol, ent, now))

    # ── 여기부터는 네이버 폴백 (KRX 미설정·조회 실패) — 종전 로직 그대로 ─────────────
    hist = ent["hist"]
    hist_fresh = (hist and ent["hist_time"] and ent["hist_day"] == today
                  and (now - ent["hist_time"]).total_seconds() < _KRX_GOLD_HIST_TTL_SEC)
    quote = ent["quote"]
    quote_fresh = (quote and ent["quote_time"]
                   and (now - ent["quote_time"]).total_seconds() < _KRX_GOLD_QUOTE_TTL_SEC)
    if hist_fresh and quote_fresh:
        return _krx_gold_frame(hist, quote)

    if ent["fail"] and (now - ent["fail"]).total_seconds() < _KRX_GOLD_NEG_TTL_SEC:
        # 음성 캐시 구간엔 만료된 성공 캐시라도 재사용한다(없으면 None)
        return _krx_gold_frame(hist, quote) if hist else None

    if not hist_fresh:
        try:
            fetched = _fetch_krx_gold_history(symbol)
            if fetched:
                hist = fetched
                with _KRX_GOLD_LOCK:
                    ent["hist"], ent["hist_time"], ent["hist_day"] = fetched, now, today
        except Exception as e:
            logger.debug(f"[NAVER] KRX 금 시계열 조회 실패: {e}")

    if not quote_fresh:
        try:
            fetched_q = _fetch_krx_gold_quote(symbol)
            if fetched_q:
                quote = fetched_q
                with _KRX_GOLD_LOCK:
                    ent["quote"], ent["quote_time"] = fetched_q, now
        except Exception as e:
            logger.debug(f"[NAVER] KRX 금 현재가 조회 실패: {e}")

    if not hist:
        # 시계열이 없으면 지표도 등락률도 만들 수 없다 → 실패로 처리한다.
        #  (현재가만 살아 있어도 한 점으로는 표를 채울 수 없다)
        logger.warning("[NAVER] KRX 금 데이터 없음 — 시계열 조회 실패")
        with _KRX_GOLD_LOCK:
            ent["fail"] = now
        return None
    with _KRX_GOLD_LOCK:
        ent["fail"] = None
    return _krx_gold_frame(hist, quote)


def _krx_gold_frame(hist, quote):
    """[(date, close)] + 현재가 → 지수 소스 공통 스키마 DataFrame(오름차순).

    현재가는 장중 갱신되므로 같은 날짜 봉이 있으면 덮어쓰고, 없으면(장중 첫 체결이 아직
    시계열에 반영되기 전) 새 봉으로 덧붙인다.
    """
    if not hist:
        return None
    rows = dict(hist)                    # 같은 날짜가 중복 페이지로 들어와도 하나로 접힌다
    if quote and quote.get('date') is not None and quote.get('close'):
        rows[quote['date']] = quote['close']

    out = pd.DataFrame({'date': list(rows.keys()), 'close': list(rows.values())})
    out = out.sort_values('date', ascending=True).reset_index(drop=True)
    # 시·고·저는 네이버가 주지 않는다(0으로 내려온다) → 종가로 평탄화한다.
    for col in ['open', 'high', 'low']:
        out[col] = out['close']
    out['volume'] = 0.0
    out = out[['date', 'open', 'high', 'low', 'close', 'volume']].copy()
    out.attrs['source'] = 'NAVER'
    return out


# [추가] 해외 종목 tvDatafeed 조회 실패(빈 응답) 음성 캐시. 익명 웹소켓은 간헐 실패가 잦고
#  실패한 종목은 대체로 계속 실패하므로, 표 렌더링마다 재시도(전역 락 직렬화)로 UI가 지연되는 것을
#  막기 위해 일정 시간 재조회를 건너뛴다. (성공 결과는 api._get_cached_chart가 6시간/디스크 캐싱)
_TVDATAFEED_OVERSEAS_NEG_CACHE = {}   # code -> 실패 기록 시각(datetime)
_TVDATAFEED_NEG_TTL_SEC = 1800        # 30분
# [추가] 종목별로 한 번 확인된 TradingView 거래소 이름. 시세가 아니라 '어디에 상장돼 있는가'라는
#  메타데이터라 재사용해도 신선도에 영향이 없다(일봉은 매번 새로 받는다). KIS 경로의
#  exchange_cache와 같은 성격으로, 재조회 때 search_symbol과 헛거래소 순회를 없앤다.
_TVDATAFEED_EXCHANGE = {}             # code -> exchange 이름
# 종목당 tvDatafeed 폴백 총 소요 상한(초). 전역 락을 쥐고 도는 경로라 상한이 없으면
#  한 종목의 연결 타임아웃이 나머지 종목의 대기로 그대로 번진다.
TVDATAFEED_FETCH_BUDGET_SEC = 12.0

def fetch_overseas_daily_via_tvdatafeed(code, n_bars=260):
    """해외 개별 종목/ETF 일봉을 TradingView(tvDatafeed)로 조회한다(토스 캔들 실패 시 폴백).

    반환 스키마는 KIS/토스 해외 일봉과 동일: columns=['date','open','high','low','close','volume']
    (date=YYYYMMDD 문자열, 오름차순, 최대 250봉). 미설치/실패/빈 응답 시 None.
    거래소는 tvDatafeed.search_symbol로 자동 해석하고, 실패 시 미국 주요 거래소를 차례로 시도한다.
    """
    if not code:
        return None
    # 음성 캐시: 최근 실패한 종목은 TTL 동안 재조회를 건너뛴다.
    hit = _TVDATAFEED_OVERSEAS_NEG_CACHE.get(code)
    if hit is not None and (datetime.now() - hit).total_seconds() < _TVDATAFEED_NEG_TTL_SEC:
        return None

    tv = _get_tvdatafeed()
    if tv is None:
        return None
    try:
        from tvDatafeed import Interval
    except Exception:
        return None

    # 거래소 자동 해석. tvDatafeed 호출은 전역 락(_TVDATAFEED_LOCK)으로 직렬화되므로 여기서
    #  헛도는 호출 하나하나가 다른 종목의 대기 시간이 된다('데이터 수신' 단계가 마지막 몇
    #  종목에서 오래 멈추는 원인). 그래서 (1) 한 번 성공한 거래소를 기억해 재조회 때 검색과
    #  헛거래소 순회를 건너뛰고, (2) 검색이 성공하면 추측 거래소는 붙이지 않는다.
    #  ※ 기억하는 것은 '거래소 이름'(메타데이터)이지 시세가 아니다 — 일봉은 매번 새로 받는다.
    known = _TVDATAFEED_EXCHANGE.get(code)
    exchanges = [known] if known else []
    if not known:
        try:
            with _TVDATAFEED_LOCK:
                matches = tv.search_symbol(code) or []
            for m in matches:
                sym = str(m.get('symbol', '')).upper()
                exch = m.get('exchange') or ''
                if sym == code.upper() and exch and exch not in exchanges:
                    exchanges.append(exch)
        except Exception as e:
            logger.debug(f"[TVDATAFEED] {code} 심볼 검색 실패: {e}")
        # 검색이 거래소를 찾아냈으면 그게 정답이다. 못 찾았을 때만 미국 주요 거래소를 추측한다.
        if not exchanges:
            exchanges = ['NASDAQ', 'NYSE', 'AMEX']

    # [시간 예산] 익명 웹소켓은 응답이 없을 때 연결 타임아웃까지 통째로 기다린다. 이 호출은
    #  전역 락을 쥐고 있어 한 종목의 지연이 나머지 종목의 대기로 번지므로(메뉴 2 '데이터 수신'
    #  단계가 마지막 한두 종목에서 오래 멈추는 현상), 종목당 총 소요에 상한을 둔다.
    #  예산을 넘기면 남은 거래소를 포기한다 — 폴백은 '있으면 더 좋은' 보강이지 필수가 아니다.
    deadline = time.monotonic() + TVDATAFEED_FETCH_BUDGET_SEC

    df = None
    for i, exch in enumerate(exchanges):
        if time.monotonic() >= deadline:
            logger.debug(f"[TVDATAFEED] {code} 조회 예산 초과 — 남은 거래소 {exchanges[i:]} 생략")
            break
        # 익명 웹소켓 간헐 실패 대비 재시도는 '가장 유력한 거래소'에만 준다.
        #  모든 거래소에 2회씩 주면 없는 종목 하나가 최대 12회 조회 + 락 점유로 번진다.
        attempts = 2 if i == 0 else 1
        for attempt in range(attempts):
            try:
                with _TVDATAFEED_LOCK:
                    df = tv.get_hist(symbol=code, exchange=exch,
                                     interval=Interval.in_daily, n_bars=n_bars)
                if df is not None and not df.empty:
                    break
            except Exception as e:
                logger.debug(f"[TVDATAFEED] {code}@{exch} 조회 오류(attempt={attempt}): {e}")
                df = None
            if attempt < attempts - 1 and time.monotonic() < deadline:
                time.sleep(0.6)
            else:
                break
        if df is not None and not df.empty:
            _TVDATAFEED_EXCHANGE[code] = exch      # 다음 조회는 곧장 이 거래소로
            break
        if known:
            # 기억한 거래소가 더 이상 맞지 않으면(상장 이전 등) 잊고 다음 기회에 다시 찾는다
            _TVDATAFEED_EXCHANGE.pop(code, None)

    if df is None or df.empty:
        _TVDATAFEED_OVERSEAS_NEG_CACHE[code] = datetime.now()  # 실패 기록(TTL 재조회 억제)
        logger.debug(f"[TVDATAFEED] {code} 해외 일봉 데이터 없음")
        return None

    try:
        out = df.reset_index().rename(columns={'datetime': 'date'})
        for col in ['open', 'high', 'low', 'close', 'volume']:
            if col not in out.columns:
                out[col] = 0.0
        # date → YYYYMMDD 문자열(해외 일봉 스키마와 동일)
        out['date'] = out['date'].apply(
            lambda x: x.strftime('%Y%m%d') if hasattr(x, 'strftime') else str(x).replace('-', '')[:8])
        out = out[['date', 'open', 'high', 'low', 'close', 'volume']].copy()
        for c in ['open', 'high', 'low', 'close', 'volume']:
            out[c] = out[c].astype(float)
        out = out.drop_duplicates(subset=['date']).sort_values('date', ascending=True)
        out = out.reset_index(drop=True).tail(250).reset_index(drop=True)
        _TVDATAFEED_OVERSEAS_NEG_CACHE.pop(code, None)  # 성공 시 음성 캐시 해제
        return out
    except Exception as e:
        logger.debug(f"[TVDATAFEED] {code} 해외 일봉 스키마 변환 실패: {e}")
        return None

def clear_smart_money_cache():
    """스마트머니 수급 캐시 초기화 (수동 갱신용)"""
    _SMART_MONEY_CACHE.clear()

def check_smart_money_turnaround(code, is_overseas=False):
    """외국인/기관 수급 턴어라운드 및 쌍끌이 발생 여부 확인"""
    if is_overseas: return False, ""

    cached = _SMART_MONEY_CACHE.get(code, ttl=3600)  # 1시간(60분) 유지
    if isinstance(cached, tuple):
        return cached

    try:
        inv_list = api.get_investor_trend(code)
        if not inv_list or len(inv_list) < 3:
            # [최적화] '수급 데이터 없음'(ETF·미제공 종목)도 부정 결과로 캐시해
            #  API 마이크로캐시(5분) 만료마다 반복되던 무의미한 재조회를 차단 (예외는 일시 장애일 수 있어 미캐시)
            _SMART_MONEY_CACHE.set(code, (False, ""))
            return False, ""

        flag = False
        reason = ""
        
        for i in range(min(2, len(inv_list))): # [수정] 최근 5일 -> 당일(0)과 전일(1) 2일만 검사
            f_net = api.safe_int(inv_list[i].get('frgn_ntby_qty', 0))
            i_net = api.safe_int(inv_list[i].get('orgn_ntby_qty', 0))
            
            if f_net > 0 and i_net > 0:
                flag, reason = True, "쌍끌이 매수"
                break
                
            if i + 2 < len(inv_list):
                f_prev1 = api.safe_int(inv_list[i+1].get('frgn_ntby_qty', 0))
                f_prev2 = api.safe_int(inv_list[i+2].get('frgn_ntby_qty', 0))
                if f_net > 0 and f_prev1 < 0 and f_prev2 < 0:
                    flag, reason = True, "외국인 턴어라운드"
                    break
                    
                i_prev1 = api.safe_int(inv_list[i+1].get('orgn_ntby_qty', 0))
                i_prev2 = api.safe_int(inv_list[i+2].get('orgn_ntby_qty', 0))
                if i_net > 0 and i_prev1 < 0 and i_prev2 < 0:
                    flag, reason = True, "기관 턴어라운드"
                    break
        
        _SMART_MONEY_CACHE.set(code, (flag, reason))
        return flag, reason
    except Exception as e:
        logger.debug(f"Smart Money Check Error for {code}: {e}")
        return False, ""

def calculate_score(price=None, ema20=None, ema60=None, ema120=None, sar=None, rsi=None, adx=None, cci=None, obv_trend=None, macd=None, macd_signal=None, weights=None, smart_money=False, plus_di=None, minus_di=None, df=None, ind=None, ema_5=None, macd_hist=None, prev_macd_hist=None, prev_cci=None, vol_spike=False, vol_trend=False, w52_pos=None, mom_ret=None, mom_ret_1m=None, mom_ret_3m=None, trend_persist=None):
    """퀀트 멀티팩터 스코어링 모델 (10점 만점)"""
    if weights is None: weights = config.SCORING_WEIGHTS
    # [안전장치] 개별 룰의 가중치는 DB에 JSON 문자열로 저장된다. 보강 단계가 실패하면
    #  문자열이 그대로 들어와 아래 weights.get()이 AttributeError를 낸다. 점수 계산은
    #  매수·매도 판정 양쪽의 심장이라, 여기서 죽으면 그 종목이 판정에서 통째로 빠진다.
    if isinstance(weights, str):
        try:
            weights = json.loads(weights)
        except Exception:
            weights = config.SCORING_WEIGHTS
    if not isinstance(weights, dict): weights = config.SCORING_WEIGHTS

    # r_* 는 '각 팩터의 세부 항목 기본배점 합(=설계상 만점)' 대비 사용자 가중치의 스케일 배수다.
    # 분모는 세부항목 기본배점 합(고정값)이며 분자만 가중치로 바뀐다.
    #   예) TREND 세부항목 기본합=4.0. 가중치 4.0이면 r_trend=1.0 → 추세 팩터가 4.0점 만점.
    # 4개 팩터 기본배점: 추세4.0 + 모멘텀2.5 + 강도1.5 + 시너지2.0 = 10.0 (총점 10점)
    r_trend = weights.get("TREND", 4.0) / 4.0
    r_mom = weights.get("MOMENTUM", 2.5) / 2.5
    r_str = weights.get("STRENGTH", 1.5) / 1.5
    r_syn = weights.get("SYNERGY", 2.0) / 2.0

    score = 0
    details = []
    
    # [Fix] df가 전달되지 않는 백테스팅 환경을 위한 변수 초기화
    vol_spike_flag = vol_spike
    vol_trend_flag = vol_trend

    if df is not None and ind is not None:
        if not df.empty: price = float(df.iloc[-1]['close'])
        ema20 = ind.get('ema_20')
        ema60 = ind.get('ema_60')
        ema120 = ind.get('ema_120')
        sar = ind.get('psar')
        rsi = ind.get('rsi')
        adx = ind.get('adx')
        cci = ind.get('cci')
        obv_trend = ind.get('obv_trend')
        macd = ind.get('macd')
        macd_signal = ind.get('macd_signal')
        if plus_di is None: plus_di = ind.get('plus_di')
        if minus_di is None: minus_di = ind.get('minus_di')
        
        # [추가] 인자로 넘어오지 않은 세부 지표들을 ind 딕셔너리에서 직접 추출하여 일관성(SSOT) 확보
        if ema_5 is None: ema_5 = ind.get('ema_5')
        if prev_cci is None: prev_cci = ind.get('prev_cci')
        if macd_hist is None: macd_hist = ind.get('macd_hist')
        if prev_macd_hist is None: prev_macd_hist = ind.get('prev_macd_hist')

    import numpy as np
    
    if df is not None and not df.empty:
        # [Early] 선행 지표 동적 계산
        if ema_5 is None:
            ema_5 = df['close'].ewm(span=config.INDICATOR_PARAMS.get('EMA_SHORT', 5), adjust=False).mean().iloc[-1]
            
        if macd is not None and macd_signal is not None and (macd_hist is None or prev_macd_hist is None):
            fast = config.INDICATOR_PARAMS.get('MACD_FAST', 12)
            slow = config.INDICATOR_PARAMS.get('MACD_SLOW', 26)
            sig = config.INDICATOR_PARAMS.get('MACD_SIGNAL', 9)
            macd_series = df['close'].ewm(span=fast, adjust=False).mean() - df['close'].ewm(span=slow, adjust=False).mean()
            signal_series = macd_series.ewm(span=sig, adjust=False).mean()
            hist_series = macd_series - signal_series
            if len(hist_series) > 0: macd_hist = hist_series.iloc[-1]
            if len(hist_series) > 1: prev_macd_hist = hist_series.iloc[-2]

        if plus_di is None or minus_di is None:
            try:
                high_diff = df['high'].diff()
                low_diff = df['low'].diff()
                pos_dm = np.where((high_diff > 0) & (high_diff > -low_diff), high_diff, 0.0)
                neg_dm = np.where((low_diff < 0) & (-low_diff > high_diff), -low_diff, 0.0)
                tr1 = df['high'] - df['low']
                tr2 = (df['high'] - df['close'].shift()).abs()
                tr3 = (df['low'] - df['close'].shift()).abs()
                tr = pd.concat([tr1, tr2, tr3], axis=1).max(axis=1)
                adx_period = config.INDICATOR_PARAMS.get('ADX_PERIOD', 14)
                atr = tr.ewm(alpha=1/adx_period, adjust=False).mean()
                atr_val = atr.iloc[-1]
                if atr_val == 0 or np.isnan(atr_val): atr_val = 1.0 # 0 나누기 방어
                plus_di = 100 * pd.Series(pos_dm).ewm(alpha=1/adx_period, adjust=False).mean().iloc[-1] / atr_val
                minus_di = 100 * pd.Series(neg_dm).ewm(alpha=1/adx_period, adjust=False).mean().iloc[-1] / atr_val
            except Exception: pass

        if prev_cci is None and len(df) > 20:
            try:
                window = config.INDICATOR_PARAMS.get('CCI_WINDOW', 20)
                # [최적화] 동일 수식의 벡터화 버전에 위임 (rolling.apply 대비 100배 이상 빠름, 결과 동일)
                cci_series = indicators.get_cci_full_series(df, window)
                if len(cci_series) > 1: prev_cci = cci_series.iloc[-2]
            except Exception: pass

        vol_ma_period = config.INDICATOR_PARAMS.get('VOLUME_MA_PERIOD', 20)
        if len(df) >= vol_ma_period:
            vol_ma20 = df['volume'].rolling(window=vol_ma_period).mean().iloc[-1]
            vol_ma5 = df['volume'].rolling(window=5).mean().iloc[-1]

            if vol_ma5 > vol_ma20:
                vol_trend_flag = True

            if not vol_spike_flag:
                vol_ratio = config.INDICATOR_PARAMS.get('VOLUME_SPIKE_RATIO', 2.0)
                vol = df['volume'].iloc[-1]
                opn = df['open'].iloc[-1]
                if price is not None and vol_ma20 > 0 and vol >= (vol_ma20 * vol_ratio) and price > opn:
                    vol_spike_flag = True

        # [추세추종] 가격 모멘텀 팩터 입력값 동적 계산 (백테스트는 사전계산 값을 인자로 전달)
        try:
            if w52_pos is None and price is not None and 'high' in df.columns and 'low' in df.columns:
                # 화면(표·상세·지수)과 동일한 365일 창. 백테스트는 w52_pos를 인자로 넘기므로
                # 여기 오지 않고, 와도 창이 비어 _w52_band가 기존 tail(250)으로 폴백한다.
                h52, l52 = _w52_band(df)
                if h52 > l52:
                    w52_pos = (price - l52) / (h52 - l52) * 100
            if mom_ret is None and price is not None:
                mom_lb = config.INDICATOR_PARAMS.get('MOMENTUM_LOOKBACK', 126)
                if len(df) > mom_lb:
                    past_close = float(df['close'].iloc[-(mom_lb + 1)])
                    if past_close > 0:
                        mom_ret = (price / past_close - 1) * 100
            # [추세추종] 다중 기간 모멘텀(1·3개월) — 6개월 가격 모멘텀 가점의 정합 게이트 입력
            if mom_ret_1m is None and price is not None:
                lb = config.INDICATOR_PARAMS.get('MOMENTUM_LOOKBACK_1M', 21)
                if len(df) > lb:
                    past_close = float(df['close'].iloc[-(lb + 1)])
                    if past_close > 0:
                        mom_ret_1m = (price / past_close - 1) * 100
            if mom_ret_3m is None and price is not None:
                lb = config.INDICATOR_PARAMS.get('MOMENTUM_LOOKBACK_3M', 63)
                if len(df) > lb:
                    past_close = float(df['close'].iloc[-(lb + 1)])
                    if past_close > 0:
                        mom_ret_3m = (price / past_close - 1) * 100
            # [추세추종] 추세 지속 이력 — 최근 N일 중 종가가 60일선 위였던 비율(%).
            #   오래 유지된 추세가 계속될 확률이 높다는 지속성 원칙의 직접 측정치.
            if trend_persist is None:
                persist_lb = config.INDICATOR_PARAMS.get('TREND_PERSIST_LOOKBACK', 120)
                if len(df) >= persist_lb:
                    ema60_series = df['close'].ewm(span=60, adjust=False).mean()
                    above = df['close'].tail(persist_lb).to_numpy(dtype=float) > ema60_series.tail(persist_lb).to_numpy(dtype=float)
                    trend_persist = float(above.mean() * 100)
        except Exception:
            pass

    if price is None:
        return 0.0, details

    # 1. Trend Factor (4.0점) — MA 군집(상한 2.0) + 추세 지속 이력 0.5 + MACD 0선 0.5 + MACD GC/확산 0.5 + SAR 0.5
    # [개선 #2] 이동평균선(EMA) 기반 신호들은 상호 상관(collinearity)이 매우 높아
    #          정배열 상승장에서 동시 충족되며 추세추종으로 점수가 편향됨.
    #          → MA 포지션 점수 합계를 상한으로 제한하고, 나머지는
    #            상대적으로 독립적인 확인 신호(MACD 0선/추세확산, SAR)로 채워
    #            'MA 군집 단독'으로는 TREND 만점을 받지 못하도록 재구성.
    # [추세추종] MA 상한 2.5 → 2.0: 절감분 0.5는 '추세 지속 이력'(아래) 가점으로 이관.
    #          현재 상태(정배열)만 보는 MA 군집과 달리 '얼마나 오래 유지됐는가'를 재는
    #          독립 신호로, 갓 골든크로스한 미검증 추세와 장기 지속 추세를 점수로 구분한다.
    ma_trend_score = 0.0
    ma_details = []
    if ema20 is not None and price > ema20:
        s = round(0.5 * r_trend, 2); ma_trend_score += s; ma_details.append(f"EMA: 현재가 > 20일선 (+{s:.2f})")
    if ema20 is not None and ema60 is not None and ema120 is not None and ema20 > ema60 and ema60 > ema120:
        s = round(1.0 * r_trend, 2); ma_trend_score += s; ma_details.append(f"EMA: 20/60/120 정배열 (+{s:.2f})")
    if ema20 is not None and ema_5 is not None and ema_5 > ema20:
        s = round(0.5 * r_trend, 2); ma_trend_score += s; ma_details.append(f"EMA: 5일선 > 20일선 (+{s:.2f})")
    if ema20 is not None and ema60 is not None:
        if ema20 <= ema60 and price > ema60:
            s = round(0.5 * r_trend, 2); ma_trend_score += s; ma_details.append(f"EMA: 60일선 돌파 [초기] (+{s:.2f})")
        elif ema_5 is not None and price > ema_5 and ema_5 > ema20 and ema20 > ema60:
            s = round(0.5 * r_trend, 2); ma_trend_score += s; ma_details.append(f"EMA: 단기 급등 추세 (+{s:.2f})")
    if ema120 is not None and price > ema120:
        s = round(0.5 * r_trend, 2); ma_trend_score += s; ma_details.append(f"EMA: 장기 지지(현재가>120일선) (+{s:.2f})")

    # [개선 #2] MA 포지션 점수 상한 적용 (상관 신호의 과대 가점 방지)
    ma_cap = round(2.0 * r_trend, 2)
    details.extend(ma_details)
    if ma_trend_score > ma_cap:
        details.append(f"[상한] EMA 군집 신호 상한 적용 ({ma_cap:.2f})")
        ma_trend_score = ma_cap
    score += ma_trend_score

    # [추세추종] 추세 지속 이력: 최근 TREND_PERSIST_LOOKBACK일 중 종가가 60일선 위였던
    #   비율이 기준(TREND_PERSIST_MIN%) 이상일 때 가점. 순간 상태가 아닌 '유지된 기간'을
    #   재는 지속성 신호로, MA 상한 축소분(0.5)을 이관받았다. (NaN은 비교식에서 자동 False)
    if trend_persist is not None and trend_persist >= config.INDICATOR_PARAMS.get('TREND_PERSIST_MIN', 70):
        s = round(0.5 * r_trend, 2); score += s
        details.append(f"추세 지속: 최근 {config.INDICATOR_PARAMS.get('TREND_PERSIST_LOOKBACK', 120)}일 중 {trend_persist:.0f}% 60일선 위 (+{s:.2f})")

    # [개선 #2] MACD 0선 위(추세 확립): MA 포지션과 독립적인 추세 확인 신호
    if macd is not None and macd > 0:
        s = round(0.5 * r_trend, 2); score += s; details.append(f"MACD: 0선 위 (추세 확립) (+{s:.2f})")

    # [수정] 단순 MACD > Signal 상태 유지가 아닌, 신규 골든크로스 또는 0선 위 확산 추세일 때만 점수 부여 (인플레이션 방지)
    if macd_hist is not None and prev_macd_hist is not None:
        if macd_hist > 0 and prev_macd_hist <= 0:
            s = round(0.5 * r_trend, 2); score += s; details.append(f"MACD: 신규 골든크로스 (+{s:.2f})")
        elif macd is not None and macd > 0 and macd_hist > prev_macd_hist and macd_hist > 0:
            s = round(0.5 * r_trend, 2); score += s; details.append(f"MACD: 상승 추세 확산 (+{s:.2f})")

    if sar is not None and price > sar:
        s = round(0.5 * r_trend, 2); score += s; details.append(f"SAR: 상승 추세 (+{s:.2f})")
    
    # 2. Momentum Factor (2.5점) — RSI 1.0 + CCI 0.5 + DMI 0.5 + 가격 모멘텀 0.5
    score_rsi_mid = config.INDICATOR_PARAMS.get('SCORE_RSI_MID', 50)
    score_rsi_strong = config.INDICATOR_PARAMS.get('SCORE_RSI_STRONG', 60)
    score_rsi_overheat = config.INDICATOR_PARAMS.get('SCORE_RSI_OVERHEAT', 80)
    score_rsi_rebound = config.INDICATOR_PARAMS.get('SCORE_RSI_REBOUND', 40)

    # [추세추종] 반등 성격 가점(RSI 상승 여력, CCI 과매도 탈출)은 추세 구조(주가>60일선) 위에서만 인정.
    #   USE_MEAN_REVERSION 기본 OFF 기조와 정합: '상승 추세 내 눌림목 회복'만 가점하고,
    #   추세 없는 약세 종목의 단순 기술적 반등이 매수 점수를 밀어올리는 것을 차단.
    in_uptrend_structure = ema60 is not None and price > ema60

    if rsi is not None:
        # [Fix] RSI 상단 제한(75) 해제로 과열 구간에서도 기본 강세 점수는 유지 (스코어 클리프 방지)
        if rsi >= score_rsi_mid:
            s = round(0.5 * r_mom, 2); score += s; details.append(f"RSI: 강세 구간 (+{s:.2f})")
            # [개선 #6] 과열 구간(>=80)에서는 추가 '모멘텀 확장' 가점을 동결하여
            #          이미 과열된 종목에 고점매수 신호가 강화되는 것을 방지.
            if score_rsi_strong <= rsi < score_rsi_overheat:
                s = round(0.5 * r_mom, 2); score += s; details.append(f"RSI: 모멘텀 확장 (+{s:.2f})")
        elif score_rsi_rebound <= rsi < score_rsi_mid and in_uptrend_structure:
            s = round(0.5 * r_mom, 2); score += s; details.append(f"RSI: 상승 여력 구간(추세 내 눌림) (+{s:.2f})")

    cci_lower = config.INDICATOR_PARAMS.get('CCI_LOWER', -100)
    cci_strong = config.INDICATOR_PARAMS.get('SCORE_CCI_STRONG', 0)
    if cci is not None:
        # [개선] 'CCI 모멘텀 심화'는 'CCI 상승 추세'·RSI 강세와 상관이 높아 정보 가치가 낮으므로 폐지하고,
        #        그 0.5점을 아래 '가격 모멘텀' 항목으로 이관. CCI는 상승/과매도탈출 중 하나만 인정(최대 0.5).
        if cci > cci_strong:
            s = round(0.5 * r_mom, 2); score += s; details.append(f"CCI: 상승 추세 (+{s:.2f})")
        elif in_uptrend_structure and prev_cci is not None and prev_cci <= cci_lower and cci > cci_lower:
            s = round(0.5 * r_mom, 2); score += s; details.append(f"CCI: 과매도권 탈출(추세 내) (+{s:.2f})")

    if plus_di is not None and minus_di is not None and plus_di > minus_di:
        s = round(0.5 * r_mom, 2); score += s; details.append(f"DMI: +DI > -DI 크로스 (+{s:.2f})")

    # [추세추종] 가격 모멘텀 (절대 모멘텀 + 52주 신고가 근접) — '강한 종목을 매수하라' 핵심 팩터.
    #   6개월(MOMENTUM_LOOKBACK) 수익률이 양수이고 52주 위치가 기준(MOMENTUM_W52_NEAR) 이상인
    #   주도주에만 가점하여, 지표만 좋은 바닥권 종목과 신고가 랠리 종목을 점수로 구분한다.
    # [추세추종] 다중 기간 모멘텀 정합: 1·3개월 수익률이 명시적으로 음수면 6개월 가점을 보류.
    #   6개월 수치만 좋고 최근 1~3개월이 꺾인 '식어가는 추세'의 고점 진입을 걸러낸다.
    #   단기 수익률을 알 수 없으면(None/NaN) 게이트를 적용하지 않는다(fail-open, 기존 호출자 호환).
    mom_align_ok = True
    if mom_ret_1m is not None and mom_ret_1m <= 0: mom_align_ok = False
    if mom_ret_3m is not None and mom_ret_3m <= 0: mom_align_ok = False

    mom_w52_near = config.INDICATOR_PARAMS.get('MOMENTUM_W52_NEAR', 80)
    if mom_ret is not None and mom_ret > 0 and w52_pos is not None and w52_pos >= mom_w52_near:
        if mom_align_ok:
            s = round(0.5 * r_mom, 2); score += s
            details.append(f"가격 모멘텀: 6개월 +{mom_ret:.1f}% & 52주 위치 {w52_pos:.0f}% (+{s:.2f})")
        else:
            details.append("가격 모멘텀 보류: 단기(1·3개월) 모멘텀 이탈 (식어가는 추세)")

    # 3. Strength & Volume Factor (1.5점)
    adx_min = config.INDICATOR_PARAMS.get('SCORE_ADX_MIN', 20)
    if adx is not None and adx >= adx_min:
        s = round(0.5 * r_str, 2); score += s; details.append(f"ADX: 추세 형성 (+{s:.2f})")
        
    if vol_spike_flag or vol_trend_flag:
        s = round(0.5 * r_str, 2); score += s
        if vol_spike_flag:
            details.append(f"VOL: 거래량 폭증/양봉 (+{s:.2f})")
        else:
            details.append(f"VOL: 거래량 추세 상승 (+{s:.2f})")
        
    if obv_trend or smart_money:
        s = round(0.5 * r_str, 2); score += s; details.append(f"수급: OBV/SM 개선 (+{s:.2f})")

    # 4. Synergy Bonus (2.0점)
    # [수정] 시너지 보너스를 확산(macd_hist > prev_macd_hist) 조건 단독에서 완화. MACD가 시그널 위에 있고(macd_hist > 0) 심한 축소가 아닐 때 유지하여 단기 노이즈로 인한 2.0점 증발 방어
    is_macd_expanding = False
    if macd_hist is not None and prev_macd_hist is not None:
        is_macd_expanding = (macd_hist > 0 and macd_hist >= prev_macd_hist * 0.8)

    # [개선] 시너지 구조 게이트: 장기 역배열(주가<120일선) '이고' MACD가 0선 아래인 '지하실'에서
    #   히스토그램만 골든크로스한 '데드캣 바운스'가 시너지 2.0점을 쓸어담는 것을 차단.
    #   ADX는 방향을 모르고(폭락장도 높음), is_macd_expanding은 0선 아래 크로스도 True가 되므로
    #   단독으로는 역추세 급반등을 걸러내지 못한다. macd>0(추세 확립) 또는 주가>120일선(장기 구조
    #   회복) 중 하나라도 충족할 때만 시너지를 인정한다. (단독 Trend 가점의 macd>0 게이트와 정합)
    synergy_structure_ok = not (macd is not None and macd <= 0 and (ema120 is None or price <= ema120))

    if ema60 is not None and price > ema60 and is_macd_expanding and (adx is not None and adx >= adx_min) and synergy_structure_ok:
        s = round(1.0 * r_syn, 2)
        score += s
        details.append(f"추세 시작: 주가>60일선+MACD확산+ADX 20↑ (+{s:.2f})")

    # Momentum Thrust
    if is_macd_expanding and (rsi is not None and rsi >= score_rsi_strong) and obv_trend and synergy_structure_ok:
        s = round(1.0 * r_syn, 2)
        score += s
        details.append(f"모멘텀 폭발: MACD확산+RSI 60↑+OBV (+{s:.2f})")

    # 5. 추세 악화 감점 (Deterioration Penalty)
    # [개선 #1] 기존 스코어는 '가산 전용'이라 추세가 꺾여도 후행 지표(EMA 정배열 등)가
    #          점수를 떠받쳐 매도 신호(SELL_SCORE 미달)가 지연되는 구조적 약점이 있음.
    #          명확한 하락 반전 신호에 대해 감점을 부여하여 점수가 악화를 적시 반영하도록 보정.
    penalty = 0.0
    if macd is not None and macd_signal is not None and macd < macd_signal:
        p = round(0.5 * r_trend, 2); penalty -= p; details.append(f"감점: MACD 데드크로스 (-{p:.2f})")
    if macd_hist is not None and prev_macd_hist is not None and macd_hist < 0 and macd_hist < prev_macd_hist:
        p = round(0.5 * r_mom, 2); penalty -= p; details.append(f"감점: MACD 하락 가속(0선 이하 확대) (-{p:.2f})")
    if plus_di is not None and minus_di is not None and minus_di > plus_di:
        p = round(0.5 * r_str, 2); penalty -= p; details.append(f"감점: -DI 우위(매도세 강화) (-{p:.2f})")

    score += penalty
    if score < 0:
        score = 0.0

    return round(score, 2), details

# [추가] 시장 지수 데이터 공유 캐시
# get_market_regime은 자동매매/체결감시/예약감시/텔레그램 등 여러 스레드에서 동시에 호출되며,
# 매 호출마다 2년치 지수차트(inquire-daily-indexchartprice)를 재조회하면 모의투자(2 TPS) 서버에
# 요청 폭주(EGW00201)가 발생한다. 이를 막기 위해 아래 3중 방어를 적용한다.
#   1) single-flight : 동일 지수의 동시 캐시 미스 시 1개 스레드만 실제 조회(스탬피드 차단).
#   2) negative cache: 조회 실패(빈 결과)도 짧게 기록해 폭주 중 재조회 폭발을 억제하고,
#                      직전 정상 데이터를 stale 폴백으로 보존한다.
#   3) stale-while-revalidate: TTL 만료 시 옛 값을 즉시 반환하고 백그라운드 1스레드로만 갱신
#                      → 5분 주기 캐시 절벽/블로킹 제거.
# 단, stale 서빙에는 상한을 둔다: 시장 기준일이 바뀌었거나(_current_market_day) 15분을 넘긴
# 데이터는 즉시 반환하지 않고 동기 재조회한다. 상한이 없던 시절, 장기 구동 시 '다음날 어제 값 +
# 등락률 0%'가 그대로 표시되고 한 번 더 조회해야 갱신되는 문제가 있었다.
_INDEX_DATA_CACHE = {}          # {market_type: {'df': df, 'time': ts, 'fail_time': ts, 'day': 'YYYYMMDD'}}
_INDEX_DATA_CACHE_LOCK = threading.Lock()   # 캐시 딕셔너리 보호
_INDEX_DATA_CACHE_TTL = 300     # 정상 데이터 유효시간(5분) - 장중 국면 판단에는 충분히 신선함
_INDEX_DATA_NEG_TTL = 30        # 실패 후 재조회 억제시간(초) - 폭주 자기증식 차단
# [Fix] stale-while-revalidate 서빙 상한. 이 시간을 넘긴 데이터는 '옛 값 즉시 반환'을 하지 않고
#  동기 재조회한다. (몇 시간~하루를 쉰 뒤 조회하면 첫 화면이 옛 지수로 나오고 재조회해야
#  갱신되던 문제 차단 — 그 정도로 오래된 값은 '빠른 응답'보다 '정확한 값'이 우선이다)
_INDEX_DATA_MAX_STALE = 900     # 15분

_INDEX_FETCH_LOCKS = {}         # {market_type: Lock} single-flight 동기 조회 잠금
_INDEX_FETCH_LOCKS_GUARD = threading.Lock()
_INDEX_REFRESH_INFLIGHT = {}    # {market_type: 기동 시각} 백그라운드 재검증 진행 중
_INDEX_REFRESH_GUARD = threading.Lock()
# [Fix] 재검증 워커가 죽거나(스레드 생성 실패 등) 비정상적으로 오래 걸리면 inflight 표시가 남아
#  이후 모든 갱신이 영구히 막혔다(재시작 전까지 지수 고착). 이 시간이 지나면 재기동을 허용한다.
_INDEX_REFRESH_STUCK_SEC = 120

def _index_cache_enabled():
    """테스트(pytest) 환경에서는 모킹된 지수 데이터가 캐시에 고착되지 않도록 캐시를 비활성화한다."""
    return "pytest" not in sys.modules and "PYTEST_CURRENT_TEST" not in os.environ

def _current_market_day():
    """국내 시장 기준일(YYYYMMDD, 휴장일이면 직전 거래일). 캐시 데이터의 '날짜 세대' 판정용."""
    try:
        return api.market_today(False)
    except Exception:
        return datetime.now().strftime('%Y%m%d')

# [재진입 가드 2026-08-19] 같은 스레드가 같은 market_type 조회 안에서 다시 들어오면
#  아래 single-flight 락을 자기 자신이 다시 잡아 **영구 교착**한다. 실제로 거래량 보강이
#  api.get_chart_data를 타고 이 함수로 되돌아와 토스 모드 코스피200·코스닥150이 멈췄다.
#  산식이 아니라 호출 그래프의 문제라 언제든 재발할 수 있으므로, 교착 대신 '보강 없이
#  진행'으로 빠지고 경고를 남긴다 — 멈추는 것보다 시끄러운 편이 낫다.
_INDEX_FETCH_INPROGRESS = threading.local()


def _index_fetch_reentrant(market_type):
    """이 스레드가 이미 같은 지수를 조회하는 중인가."""
    return market_type in getattr(_INDEX_FETCH_INPROGRESS, "types", ())


def _get_index_fetch_lock(market_type):
    """market_type별 single-flight 잠금을 반환한다(없으면 생성)."""
    with _INDEX_FETCH_LOCKS_GUARD:
        lock = _INDEX_FETCH_LOCKS.get(market_type)
        if lock is None:
            lock = threading.Lock()
            _INDEX_FETCH_LOCKS[market_type] = lock
        return lock

def _lookup_index_cache(market_type):
    """캐시 상태를 (status, df)로 반환한다.
    - 'fresh'   : TTL 이내 정상 데이터 → 즉시 반환
    - 'suppress': 최근 실패(음성 TTL 이내) → 재조회 없이 직전 정상 데이터(또는 None) 반환
    - 'stale'   : TTL 경과한 정상 데이터 존재 → 즉시 반환 후 백그라운드 갱신
    - 'expired' : 날짜(거래일)가 바뀌었거나 너무 오래된 데이터 → 동기 조회 필요(실패 시 폴백용)
    - 'miss'    : 쓸만한 데이터 없음 → 동기 조회 필요
    """
    now = time.time()
    today = _current_market_day()   # 네트워크(휴장일 조회) 가능 → 캐시 락 밖에서 계산
    with _INDEX_DATA_CACHE_LOCK:
        entry = _INDEX_DATA_CACHE.get(market_type)
        if not entry:
            return 'miss', None
        df = entry.get('df')
        has_good = df is not None and not df.empty
        # 최근 실패가 음성 TTL 이내면 재조회를 억제(폭주 차단)하고 가진 데이터를 반환한다.
        # 단, 서빙할 정상 데이터가 있을 때만 억제한다. 한 번도 성공한 적 없으면(df=None)
        # 억제해도 '-'만 반복되고 사용자 재시도가 무의미해지므로, 'miss'로 넘겨 실제 재조회를
        # 허용한다(tvDatafeed 익명 조회는 간헐 실패가 있어 재시도로 대부분 성공).
        if has_good and (now - entry.get('fail_time', 0)) < _INDEX_DATA_NEG_TTL:
            return 'suppress', df
        if has_good:
            age = now - entry.get('time', 0)
            # [Fix] 시장 기준일이 바뀌면(하루 이상 구동 후 다음날) 옛 데이터를 stale로 즉시
            #  서빙하면 안 된다 — 어제 종가·등락률 0%가 그대로 화면에 남는다. 동기 재조회한다.
            if entry.get('day') == today:
                if age < _INDEX_DATA_CACHE_TTL:
                    return 'fresh', df
                if age < _INDEX_DATA_MAX_STALE:
                    return 'stale', df
            return 'expired', df
        return 'miss', None

def _store_index_cache(market_type, df):
    """조회 결과를 캐시에 반영한다. 정상 데이터는 갱신하고, 실패(빈 결과)는
    직전 정상 데이터를 보존한 채 실패 시각만 기록(음성 캐시)한다."""
    if not _index_cache_enabled():
        return
    now = time.time()
    day = _current_market_day()
    with _INDEX_DATA_CACHE_LOCK:
        entry = _INDEX_DATA_CACHE.get(market_type) or {'df': None, 'time': 0, 'fail_time': 0, 'day': None}
        if df is not None and not df.empty:
            entry['df'] = df
            entry['time'] = now
            entry['day'] = day      # 데이터의 '거래일 세대' — 날짜가 바뀌면 stale 서빙 금지
            entry['fail_time'] = 0
        else:
            # 실패: 직전 정상 df/time/day는 보존, 실패 시각만 갱신
            entry['fail_time'] = now
        _INDEX_DATA_CACHE[market_type] = entry

def _trigger_async_refresh(market_type):
    """stale 데이터 제공 후 백그라운드에서 1스레드로만 캐시를 재검증한다."""
    if not _index_cache_enabled():
        return
    now = time.time()
    with _INDEX_REFRESH_GUARD:
        started = _INDEX_REFRESH_INFLIGHT.get(market_type)
        # 이미 갱신 중이면 중복 기동 방지. 단, 워커가 죽어 표시만 남은 경우를 대비해
        # 일정 시간(_INDEX_REFRESH_STUCK_SEC)이 지나면 재기동을 허용한다(영구 고착 방지).
        if started is not None and (now - started) < _INDEX_REFRESH_STUCK_SEC:
            return
        _INDEX_REFRESH_INFLIGHT[market_type] = now

    def _release():
        with _INDEX_REFRESH_GUARD:
            if _INDEX_REFRESH_INFLIGHT.get(market_type) == now:  # 내 기동분만 해제
                _INDEX_REFRESH_INFLIGHT.pop(market_type, None)

    def _worker():
        try:
            fresh = _fetch_domestic_index_data(market_type)
            _store_index_cache(market_type, fresh)
        except Exception as e:
            logger.debug(f"[MARKET_INDEX_DEBUG] {market_type} 비동기 재검증 실패: {e}")
            _store_index_cache(market_type, None)  # 실패 기록(음성 캐시)
        finally:
            _release()

    try:
        threading.Thread(target=_worker, name=f"IndexRefresh-{market_type}", daemon=True).start()
    except Exception as e:
        # 스레드 생성 실패(라즈베리파이 메모리 압박 등). inflight 표시를 반드시 되돌린다.
        logger.warning(f"[MARKET_INDEX] {market_type} 재검증 스레드 기동 실패: {e}")
        _release()

def _fetch_domestic_index_data(market_type):
    """국내 지수 데이터를 실제 조회한다(캐시 미적용).

    폴백 체인(각 단계는 '데이터 없음/부족(< ma_period)'일 때만 다음으로 내려간다):
      - 모드 1/2(KIS): KIS API → tvDatafeed → yfinance
      - 모드 3(토스):  토스 시장지표(코스피·코스닥) → tvDatafeed → yfinance   (KIS 미사용)
    토스 시장지표(/api/v1/market-indicators, API 1.2.4)는 KRX 공식 지수를 인증된 채널로 주므로
    익명 tvDatafeed(간헐 빈 응답)보다 안정적이다. 다만 코스피200·코스닥150은 심볼 카탈로그에
    없어 토스로 조회할 수 없다 → 그 둘은 종전대로 tvDatafeed가 1순위다.
    tvDatafeed는 4종 지수(코스피·코스닥·코스피200·코스닥150) 모두 지원하며 KRX 정확·당일 종가를
    준다. yfinance(^KS11/^KQ11 등)는 최신 거래일 종가를 NaN으로 주는 일이 잦아 최후 폴백으로 둔다.
    """
    # [추가] 코스피200 선물 (주간 K200FUT_F / 야간 K200FUT_CM): KIS 전용, 폴백 없음.
    #  세션별 market_type을 분리해 TTL 캐시가 주간/야간 데이터를 섞지 않게 한다.
    if market_type in ("K200FUT_F", "K200FUT_CM"):
        # [KIS 전용 · 2026-08-25 재확인] KRX 공식 경로는 **마감 후 확정 봉만** 준다. 선물은
        #  세션이 하루를 거의 덮어(주간 09:00~15:45 · 야간 18:00~익일 06:00) 장중 내내 최대
        #  하루 묵은 값이 나온다 — 야간장 01:47 실측에서 KIS 실시간 1,034.85 vs KRX 확정
        #  1,074.55 로 40포인트 벌어졌고 등락률 기준도 어긋났다. 토스는 선물을 제공하지 않고
        #  tvDatafeed 는 심볼 검색이 막혀 있어 살릴 방법이 없다 → **토스 모드에서는 아예
        #  목록에서 뺀다**(market.blocked_kis_only_indices). 여기서 KRX 로 폴백하지 않는 이유도
        #  같다: 부정확하거나 오해를 부르는 값은 보여 주지 않는다.
        if config.session.is_toss:
            return None
        div = "F" if market_type == "K200FUT_F" else "CM"
        try:
            iscd = api.get_k200_futures_front_code()
            if not iscd:
                return None
            df = api.get_k200_futures_chart(div, iscd)
            return df if df is not None and not df.empty else None
        except Exception as e:
            logger.debug(f"[MARKET_INDEX_DEBUG] {market_type} K200선물 조회 실패: {e}")
            return None

    kis_code = "0001"
    yf_ticker = "^KS11"

    if market_type == "KOSDAQ":
        kis_code = "1001"
        yf_ticker = "^KQ11"
    elif market_type == "KOSPI200":
        kis_code = "2001"
        yf_ticker = "^KS200"
    elif market_type == "KOSDAQ150":
        kis_code = "2203"
        yf_ticker = "^KQ150"
    elif market_type == "VKOSPI":
        # V코스피200(코스피200 변동성지수): KIS 업종코드 0503 전용.
        #  yfinance/tvDatafeed 미제공이고, KRX 공식은 **마감 후 확정 봉만** 주므로 장중에는
        #  직전 확정치가 나온다. 변동성지수를 묵은 값으로 보는 것은 안 보는 것보다 낫지 않다 —
        #  하필 변동성이 튀는 순간에 화면만 조용하기 때문이다. 그래서 폴백하지 않고,
        #  토스 모드에서는 목록에서도 뺀다(market.blocked_kis_only_indices). 선물과 같은 정책.
        if config.session.is_toss:
            return None
        kis_code = "0503"
        yf_ticker = None

    # 지수 데이터는 국면 판단(이중 EMA의 느린 기간)과 시장 필터링(SMA, MARKET_FILTER_MA)이 함께
    #  사용하므로 두 기간 중 큰 값을 '충분성' 기준으로 삼는다(부족하면 다음 소스로 폴백).
    ma_period = max(config.MARKET_REGIME_PARAMS.get("REGIME_EMA_SLOW", 41),
                    config.MARKET_REGIME_PARAMS.get("REGIME_MA_PERIOD", 5))
    if getattr(config, 'USE_MARKET_FILTER', True):
        ma_period = max(ma_period, getattr(config, 'MARKET_FILTER_MA', 80))

    def _insufficient(d):
        return d is None or d.empty or len(d) < ma_period

    df = None

    # KRX 공식 확정 봉 — 지표의 뼈대. 실시간 소스를 **대체하지 않고** 이력만 바꾼다(맨 아래 병합).
    #  ① 코스피200·코스닥150은 여태 tvDatafeed 단일 소스였다(야후·FDR에 티커가 없다) →
    #    간헐 빈 응답이 곧 시장필터·국면 판정의 공백이었다. KRX가 그 바닥을 받친다.
    #  ② tvDatafeed는 지수 거래량을 0으로 준다 → KRX 거래량으로 지수 OBV가 성립한다.
    krx_hist = (_fetch_index_via_krx(market_type)
                if market_type in _KRX_INDEX_MERGE_TYPES else None)

    # 0) 토스 시장지표 (모드 3의 1순위) — 코스피·코스닥만 지원(그 외는 아래 폴백으로)
    if config.session.is_toss and market_type in _TOSS_INDEX_MARKET_TYPES:
        try:
            toss_df = api.get_domestic_index_chart(kis_code)
            if toss_df is not None and not toss_df.empty:
                toss_df.attrs['source'] = 'TOSS'
                df = toss_df
        except Exception as e:
            logger.debug(f"[MARKET_INDEX_DEBUG] {market_type} 토스 시장지표 조회 실패: {e}")

    # 1) KIS API (모드 1/2 전용; 토스 모드는 KIS 미사용)
    if _insufficient(df) and not config.session.is_toss:
        try:
            kis_df = api.get_domestic_index_chart(kis_code)
            if kis_df is not None and not kis_df.empty:
                # [Fix] KIS API 컬럼명 표준화 및 타입 변환
                rename_map = {
                    'stck_bsop_date': 'date',
                    'bstp_nmix_prpr': 'close',
                    'bstp_nmix_oprc': 'open',
                    'bstp_nmix_hgpr': 'high',
                    'bstp_nmix_lwpr': 'low',
                    'acml_vol': 'volume'
                }
                kis_df.rename(columns=rename_map, inplace=True)
                for col in ['close', 'open', 'high', 'low', 'volume']:
                    if col in kis_df.columns:
                        kis_df[col] = pd.to_numeric(kis_df[col], errors='coerce')
                kis_df.attrs['source'] = 'KIS'
                df = kis_df
        except Exception as e:
            logger.debug(f"[MARKET_INDEX_DEBUG] {market_type} KIS API 조회 실패: {e}")

    # 2) tvDatafeed (모드 1/2의 1차 폴백 / 토스 모드의 1순위)
    if _insufficient(df) and market_type in _TVDATAFEED_INDEX_SYMBOLS:
        logger.debug(f"[MARKET_INDEX_DEBUG] {market_type} → tvDatafeed 조회 시도")
        tv_df = _fetch_index_via_tvdatafeed(market_type)
        if tv_df is not None and not tv_df.empty:
            df = tv_df  # attrs['source']='TVDATAFEED' (fetch 함수가 설정)

    # 3) yfinance (최후 폴백) — ^KS200, ^KQ150은 야후 미제공이므로 무한 루프 방지를 위해 제외
    #    (VKOSPI는 yfinance 미제공 → yf_ticker=None이면 건너뜀)
    if _insufficient(df) and yf_ticker and yf_ticker not in ['^KS200', '^KQ150']:
        logger.debug(f"[MARKET_INDEX_DEBUG] {market_type} → yfinance({yf_ticker}) 폴백 시도")
        try:
            yf_df = api.get_chart_data(yf_ticker, is_overseas=True)
            if yf_df is not None and not yf_df.empty:
                yf_df.attrs['source'] = 'YFINANCE'
                df = yf_df
        except Exception as e:
            logger.debug(f"[MARKET_INDEX_DEBUG] {market_type} yfinance 폴백 실패: {e}")

    return _merge_index_history(krx_hist, df)

def get_domestic_index_data(market_type, force_refresh=False):
    """국내 지수 데이터 조회 (KIS API -> yfinance Fallback, 공유 캐시 적용)"""
    # 테스트 환경: 캐시/스탬피드 방어 없이 매번 직접 조회(모킹 데이터 고착 방지)
    if not _index_cache_enabled():
        return _fetch_domestic_index_data(market_type)

    # 1. 빠른 경로: 락 없이 캐시 상태만 판정
    if not force_refresh:
        status, df = _lookup_index_cache(market_type)
        if status in ('fresh', 'suppress'):
            return df
        if status == 'stale':
            _trigger_async_refresh(market_type)  # stale 즉시 반환 + 백그라운드 갱신
            return df
        # status == 'miss'/'expired': 동기 조회 필요

    # 2. single-flight: market_type별 1개 스레드만 실제 조회, 나머지는 대기 후 결과 공유
    if _index_fetch_reentrant(market_type):
        # 같은 스레드가 이미 이 지수를 조회 중이다 — 락을 다시 잡으면 영구 교착이다.
        logger.warning(f"[MARKET_INDEX] {market_type} 조회가 자기 자신을 다시 불렀다 "
                       f"(호출 그래프 순환) — 재귀 호출은 캐시값으로 되돌린다")
        return _lookup_index_cache(market_type)[1]

    fetch_lock = _get_index_fetch_lock(market_type)
    with fetch_lock:
        if not hasattr(_INDEX_FETCH_INPROGRESS, "types"):
            _INDEX_FETCH_INPROGRESS.types = set()
        _INDEX_FETCH_INPROGRESS.types.add(market_type)
        try:
            return _fetch_index_locked(market_type, force_refresh)
        finally:
            _INDEX_FETCH_INPROGRESS.types.discard(market_type)


def _fetch_index_locked(market_type, force_refresh):
    """single-flight 락을 잡은 상태의 실제 조회 본문(get_domestic_index_data 전용)."""
    stale_df = None
    # 대기 중 다른 스레드가 캐시를 채웠을 수 있으니 재확인(강제 갱신 제외)
    if not force_refresh:
        status, df = _lookup_index_cache(market_type)
        if status in ('fresh', 'suppress'):
            return df
        if status == 'stale':
            _trigger_async_refresh(market_type)
            return df
        stale_df = df  # 'miss' → None / 'expired' → 날짜 지난 옛 데이터(폴백용)

    df = _fetch_domestic_index_data(market_type)
    _store_index_cache(market_type, df)

    # 조회 실패(빈 결과) 시 직전 정상 데이터(stale)로 폴백
    if (df is None or df.empty) and stale_df is not None and not stale_df.empty:
        # 날짜가 지난 데이터를 계속 서빙하는 상황은 조용히 넘기면 안 된다(등락률 고착·국면 오판).
        logger.warning(f"[MARKET_INDEX] {market_type} 지수 재조회 실패 — 직전(과거) 데이터로 폴백")
        return stale_df
    return df

# 국면 문자열 -> 점수 보정 설정 키. 국면 상태를 추가할 때 이 표만 갱신하면 된다.
REGIME_SCORE_ADJ_KEYS = {
    "Bull": ("BULL_SCORE_ADJ", -0.5),
    "PendUp": ("PENDING_UP_SCORE_ADJ", 0.0),
    "PendDown": ("PENDING_DOWN_SCORE_ADJ", 0.5),
    "Bear": ("BEAR_SCORE_ADJ", 0.5),
    "Sideways": ("SIDEWAYS_SCORE_ADJ", 0.0),
}

# 국면 문자열 -> (한글 라벨, rich 색상). 화면·텔레그램·자동매매 로그가 공유한다.
#  (표시부가 제각각 하드코딩돼 상태 추가 시 일부만 누락되던 문제 방지)
REGIME_DISPLAY = {
    "Bull": ("강세장", "red"),
    "PendUp": ("상승 미확정", "orange3"),
    # PendDown은 하락축의 '옅은' 색 — Bull/PendUp(빨강/주황)과 대칭을 이루게 한다.
    #  magenta는 팔레트 전반에서 '극단/과열'(VIX≥40, WTI≥100, RSI 과열 등)로 굳어 있어 쓰지 않는다.
    "PendDown": ("하락 미확정", "sky_blue3"),
    "Bear": ("약세장", "blue"),
    "Sideways": ("판정 보류", "yellow"),
}


# 국면 문자열 -> 이모지. 위 REGIME_DISPLAY 의 색 계열과 짝을 맞춘다.
#  [왜 한 곳인가 · 2026-08-29] 같은 if-elif 사다리가 메뉴 헤더(main), 텔레그램 하단 버튼
#   (telegram_bot), /status 본문(trader) 세 곳에 복제돼 있었다. 게다가 텔레그램 버튼
#   매핑(button_map)은 그 출력과 손으로 맞춰야 해서, 국면이 하나 늘거나 이모지를 바꾸면
#   **네 곳**을 함께 고쳐야 했다 — 한 곳만 빠지면 버튼이 조용히 안 먹는다.
#  하늘색 원형 이모지는 유니코드에 없어 PendDown 만 밝은 파랑 마름모로 대체한다
#  (REGIME_DISPLAY 의 sky_blue3 에 대응 — 하락축의 '옅은' 단계).
REGIME_EMOJI = {
    "Bull": "🔴",
    "PendUp": "🟠",
    "PendDown": "🔷",
    "Bear": "🔵",
    "Sideways": "🟡",
}
# 국면을 알 수 없을 때(조회 실패·미지의 값). 판정 보류(🟡)와 구분한다.
REGIME_EMOJI_UNKNOWN = "⚪"


def regime_emoji(regime):
    """국면 문자열 -> 이모지. 모르는 값이면 ⚪."""
    return REGIME_EMOJI.get(regime, REGIME_EMOJI_UNKNOWN)


def all_regime_emojis():
    """regime_emoji 가 낼 수 있는 모든 값.

    텔레그램 버튼 매핑을 이 목록에서 만들어, 이모지를 바꿔도 손으로 맞출 일이 없게 한다.
    """
    return list(REGIME_EMOJI.values()) + [REGIME_EMOJI_UNKNOWN]


def format_regime(regime, markup=True):
    """국면 문자열을 한글 라벨로 변환. markup=True면 rich 색상 태그를 입힌다."""
    label, color = REGIME_DISPLAY.get(regime, (regime, "yellow"))
    return f"[{color}]{label}[/]" if markup else label


# DMI 방향 표기(▲/▼/●)에서 '중립'으로 볼 DX 상한.
#  절대 격차(|+DI - -DI|)는 종목 변동성에 따라 스케일이 달라져 기준이 흔들리므로,
#  ADX의 재료인 정규화 값 DX = 100*|+DI - -DI|/(+DI + -DI)로 판정한다.
DMI_NEUTRAL_DX = 10.0


def dmi_direction_icon(plus_di, minus_di):
    """DMI 우위 방향 아이콘. +DI 우위 ▲(빨강) / -DI 우위 ▼(파랑) / 중립 ●(회백색).

    데이터가 없으면 빈 문자열을 돌려 ADX 셀이 기존 표기를 유지하게 한다.
    """
    if plus_di is None or minus_di is None:
        return ""
    try:
        if math.isnan(plus_di) or math.isnan(minus_di):
            return ""
    except TypeError:
        return ""

    di_sum = plus_di + minus_di
    if di_sum <= 0:
        return "[dmi.neutral]●[/]"
    dx = 100 * abs(plus_di - minus_di) / di_sum
    if dx < DMI_NEUTRAL_DX:
        return "[dmi.neutral]●[/]"
    return "[red]▲[/]" if plus_di > minus_di else "[blue]▼[/]"


def format_adx_cell(adx_val, plus_di=None, minus_di=None, digits=0):
    """표용 ADX 셀: 값(강도 색상) + DMI 방향 아이콘.

    색상 5단계(magenta/red/orange3/yellow/white)는 지수표·전체분석표와 동일 기준.
    목록 표에서는 아이콘 자리를 만들기 위해 정수로 표기한다(소수 2자리는 개별 분석 화면).
    """
    if adx_val is None:
        return "[dim]-[/dim]"

    adx_str = f"{adx_val:.{digits}f}"
    if adx_val >= 40: adx_str = f"[magenta]{adx_str}[/]"
    elif adx_val >= 30: adx_str = f"[red]{adx_str}[/]"
    elif adx_val >= 20: adx_str = f"[orange3]{adx_str}[/]"
    elif adx_val >= 15: adx_str = f"[yellow]{adx_str}[/]"
    else: adx_str = f"[white]{adx_str}[/]"

    icon = dmi_direction_icon(plus_di, minus_di)
    return f"{adx_str} {icon}" if icon else adx_str


def price_trend_color(price, ema20, ema60, ind=None):
    """현재가 색상: 중장기 추세(EMA20 vs EMA60) × 단기 위치(현재가 vs EMA20).

    지수 화면(market.py)과 종목 표(analysis.print_table)가 같은 규칙을 쓰도록 분리했다.
    (동일 로직이 양쪽에 복제돼 한쪽만 고치면 어긋나던 문제 방지)

    자산 종류와 무관하게 '값 자체의 방향'만 나타낸다 — VIX·금리·달러를 반전하던
    규칙은 같은 줄의 등락률(반전 없음)과 색이 엇갈려 폐기했다(config 색상 규칙 주석 참조).

    `ind`(calculate_indicators 결과)를 주면 5일선·20일선 기울기·120일선까지 반영한 5단계로
    판정하고, 없으면 20일선 단면만 보는 폴백을 쓴다. 두 경로 모두 중장기 구조를 먼저
    가르므로 상승 구조가 파랑으로, 하락 구조가 빨강으로 뒤집히는 일은 없다.

    Returns: rich 색상 태그 문자열 ("[red]" 등).
      혼조(ema20 == ema60)는 "[white]", 산출 불가(값 없음)는 "[dim]"로 구분한다.
    """
    if price is None or ema20 is None or ema60 is None:
        return "[dim]"  # 데이터 부족

    # [추세추종] ind 가 오면 5일선과 20일선 기울기(5봉 차분)까지 본다.
    #  [구조] 바깥은 **중장기 구조**(ema20 vs ema60)가 가르고, 안쪽에서 단기 상태를 나눈다.
    #   구조가 정해지면 색은 그 구조 안에서만 움직인다 — 상승 구조는 red/yellow(과열이면
    #   magenta), 하락 구조는 orange/blue. **white 는 ema20 == ema60 에만 남긴다.**
    #
    #  [왜 이렇게 닫았나 · 2026-08-29] 종전 구현은 조건을 평평하게 늘어놓고 아무 데도
    #   걸리지 않으면 white 로 떨어뜨렸는데, 그 '아무 데도'가 매우 넓었다:
    #    · 상승 구조 + 5일선이 20일선 아래 + 현재가가 5일선 위 → white
    #      (장기 상승 추세에서 5일선을 막 회복한 눌림목 반등 초입 — 추세추종에서 가장
    #       값진 진입 후보 구간이 '방향 판단 보류'로 찍혔다)
    #    · 완전 정배열인데 20일선이 하루 눌림 + 현재가가 5일선 위 → white
    #      (휩소를 줄이려던 규칙이 오히려 하루짜리 기울기 스위치로 번복을 만들었다)
    #   합성 분포에서 white 가 33% 를 먹었고 그중 10%p 는 종전 red, 10%p 는 종전 orange 였다.
    #   화면 색은 국면을 읽는 1차 수단이라 '보류'가 최빈값이 되면 그 자체로 정보가 죽는다.
    if ind is not None:
        ema5 = ind.get('ema_5')
        slope_ref = ind.get('ema_20_slope_ref')

        if ema5 is not None and slope_ref is not None:
            # 기울기는 전일이 아니라 EMA20_SLOPE_LOOKBACK(5)봉 전과 비교한다 —
            #  하루 등락으로 색이 뒤집히던 번복을 줄인다(core.indicators 주석 참조).
            is_ema20_up = ema20 > slope_ref
            disp_ratio = (price / ema20 * 100) if ema20 > 0 else 0
            # [단일 소스] 같은 화면의 개별 분석(이격도 행)이 쓰는 값과 같아야 한다.
            #  종전에는 여기만 110 을 박아 둬, 설정을 바꾸면 표와 색이 갈렸다.
            disp_upper = config.ANALYSIS_THRESHOLDS.get("DISPARITY_UPPER", 110)

            if ema20 > ema60:
                # 상승 구조 — 정배열이고 20일선이 우상향일 때만 '강세(red)'다.
                #
                # [장기선 확인 · 2026-08-29] 여기에 `현재가 > EMA120` 을 건다.
                #  종전에는 빨강을 "완벽한 정배열"이라 부르면서 정작 120일선을 보지 않았다 —
                #  라벨과 조건이 어긋나 있었고, 추세추종에서 정배열은 장기선 위에서 성립한다.
                #  실제로 걸리는 것은 **성숙한 베어마켓 랠리**다. 단기는 5>20>60 으로 완전히
                #  정렬됐는데 장기 추세는 아직 하락인 구간.
                #  [실측 2026-08-29 · 44종목 10년] 빨강 32,124봉 중 이 조건에 걸리는 1,476봉
                #   (4.6%)의 60일 전방수익은 +1.63% 로, 나머지 +6.52% 의 4분의 1이다.
                #   격차 +4.90%p 는 5개 창에서 **전부 같은 부호**로, 이 감사에서 찾은 보조축
                #   중 유일하게 5/5 다(52주 위치 4/5, ADX·DMI 는 사실상 못 가른다).
                #   `120일선 위` 축은 빨강·주황·노랑·파랑 **네 색 모두에서 부호가 같다**.
                #  [왜 색만 고치고 진입은 안 막나] 같은 조건을 진입 게이트로 걸면(G1/G2)
                #   포트폴리오 짝비교에서 2년 창 5개 중 3승2패, 무작위 대조 순위가 창별로
                #   6/6(꼴찌)까지 섞여 기각됐다(tools/audit_state_ma120_gate.py).
                #   게이트는 슬롯 경쟁과 순위 흔들림을 타지만, 색은 매매하지 않고 상태에
                #   이름만 붙인다 — 표시 규칙의 판단 기준은 신호 분리도다. 모순이 아니다.
                #  [폴백] EMA120 이 없으면(상장 초기·데이터 부족) 조건을 묻지 않는다.
                #   자료가 없다고 강세를 조정으로 낮추면 없는 정보를 근거로 색을 바꾸는 것이다.
                ema120 = ind.get('ema_120')
                long_ok = not (ema120 is not None and price < ema120)
                if ema5 > ema20 and long_ok:
                    if disp_ratio >= disp_upper:
                        # 과열 = **신규 진입을 자제할 구간**이지, 파는 구간이 아니다.
                        #  [실측 2026-08-29 · 39종목 5년, 43,792관측] 보라 구간의 전방
                        #   수익은 전 색 중 가장 높다 — 60일 평균 +14.21%, 승률 58.3%
                        #   (빨강 +9.32%, 주황 +11.51%, 파랑 +4.91%). 평균이 중앙값
                        #   (+1.12%)보다 훨씬 큰 것은 이 구간이 꼬리를 물고 있다는 뜻이고,
                        #   추세추종이 먹는 것이 바로 그 꼬리다.
                        #  그래서 여기서 '익절'을 권하면 시스템 정책(고정 익절 기본 OFF,
                        #   샹들리에 TS 주청산)과도, 실측과도 반대가 된다.
                        #  진입 차단 쪽은 근거가 있다 — TREND_QUALITY_MAX(300) 상한과
                        #   BUY_RSI_MAX(70) 가 같은 취지다.
                        return "[magenta]"  # 과열: 신규 진입 자제 (보유는 TS 에 맡긴다)
                    if is_ema20_up:
                        return "[red]"      # 강세: 완전 정배열 + 20일선 우상향
                # 정배열이 아니거나(5일선이 20일선 아래) 20일선이 꺾였으면, 또는 장기선
                #  아래면 '조정'이다.
                #  구조는 살아 있으므로 관망(blue)도 보류(white)도 아니다.
                #  [색 배정] 상승 구조 쪽이 주황이다 — 빨강(강세)에 인접한 난색이라
                #   '추세는 살아 있고 잠시 쉬는 중'이 색만 보고도 읽힌다.
                return "[orange3]"          # 눌림목: 장기 상승 추세 속 단기 휴식

            if ema20 < ema60:
                # 하락 구조 — 20일선을 되찾고 그 20일선이 턴했을 때만 '반등 시도(yellow)'다.
                if price > ema20 and is_ema20_up:
                    # [주의] '전환 시도'는 기대일 뿐 우위가 아니다.
                    #  실측(위와 같은 표본): 이 구간의 60일 전방 수익은 평균 +2.71%,
                    #  승률 49.0% 로 **전 색 중 가장 나쁘다**(관망인 파랑조차 +4.91%,
                    #  53.7%). 역배열 되돌림이 열위라는 기존 결론과 같은 방향이다
                    #  (역배열 데드캣 매수 승률 18% — 시너지 게이트 주석 참조).
                    #  색은 남기되 '진입 후보'로 읽히지 않게 라벨을 낮춰 둔다.
                    #
                    #  [교차 실측 2026-08-29 · 44종목 5년, 51,850관측] 위 수치는 **색만으로
                    #   나눈 전 봉**의 기저율이다. 상태 판정(점수 7.0 + RSI 캡 + SAR·MACD·DMI
                    #   방어)이 붙은 부분집합은 다르다 — 노랑 전체 60일 +2.75%(승률 48.9%)가
                    #   '노랑 & 매수'에서는 +4.14%(51.6%)로 올라간다. 상태 필터가 이 구간
                    #   안에서 실제로 고르고 있다는 뜻이므로, 같은 줄에 노랑과 '매수'가
                    #   함께 찍히는 것은 모순이 아니다(도움말 문구도 이에 맞춰 정정).
                    #   다만 '비노랑 & 매수' +10.48%(57.7%)에는 여전히 크게 못 미치고
                    #   꼬리(상위10%)가 +29.4% vs +47.4% 로 얇다 — 색의 경고 자체는 유효하다.
                    #
                    #  [노랑이 섞고 있는 것] 이 조건 하나에 **데드캣 반등**과 **바닥 탈출
                    #   초기**가 함께 들어온다. 새 추세가 시작돼도 EMA20 이 EMA60 을 되찾기
                    #   전 몇 주는 모습이 데드캣과 똑같아서, 구조만으로는 갈 수 없다.
                    #   둘을 가르는 축을 전 봉(44종목 10년, 노랑 10,794봉)에서 재보면
                    #   **52주 위치**가 가장 크고 일관되다 — 60% 이상 +6.91% vs 미만 +2.00%
                    #   (5개 창 중 4개 부호 일치). 120일선은 같은 자리에서 +1.52%p·3/5 로
                    #   약하고, ADX(+0.10)·DMI(+0.67)는 사실상 못 가른다.
                    #   색을 쪼개는 대신 도움말에 이 갈림을 적어 두고, 판단은 상태 열과
                    #   52주 열에 맡긴다(색 단계를 늘리면 화면의 1차 판독이 무거워진다).
                    #
                    #  [한 색으로 둔다 · 2026-08-29] 이 구간을 52주 위치로 쪼개 두 색으로
                    #   찍어 봤다가 되돌렸다. 실측 격차 자체는 크지만(60% 이상 +6.91% vs
                    #   미만 +2.00%, 5개 창 중 4개 부호 일치) 색을 늘려서 얻을 것이 아니다 —
                    #   가르는 값이 이미 같은 표의 `52W%` 열에 숫자로 있고, 값 색상의 축은
                    #   EMA 구조 하나여야 한다(52주는 구조가 아니라 위치다).
                    #   대신 **모호하다는 사실 자체를 도움말에 적는다** — 색이 두 사건을
                    #   합쳐 보여준다는 것을 알고 보는 것과 모르고 보는 것은 다르다.
                    return "[yellow]"       # 되돌림: 데드캣 반등 / 바닥 탈출 초기가 섞인 구간
                return "[blue]"             # 약세: 하락 추세 관망

            return "[white]"                # ema20 == ema60 — 방향 판단 보류

    # [기본 로직] ind 파라미터가 없거나 데이터가 부족한 경우 폴백
    if ema20 > ema60:
        # 상승 추세: 20일선 위면 강세, 아래면 눌림목 조정
        color = "[red]" if price > ema20 else "[yellow]"
    elif ema20 < ema60:
        # 하락 추세: 20일선 아래면 약세, 위면 반등 시도
        color = "[blue]" if price < ema20 else "[orange3]"
    else:
        color = "[white]"  # ema20 == ema60 (혼조) — 방향 판단 보류
    return color


def classify_regime_from_df(df, params=None):
    """[국면 판정 핵심] 이중 EMA 교차 + 추종 확인(follow-through) 규칙.

    에드 세이코타 방식: 빠른 EMA(9일, β=0.25)와 느린 EMA(41일, β=0.05)가 교차하면
    방향 전환으로 보되, 교차 이후 지수가 확인 기준(기본 5%)만큼 진행해야 '확정 추세'로
    인정한다. 미달 상태에서 다시 교차하면 그 구간은 휩소(실패한 추세)로 집계한다.

    Returns:
        dict: {
          'regime': 'Bull'|'PendUp'|'PendDown'|'Bear'|'Sideways',
          'moved_pct': 교차 이후 현재까지 진행률(%),
          'whipsaw_ratio': 직전 N개 완료 교차 중 확인 기준 미달 비율(0~1, 산출 불가 시 None),
          'segments': 집계에 사용한 완료 교차 구간 수,
        }
        데이터가 부족하면 regime='Sideways'를 돌려준다(호출부는 중립으로 취급).

    지수 화면(market.py)과 자동매매(analysis.get_market_regime)가 같은 판정을 쓰도록
    데이터프레임만 받는 순수 함수로 분리했다.
    """
    params = params or getattr(config, 'MARKET_REGIME_PARAMS', {}) or {}
    slow = int(params.get("REGIME_EMA_SLOW", 41))

    blank = {'regime': "Sideways", 'moved_pct': 0.0, 'whipsaw_ratio': None, 'segments': 0}
    if df is None or df.empty or 'close' not in df.columns or len(df) < slow:
        return blank

    close = pd.to_numeric(df['close'], errors='coerce').dropna()
    if len(close) < slow:
        return blank

    # 판정식은 indicators.get_regime_series 하나만 둔다 — 시장 필터(국면 동시 확인)도 같은 함수를
    # 쓰므로, 여기서 로직을 따로 들고 있으면 두 판정이 소리 없이 갈라진다.
    ser = indicators.get_regime_series(close, params)
    whipsaw = ser['whipsaw'][-1]
    return {'regime': ser['regime'][-1], 'moved_pct': float(ser['moved_pct'][-1]),
            'whipsaw_ratio': None if np.isnan(whipsaw) else float(whipsaw),
            'segments': int(ser['segments'][-1])}


def get_market_regime_detail(market_type="KOSPI"):
    """시장 국면 상세 — classify_regime_from_df 결과에 점수 보정(score_adj)을 더해 반환.

    자동매매는 매 주기(매도검사·매수스캔 각각) 시장별로 호출하므로, 짧은 TTL 캐시로
    지수 차트 조회 + 지표 재계산 중복을 제거한다. (국면은 초 단위로 변하지 않음)
    데이터 부족/오류 폴백은 일시 장애일 수 있어 캐시하지 않는다.
    """
    cached = _MARKET_REGIME_CACHE.get(market_type, ttl=_MARKET_REGIME_TTL_SEC)
    if cached is not None:
        return cached

    neutral = {'regime': "Sideways", 'score_adj': 0.0, 'moved_pct': 0.0,
               'whipsaw_ratio': None, 'segments': 0}
    try:
        params = getattr(config, 'MARKET_REGIME_PARAMS', {}) or {}
        df = get_domestic_index_data(market_type)
        info = classify_regime_from_df(df, params)
        if info['regime'] == "Sideways" and info['segments'] == 0:
            return neutral  # 데이터 부족 — 캐시하지 않음

        key, default = REGIME_SCORE_ADJ_KEYS.get(info['regime'], ("SIDEWAYS_SCORE_ADJ", 0.0))
        try:
            info['score_adj'] = float(params.get(key, default))
        except (TypeError, ValueError):
            info['score_adj'] = default

        _MARKET_REGIME_CACHE.set(market_type, info)
        return info

    except Exception as e:
        logger.error(f"시장 국면 판단 오류: {e}")
        return neutral


def get_market_regime(market_type="KOSPI"):
    """시장 국면 판단 — (국면 문자열, 매수 점수 보정) 튜플.

    상세 정보(휩소율 등)가 필요하면 get_market_regime_detail을 쓴다.
    """
    info = get_market_regime_detail(market_type)
    return info['regime'], info['score_adj']

def get_index_momentum(market_type="KOSPI", lookback=None):
    """[추세추종] 지수 모멘텀(%) — 최근 lookback 거래일 지수 수익률 (상대강도 RS 필터 기준선).

    종목의 절대 모멘텀(스코어링 '가격 모멘텀')과 같은 룩백(MOMENTUM_LOOKBACK)을 사용해
    "같은 기간 동안 지수보다 강했는가"를 비교할 수 있게 한다.
    지수 데이터는 get_domestic_index_data 공유 캐시를 그대로 사용하므로 추가 API 부담 없음.
    데이터 부족/조회 실패 시 None을 반환한다 — 호출부는 필터를 통과시켜야 한다(fail-open,
    지수 조회 장애가 매수 전면 중단으로 번지는 것을 방지).
    """
    if lookback is None:
        lookback = config.INDICATOR_PARAMS.get('MOMENTUM_LOOKBACK', 126)
    try:
        df = get_domestic_index_data(market_type)
        if df is None or df.empty or len(df) <= lookback:
            return None
        closes = pd.to_numeric(df['close'], errors='coerce')
        cur = float(closes.iloc[-1])
        past = float(closes.iloc[-(lookback + 1)])
        if not (past > 0) or pd.isna(cur):
            return None
        return (cur / past - 1) * 100
    except Exception as e:
        logger.debug(f"지수 모멘텀 계산 실패({market_type}): {e}")
        return None

def classify_stock_state(price=None, ema20=None, ema60=None, ema120=None, sar=None, rsi=None, prev_rsi=None, adx=None, cci=None, obv_trend=None, macd=None, macd_signal=None, thresholds=None, w52_pos=None, smart_money=False, plus_di=None, minus_di=None, df=None, ind=None, ema_5=None, macd_hist=None, prev_macd_hist=None, prev_cci=None, vol_spike=False, vol_trend=False, is_yangbong=False, mom_ret=None, mom_ret_1m=None, mom_ret_3m=None, trend_persist=None):
    if df is not None and ind is not None:
        if not df.empty: price = float(df.iloc[-1]['close'])
        ema20 = ind.get('ema_20')
        ema60 = ind.get('ema_60')
        ema120 = ind.get('ema_120')
        sar = ind.get('psar')
        rsi = ind.get('rsi')
        adx = ind.get('adx')
        cci = ind.get('cci')
        obv_trend = ind.get('obv_trend')
        macd = ind.get('macd')
        macd_signal = ind.get('macd_signal')
        if plus_di is None: plus_di = ind.get('plus_di')
        if minus_di is None: minus_di = ind.get('minus_di')

        # [추가] 인자로 넘어오지 않은 세부 지표들을 ind 딕셔너리에서 직접 추출
        if ema_5 is None: ema_5 = ind.get('ema_5')
        if prev_cci is None: prev_cci = ind.get('prev_cci')
        if macd_hist is None: macd_hist = ind.get('macd_hist')
        if prev_macd_hist is None: prev_macd_hist = ind.get('prev_macd_hist')

        if plus_di is None or minus_di is None:
            import numpy as np
            try:
                high_diff = df['high'].diff()
                low_diff = df['low'].diff()
                pos_dm = np.where((high_diff > 0) & (high_diff > -low_diff), high_diff, 0.0)
                neg_dm = np.where((low_diff < 0) & (-low_diff > high_diff), -low_diff, 0.0)
                tr1 = df['high'] - df['low']
                tr2 = (df['high'] - df['close'].shift()).abs()
                tr3 = (df['low'] - df['close'].shift()).abs()
                tr = pd.concat([tr1, tr2, tr3], axis=1).max(axis=1)
                adx_period = config.INDICATOR_PARAMS.get('ADX_PERIOD', 14)
                atr = tr.ewm(alpha=1/adx_period, adjust=False).mean()
                atr_val = atr.iloc[-1]
                if atr_val == 0 or np.isnan(atr_val): atr_val = 1.0 # 0 나누기 방어
                plus_di = 100 * pd.Series(pos_dm).ewm(alpha=1/adx_period, adjust=False).mean().iloc[-1] / atr_val
                minus_di = 100 * pd.Series(neg_dm).ewm(alpha=1/adx_period, adjust=False).mean().iloc[-1] / atr_val
            except Exception: pass
            
    if price is None or ema60 is None or sar is None or rsi is None:
        return "-", "[dim]", "데이터 부족"
    
    # [수정] 1순위 절대 방어 필터를 가장 위로 끌어올림 (역추세 매수보다 우선 적용하여 떨어지는 칼날 완벽 방어)
    is_panic_regime = False
    if plus_di is not None and minus_di is not None and minus_di > plus_di:
        if adx is not None and adx >= 45: # ADX 45 이상의 초강력 하락장
            is_panic_regime = True
            # [추세추종] ADX는 방향 무관 지표라 강한 상승추세 직후의 급락 눌림에서도 45 이상이
            #  유지된 채 -DI가 역전될 수 있음. 추세 구조(60일선)가 살아있는 동안은 광폭 손절
            #  (ATR 손절/샹들리에 TS)에 청산을 맡기고, 구조 훼손을 동반한 경우에만 즉시 '매도'로 분류
            #  (승자 포지션의 반추세성 조기 전량청산 방지. 매수 게이트는 -DI 우위 hard_caution이 계속 차단)
            if price < ema60:
                return "매도", "[blue]", "초강력 투매 패닉 구간 (ADX 과열 및 -DI 우위)"

    # [추가] 2. 낙폭과대(역추세) 반등 조건 확인
    #  패닉 구간(is_panic_regime)에서는 60일선 위라도 역매수를 봉쇄해 '떨어지는 칼날 방어'를 유지한다
    #  (구조 게이트 도입으로 패닉 체크가 return하지 않게 되면서 MR 경로가 열리는 회귀 방지)
    use_mr = thresholds.get("USE_MEAN_REVERSION", config.ANALYSIS_THRESHOLDS.get("USE_MEAN_REVERSION", False)) if thresholds else config.ANALYSIS_THRESHOLDS.get("USE_MEAN_REVERSION", False)
    if use_mr and not is_panic_regime and ema20 is not None and prev_rsi is not None and rsi is not None:
        mr_rsi = thresholds.get("MR_RSI_MAX", config.ANALYSIS_THRESHOLDS.get("MR_RSI_MAX", 40.0)) if thresholds else config.ANALYSIS_THRESHOLDS.get("MR_RSI_MAX", 40.0)
        mr_disp = thresholds.get("MR_DISPARITY_MAX", config.ANALYSIS_THRESHOLDS.get("MR_DISPARITY_MAX", 90.0)) if thresholds else config.ANALYSIS_THRESHOLDS.get("MR_DISPARITY_MAX", 90.0)
        
        disparity = (price / ema20) * 100
        
        is_yangbong_flag = is_yangbong
        if df is not None and not df.empty:
            is_yangbong_flag = df.iloc[-1]['close'] > df.iloc[-1]['open']

        # [개선 #3] 떨어지는 칼날(데드캣) 방어 강화 — 하락 가속도 둔화 확인.
        #          기존 조건(RSI 반등+양봉+수급)만으로는 ADX<45 의 추세적 하락 중에도
        #          역매수가 발동될 수 있어 방어가 얇음. df가 있을 때, 당일 저가가
        #          직전 5거래일 저점을 추가로 경신(=신저가 갱신, 낙하 지속)하면 진입 보류.
        mr_decel_ok = True
        if df is not None and not df.empty and len(df) >= 6:
            try:
                prior_low = df['low'].iloc[-6:-1].min()
                today_low = df['low'].iloc[-1]
                if today_low < prior_low:
                    mr_decel_ok = False
            except Exception:
                pass

        # [수정] RSI 반등 + 양봉 + 수급 확인(OBV 상승 또는 스마트머니 유입) + 하락 둔화를 모두 만족해야 역매수 발동
        # 조건 미충족 시 단순 기술적 반등(데드캣)으로 간주하여 진입 보류
        if rsi <= mr_rsi and rsi > prev_rsi and disparity <= mr_disp and is_yangbong_flag and (obv_trend or smart_money) and mr_decel_ok:
            return "역매수", "[magenta]", "낙폭과대 (역매수 반등 신호 + 양봉 + 수급 확인 + 하락 둔화)"

    # [수정] 가중치 적용 점수 계산 (thresholds에 weights가 포함되어 있을 수 있음)
    weights = thresholds.get("WEIGHTS") if thresholds else None
    score, _ = calculate_score(
        price=price, ema20=ema20, ema60=ema60, ema120=ema120, sar=sar, rsi=rsi, adx=adx, cci=cci,
        obv_trend=obv_trend, macd=macd, macd_signal=macd_signal, weights=weights, smart_money=smart_money,
        plus_di=plus_di, minus_di=minus_di, df=df, ind=ind, ema_5=ema_5, macd_hist=macd_hist,
        prev_macd_hist=prev_macd_hist, prev_cci=prev_cci, vol_spike=vol_spike, vol_trend=vol_trend,
        w52_pos=w52_pos, mom_ret=mom_ret,  # [추세추종] 가격 모멘텀 팩터 입력 전달 (df 없이 호출되는 백테스트 경로 패리티)
        mom_ret_1m=mom_ret_1m, mom_ret_3m=mom_ret_3m, trend_persist=trend_persist  # [추세추종] 다중 기간 정합·추세 지속 이력 동일 전달
    )

    # [수정] config.py의 설정값을 사용하여 상태 판정
    if thresholds:
        buy_score = thresholds.get("BUY_SCORE", config.ANALYSIS_THRESHOLDS["BUY_SCORE"])
        rise_score = thresholds.get("RISE_SCORE", config.ANALYSIS_THRESHOLDS["RISE_SCORE"])
        buy_rsi_max = thresholds.get("BUY_RSI_MAX", config.ANALYSIS_THRESHOLDS["BUY_RSI_MAX"])
        interest_min = thresholds.get("INTEREST_SIGNAL_MIN", config.ANALYSIS_THRESHOLDS.get("INTEREST_SIGNAL_MIN", 3))
        interest_ma60_near = thresholds.get("INTEREST_MA60_NEAR", config.ANALYSIS_THRESHOLDS.get("INTEREST_MA60_NEAR", 0.97))

        use_super = thresholds.get("SUPER_MOMENTUM_USE", config.ANALYSIS_THRESHOLDS.get("SUPER_MOMENTUM_USE", True))
        super_score = thresholds.get("SUPER_MOMENTUM_SCORE", config.ANALYSIS_THRESHOLDS.get("SUPER_MOMENTUM_SCORE", 8.0))
        super_w52 = thresholds.get("SUPER_MOMENTUM_W52_POS", config.ANALYSIS_THRESHOLDS.get("SUPER_MOMENTUM_W52_POS", 90.0))
        super_rsi = thresholds.get("SUPER_BUY_RSI_MAX", config.ANALYSIS_THRESHOLDS.get("SUPER_BUY_RSI_MAX", 80.0))
    else:
        buy_score = config.ANALYSIS_THRESHOLDS["BUY_SCORE"]
        rise_score = config.ANALYSIS_THRESHOLDS["RISE_SCORE"]
        buy_rsi_max = config.ANALYSIS_THRESHOLDS["BUY_RSI_MAX"]
        
        interest_min = config.ANALYSIS_THRESHOLDS.get("INTEREST_SIGNAL_MIN", 3)
        interest_ma60_near = config.ANALYSIS_THRESHOLDS.get("INTEREST_MA60_NEAR", 0.97)

        use_super = config.ANALYSIS_THRESHOLDS.get("SUPER_MOMENTUM_USE", True)
        super_score = config.ANALYSIS_THRESHOLDS.get("SUPER_MOMENTUM_SCORE", 8.0)
        super_w52 = config.ANALYSIS_THRESHOLDS.get("SUPER_MOMENTUM_W52_POS", 90.0)
        super_rsi = config.ANALYSIS_THRESHOLDS.get("SUPER_BUY_RSI_MAX", 80.0)

    reasons = []
    is_severe_danger = False
    
    # [수정] 위험 조건 완화: 장기(120)와 중기(60) 이평선을 모두 이탈해야 '위험'으로 간주 (변동성 감소)
    if ema120 is not None and price < ema120 and price < ema60: 
        is_severe_danger = True
        reasons.append("이평선 완전 이탈(60&120)")
    elif rsi <= (config.INDICATOR_PARAMS["RSI_LOWER"] - 10): 
        is_severe_danger = True # 위험 기준은 하한선보다 더 낮게 설정 (예: 20)
        reasons.append(f"RSI 초과매도({rsi:.1f})")
        
    # [수정] ADX 과열 중 RSI 하락은 '위험'보다는 '주의'로 이동
    if is_severe_danger: return "매도", "[blue]", ", ".join(reasons)
    
    is_caution = False
    # [추가] 위험형 주의(hard): '관심(태동)'으로도 분류 불가한 명백한 하락/과열 신호.
    #        약세형(soft: 이평선 한쪽 이탈)과 분리하여, 60일선 또는 120일선 중 한쪽 아래여도
    #        추세 전환 초기 신호가 있으면 '관심'으로 건질 수 있도록 한다.
    #  [주의] 60·120일선을 '동시에' 이탈한 종목은 위 is_severe_danger에서 이미 '매도'로 확정
    #        되어 여기까지 오지 않는다(추세추종 원칙상 의도된 차단 — 역배열 데드캣 반등을
    #        관심으로도 올리지 않는다). 따라서 이 아래 약세형 경로가 실제로 커버하는 것은
    #        '한쪽만 이탈'(또는 120일선 데이터 부족으로 위험 판정이 불가한) 경우다.
    hard_caution = False
    # [수정] 주의 조건: 60일선 이탈 또는 120일선 이탈 → 약세형(soft, 관심 후보 가능)
    if price < ema60:
        is_caution = True
        reasons.append("60일선 이탈")
    if ema120 is not None and price < ema120:
        is_caution = True
        reasons.append("120일선 이탈")
    if sar is not None and sar > price:
        is_caution = True; hard_caution = True
        reasons.append("SAR 매도신호")
    if rsi >= (config.INDICATOR_PARAMS["RSI_UPPER"] + 10):
        is_caution = True; hard_caution = True
        reasons.append(f"RSI 과열({rsi:.1f})")
    elif rsi <= config.INDICATOR_PARAMS["RSI_LOWER"]:
        is_caution = True; hard_caution = True
        reasons.append(f"RSI 침체({rsi:.1f})")
    elif adx is not None and prev_rsi is not None and adx >= 40 and rsi < prev_rsi:
        is_caution = True; hard_caution = True
        reasons.append(f"ADX과열({adx:.1f})+RSI하락")

    if macd is not None and macd_signal is not None and macd < macd_signal:
        is_caution = True; hard_caution = True
        reasons.append("MACD 데드크로스")
    # [추가] DMI 매도세 우위 필터
    if plus_di is not None and minus_di is not None and minus_di > plus_di:
        is_caution = True; hard_caution = True
        reasons.append("-DI 우위(매도세 강함)")

    # 2순위: 얼리 스테이지 및 기본 매수 조건 (위험 필터를 모두 통과해야 함)
    is_super = use_super and score >= super_score and w52_pos is not None and w52_pos >= super_w52
    actual_buy_rsi_max = super_rsi if is_super else buy_rsi_max

    if score >= buy_score and rsi < actual_buy_rsi_max:
        # [핵심 방어 로직] 하락 반전 신호(고점 꺾임)를 방어하기 위한 강력 필터
        # 다음 조건 중 하나라도 해당되면 아무리 점수가 높아도 "매수"가 아닌 "주의"로 강등
        down_trend_flags = []
        if sar is not None and sar > price: down_trend_flags.append("SAR 매도신호")
        if macd is not None and macd_signal is not None and macd < macd_signal: down_trend_flags.append("MACD 데드크로스")
        if plus_di is not None and minus_di is not None and minus_di > plus_di: down_trend_flags.append("-DI 우위")
        
        if down_trend_flags:
            is_caution = True; hard_caution = True
        else:
            if is_super: return "강매수", "[magenta]", "매수 조건 충족 (슈퍼 모멘텀 적용)"
            else: return "매수", "[red]", "매수 조건 충족 (얼리 스테이지 반등 포함)"

    # [상승/대기] 주의(약세+위험) 신호가 전혀 없고 점수가 양호 → 추세 정렬이 완성된 강한 상태.
    #   두 하위 케이스를 구분하되 색상·매수 로직상 취급은 '상승'과 동급(sibling)으로 둔다
    #   (아래 return 지점에서 이 함수를 소비하는 멤버십 목록들이 '상승'과 함께 '대기'도 포함해야 함):
    #   - 대기: 점수는 매수 기준(buy_score) 충족이나 단기 RSI 과열(actual_buy_rsi_max 이상,
    #           단 과열 주의선 RSI_UPPER+10 미만)로 진입만 보류 → 눌림목 매수 대기(매수 직전).
    #   - 상승: 점수가 rise~buy 사이(아직 폭발적 강도 미달) → 점수 축적 대기.
    if not is_caution and score >= rise_score:
        if score >= buy_score:
            return "대기", "[orange3]", f"매수 직전 (점수 충족, RSI 과열 눌림 대기 · RSI {rsi:.0f}≥{actual_buy_rsi_max:.0f})"
        return "상승", "[orange3]", "상승 추세 (점수 양호, 점수 축적 대기)"

    # [관심(태동)] 추세 정렬은 미완성('상승' 미달)이나, 위험형 신호가 없고
    #   추세 전환 초기 신호가 INTEREST_SIGNAL_MIN개 이상 포착된 종목.
    #   60일선 또는 120일선 중 한쪽 아래(약세형 주의)여도 초기 반등 여력을 빠르게 포착하기 위함.
    #   (양쪽 동시 이탈은 위 is_severe_danger에서 '매도'로 확정되어 여기 도달하지 않는다)
    #   수동 스윙 매매 모니터링 대상으로 표시한다. 자동매매 진입 상태가 아니다.
    if not hard_caution and interest_min > 0:
        interest_signals = []
        if ema_5 is not None and ema20 is not None and ema_5 > ema20:
            interest_signals.append("단기 골든크로스(5>20)")
        if macd_hist is not None and prev_macd_hist is not None and macd_hist > prev_macd_hist:
            interest_signals.append("MACD 히스토그램 개선")
        elif macd is not None and macd_signal is not None and macd > macd_signal:
            interest_signals.append("MACD 골든크로스")
        if plus_di is not None and minus_di is not None and plus_di > minus_di:
            interest_signals.append("+DI 우위")
        if rsi is not None and prev_rsi is not None and rsi >= config.INDICATOR_PARAMS.get("RSI_MID", 50) and rsi > prev_rsi:
            interest_signals.append("RSI 50선 상향")
        if cci is not None and prev_cci is not None and cci > prev_cci and cci > -100:
            interest_signals.append("CCI 개선")
        if vol_trend or vol_spike or obv_trend or smart_money:
            interest_signals.append("수급 유입")
        if price is not None and ema60 is not None and price >= ema60 * interest_ma60_near:
            interest_signals.append("60일선 근접/돌파")

        if len(interest_signals) >= interest_min:
            return "관심", "[green]", "추세 전환 초기 신호 " + str(len(interest_signals)) + "개 (" + ", ".join(interest_signals) + ")"

    if is_caution: return "주의", "[yellow]", ", ".join(reasons)
    return "관망", "[white]", "방향성 탐색 구간"

@contextlib.contextmanager
def _get_db_connection():
    """SQLite 연결 컨텍스트 매니저.

    sqlite3의 기본 `with conn:` 구문은 트랜잭션(commit/rollback)만 관리하고
    연결 자체는 닫지 않아 ResourceWarning(unclosed database)이 발생한다.
    이 컨텍스트 매니저로 감싸 `with _get_db_connection() as conn:` 종료 시
    연결을 확실히 닫는다.
    """
    conn = sqlite3.connect(config.DB_FILE_PATH)
    try:
        yield conn
    finally:
        conn.close()

def _init_analysis_db_logic():
    try:
        with _get_db_connection() as conn:
            cursor = conn.cursor()
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS market_analysis_cache (
                    market_type TEXT PRIMARY KEY,
                    updated_at TEXT,
                    params TEXT,
                    data TEXT
                )
            """)
            conn.commit()
    except Exception: pass

def _save_analysis_result_logic(market_type, results, params):
    try:
        _init_analysis_db()
        with _get_db_connection() as conn:
            cursor = conn.cursor()
            now_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            data_json = json.dumps(results, ensure_ascii=False)
            params_json = json.dumps(params, ensure_ascii=False)
            
            cursor.execute("""
                INSERT OR REPLACE INTO market_analysis_cache (market_type, updated_at, params, data)
                VALUES (?, ?, ?, ?)
            """, (market_type, now_str, params_json, data_json))
            conn.commit()
    except Exception as e:
        config.console.print(f"[dim red]분석 결과 저장 실패: {e}[/dim red]")

def _load_analysis_result_logic(market_type):
    try:
        _init_analysis_db()
        with _get_db_connection() as conn:
            cursor = conn.cursor()
            cursor.execute("SELECT updated_at, params, data FROM market_analysis_cache WHERE market_type = ?", (market_type,))
            row = cursor.fetchone()
            if row:
                return {
                    'updated_at': row[0],
                    'params': json.loads(row[1]),
                    'data': json.loads(row[2])
                }
    except Exception as e:
        config.console.print(f"[dim red]분석 결과 로드 실패: {e}[/dim red]")
    return None

# [수정] 큐 시스템을 통한 실행 래퍼 함수들
def _init_analysis_db():
    _init_analysis_db_logic()

def _save_analysis_result(market_type, results, params):
    if hasattr(db_manager.db, 'execute_custom'):
        db_manager.db.execute_custom(_save_analysis_result_logic, market_type, results, params)
    else:
        _save_analysis_result_logic(market_type, results, params)

def _load_analysis_result(market_type):
    if hasattr(db_manager.db, 'execute_custom'):
        return db_manager.db.execute_custom(_load_analysis_result_logic, market_type)
    else:
        return _load_analysis_result_logic(market_type)

def _get_master_stock_list(market_type):
    """(내부함수) 마스터 파일 다운로드 및 파싱하여 종목 리스트 반환"""
    base_dir = getattr(config, 'DATA_DIR', 'data')
    if not os.path.exists(base_dir):
        os.makedirs(base_dir)

    if market_type == 'KOSPI':
        url = "https://new.real.download.dws.co.kr/common/master/kospi_code.mst.zip"
        filename = "kospi_code.mst"
        _MST_GRP_OFFSET = 227    # 개행 제거 후 뒤에서 227바이트 지점이 증권그룹구분코드(아래 주석)
    else:
        url = "https://new.real.download.dws.co.kr/common/master/kosdaq_code.mst.zip"
        filename = "kosdaq_code.mst"
        _MST_GRP_OFFSET = 221

    zip_path = os.path.join(base_dir, f"{filename}.zip")
    extract_path = os.path.join(base_dir, filename)
    
    stock_list = []

    try:
        # [수정] 파일이 존재하고 오늘 다운로드된 것이라면 다운로드 스킵
        need_download = True
        if os.path.exists(zip_path) and os.path.exists(extract_path):
            file_time = datetime.fromtimestamp(os.path.getmtime(zip_path))
            if file_time.date() == datetime.now().date():
                need_download = False
                config.console.print(f"[dim]{market_type} 마스터 파일이 최신입니다. (기존 파일 사용)[/dim]")

        if need_download:
            with Progress(
                SpinnerColumn(),
                TextColumn("[progress.description]{task.description}"),
                BarColumn(),
                TextColumn("[progress.percentage]{task.percentage:>3.0f}%"),
                "•",
                DownloadColumn(),
                "•",
                TransferSpeedColumn(),
                "•",
                TimeRemainingColumn(),
                console=config.console
            ) as progress:
                task_id = progress.add_task(f"[cyan]{market_type} 마스터 파일 다운로드...[/cyan]", total=None)
                
                def report_hook(block_num, block_size, total_size):
                    progress.update(task_id, total=total_size, completed=block_num * block_size)
                
                urllib.request.urlretrieve(url, zip_path, reporthook=report_hook)

            with Progress(
                SpinnerColumn(),
                TextColumn("[progress.description]{task.description}"),
                BarColumn(),
                console=config.console,
                transient=True
            ) as progress:
                progress.add_task(f"[cyan]{market_type} 데이터 압축 해제 중...[/cyan]", total=None)
                with zipfile.ZipFile(zip_path, 'r') as zip_ref:
                    zip_ref.extractall(base_dir)
        
        # [수정] 파일 파싱 시 Progress Bar 적용 (파일 크기 기준)
        file_size = os.path.getsize(extract_path)
        with Progress(
            SpinnerColumn(),
            TextColumn("[progress.description]{task.description}"),
            BarColumn(),
            TextColumn("[progress.percentage]{task.percentage:>3.0f}%"),
            console=config.console,
            transient=True
        ) as progress:
            task = progress.add_task(f"[cyan]{market_type} 종목 리스트 파싱 중...[/cyan]", total=file_size)
            
            with open(extract_path, 'rb') as f:
                for line in f:
                    progress.advance(task, advance=len(line))
                    try:
                        code = line[0:9].decode('cp949').strip()
                        name = line[21:61].decode('cp949').strip()

                        # [추가 2026-08-11] 증권그룹구분코드(ST=주권, EF=ETF, RT=리츠, IF=인프라 …).
                        #  종목명 브랜드 접두어로 ETF를 거르던 방식이 신규 운용사를 못 잡아
                        #  ('DS 코스닥액티브' 0220B0), 마스터가 이미 들고 있는 정본 필드로 바꾼다.
                        #  마스터 앞단(코드+표준코드+한글명)은 가변폭이라 앞에서 못 세고, 뒷단은
                        #  고정폭이라 뒤에서 센다. 개행을 먼저 떼야 한 칸씩 밀리지 않는다
                        #  (오프셋은 실측 확정 — KOSPI 227 / KOSDAQ 221에서 전 레코드 유효:
                        #   코스피 2,559건·코스닥 1,820건 모두 정상 코드로 파싱된다).
                        try:
                            body = line.rstrip(b'\r\n')
                            grp = body[-_MST_GRP_OFFSET:][0:2].decode('cp949', 'ignore').strip()
                        except Exception:
                            grp = ""

                        # [수정] 영문이 포함된 최신 ETF(예: 0080G0)를 지원하기 위해 숫자로 시작하는 6자리 영문/숫자 코드로 완화
                        if len(code) == 6 and code[0].isdigit() and code.isalnum():
                            stock_list.append({'code': code, 'name': name, 'grp': grp})
                    except Exception:
                        continue
    except Exception as e:
        config.console.print(f"[red]{market_type} 마스터 파일 처리 실패: {e}[/red]")
        
    return stock_list

# [추가] 마스터 코드 기반 시장 구분 캐시
_MASTER_KOSDAQ_CODES = None
_MASTER_KOSPI_CODES = None

def _get_market_type_by_master(code):
    """마스터 파일(KOSPI/KOSDAQ)을 참조하여 종목의 시장 구분을 정확히 반환합니다.

    KOSDAQ/KOSPI 마스터 양쪽을 모두 조회한다.
      - KOSDAQ 마스터에 있으면 'KOSDAQ'
      - KOSPI 마스터에 있으면 'KOSPI'
      - 둘 다 없으면(신규상장/누락 등) 보수적으로 'KOSPI'로 폴백
    """
    global _MASTER_KOSDAQ_CODES, _MASTER_KOSPI_CODES
    if _MASTER_KOSDAQ_CODES is None:
        try:
            k_list = _get_master_stock_list("KOSDAQ")
            _MASTER_KOSDAQ_CODES = set(s['code'] for s in k_list)
        except Exception as e:
            logger.debug(f"KOSDAQ 마스터 목록 로드 실패: {e}")
            _MASTER_KOSDAQ_CODES = set()
    if _MASTER_KOSPI_CODES is None:
        try:
            p_list = _get_master_stock_list("KOSPI")
            _MASTER_KOSPI_CODES = set(s['code'] for s in p_list)
        except Exception as e:
            logger.debug(f"KOSPI 마스터 목록 로드 실패: {e}")
            _MASTER_KOSPI_CODES = set()

    if code in _MASTER_KOSDAQ_CODES:
        return "KOSDAQ"
    if code in _MASTER_KOSPI_CODES:
        return "KOSPI"
    # 어느 마스터에도 없으면(신규상장 등) 보수적 폴백
    logger.debug(f"마스터 미발견 종목({code}) → KOSPI로 폴백 처리")
    return "KOSPI"

def diagnose_stock(target_code=None, target_name=None, target_is_overseas=False):
    """특정 종목에 대해 시스템 트레이딩 로직을 진단(시뮬레이션)합니다."""
    
    code, name, is_overseas = None, None, False

    if target_code:
        code, name, is_overseas = target_code, target_name, target_is_overseas
    else:
        menu_items = [
            ("1", "국내 주식", "Domestic Stock"), ("2", "국내 ETF", "Domestic ETF"),
            ("3", "미국 주식", "US Stock"), ("4", "미국 ETF", "US ETF"), ("5", "직접 입력", "Direct Input")
        ]
        choice = utils.show_menu("개별 종목 분석 (Individual Analysis)", menu_items, default_choice="1")
        if choice.lower() in ['b', 'q']: return False
        
        menu_map = dict((k, v) for k, v, _ in menu_items)
        context.USER_ACTION_BREADCRUMB.append(f"[{choice}] {menu_map.get(choice, '')}")

        if choice == '5':
            # 직접 입력 로직
            from modules import manage
            # manage.get_current_price는 출력을 포함하므로, 여기서는 간단히 입력만 받음
            utils.print_breadcrumb()
            raw_input = Prompt.ask("종목코드(6자리/티커) 또는 종목명 [dim](이전: b, 메인: q)[/dim]")
            config.console.print()
            if not raw_input or raw_input.lower() in ['b', 'q']: return
            context.USER_ACTION_BREADCRUMB.append(f"[직접입력] {raw_input}")
            
            # manage 모듈의 _resolve_stock 로직과 유사하게 처리하거나 utils 활용
            # 여기서는 utils가 없으므로 telegram_bot의 로직을 참고하여 간단히 구현
            if len(raw_input) == 6 and raw_input[0].isdigit() and raw_input.isalnum():
                code = raw_input
                name = api.get_stock_name_by_code(code, False) or code
                is_overseas = False
            elif all(ord(c) < 128 for c in raw_input): # 해외 티커 가정
                code = raw_input.upper()
                name = api.get_stock_name_by_code(code, True) or code
                is_overseas = True
            else:
                # 한글 종목명 검색 시도 (config.session.stock_data 활용)
                found = False
                for key in ['stocks_kr', 'etfs_kr']:
                    for item in config.session.stock_data.get(key, []):
                        if item['name'] == raw_input:
                            code, name, is_overseas = item['code'], item['name'], False
                            found = True; break
                    if found: break
                if not found:
                    for key in ['stocks_us', 'etfs_us']:
                        for item in config.session.stock_data.get(key, []):
                            if item['name'].lower() == raw_input.lower():
                                code, name, is_overseas = item['code'], item['name'], True
                                found = True; break
                        if found: break
                
                if not found:
                    config.console.print(f"[red]'{raw_input}'을(를) 찾을 수 없습니다. 코드로 입력해주세요.[/red]")
                    return False
                    
            if not utils.validate_and_confirm_stock(code, name, is_overseas, "이 종목으로 분석을 진행하시겠습니까?"):
                return False
        else:
            # 리스트 선택
            key_map = {"1": "stocks_kr", "2": "etfs_kr", "3": "stocks_us", "4": "etfs_us"}
            target_key = key_map.get(choice)
            stock_list = config.session.stock_data.get(target_key, [])
            
            if not stock_list:
                config.console.print("[yellow]등록된 종목이 없습니다.[/yellow]")
                return False
                
            idx, item = utils.search_stock_in_list(stock_list, title=f"{menu_map[choice]} 목록")
            if not item: return False
            code, name = item['code'], item['name']
            is_overseas = (choice in ["3", "4"])
            context.USER_ACTION_BREADCRUMB.append(f"[종목선택] {name}")

    if not code: return False

    logger.info(f"운영자 실행: {' - '.join(context.USER_ACTION_BREADCRUMB)}")

    # [수정] 종목 선택 후 데이터 분석 시작 (UI 응답성 개선)
    with Progress(
        SpinnerColumn(),
        TextColumn("[progress.description]{task.description}"),
        BarColumn(),
        console=config.console,
        transient=True
    ) as progress:
        # [추가] 개별 룰 로드 및 설정 준비
        custom_rule = db_manager.db.get_stock_strategy(code)
        rule_applied = False
        
        # 기본값 설정
        buy_score = config.ANALYSIS_THRESHOLDS["BUY_SCORE"]
        buy_rsi = config.ANALYSIS_THRESHOLDS["BUY_RSI_MAX"]
        rise_score = config.ANALYSIS_THRESHOLDS["RISE_SCORE"]
        buy_vol = config.ANALYSIS_THRESHOLDS.get("BUY_VOL_STRENGTH", 100.0)
        weights = config.SCORING_WEIGHTS
        
        if custom_rule:
            rule_applied = True
            buy_score = custom_rule['buy_score']
            buy_rsi = custom_rule['buy_rsi']
            if custom_rule.get('weights'):
                try:
                    w_data = custom_rule['weights']
                    if isinstance(w_data, str): weights = json.loads(w_data)
                    elif isinstance(w_data, dict): weights = w_data
                except Exception: pass
            if custom_rule.get('buy_vol_strength'):
                buy_vol = custom_rule['buy_vol_strength']

        # [추가] stock.json에서 표준 시장(거래소) 이름 가져오기
        std_market = None
        for key in ["stocks_kr", "etfs_kr", "stocks_us", "etfs_us"]:
            for item in config.session.stock_data.get(key, []):
                if item.get('code') == code and "exchange" in item:
                    std_market = item['exchange'].upper()
                    break
            if std_market: break

        foreign_rate_str = "[dim]-[/dim]"
        market_str = std_market if std_market else ("해외" if is_overseas else "KOSPI")
        # [추가] 적응형 임계값 적용 (시장 국면 보정)
        score_adj = 0.0
        is_domestic_index = not is_overseas and code in ["KOSPI", "KOSDAQ", "KOSPI200", "KOSDAQ150", "VKOSPI", "K200FUT_F", "K200FUT_CM"]

        from modules import market
        all_idx_codes = [c for n, c in market.ALL_INDICES]
        is_index = is_domestic_index or (is_overseas and code in all_idx_codes)
        
        task = progress.add_task("[cyan]분석 준비 중...[/cyan]", total=None)

        if not is_overseas and not is_domestic_index:
            progress.update(task, description="[cyan]시장 국면 및 수급 정보 조회 중...[/cyan]")
            try:
                # API로 시장 구분 및 외인 소진율 확인
                cp = api.get_current_price_data(code, False)
                if cp.get('rt_cd') == '0':
                    foreign_rate_str = f"{cp['output'].get('hts_frgn_ehrt', '-')}%"
                    
                    # [수정] 국내 주식 현재가 API 응답에는 시장구분 필드가 없으므로 마스터 파일을 이용해 판별
                    market_type = _get_market_type_by_master(code)
                        
                    # [수정] std_market이 잘못 캐시되어 있는 경우를 대비하여 실시간 API 조회값을 최우선 반영
                    market_str = market_type
                    
                    if config.MARKET_REGIME_PARAMS.get("USE_ADAPTIVE_THRESHOLD", True):
                        regime, score_adj = get_market_regime(market_type)
                        if score_adj != 0 and not rule_applied: # [수정] 개별 룰이 없을 때만 보정 적용
                            buy_score += score_adj
            except Exception: pass
        elif is_domestic_index:
            market_type = "KOSDAQ" if "KOSDAQ" in code else "KOSPI"
            market_str = code
            if config.MARKET_REGIME_PARAMS.get("USE_ADAPTIVE_THRESHOLD", True):
                try:
                    regime, score_adj = get_market_regime(market_type)
                    if score_adj != 0 and not rule_applied:
                        buy_score += score_adj
                except Exception: pass
        else:
            if not std_market:
                cached_ex = config.session.exchange_cache.get(code)
                if cached_ex:
                    if cached_ex in ["NAS", "NASD"]: market_str = "NASDAQ"
                    elif cached_ex in ["NYS", "NYSE"]: market_str = "NYSE"
                    elif cached_ex in ["AMS", "AMEX"]: market_str = "AMEX"
                    else: market_str = cached_ex
                elif is_index:
                    try:
                        import yfinance as yf
                        config.silence_yfinance_numpy_warning()  # import 뒤에 걸어야 억제 유효
                        tk = yf.Ticker(code)
                        ex = getattr(tk.fast_info, 'exchange', None)
                        if not ex:
                            ex = tk.info.get('exchange')
                        if ex:
                            ex_str = str(ex).upper()
                            ex_map = {
                                "NMS": "NASDAQ", "NYQ": "NYSE", "ASE": "AMEX",
                                "PNK": "OTC", "CMX": "COMEX", "NYM": "NYMEX",
                            "CBT": "CBOT", "CCY": "Currency (통화/환율)",
                            "CCC": "Crypto (암호화폐)",
                                "PHX": "PHLX (필라델피아)", "SNP": "S&P Index",
                            "CBOE": "CBOE (시카고옵션)", "DJI": "Dow Jones",
                            "CME": "CME (시카고상품거래소)", "NYB": "NYBOT (뉴욕상품거래소)",
                            "OSA": "Osaka (오사카)", "TAI": "Taiwan (대만)",
                            "HKG": "Hong Kong (홍콩)", "SHH": "Shanghai (상해)",
                            "FRA": "Frankfurt (프랑크푸르트)", "STO": "STOXX (유럽)"
                            }
                            market_str = ex_map.get(ex_str, ex_str)
                    except Exception:
                        pass

        # [추가] 임계값 및 가중치 딕셔너리 구성
        thresholds = {
            "BUY_SCORE": buy_score,
            "BUY_RSI_MAX": buy_rsi,
            "RISE_SCORE": rise_score,
            "WEIGHTS": weights
        }

        # 1. [최적화] 데이터 병렬 조회 (차트 캐시 확인 및 체결강도 동시 호출)
        progress.update(task, description=f"[cyan]{name}({code}) 지표 및 수급 데이터 병렬 수집 중...[/cyan]")
        
        df = None
        vol_strength = None
        inv_data = None
        ask_bid_ratio = None

        # [추가] 미국채 금리는 지수 화면과 동일하게 tvDatafeed 현물(TVC:USxxY)을 1차 소스로
        #  사용한다 — yfinance 경로만 쓰면 표(현물)와 심층 분석(2년물은 죽은 선물) 값이 어긋난다.
        treasury_sym = config.US_TREASURY_SPOT_SYMBOLS.get(name) if is_overseas else None

        with concurrent.futures.ThreadPoolExecutor(max_workers=4) as ex:
            if is_domestic_index:
                fut_chart = ex.submit(get_domestic_index_data, code)
                fut_vol = None
                fut_inv = None
                fut_ab = None
            elif treasury_sym:
                fut_chart = ex.submit(get_us_treasury_spot_data, treasury_sym)
                fut_vol = None
                fut_inv = None
                fut_ab = None
            else:
                fut_chart = ex.submit(api.get_chart_data, code, is_overseas=is_overseas)
                fut_vol = ex.submit(api.get_realtime_vol_strength, code) if not is_overseas else None
                fut_inv = ex.submit(api.get_investor_trend, code) if not is_overseas else None
                # 수급 게이트용 비율만 필요 → WS 우선(get_ask_bid_ratio)으로 호가 REST 절감
                fut_ab = ex.submit(api.get_ask_bid_ratio, code, False) if not is_overseas else None

            df = fut_chart.result()
            vol_strength = fut_vol.result() if fut_vol else None
            inv_data = fut_inv.result() if fut_inv else None
            ask_bid_ratio = fut_ab.result() if fut_ab else None

        # [추가] 국채 현물 실패 폴백: 5/10/30년은 야후 현물 지수(^FVX류)로 대체(죽은 값 아님),
        #  2년물은 대체 소스가 없어(자리표시자 ^US02Y) 아래 공통 오류 처리로 종료된다.
        if treasury_sym and (df is None or df.empty) and code != "^US02Y":
            df = api.get_chart_data(code, is_overseas=True)
            treasury_sym = None  # yfinance 소스로 진행 (실시간 보정도 야후 기준)

        # [Fix] 국내 지수·국채 현물 df는 공유 캐시 객체이므로 복사 후 사용
        #  (apply_realtime_price의 당일 봉 덮어쓰기/추가가 캐시를 오염시키지 않도록)
        if (is_domestic_index or treasury_sym) and df is not None:
            df = df.copy()

        if df is None or df.empty:
            config.console.print("[red]차트 데이터를 불러올 수 없습니다.[/red]")
            return

        # [추가] 실시간 현재가 조회 및 차트 데이터 최신화 (점수 불일치 방지)
        #  (국채 현물은 TV 일봉 마지막 봉이 실시간 값 — 야후 시세를 덮어쓰면 소스가 섞이므로 제외)
        #  [Fix 2026-07-28] 봉 반영(지표)은 KRX 정규장에만, 실시간가 조회는 '표시에 쓸 일이 있으면'
        #   한다. 모든 장 마감 후 KRX 고정(설정 True)이면 조회 자체를 생략해 종전처럼 TPS를 아낀다.
        rt_price = 0.0
        if not treasury_sym:
            try:
                if api.chart_overlay_enabled(is_overseas) or not api.display_price_krx_fixed(is_overseas):
                    rt_price = float(api.get_current_price(code, is_overseas=is_overseas) or 0)
                # [SSOT] 봉 반영 여부는 api.chart_overlay_price 가 단독으로 정한다.
                #  종전에는 그 함수의 내부 조건(chart_overlay_enabled)을 여기서 손으로
                #  한 번 더 썼다 — 값은 같았지만 게이트에 조건이 하나 늘면 이 자리만
                #  따라오지 않는다. 위 조회 여부 판단에는 여전히 enabled 가 필요하다
                #  (표시에도 쓸 일이 없으면 API 호출 자체를 아낀다).
                _overlay = api.chart_overlay_price(rt_price, is_overseas)
                if _overlay > 0:
                    indicators.apply_realtime_price(df, _overlay, market_date=utils.market_today(is_overseas))
            except Exception: pass

        # 2. 지표 계산
        progress.update(task, description="[cyan]기술적 지표 계산 및 상태 분류 중...[/cyan]")
        ind = indicators.calculate_indicators(df)

        # 전일 RSI — calculate_indicators가 계산한 값 재사용 (중복 계산 제거·SSOT)
        prev_rsi = ind.get('prev_rsi') if df is not None and not df.empty and len(df) >= 16 else None

        current_price = float(df.iloc[-1]['close'])   # [판단 기준] 지표·상태·점수·이격도·52주 위치
        # [표시 전용] NXT 거래시간(프리 08:00~09:00 / 애프터 15:30~20:00)에는 살아있는 실시간가를
        #  현재가로 보여준다. 다만 지표는 KRX 확정 봉으로 계산하므로, 같은 화면에서 기준이 갈리는
        #  것을 감추지 않도록 (NXT) 표기를 함께 붙인다.
        display_price = current_price
        display_tag = ""
        if rt_price > 0 and not is_overseas and not api.chart_overlay_enabled(False) \
                and not api.display_price_krx_fixed(False):
            display_price = rt_price
            display_tag = " [dim](NXT)[/dim]"
        # 등락 기준봉: 마지막 봉이 '최신 확정 세션'이면 직전 봉과 비교하고, 아직 그 봉이 없으면
        #  (프리마켓 등) 마지막 봉 자체가 직전 종가다 — 그대로 두면 등락이 하루 밀린다.
        #  [Fix 2026-07-28] 국내는 market_today 대신 krx_last_settled_day를 쓴다. 자정~개장 전에는
        #   market_today가 아직 열리지도 않은 '오늘'이라, 확정 봉을 갖고도 등락이 0%로 나왔다.
        try:
            _ref_day = utils.market_today(True) if is_overseas else api.krx_last_settled_day()
            _last_bar_is_today = str(df.iloc[-1]['date']).replace('-', '')[:8] >= _ref_day
        except Exception:
            _last_bar_is_today = True
        if _last_bar_is_today:
            prev_price = float(df.iloc[-2]['close']) if len(df) > 1 else current_price
        else:
            prev_price = current_price
        diff = display_price - prev_price
        rate = (diff / prev_price) * 100 if prev_price > 0 else 0.0
        
        # 52주 위치 계산 (슈퍼 모멘텀 판정용)
        w52_pos = 0.0
        h52 = 0.0
        l52 = 0.0
        high_52_rate = 0.0
        if len(df) > 0:
            h52, l52 = _w52_band(df)   # 표(_analyze_table_row)와 동일한 365일 창
            if h52 > l52:
                w52_pos = (current_price - l52) / (h52 - l52) * 100
            if h52 > 0:
                high_52_rate = ((current_price - h52) / h52) * 100
        
        sm_flag, sm_reason = (False, "") if is_domestic_index else check_smart_money_turnaround(code, is_overseas)
        
        # 3. 상태 분류 및 점수 계산
        state, state_color, state_reason = classify_stock_state(
            df=df, ind=ind, prev_rsi=prev_rsi, thresholds=thresholds, w52_pos=w52_pos, smart_money=sm_flag
        )
        
        score, details = calculate_score(
            df=df, ind=ind, weights=weights, smart_money=sm_flag
        )

    # 4. 결과 출력
    config.console.print()

    # [경고] KRX 공식 일봉 실패로 토스 캔들(NXT 포함)이 쓰였으면 결과 앞에 알린다
    #  부가 정보이므로 실패해도 분석 결과 출력을 막지 않는다.
    if not is_overseas:
        try:
            utils.print_krx_fallback_warning({str(code): name})
        except Exception as e:
            logger.debug(f"KRX 폴백 경고 출력 실패: {e}")

    # [추가] 종목 메모 출력 (존재 시 패널 형태로 상단에 표시)
    memo_data_list = utils.get_stock_memos(code)
    if memo_data_list:
        from rich.panel import Panel
        from rich.rule import Rule
        from rich.console import Group
        from rich.text import Text
        
        renderables = []
        for i, m in enumerate(memo_data_list):
            if i > 0:
                renderables.append(Rule(style="dim"))
            text = Text.from_markup(f"[dim]{m['updated_at']}[/dim]\n{m['memo']}")
            renderables.append(text)
            
        config.console.print(Panel(Group(*renderables), title=f"{name} ({code})", border_style="cyan", expand=False))
        config.console.print()

    # [추가] TradingView 종합 평가 및 배당 수익률 등 추가 데이터 조회
    tv_rating_str = "조회 불가"
    div_yield_str = "[dim]-[/dim]"
    try:
        from tradingview_screener import Query, Column
        tv_market = 'america' if is_overseas else 'korea'
        _, tv_df = Query().set_markets(tv_market).select('Recommend.All', 'dividend_yield_recent').where(Column('name') == code).limit(1).get_scanner_data()
        if tv_df is not None and not tv_df.empty:
            rating_val = tv_df.iloc[0].get('Recommend.All')
            if pd.notna(rating_val):
                if rating_val >= 0.5: tv_rating_str = f"[bold red]Strong Buy (강력 매수)[/bold red] ({rating_val:+.2f})"
                elif rating_val >= 0.1: tv_rating_str = f"[red]Buy (매수)[/red] ({rating_val:+.2f})"
                elif rating_val > -0.1: tv_rating_str = f"[white]Neutral (중립)[/white] ({rating_val:+.2f})"
                elif rating_val > -0.5: tv_rating_str = f"[blue]Sell (매도)[/blue] ({rating_val:+.2f})"
                else: tv_rating_str = f"[bold blue]Strong Sell (강력 매도)[/bold blue] ({rating_val:+.2f})"
            
            div_val = tv_df.iloc[0].get('dividend_yield_recent')
            if pd.notna(div_val) and div_val > 0:
                div_yield_str = f"{div_val:.2f}%"
    except Exception: pass

    # [테이블 1] 기술적 지표 분석
    tech_title = f"기술적 지표 분석: {name} ({code})"

    table_tech = Table(title=tech_title, box=box.HORIZONTALS, header_style="dim", border_style="dim")
    table_tech.add_column("지표", justify="left", style="cyan", width=15)
    table_tech.add_column("값 (상태)", justify="left")
    table_tech.add_column("해석/기준", justify="left", style="dim")

    # 시장 정보
    table_tech.add_row("시장", market_str, "소속 거래소")

    # 현재가 ([통일] 지수 화면·종목 표와 동일 규칙 — price_trend_color 단일 소스)
    #  [표시] display_price는 NXT 거래시간에만 current_price와 갈린다(그 외에는 동일 값).
    curr_price_color = price_trend_color(display_price, ind.get('ema_20'), ind.get('ema_60'), ind=ind)

    if is_index:
        price_str_tech = f"{display_price:,.0f}" if display_price >= 1000 else f"{display_price:,.2f}"
        h52_str = f"{h52:,.0f}" if h52 >= 1000 else f"{h52:,.2f}"
        l52_str = f"{l52:,.0f}" if l52 >= 1000 else f"{l52:,.2f}"
    else:
        price_str_tech = f"${display_price:,.2f}" if is_overseas else f"{display_price:,.0f}원"
        h52_str = f"${h52:,.2f}" if is_overseas else f"{int(h52):,}원"
        l52_str = f"${l52:,.2f}" if is_overseas else f"{int(l52):,}원"
        
    table_tech.add_row("현재가", f"{curr_price_color}{price_str_tech}[/]{display_tag}",
                       "이평선 배열 및 위치 기반" + (" (지표는 KRX 종가 기준)" if display_tag else ""))

    # ATR (변동성)
    atr_val = ind.get('atr', 0)
    vol_str = "-"
    if atr_val > 0 and current_price > 0:
        annual_vol = (atr_val / current_price) * math.sqrt(252) * 100
        vol_str = f"{annual_vol:.1f}%"
    
    table_tech.add_row("변동성 (ATR)", f"{int(atr_val):,} ({vol_str})", "연환산 변동성 (리스크)")

    # 52주 위치
    w_color = "[white]"
    if w52_pos >= 90: w_color = "[red]"
    elif w52_pos >= 80: w_color = "[orange3]"
    elif w52_pos <= 30: w_color = "[blue]"
    elif w52_pos <= 50: w_color = "[yellow]"
    
    table_tech.add_row("52주 위치", f"{w_color}{w52_pos:.1f}%[/] [dim]({l52_str} ~ {h52_str})[/dim]", "최고가/최저가 밴드 내 현 위치")

    # 추세품질(TQ) — 진입 순위에서 점수 동점을 가르는 값(연환산 회귀 기울기 × R²).
    #  종목 표(분류 옆)와 자동매매 후보 로그에는 이미 나오는데 개별 분석 표에만 없었다.
    #  같은 '매수'라도 검증된 추세인지가 여기서 갈리므로 추세 지표들 맨 위에 둔다.
    #  색·밴드는 indicators.TREND_QUALITY_COLORS/BANDS 단일 소스 — 도움말 표와 같다.
    tq_lb = config.INDICATOR_PARAMS.get("TREND_QUALITY_LOOKBACK", 90)
    tq_val = indicators.get_trend_quality(df, lookback=tq_lb)
    tq_band = indicators.describe_trend_quality(tq_val)
    tq_color = indicators.TREND_QUALITY_COLORS.get(tq_band, "white")
    if tq_val is None:
        tq_str = f"[dim]- ({tq_band})[/dim]"
    else:
        tq_str = f"[{tq_color}]{tq_val:,.1f} ({tq_band})[/]"
    table_tech.add_row(f"추세품질 ({tq_lb}일)", tq_str,
                       "연환산 기울기 × R² (진입 순위 동점 가름)")

    # SAR
    sar_val = ind.get('psar')
    if sar_val is not None and not math.isnan(sar_val):
        sar_pos = "주가 아래 (상승)" if sar_val < current_price else "주가 위 (하락)"
        sar_color = "[red]" if sar_val < current_price else "[blue]"
    else:
        sar_pos = "-"
        sar_color = "[dim]"
    table_tech.add_row("SAR 위치", f"{sar_color}{sar_pos}[/]", "파라볼릭 추세 전환")

    # MACD
    macd_val = ind.get('macd')
    sig_val = ind.get('macd_signal')
    
    macd_str = "[dim]-[/dim]"
    macd_desc = "추세 확인"
    if macd_val is not None and sig_val is not None and not math.isnan(macd_val) and not math.isnan(sig_val):
        m_color = "[red]" if macd_val >= sig_val else "[blue]"
        macd_str = f"{m_color}{macd_val:+.2f}[/]"
        
        cross_desc = "골든 (매수 우위)" if macd_val >= sig_val else "데드 (매도 우위)"
        phase_desc = "상승 국면" if macd_val > 0 else "하락 국면"
        macd_desc = f"{cross_desc} / {phase_desc}"
        
        # 시그널 값도 참고용으로 작게 표시
        macd_str += f" [dim](Sig: {sig_val:+.2f})[/dim]"
            
    table_tech.add_row("MACD (12/26/9)", macd_str, macd_desc)

    # OBV
    obv_trend = ind.get('obv_trend')
    obv_val = ind.get('obv')
    vol_sum = df['volume'].tail(5).sum() if df is not None and 'volume' in df.columns else 0
    
    if df is None or len(df) < config.INDICATOR_PARAMS.get("OBV_MA_PERIOD", 5):
        obv_trend = None
        obv_val = None
        
    if vol_sum == 0 or obv_trend is None or obv_val is None or math.isnan(obv_val):
        obv_trend_str = '-'
        obv_color = "[dim]"
        obv_desc = "거래량 데이터 없음"
        obv_formatted = f"{obv_color}{obv_trend_str}[/]"
    else:
        obv_trend_str = '상승' if obv_trend else '하락'
        obv_color = "[red]" if obv_trend else "[blue]"
        obv_desc = "이동평균 상회 여부"
        abs_val = abs(obv_val)
        if abs_val >= 999_950_000_000: obv_str = f"{obv_val/1_000_000_000_000:,.1f}T"
        elif abs_val >= 999_950_000: obv_str = f"{obv_val/1_000_000_000:,.1f}B"
        elif abs_val >= 999_500: obv_str = f"{obv_val/1_000_000:,.1f}M"
        elif abs_val >= 999.5: obv_str = f"{obv_val/1_000:,.0f}K"
        else: obv_str = f"{obv_val:,.0f}"
        obv_formatted = f"{obv_color}{obv_trend_str}[/] [dim]({obv_str})[/dim]"
        
    table_tech.add_row("OBV 추세", obv_formatted, obv_desc)
    
    # RSI
    rsi_val = ind.get('rsi')
    if rsi_val is not None:
        rsi_str = f"{rsi_val:.2f}"
        rsi_desc = ""
        if rsi_val >= config.INDICATOR_PARAMS["RSI_UPPER"]: 
            rsi_str = f"[magenta]{rsi_str}[/]"
            rsi_desc = "과열 (추격금지)"
        elif 55 <= rsi_val < config.INDICATOR_PARAMS["RSI_UPPER"]: 
            rsi_str = f"[red]{rsi_str}[/]"
            rsi_desc = "강세 유지"
        elif 45 <= rsi_val < 55: 
            rsi_str = f"[orange3]{rsi_str}[/]"
            rsi_desc = "강세 조정 (진입후보)"
        elif config.INDICATOR_PARAMS["RSI_LOWER"] < rsi_val < 45: 
            rsi_str = f"[yellow]{rsi_str}[/]"
            rsi_desc = "약세/하락전환 가능"
        else: 
            rsi_str = f"[blue]{rsi_str}[/]"
            rsi_desc = "침체 (과매도)"
    else:
        rsi_str = "[dim]-[/dim]"
        rsi_desc = "데이터 부족"
    table_tech.add_row("RSI (14)", f"{rsi_str} [dim]({rsi_desc})[/dim]", "과매수(70)/과매도(30)")

    # CCI
    cci_val = ind.get('cci')
    if cci_val is not None:
        cci_str = f"{cci_val:.2f}"
        cci_desc = ""
        if cci_val >= config.INDICATOR_PARAMS["CCI_UPPER"]: 
            cci_str = f"[red]{cci_str}[/]"
            cci_desc = "과열 (추격 금물)"
        elif 0 < cci_val < config.INDICATOR_PARAMS["CCI_UPPER"]: 
            cci_str = f"[orange3]{cci_str}[/]"
            cci_desc = "상승 추세"
        elif config.INDICATOR_PARAMS["CCI_LOWER"] < cci_val <= 0: 
            cci_str = f"[yellow]{cci_str}[/]"
            cci_desc = "반등 시도"
        else: 
            cci_str = f"[blue]{cci_str}[/]"
            cci_desc = "과매도 (저점 탐색)"
    else:
        cci_str = "[dim]-[/dim]"
        cci_desc = "데이터 부족"
    table_tech.add_row("CCI (20)", f"{cci_str} [dim]({cci_desc})[/dim]", "추세 및 과매수/매도")

    # ADX
    adx_val = ind.get('adx')
    if adx_val is not None:
        adx_str = f"{adx_val:.2f}"
        adx_desc = ""
        if adx_val >= 40: 
            adx_str = f"[magenta]{adx_str}[/]" 
            adx_desc = "과열 (조정 주의)"
        elif adx_val >= 30: 
            adx_str = f"[red]{adx_str}[/]"     
            adx_desc = "강한 추세"
        elif adx_val >= 20: 
            adx_str = f"[orange3]{adx_str}[/]"
            adx_desc = "안정적 추세"
        elif adx_val >= 15: 
            adx_str = f"[yellow]{adx_str}[/]"
            adx_desc = "추세 형성 중"
        else: 
            adx_str = f"[white]{adx_str}[/]"
            adx_desc = "추세 없음 (횡보)"
    else:
        adx_str = "[dim]-[/dim]"
        adx_desc = "데이터 부족"
    table_tech.add_row("ADX (14)", f"{adx_str} [dim]({adx_desc})[/dim]", "추세 강도 (25 이상 강세)")

    # DMI
    plus_di = ind.get('plus_di')
    minus_di = ind.get('minus_di')
    if plus_di is not None and minus_di is not None:
        if plus_di > minus_di:
            dmi_str = f"[red]{plus_di:.1f}[/] / [dim]{minus_di:.1f}[/]"
            dmi_desc = "+DI 우위"
        elif minus_di > plus_di:
            dmi_str = f"[dim]{plus_di:.1f}[/] / [blue]{minus_di:.1f}[/]"
            dmi_desc = "-DI 우위"
        else: 
            dmi_str = f"{plus_di:.1f} / {minus_di:.1f}"
            dmi_desc = "중립"
    else:
        dmi_str = "[dim]- / -[/dim]"
        dmi_desc = "데이터 부족"
    table_tech.add_row("DMI (+DI/-DI)", f"{dmi_str} [dim]({dmi_desc})[/dim]", "매수/매도 세력 강도 (+DI 상승)")

    # [수정] 외인 소진율 및 배당 수익률 위치 변경 (이평 배열 위로 이동)
    if not is_overseas and not is_domestic_index:
        table_tech.add_row("외인 소진율", foreign_rate_str, "외국인 보유 비중")

    if not is_domestic_index:
        table_tech.add_row("배당 수익률", div_yield_str, "최근 연환산 배당수익률")

    # 이평 배열
    ema_align = "알 수 없음"
    ema_color = "[white]"
    if ind['ema_20'] is not None and ind['ema_60'] is not None and ind['ema_120'] is not None:
        if ind['ema_20'] > ind['ema_60'] > ind['ema_120']: 
            ema_align = "정배열 (20>60>120)"; ema_color = "[red]"
        elif ind['ema_20'] < ind['ema_60'] < ind['ema_120']: 
            ema_align = "역배열 (20<60<120)"; ema_color = "[blue]"
        else: 
            ema_align = "혼조세"; ema_color = "[yellow]"
    else:
        ema_align = "데이터 부족"; ema_color = "[dim]"
    table_tech.add_row("이평 배열", f"{ema_color}{ema_align}[/]", "5/20/60/120일선 배열")
    
    # 가격·52주와 같은 기준으로 가른다 — 지수는 통화 기호 없이, 원화 종목(KRX 금현물 등)은 원.
    _usd_ema = utils.is_usd_quoted(code) and not is_index
    def _fmt_ema(v): return f"${v:,.2f}" if _usd_ema else f"{int(v):,}"
    
    v5 = ind.get('ema_5')
    v20 = ind.get('ema_20')
    v60 = ind.get('ema_60')
    v120 = ind.get('ema_120')

    c5 = "[red]" if (v5 is not None and v20 is not None and v5 > v20) else ("[blue]" if v5 is not None else "")
    c20 = "[red]" if (v20 is not None and v60 is not None and v20 > v60) else ("[blue]" if v20 is not None else "")
    c60 = "[red]" if (v60 is not None and v120 is not None and v60 > v120) else ("[blue]" if v60 is not None else "")
    
    c120 = ""
    if df is not None and not df.empty and len(df) > 121:
        try:
            ema120_series = df['close'].ewm(span=120, adjust=False).mean()
            if ema120_series.iloc[-1] > ema120_series.iloc[-2]: c120 = "[red]"
            else: c120 = "[blue]"
        except Exception: pass
    elif v120 is not None:
        c120 = "[blue]"
        
    e5_disp = f"{c5}{_fmt_ema(v5)}[/]" if v5 is not None else "-"
    e20_disp = f"{c20}{_fmt_ema(v20)}[/]" if v20 is not None else "-"
    e60_disp = f"{c60}{_fmt_ema(v60)}[/]" if v60 is not None else "-"
    e120_disp = f"{c120}{_fmt_ema(v120)}[/]" if v120 is not None else "-"

    table_tech.add_row("주요 이평선", f"5선: {e5_disp} | 20선: {e20_disp} | 60선: {e60_disp} | 120선: {e120_disp}", "지수이동평균(EMA) 가격")

    # 이격도
    d_5 = (current_price / ind['ema_5'] * 100) if ind.get('ema_5') is not None else 0
    d_20 = (current_price / ind['ema_20'] * 100) if ind['ema_20'] is not None else 0
    d_60 = (current_price / ind['ema_60'] * 100) if ind['ema_60'] is not None else 0
    d_120 = (current_price / ind['ema_120'] * 100) if ind['ema_120'] is not None else 0
    
    def dc(val): return "[red]" if val >= 100 else "[blue]"
    
    disp_msg = f"5선({dc(d_5)}{d_5:.1f}%[/]) 20선({dc(d_20)}{d_20:.1f}%[/]) 60선({dc(d_60)}{d_60:.1f}%[/]) 120선({dc(d_120)}{d_120:.1f}%[/])"
    
    disp_upper = config.ANALYSIS_THRESHOLDS.get("DISPARITY_UPPER", 110)
    disp_lower = config.ANALYSIS_THRESHOLDS.get("DISPARITY_LOWER", 90)

    disp_eval = ""
    if d_20 >= disp_upper: disp_eval = "[bold red]단기 과열[/]"
    elif d_20 <= disp_lower: disp_eval = "[bold blue]과매도[/]"
    else: disp_eval = "[white]적정 범위[/]"
    
    table_tech.add_row("이격도", disp_msg, f"{disp_eval} [dim](현재가/이평선)[/dim]")

    # [수정] 스마트머니를 표의 가장 아래로 이동
    if not is_overseas and not is_domestic_index:
        inv_str = "[dim]-[/dim]"
        if inv_data and len(inv_data) > 0:
            item = inv_data[0]
            p = api.safe_int(item.get('prsn_ntby_qty'))
            f = api.safe_int(item.get('frgn_ntby_qty'))
            o = api.safe_int(item.get('orgn_ntby_qty'))
            def _fmt_i(val):
                if val == 0: return "[dim]-[/dim]"
                abs_val = abs(val)
                if abs_val >= 1_000_000: s = f"{val/1_000_000:+.1f}M"
                elif abs_val >= 1000: s = f"{val/1000:+.0f}K"
                else: s = f"{val:+,}"
                return f"[red]{s}[/]" if val > 0 else f"[blue]{s}[/]"
            inv_str = f"개인: {_fmt_i(p)} | 외인: {_fmt_i(f)} | 기관: {_fmt_i(o)}"
            
        table_tech.add_row("수급", inv_str, "당일 개인/외국인/기관 순매수량")
        
        sm_str = f"[red]{sm_reason}[/]" if sm_flag else "[dim]특이사항 없음[/]"
        table_tech.add_row("스마트머니", sm_str, "외인/기관 쌍끌이 및 순매수 전환")

    config.console.print(table_tech)
    config.console.print()
    
    # [테이블 2] 시스템 트레이딩 판단 결과
    if not is_index:
        logic_title = "시스템 트레이딩 판단 결과"
        changes_summary = None
            
        if rule_applied:
            changes = []
            
            def_buy_score = config.ANALYSIS_THRESHOLDS["BUY_SCORE"]
            def_buy_rsi = config.ANALYSIS_THRESHOLDS["BUY_RSI_MAX"]
            def_buy_vol = config.ANALYSIS_THRESHOLDS.get("BUY_VOL_STRENGTH", 100.0)
            def_sell_score = config.SELL_STRATEGY["SELL_SCORE"]
            def_tp = config.SELL_STRATEGY["TAKE_PROFIT_RATE"]
            def_sl = config.SELL_STRATEGY["STOP_LOSS_RATE"]
            
            if custom_rule.get('buy_score') != def_buy_score:
                changes.append(f"매수점수({def_buy_score}->{custom_rule['buy_score']})")
            if custom_rule.get('buy_rsi') != def_buy_rsi:
                changes.append(f"매수RSI({def_buy_rsi}->{custom_rule['buy_rsi']})")
            if custom_rule.get('buy_vol_strength') and custom_rule['buy_vol_strength'] != def_buy_vol:
                changes.append(f"체결강도({def_buy_vol}%->{custom_rule['buy_vol_strength']}%)")
            
            if custom_rule.get('sell_score') != def_sell_score:
                changes.append(f"매도점수({def_sell_score}->{custom_rule['sell_score']})")
            if custom_rule.get('take_profit') != def_tp:
                changes.append(f"익절({def_tp}%->{custom_rule['take_profit']}%)")
            if custom_rule.get('stop_loss') != def_sl:
                changes.append(f"손절({def_sl}%->{custom_rule['stop_loss']}%)")
                
            if custom_rule.get('weights'):
                changes.append("가중치")

            if changes:
                changes_summary = ", ".join(changes)

        table_logic = Table(title=logic_title, box=box.HORIZONTALS, header_style="dim", border_style="dim")
        table_logic.add_column("항목", justify="center", style="cyan", width=15)
        table_logic.add_column("결과", justify="center", width=30)
        table_logic.add_column("상세 내용 / 사유", justify="left", style="dim")

        s_color = state_color.replace('[', '').replace(']', '')
        score_str = f"[bold {s_color}]{score:.2f}점[/]"
        
        details_str = ""
        if details:
            details_str = "\n".join([f"[green]* {d}[/green]" for d in details])
        else:
            details_str = "[dim]획득한 점수가 없습니다.[/dim]"
        
        table_logic.add_row("종합 점수", score_str, details_str)
        
        w_val = f"{weights.get('TREND', 4.0):.1f} / {weights.get('MOMENTUM', 2.5):.1f} / {weights.get('STRENGTH', 1.5):.1f} / {weights.get('SYNERGY', 2.0):.1f}"
        w_desc = "추세 / 모멘텀 / 강도 / 시너지"
        if rule_applied and custom_rule.get('weights'):
            w_desc += " [magenta](개별 설정)[/]"
        else:
            w_desc += " [dim](시스템 설정)[/dim]"
        table_logic.add_row("적용 가중치", w_val, w_desc)
        
        table_logic.add_row("상태 분류", f"[bold {s_color}]{state}[/]", state_reason)
        
        buy_score_limit = buy_score
        
        is_mr_state = (state == "역매수")
        if is_mr_state:
            buy_vol_limit = thresholds.get("MR_VOL_STRENGTH", config.ANALYSIS_THRESHOLDS.get("MR_VOL_STRENGTH", 120.0))
        else:
            buy_vol_limit = config.ANALYSIS_THRESHOLDS.get("BUY_VOL_STRENGTH", 100.0)
            if rule_applied and custom_rule.get('buy_vol_strength'):
                buy_vol_limit = custom_rule['buy_vol_strength']
                
        min_ask_bid_ratio = config.ANALYSIS_THRESHOLDS.get("BUY_ASK_BID_RATIO", 1.0)
        if rule_applied and custom_rule.get('buy_ask_bid_ratio') is not None:
            min_ask_bid_ratio = custom_rule['buy_ask_bid_ratio']
            
        auto_adjust = config.ANALYSIS_THRESHOLDS.get("AUTO_ADJUST_ASK_BID_RATIO", True)
        if rule_applied and custom_rule.get('auto_adjust_ask_bid_ratio') is not None:
            auto_adjust = bool(custom_rule['auto_adjust_ask_bid_ratio'])

        if auto_adjust and min_ask_bid_ratio > 0 and buy_vol_limit > 0:
            ratio_multiplier = buy_vol_limit / 100.0
            min_ask_bid_ratio = round(min_ask_bid_ratio * ratio_multiplier, 2)
                
        use_super = thresholds.get("SUPER_MOMENTUM_USE", config.ANALYSIS_THRESHOLDS.get("SUPER_MOMENTUM_USE", True))
        super_score = thresholds.get("SUPER_MOMENTUM_SCORE", config.ANALYSIS_THRESHOLDS.get("SUPER_MOMENTUM_SCORE", 8.0))
        super_w52 = thresholds.get("SUPER_MOMENTUM_W52_POS", config.ANALYSIS_THRESHOLDS.get("SUPER_MOMENTUM_W52_POS", 90.0))
        super_rsi = thresholds.get("SUPER_BUY_RSI_MAX", config.ANALYSIS_THRESHOLDS.get("SUPER_BUY_RSI_MAX", 80.0))
        
        is_super = use_super and score >= super_score and w52_pos >= super_w52
        buy_rsi_limit = super_rsi if is_super else thresholds["BUY_RSI_MAX"]

        is_buy_score = score >= buy_score_limit
        is_buy_rsi = (ind['rsi'] is not None) and (ind['rsi'] < buy_rsi_limit)
        is_safe_state = state not in ["매도", "주의", "-"]
        is_buy_vol = True
        is_ask_bid_ok = True
        if vol_strength is not None:
            is_buy_vol = vol_strength >= buy_vol_limit
        if ask_bid_ratio is not None and min_ask_bid_ratio > 0:
            is_ask_bid_ok = ask_bid_ratio >= min_ask_bid_ratio
            
        is_buy_all_ok = is_buy_score and is_buy_rsi and is_safe_state and is_buy_vol and is_ask_bid_ok
        
        if is_mr_state:
            buy_result = "[bold magenta]매수 가능 (역추세)[/]" if (is_buy_vol and is_ask_bid_ok) else "[bold blue]매수 불가 (체결/잔량 미달)[/]"
        else:
            if state == "강매수":
                buy_result = "[bold magenta]매수 가능 (슈퍼모멘텀)[/]" if is_buy_all_ok else "[bold blue]매수 불가[/]"
            else:
                buy_result = "[bold red]매수 가능[/]" if is_buy_all_ok else "[bold blue]매수 불가[/]"
        
        buy_reason_list = []
        if not is_safe_state: buy_reason_list.append(f"진입 불가 상태 ({state})")
        if not is_buy_score and not is_mr_state:
            if score_adj != 0 and not rule_applied:
                origin_score = round(buy_score_limit - score_adj, 2)
                buy_reason_list.append(f"점수 미달 (기준: {buy_score_limit} 이상 [설정: {origin_score}, 시장보정 {score_adj:+.1f}점])")
            else:
                buy_reason_list.append(f"점수 미달 (기준: {buy_score_limit}점 이상)")
        if not is_buy_rsi and not is_mr_state:
            if ind['rsi'] is None: buy_reason_list.append("RSI 데이터 부족")
            else:
                rsi_reason = f"RSI 과열 (기준: {buy_rsi_limit} 미만)"
                if is_super: rsi_reason += " [슈퍼모멘텀 완화 적용됨]"
                buy_reason_list.append(rsi_reason)
        if not is_buy_vol: buy_reason_list.append(f"체결강도 미달 ({vol_strength:.1f}% < {buy_vol_limit}%)")
        if not is_ask_bid_ok: buy_reason_list.append(f"매도잔량비 미달 ({ask_bid_ratio:.2f}배 < {min_ask_bid_ratio}배)")
        
        buy_reason = ", ".join(buy_reason_list) if buy_reason_list else ("역추세 반등 확인" if is_mr_state else "모든 매수 조건 충족")
        
        table_logic.add_row("매수 판단", buy_result, buy_reason)
        
        sell_score_limit = config.SELL_STRATEGY["SELL_SCORE"]
        # [추세추종] 점수 하락 매도는 추세 구조 훼손(현재가<60일선) 동반 시에만 발동 (실매매 analyze_sell과 동일 기준)
        ema60_now = ind.get('ema_60')
        structure_broken = ema60_now is None or current_price < ema60_now
        is_sell_signal = (state == "매도") or (score < sell_score_limit and structure_broken)

        sell_result = "[bold blue]매도(추세이탈)[/]" if is_sell_signal else "[bold green]보유(추세유지)[/]"

        if state == "매도":
            sell_reason = "매도 상태 진입 (필터링 조건)"
        elif score < sell_score_limit and structure_broken:
            sell_reason = f"점수 하락 (기준: {sell_score_limit}점 미만) + 60일선 이탈"
        elif score < sell_score_limit:
            sell_reason = f"점수 미달이나 60일선 위 추세 구조 유지 (주청산은 트레일링 스탑에 위임)"
        else:
            sell_reason = "추세 유지 중 (주의/관망 상태라도 점수 유지 시 보유)"
        
        table_logic.add_row("보유 판단", sell_result, sell_reason)
        
        # [토스] 체결강도 미제공 → '체결강도' 행을 숨기고 아래 '매도잔량 비율'을 수급 지표로 사용
        if not config.session.is_toss:
            vol_str = "-"
            vol_eval = ""
            if vol_strength is not None:
                v_color = "[red]" if is_buy_vol else "[blue]"
                vol_str = f"{v_color}{vol_strength:.1f}%[/]"
                vol_eval = "[bold red]양호[/]" if is_buy_vol else "[bold blue]미달[/]"
            table_logic.add_row("체결강도", vol_str, f"{vol_eval} (기준: {buy_vol_limit}% 이상)")

        if not is_overseas and not is_index:
            ask_bid_str = "-"
            ask_bid_eval = ""
            if ask_bid_ratio is not None and min_ask_bid_ratio > 0:
                ab_color = "[red]" if is_ask_bid_ok else "[blue]"
                ask_bid_str = f"{ab_color}{ask_bid_ratio:.2f}배[/]"
                ask_bid_eval = "[bold red]양호[/]" if is_ask_bid_ok else "[bold blue]미달[/]"
                table_logic.add_row("매도잔량 비율", ask_bid_str, f"{ask_bid_eval} (기준: {min_ask_bid_ratio}배 이상)")
            elif min_ask_bid_ratio <= 0:
                table_logic.add_row("매도잔량 비율", "[dim]미사용[/]", "-")
            else:
                table_logic.add_row("매도잔량 비율", "-", "데이터 확인 불가")

        rule_res = "[bold magenta]적용[/]" if rule_applied else "[dim]미적용[/]"
        rule_desc = f"[dim]{changes_summary}[/dim]" if changes_summary else "-"
        table_logic.add_row("개별 룰", rule_res, rule_desc)

        table_logic.add_section()
        table_logic.add_row("TradingView 의견", tv_rating_str, "TradingView Technical Rating (-1~1)")

        config.console.print(table_logic)
    
    config.console.print()

    # [추가] 기간별 시세 30일치 출력
    _print_period_price_30(code, is_overseas)
    
    # [추가] 상세 차트 분석 여부 확인
    config.console.print()
    if Prompt.ask("📊 상세 차트 분석 데이터를 출력하시겠습니까?", choices=["y", "n"], default="n") == 'y':
        from modules import chart
        chart.generate_visual_chart(code, name, is_overseas=is_overseas)

    config.console.print()
    ai_prompt_msg = "🤖 AI 지수 심층 진단을 수행하시겠습니까?" if is_index else "🤖 AI 종목 심층 진단을 수행하시겠습니까?"
    if Prompt.ask(ai_prompt_msg, choices=["y", "n"], default="n") == 'y':
        from modules import theme_analysis
        from rich.markdown import Markdown
        from rich.panel import Panel
        from rich.padding import Padding
        
        rsi_val_str = f"{ind['rsi']:.1f}" if ind['rsi'] is not None else "-"
        adx_val_str = f"{ind['adx']:.1f}" if ind['adx'] is not None else "-"
        cci_val_str = f"{ind['cci']:.1f}" if ind['cci'] is not None else "-"
        
        plus_di = ind.get('plus_di')
        minus_di = ind.get('minus_di')
        dmi_str = "-"
        if plus_di is not None and minus_di is not None:
            if plus_di > minus_di:
                dmi_str = f"+DI 우위 ({plus_di:.1f} / {minus_di:.1f})"
            elif minus_di > plus_di:
                dmi_str = f"-DI 우위 ({plus_di:.1f} / {minus_di:.1f})"
            else:
                dmi_str = f"중립 ({plus_di:.1f} / {minus_di:.1f})"

        if is_index:
            tech_info = (
                f"• 현재가: {price_str_tech}\n"
                f"• 시스템 상태: {state} (사유: {state_reason})\n"
                f"• 핵심 지표: RSI {rsi_val_str} | ADX {adx_val_str} | CCI {cci_val_str} | DMI {dmi_str}"
            )
        else:
            tech_info = (
                f"• 현재가: {price_str_tech}\n"
                f"• 시스템 상태: {state} (사유: {state_reason})\n"
                f"• 퀀트 점수: {score}점 / 10점 만점\n"
                f"• 핵심 지표: RSI {rsi_val_str} | ADX {adx_val_str} | CCI {cci_val_str} | DMI {dmi_str}"
            )
        
        title_str = f"🤖 AI 지수 심층 진단: {name}({code})" if is_index else f"🤖 AI 종목 심층 진단: {name}({code})"

        with Progress(
            SpinnerColumn(),
            TextColumn("[progress.description]{task.description}"),
            BarColumn(),
            console=config.console,
            transient=True
        ) as progress:
            if is_index:
                progress.add_task(f"[cyan]Google Gemini가 매크로 모멘텀을 결합하여 지수 심층 진단 중...[/cyan]\n[dim]  (모델: {config.GEMINI_MODEL})[/dim]", total=None)
                answer = theme_analysis.analyze_index_with_gemini(code, name, tech_info)
            else:
                progress.add_task(f"[cyan]Google Gemini가 기업 모멘텀을 결합하여 종목 심층 진단 중...[/cyan]\n[dim]  (모델: {config.GEMINI_MODEL})[/dim]", total=None)
                answer = theme_analysis.analyze_stock_with_gemini(code, name, tech_info)
            
        if answer:
            if answer.startswith("⚠️"):
                config.console.print(f"\n{answer}")
            else:
                md = Markdown(answer)
                
                if is_index:
                    idx_table = Table(title="지수 분석 정보", box=box.HORIZONTALS, show_header=True, header_style="dim", border_style="dim")
                    idx_table.add_column("지수명", justify="left", style="white", no_wrap=True)
                    idx_table.add_column("지수", justify="right")
                    idx_table.add_column("등락폭 (등락률)", justify="right")
                    idx_table.add_column("52주 고점", justify="right")
                    idx_table.add_column("EMA(5)", justify="right")
                    idx_table.add_column("EMA(20)", justify="right")
                    idx_table.add_column("EMA(60)", justify="right")
                    idx_table.add_column("EMA(120)", justify="right")
                    idx_table.add_column("추세SMO", justify="center")
                    idx_table.add_column("ADX", justify="right")
                    idx_table.add_column("RSI", justify="right")
                    idx_table.add_column("CCI", justify="right")
                    idx_table.add_column("OBV", justify="right")
                    
                    curr_fmt = f"{current_price:,.2f}"
                    curr_str = f"{curr_price_color}{curr_fmt}[/]"
                    
                    diff_color = "[red]" if diff > 0 else ("[blue]" if diff < 0 else "[white]")
                    diff_str_fmt = f"{diff:+.2f}"
                    change_str = f"{diff_color}{diff_str_fmt} ({rate:+.2f}%)[/]"

                    h_color = "[white]"
                    if high_52_rate > -3.0: h_color = "[red]"
                    elif high_52_rate < -20.0: h_color = "[blue]"
                    h52_fmt = f"{h52:,.0f}" if h52 >= 1000 else f"{h52:,.2f}"
                    high_52_str = f"[dim]{h52_fmt}[/] ({h_color}{high_52_rate:.1f}%[/])"

                    def fmt_val(val, color_tag):
                        if val is None or math.isnan(val): return "[dim]-[/dim]"
                        s = f"{val:,.0f}" if val >= 1000 else f"{val:,.2f}"
                        return f"{color_tag}{s}[/]" if color_tag else s
                        
                    ema5_color = "[white]"
                    if ind.get('ema_5') is not None and ind.get('ema_20') is not None:
                        ema5_color = "[red]" if ind['ema_5'] > ind['ema_20'] else "[blue]"

                    ema20_color = "[white]"
                    if ind.get('ema_20') is not None and ind.get('ema_60') is not None:
                        ema20_color = "[red]" if ind['ema_20'] > ind['ema_60'] else "[blue]"

                    ema60_color = "[white]"
                    if ind.get('ema_60') is not None and ind.get('ema_120') is not None:
                        ema60_color = "[red]" if ind['ema_60'] > ind['ema_120'] else "[blue]"

                    ema120_color = "[white]"
                    if df is not None and not df.empty and len(df) > 121:
                        try:
                            ema120_series = df['close'].ewm(span=120, adjust=False).mean()
                            if ema120_series.iloc[-1] > ema120_series.iloc[-2]:
                                ema120_color = "[red]"
                            else:
                                ema120_color = "[blue]"
                        except Exception: pass
                    
                    t_sar = "[dim]-[/dim]"
                    if ind.get('psar') is not None:
                        t_sar = "[red]▲[/]" if current_price > ind['psar'] else "[blue]▼[/]"
                    m_val = ind.get('macd')
                    s_val = ind.get('macd_signal')
                    t_macd = "[dim]-[/dim]"
                    if m_val is not None and s_val is not None:
                        zs = "+" if m_val > 0 else "-"
                        cc = "G" if m_val > s_val else "D"
                        mc = "red" if m_val > s_val else "blue"
                        t_macd = f"[{mc}]{zs}{cc}[/]"
                        
                    obv_trend = ind.get('obv_trend')
                    obv_val = ind.get('obv')
                    vol_sum = df['volume'].tail(5).sum() if df is not None and 'volume' in df.columns else 0
                    
                    if df is None or len(df) < config.INDICATOR_PARAMS.get("OBV_MA_PERIOD", 5):
                        obv_trend = None
                        obv_val = None
                        
                    if vol_sum == 0 or obv_trend is None or obv_val is None or math.isnan(obv_val):
                        obv_icon = "[dim]-[/dim]"
                        obv_disp = "[dim]-[/dim]"
                    else:
                        obv_icon = "[red]▲[/]" if obv_trend else "[blue]▼[/]"
                        obv_c = "red" if obv_trend else "blue"
                        abs_val = abs(obv_val)
                        if abs_val >= 999_950_000_000: obv_str = f"{obv_val/1_000_000_000_000:,.1f}T"
                        elif abs_val >= 999_950_000: obv_str = f"{obv_val/1_000_000_000:,.1f}B"
                        elif abs_val >= 999_500: obv_str = f"{obv_val/1_000_000:,.1f}M"
                        elif abs_val >= 999.5: obv_str = f"{obv_val/1_000:,.0f}K"
                        else: obv_str = f"{obv_val:,.0f}"
                        obv_disp = f"[{obv_c}]{obv_str}[/]"
                        
                    trend_str = f"{t_sar} {t_macd} {obv_icon}"
                    
                    val_rsi = ind.get('rsi')
                    t_rsi_str = f"{val_rsi:.1f}" if val_rsi is not None else "[dim]-[/dim]"
                    if val_rsi is not None:
                        if val_rsi >= config.INDICATOR_PARAMS["RSI_UPPER"]: t_rsi_str = f"[magenta]{t_rsi_str}[/]"
                        elif 55 <= val_rsi < config.INDICATOR_PARAMS["RSI_UPPER"]: t_rsi_str = f"[red]{t_rsi_str}[/]"
                        elif 45 <= val_rsi < 55: t_rsi_str = f"[orange3]{t_rsi_str}[/]"
                        elif config.INDICATOR_PARAMS["RSI_LOWER"] < val_rsi < 45: t_rsi_str = f"[yellow]{t_rsi_str}[/]"
                        else: t_rsi_str = f"[blue]{t_rsi_str}[/]"

                    val_cci = ind.get('cci')
                    t_cci_str = f"{val_cci:.1f}" if val_cci is not None else "[dim]-[/dim]"
                    if val_cci is not None:
                        if val_cci >= config.INDICATOR_PARAMS["CCI_UPPER"]: t_cci_str = f"[red]{t_cci_str}[/]"
                        elif 0 < val_cci < config.INDICATOR_PARAMS["CCI_UPPER"]: t_cci_str = f"[orange3]{t_cci_str}[/]"
                        elif config.INDICATOR_PARAMS["CCI_LOWER"] < val_cci <= 0: t_cci_str = f"[yellow]{t_cci_str}[/]"
                        else: t_cci_str = f"[blue]{t_cci_str}[/]"

                    # ADX 값 뒤에 DMI 우위 방향(▲/▼/●)을 함께 표기
                    val_adx = ind.get('adx')
                    t_adx_str = format_adx_cell(val_adx, ind.get('plus_di'), ind.get('minus_di'))

                    idx_table.add_row(
                        name, curr_str, change_str, high_52_str, 
                        fmt_val(ind.get('ema_5'), ema5_color), 
                        fmt_val(ind.get('ema_20'), ema20_color), 
                        fmt_val(ind.get('ema_60'), ema60_color), 
                        fmt_val(ind.get('ema_120'), ema120_color), 
                        trend_str, t_adx_str, t_rsi_str, t_cci_str, obv_disp
                    )
                    config.console.print()
                    config.console.print(idx_table)
                else:
                    table_title = "미국 주식 분석 정보" if is_overseas else "국내 주식 분석 정보"
                    # 이 경로는 ETF도 '주식 분석 정보' 제목으로 출력하므로, 세션 표기가
                    #  제목만 보고 오판하지 않도록 종목 단위로 ETF 여부를 넘긴다.
                    is_etf = (not is_overseas) and api.is_domestic_etf_etn(code, name)
                    print_table(table_title, [(name, code)], is_overseas=is_overseas, is_etf=is_etf)
                
                panel = Panel(md, title=title_str, border_style="cyan", padding=(1, 2), width=120)
                config.console.print()
                config.console.print(Padding(panel, (0, 4)))
        else:
            config.console.print("[red]분석 결과를 생성하지 못했습니다.[/red]")

def _diagnose_group_stock_worker(item, market_filter, restricted_stocks, rules_map, reserved_codes=None, m_codes=None):
    """(내부함수) 관심 종목 일괄 분석용 단일 워커 (병렬 처리용)"""
    if reserved_codes is None: reserved_codes = set()
    if m_codes is None: m_codes = set()
    try:
        code = item['code']
        name = item['name']
        
        # 1. [최적화] 시장 구분 확인(선택), 차트 데이터, 체결강도 병렬(동시) 조회 (누락 방지 포함)
        for attempt in range(2):
            with concurrent.futures.ThreadPoolExecutor(max_workers=3) as ex:
                fut_cp = ex.submit(api.get_current_price_data, code, False) if market_filter else None
                fut_chart = ex.submit(api.get_chart_data, code, is_overseas=False)
                fut_vol = ex.submit(api.get_realtime_vol_strength, code)
                
                cp_data = fut_cp.result() if fut_cp else None
                df = fut_chart.result()
                try: vol_strength = fut_vol.result()
                except Exception: vol_strength = None
            
            is_cp_valid = True if not market_filter else (cp_data and cp_data.get('rt_cd') == '0')
            if is_cp_valid and df is not None and not df.empty:
                break
            time.sleep(0.5)

        if market_filter and cp_data:
            if cp_data.get('rt_cd') != '0': return None
            
            # [수정] 현재가 데이터에는 시장 정보가 없으므로 마스터 파일 기반으로 필터링
            m_type = _get_market_type_by_master(code)
            if market_filter == "KOSPI" and m_type != "KOSPI": return None
            if market_filter == "KOSDAQ" and m_type != "KOSDAQ": return None

        if df is None or df.empty: return None
        
        # [추가] 실시간 현재가 조회 및 차트 당일 고가/저가/종가 최신화
        #  (봉 반영은 KRX 정규장에만 — chart_overlay_price가 정규장 밖에서 0을 돌려준다)
        rt_price = 0.0
        try:
            if cp_data and cp_data.get('rt_cd') == '0':
                # [수정] NXT 장 현재가(ats_prpr)가 있으면 우선 반영, 없으면 정규장 현재가 반영
                nxt_price = float(cp_data['output'].get('ats_prpr', 0) or 0)
                krx_price = float(cp_data['output'].get('stck_prpr', 0) or 0)
                rt_price = nxt_price if nxt_price > 0 else krx_price
            else:
                rt_price = float(api.get_current_price(code, is_overseas=False) or 0)

            indicators.apply_realtime_price(df, api.chart_overlay_price(rt_price, False),
                                            market_date=utils.market_today(False))
        except Exception: pass

        ind = indicators.calculate_indicators(df)
        current_price = float(df.iloc[-1]['close'])   # [판단 기준] 지표·상태·점수·52주 위치
        # [표시 전용] NXT 거래시간에는 살아있는 실시간가를 현재가로 내보낸다(지표 기준은 위 값 유지).
        display_price = current_price
        if rt_price > 0 and not api.chart_overlay_enabled(False) and not api.display_price_krx_fixed(False):
            display_price = rt_price

        # 전일 RSI (상태 분류용) — calculate_indicators가 계산한 값 재사용 (중복 계산 제거·SSOT)
        prev_rsi = ind.get('prev_rsi') if len(df) >= 16 else None

        # 52주 위치 계산 (슈퍼 모멘텀 판정용)
        w52_pos = 0.0
        if len(df) > 0:
            h52, l52 = _w52_band(df)   # 표(_analyze_table_row)와 동일한 365일 창
            if h52 > l52:
                w52_pos = (current_price - l52) / (h52 - l52) * 100

        sm_flag, sm_reason = check_smart_money_turnaround(code, is_overseas=False)

        # 3. 점수 및 상태 계산
        rule = rules_map.get(code)
        thresholds = None
        weights = None
        
        if rule:
            # [SSOT] 매매 경로와 같은 규약 — 룰의 NULL 컬럼은 전역 기본값으로 되돌리고
            #  가중치는 dict로 확정한다. (지연 import: engine이 analysis를 import한다)
            from modules.auto_trade.engine import normalize_weights, rule_value
            thresholds = {
                "BUY_SCORE": rule_value(rule, 'buy_score', config.ANALYSIS_THRESHOLDS["BUY_SCORE"]),
                "BUY_RSI_MAX": rule_value(rule, 'buy_rsi', config.ANALYSIS_THRESHOLDS["BUY_RSI_MAX"]),
                "BUY_VOL_STRENGTH": rule_value(rule, 'buy_vol_strength', config.ANALYSIS_THRESHOLDS.get("BUY_VOL_STRENGTH", 100.0)),
                "WEIGHTS": normalize_weights(rule.get('weights')),
                "RISE_SCORE": config.ANALYSIS_THRESHOLDS["RISE_SCORE"]
            }
            weights = thresholds["WEIGHTS"]

        state, state_color, state_reason = classify_stock_state(
            df=df, ind=ind, prev_rsi=prev_rsi, thresholds=thresholds, w52_pos=w52_pos, smart_money=sm_flag
        )
        
        score, _ = calculate_score(
            df=df, ind=ind, weights=weights, smart_money=sm_flag
        )
        
        # [추가] 개별 룰 여부 확인
        is_custom_rule = code in rules_map
        is_restricted = code in restricted_stocks
        is_reserved = code in reserved_codes
        is_memo = code in m_codes
        
        return {
            'code': code, 'name': name, 'price': display_price,
            'score': score, 'state': state, 'state_color': state_color,
            'rsi': ind['rsi'], 'adx': ind['adx'], 'cci': ind['cci'],
            'plus_di': ind.get('plus_di'), 'minus_di': ind.get('minus_di'),
            'vol_strength': vol_strength, 'is_custom_rule': is_custom_rule,
            'is_restricted': is_restricted,
            'is_reserved': is_reserved,
            'is_memo': is_memo
        }
    except Exception:
        return None

def diagnose_group_stocks(market_filter=None):
    """등록된 종목들에 대해 일괄 분석을 수행합니다."""
    # 대상: 국내 주식 + 국내 ETF
    targets = config.session.stock_data.get('stocks_kr', []) + config.session.stock_data.get('etfs_kr', [])
    
    if not targets:
        config.console.print("[yellow]등록된 국내 종목이 없습니다.[/yellow]")
        return
        
    # [추가] 개별 룰 로드 (전체 조회 최적화)
    custom_rules = db_manager.db.get_all_stock_strategies()
    rules_map = {}
    for r in custom_rules:
        r_dict = dict(r)
        if r_dict.get('weights') and isinstance(r_dict['weights'], str):
            try: r_dict['weights'] = json.loads(r_dict['weights'])
            except Exception: r_dict['weights'] = None
        rules_map[r_dict['code']] = r_dict

    # [추가] 트레이딩 제한 종목 로드
    from modules import auto_trade
    restricted_stocks = auto_trade.get_restricted_stocks()
    any_restricted = False
    
    # [추가] 예약 매매 및 메모 마커 조회
    try:
        pending_reserves = db_manager.db.get_pending_reserved_orders()
        reserved_codes = set(o['code'] for o in pending_reserves)
    except Exception:
        reserved_codes = set()
    m_codes = utils.get_memo_codes()

    results = []
    
    title_suffix = f" ({market_filter})" if market_filter else " (전체)"
    
    with Progress(
        SpinnerColumn(),
        TextColumn("[progress.description]{task.description}"),
        BarColumn(),
        TextColumn("[progress.percentage]{task.percentage:>3.0f}%"),
        TimeRemainingColumn(),
        console=config.console,
        transient=True
    ) as progress:
        task = progress.add_task(f"[cyan]등록된 종목 병렬 분석 중{title_suffix}...[/cyan]", total=len(targets))
        
        # [최적화] ThrottledSession 제어 기반으로 모의투자(2) / 실전(4) 통합 병렬 처리 허용
        max_w = config.analysis_max_workers()
        with concurrent.futures.ThreadPoolExecutor(max_workers=max_w) as executor:
            futures = [executor.submit(_diagnose_group_stock_worker, item, market_filter, restricted_stocks, rules_map, reserved_codes, m_codes) for item in targets]
            for future in concurrent.futures.as_completed(futures):
                res = future.result()
                if res: results.append(res)
                progress.advance(task)

    # 결과 출력
    if not results:
        config.console.print(f"[yellow]해당 조건({market_filter})에 맞는 종목이 없거나 데이터를 불러올 수 없습니다.[/yellow]")
        return

    used_marks = set()
    # 정렬 기준 개선: 1. 점수 높은 순, 2. RSI 낮은 순 (상승 여력)
    # RSI가 None인 경우 맨 뒤로 보내기 위해 999 처리
    results.sort(key=lambda x: (-x['score'], x['rsi'] if x['rsi'] is not None else 999))
    
    table_title = f"전체 종목 분석 결과{title_suffix}"
    
    # [추가] 적용된 가중치 정보 표시 (검증용)
    if config.SCORING_WEIGHTS:
        w = config.SCORING_WEIGHTS
        w_str = f"{w.get('TREND', 4.0)}/{w.get('MOMENTUM', 2.5)}/{w.get('STRENGTH', 1.5)}/{w.get('SYNERGY', 2.0)}"
        table_title += f" [dim](가중치: {w_str})[/dim]"

    table = Table(title=table_title, box=box.HORIZONTALS, header_style="dim", border_style="dim")
    table.add_column("종목명(코드)", justify="left")
    table.add_column("현재가", justify="right")
    table.add_column("점수", justify="center")
    table.add_column("상태", justify="center")
    table.add_column("ADX", justify="right")
    table.add_column("RSI", justify="right")
    table.add_column("CCI", justify="right")
    table.add_column("체결강도", justify="right")
    
    for r in results:
        s_color = r['state_color'].replace('[', '').replace(']', '')
        score_str = f"[{s_color}]{r['score']:.2f}점[/]"
        state_str = f"[{s_color}]{r['state']}[/]"
        
        rsi_val = r['rsi']
        rsi_str = f"{rsi_val:.1f}" if rsi_val is not None else "-"
        if rsi_val is not None:
            if rsi_val >= 70: rsi_str = f"[magenta]{rsi_str}[/]"
            elif rsi_val <= 30: rsi_str = f"[blue]{rsi_str}[/]"
            
        # ADX 값 뒤에 DMI 우위 방향(▲/▼/●)을 함께 표기
        adx_str = format_adx_cell(r['adx'], r.get('plus_di'), r.get('minus_di'))
        cci_str = f"{r['cci']:.1f}" if r['cci'] is not None else "-"
        
        vol_val = r.get('vol_strength')
        vol_str = f"{vol_val:.1f}%" if vol_val else "-"
        
        name_display = r['name']
        marks = []
        if r.get('is_restricted'):
            marks.append("-")
            used_marks.add('-')
        if r.get('is_custom_rule'):
            marks.append("+")
            used_marks.add('+')
        if r.get('is_memo'):
            marks.append("=")
            used_marks.add('=')
        if r.get('is_reserved'):
            marks.append("[magenta]*[/magenta]")
            used_marks.add('*')
            
        if marks:
            name_display += f"[dim]{''.join(marks)}[/dim]"
        
        table.add_row(
            f"{name_display}({r['code']})",
            f"{int(r['price']):,}원",
            score_str,
            state_str,
            adx_str,
            rsi_str,
            cci_str
        )
        
    config.console.print(table, crop=False)
    sys.stdout.flush()
    config.console.print()
    

def get_analysis_params(use_vol=True):
    """분석에 사용할 파라미터를 사용자로부터 입력받습니다. (매수 체결강도 옵션 연동)"""
    params = {
        "BUY_SCORE": config.ANALYSIS_THRESHOLDS["BUY_SCORE"],
        "BUY_RSI_MAX": config.ANALYSIS_THRESHOLDS["BUY_RSI_MAX"],
        "RISE_SCORE": config.ANALYSIS_THRESHOLDS["RISE_SCORE"],
        "WEIGHTS": config.SCORING_WEIGHTS.copy() # [추가] 가중치 포함 (복사본 사용)
    }
    
    config.console.print("\n[bold]분석 파라미터 설정 (Enter: 현재값 유지, 이전: b, 메인: q)[/bold]")
    
    config.console.print("\n[bold]1. 기본 매수 타점 설정[/bold]")
    val = Prompt.ask(f"매수 기준 점수 (기본: {params['BUY_SCORE']}점)\n[dim]이 점수 이상일 때 매수 진입 (지표 종합 점수)[/dim]", default=str(params['BUY_SCORE']))
    if val.lower() in ['b', 'q']: return None
    try: params['BUY_SCORE'] = float(val)
    except Exception: pass
    
    val = Prompt.ask(f"매수 허용 RSI 상한 (기본: {params['BUY_RSI_MAX']})\n[dim]RSI가 이 값보다 낮아야 매수 (과열 방지)[/dim]", default=str(params['BUY_RSI_MAX']))
    if val.lower() in ['b', 'q']: return None
    if val.isdigit(): params['BUY_RSI_MAX'] = int(val)
    
    if use_vol:
        current_vol = config.ANALYSIS_THRESHOLDS.get("BUY_VOL_STRENGTH", 100.0)
        val = Prompt.ask(f"매수 체결강도 기준(%) (기본: {current_vol}, 0: 미사용)\n[dim]수급 확인 (이 값 이상이어야 매수)[/dim]", default=str(current_vol))
        if val.lower() in ['b', 'q']: return None
        try: params['BUY_VOL_STRENGTH'] = float(val)
        except Exception: params['BUY_VOL_STRENGTH'] = current_vol
    else:
        params['BUY_VOL_STRENGTH'] = 0.0

    config.console.print("\n[bold]2. 스캐닝 필터 설정[/bold]")
    val = Prompt.ask(f"상승 추세 기준 점수 (기본: {params['RISE_SCORE']}점)\n[dim]매수에는 미달하지만 관망/상승으로 판단할 점수 기준[/dim]", default=str(params['RISE_SCORE']))
    if val.lower() in ['b', 'q']: return None
    try: params['RISE_SCORE'] = float(val)
    except Exception: pass

    config.console.print("\n[bold]3. 스코어링 가중치 설정[/bold]")
    curr_weights = params['WEIGHTS'].copy()
    while True:
        config.console.print("[dim]순서: 추세 / 모멘텀 / 강도 / 시너지 (합계 10점 권장)[/dim]")
        config.console.print()
        
        try:
            def ask_w(key, desc, default_v):
                v = Prompt.ask(f"{desc} [dim](현재: {default_v})[/dim]", default=str(default_v))
                if v.lower() in ['b', 'q']: raise ValueError("quit")
                return float(v)

            w_trend = ask_w("TREND", "추세 (TREND)", curr_weights.get('TREND', 4.0))
            w_mom = ask_w("MOMENTUM", "모멘텀 (MOMENTUM)", curr_weights.get('MOMENTUM', 2.5))
            w_str = ask_w("STRENGTH", "강도 (STRENGTH)", curr_weights.get('STRENGTH', 1.5))
            w_syn = ask_w("SYNERGY", "시너지 (SYNERGY)", curr_weights.get('SYNERGY', 2.0))
            
            total_score = w_trend + w_mom + w_str + w_syn
            
            if abs(total_score - 10.0) > 0.01:
                config.console.print(f"\n[bold red]경고: 가중치 합계가 {total_score:.1f}점입니다. (권장: 10.0점)[/bold red]")
                config.console.print("[yellow]합계가 10점이 되도록 다시 입력해주세요.[/yellow]")
                curr_weights = {"TREND": w_trend, "MOMENTUM": w_mom, "STRENGTH": w_str, "SYNERGY": w_syn}
                continue
            
            params['WEIGHTS'] = {"TREND": w_trend, "MOMENTUM": w_mom, "STRENGTH": w_str, "SYNERGY": w_syn}
            break
        except ValueError as e:
            if str(e) == "quit": return None
            config.console.print("[red]잘못된 입력입니다. 숫자를 입력해주세요.[/red]")
            continue

    # [안내] 이 화면에서 묻지 않는 값 — 잠긴 항목은 '현재값 그대로' 쓰인다는 사실을 명시한다.
    #  params에 없는 키는 평가 시 config 값으로 폴백하므로, 운용자가 여기서 본 조건만으로
    #  결과가 결정된다고 오해하면 안 된다(백테스트 조건변경 화면의 잠금 안내와 같은 취지).
    try:
        from modules import settings as _settings
        _locked = set(_settings.ANTI_TREND_HIDDEN_KEYS) | set(_settings.BACKTESTED_HIDDEN_KEYS)
    except Exception:
        _locked = set()
    _fixed = []
    if "SELL_SCORE" in _locked:
        _fixed.append(f"매도(추세이탈) 점수 {config.SELL_STRATEGY.get('SELL_SCORE', 4.0)}")
    if "SUPER_MOMENTUM_USE" in _locked and config.ANALYSIS_THRESHOLDS.get("SUPER_MOMENTUM_USE", True):
        _fixed.append(
            f"슈퍼 모멘텀 ON (점수 {config.ANALYSIS_THRESHOLDS.get('SUPER_MOMENTUM_SCORE', 8.0)}↑ "
            f"& 52주 {config.ANALYSIS_THRESHOLDS.get('SUPER_MOMENTUM_W52_POS', 90.0)}%↑ 이면 "
            f"RSI 상한 {config.ANALYSIS_THRESHOLDS.get('SUPER_BUY_RSI_MAX', 80.0)}로 완화)")
    if _fixed:
        config.console.print("\n[dim]※ 추세추종 보호로 잠긴 항목은 현재값을 그대로 사용합니다 — "
                             + " · ".join(_fixed) + "[/dim]")

    config.console.print("\n[bold]4. 최종 출력 대상 선택[/bold]")
    filter_choice = Prompt.ask("출력 대상 선택 (1: 매수, 2: 상승, 3: 매수+상승) [dim](이전: b, 메인: q)[/dim]", choices=["1", "2", "3", "b", "q"], default="1")
    if filter_choice.lower() in ['b', 'q']: return None
    if filter_choice == '1': params['OUTPUT_FILTER'] = 'BUY'
    elif filter_choice == '2': params['OUTPUT_FILTER'] = 'RISE'
    else: params['OUTPUT_FILTER'] = 'ALL'

    return params

def _analyze_stock_worker(stock, params=None, restricted_stocks=None, rules_map=None, reserved_codes=None, m_codes=None):
    """(내부함수) 단일 종목 분석 워커 (멀티스레드용)"""
    if restricted_stocks is None: restricted_stocks = {}
    if rules_map is None: rules_map = {}
    if reserved_codes is None: reserved_codes = set()
    if m_codes is None: m_codes = set()
    
    code = stock['code']
    name = stock['name']
    is_custom_rule = stock.get('is_custom_rule', False) # [추가]
    
    # [최적화] 시스템 트레이딩(AutoTrader)과의 API 대역폭 경합 방지 (유동적 Pacing)
    # 모의투자는 0.3초, 실전은 0.05초의 지연을 주어 백그라운드 자동매매가 즉시 호출될 수 있는 틈을 양보합니다.
    delay = 0.05
    time.sleep(delay)

    try:
        # API 호출 (KIS API Rate Limit 처리 및 누락 방지 재시도)
        df = None
        for attempt in range(2):
            df = api.get_chart_data(code, is_overseas=False)
            if df is not None and not df.empty:
                break
            time.sleep(0.5)
            
        if df is None: return {'error': 'API 응답 없음'}
        if df.empty: return {'error': '차트 데이터 없음 (거래정지 등)'}

        # [Fix] 이 워커만 실시간 갱신이 빠져 있어, 현재가·점수가 차트 캐시 히트/미스에 따라
        #  달라졌다(히트=오버레이 반영, 미스=원본 봉). 개별 분석·관심종목 진단과 같은 기준으로 통일한다.
        #  (봉 반영은 KRX 정규장에만 — 정규장 밖 현재가는 NXT 체결가라 지표를 흔든다)
        rt_price = 0.0
        try:
            # 모든 장 마감 후 KRX 고정(설정 True)이면 표시에도 쓸 일이 없어 조회 자체를 생략한다.
            if api.chart_overlay_enabled(False) or not api.display_price_krx_fixed(False):
                rt_price = float(api.get_current_price(code, is_overseas=False) or 0)
            # [SSOT] 위와 같은 이유로 chart_overlay_price 를 지난다.
            _overlay = api.chart_overlay_price(rt_price, False)
            if _overlay > 0:
                indicators.apply_realtime_price(df, _overlay, market_date=utils.market_today(False))
        except Exception: pass

        current_price = float(df.iloc[-1]['close'])   # [판단 기준] 지표·상태·점수·52주 위치
        # [표시 전용] NXT 거래시간에는 살아있는 실시간가를 현재가로 내보낸다(지표 기준은 위 값 유지).
        display_price = current_price
        if rt_price > 0 and not api.chart_overlay_enabled(False) and not api.display_price_krx_fixed(False):
            display_price = rt_price
        ind = indicators.calculate_indicators(df)
        
        # 전일 RSI — calculate_indicators가 계산한 값 재사용 (중복 계산 제거·SSOT)
        prev_rsi = ind.get('prev_rsi') if len(df) >= 16 else None

        # 52주 위치 계산 (슈퍼 모멘텀 판정용)
        w52_pos = 0.0
        if len(df) > 0:
            h52, l52 = _w52_band(df)   # 표(_analyze_table_row)와 동일한 365일 창
            if h52 > l52:
                w52_pos = (current_price - l52) / (h52 - l52) * 100

        sm_flag, sm_reason = check_smart_money_turnaround(code, is_overseas=False)

        # 상태 분류 및 점수 계산
        state, state_color, state_reason = classify_stock_state(
            df=df, ind=ind, prev_rsi=prev_rsi, thresholds=params, w52_pos=w52_pos, smart_money=sm_flag
        )
        
        if state == "-": return {'error': '지표 계산용 데이터 부족 (신규상장 등)'}

        # [추가] 초기 상태 보존 (로그 출력 시 체결강도 미달로 관망으로 변경되더라도 원본 상태 표시)
        initial_state = state
        initial_state_color = state_color

        # [수정] 사용자 설정 가중치 적용
        weights = params.get('WEIGHTS') if params else None
        score, _ = calculate_score(df=df, ind=ind, weights=weights, smart_money=sm_flag)

        # [수정] 체결강도 조회 최적화: 필터 조건에 맞는 종목만 조회
        vol_strength = None
        
        # 조회 대상 상태 정의 (기본: 매수, 상승/대기)
        target_vol_states = ["매수", "강매수", "역매수", "상승", "대기"]

        if params:
            filter_mode = params.get("OUTPUT_FILTER", "ALL")
            if filter_mode == "BUY": target_vol_states = ["매수", "강매수"]
            elif filter_mode == "RISE": target_vol_states = ["상승", "대기"]
        
        # 현재 상태가 조회 대상에 포함될 때만 체결강도 API 호출
        use_vol = True
        if params is not None and not params.get("USE_VOL", True):
            use_vol = False
            
        if use_vol and state in target_vol_states:
            # [추가] 조회 실패 시 재시도 로직 (최대 2회)
            for _ in range(2):
                try:
                    vol_strength = api.get_realtime_vol_strength(code)
                    if vol_strength is not None: break
                except Exception: time.sleep(0.1)

        # [수정] 매수(강매수, 역추세포함) 또는 상승/대기 상태일 경우 체결강도 기준 엄격히 체크 (필터링)
        if state in ["매수", "강매수", "역매수", "상승", "대기"]:
            try:
                if not use_vol:
                    min_vol = 0.0
                elif state == "역매수":
                    min_vol = params.get("MR_VOL_STRENGTH", config.ANALYSIS_THRESHOLDS.get("MR_VOL_STRENGTH", 120.0)) if params else config.ANALYSIS_THRESHOLDS.get("MR_VOL_STRENGTH", 120.0)
                elif params and 'BUY_VOL_STRENGTH' in params:
                    min_vol = params.get('BUY_VOL_STRENGTH', 100.0)
                else:
                    min_vol = config.ANALYSIS_THRESHOLDS.get("BUY_VOL_STRENGTH", 100.0)
                
                if min_vol > 0:
                    if vol_strength is None:
                        state = "관망"
                        state_color = "[white]"
                        state_reason = "체결강도 확인 불가 (API 응답 지연)"
                    elif vol_strength < min_vol:
                        state = "관망"
                        state_color = "[white]"
                        state_reason = f"체결강도 미달({vol_strength:.1f}% < {min_vol}%)"
            except Exception: pass

        # 필터링 조건 확인
        is_target = False
        if params:
            filter_mode = params.get("OUTPUT_FILTER", "BUY")
            target_states = []
            if filter_mode == "BUY": target_states = ["매수", "강매수"]
            elif filter_mode == "RISE": target_states = ["상승", "대기", "관심"]
            elif filter_mode == "ALL": target_states = ["매수", "강매수", "상승", "대기", "관심"]
            if state in target_states:
                is_target = True
        else:
            is_target = True # params가 없으면(엑셀 저장 등) 모두 유효

        obv_val = ind.get('obv')
        obv_trend = ind.get('obv_trend')
        vol_sum = df['volume'].tail(5).sum() if df is not None and 'volume' in df.columns else 0
        
        if df is None or len(df) < config.INDICATOR_PARAMS.get("OBV_MA_PERIOD", 5):
            obv_trend = None
            obv_val = None
            
        if vol_sum == 0 or obv_val is None or math.isnan(obv_val):
            obv_trend = None
            obv_val = None

        return {
            'code': code, 'name': name, 'price': display_price,
            'score': score, 'state': initial_state, 'state_color': initial_state_color, 'state_reason': state_reason,
            'rsi': ind['rsi'], 'adx': ind['adx'], 'cci': ind['cci'], 'obv_trend': obv_trend,
            'plus_di': ind.get('plus_di'), 'minus_di': ind.get('minus_di'),
            'psar': ind['psar'], 'macd': ind.get('macd'), 'macd_signal': ind.get('macd_signal'),
            'is_target': is_target, 
            'vol_strength': vol_strength,
            'w52_pos': w52_pos,
            'is_custom_rule': is_custom_rule # [추가]
        }
    except Exception as e:
        return {'error': f'분석 중 예외 발생: {e}'}

def analyze_market_stocks(market_type):
    """선택한 시장의 전체 종목을 분석하고 매수 가능 종목을 출력합니다."""
    
    # 1. DB에서 기존 분석 결과 확인
    cached_data = _load_analysis_result(market_type)
    buy_candidates = []
    params = None
    use_cache = False
    
    # [추가] 개별 룰 로드
    custom_rules = db_manager.db.get_all_stock_strategies()
    rules_map = {r['code']: True for r in custom_rules}
    
    # [추가] 트레이딩 제한 종목 로드
    from modules import auto_trade
    restricted_stocks = auto_trade.get_restricted_stocks()
    
    if cached_data:
        updated_at = cached_data['updated_at']
        c_params = cached_data['params']
        
        config.console.print(f"\n[bold cyan]기존 분석 결과가 존재합니다.[/bold cyan]")
        config.console.print(f"• 분석 일시: {updated_at}")
        
        w = c_params.get('WEIGHTS', config.SCORING_WEIGHTS)
        w_str = f"{w.get('TREND', 4.0)}/{w.get('MOMENTUM', 2.5)}/{w.get('STRENGTH', 1.5)}/{w.get('SYNERGY', 2.0)}"
        
        # [수정] 매수 점수 표시 (보정 정보 포함)
        buy_score_val = c_params.get('BUY_SCORE')
        buy_score_str = f"{buy_score_val}점"
        if c_params.get('SCORE_ADJ'):
            buy_score_str += f" (시장보정 {c_params['SCORE_ADJ']:+.1f}점)"

        # [표시] 체결강도 0%는 '기준이 0'이 아니라 '수급 조건 미사용'으로 돌렸다는 뜻이다.
        #  아래 '현재 설정'(config 값 100%)과 나란히 보이면 값이 바뀐 것처럼 읽히므로 구분해 적는다.
        _cached_vol = c_params.get('BUY_VOL_STRENGTH', 100)
        _vol_str = "미사용" if not _cached_vol else f"{_cached_vol}%"
        config.console.print(f"• 분석 조건: 매수 {buy_score_str}, RSI {c_params.get('BUY_RSI_MAX')}, 체결 {_vol_str}, 상승 {c_params.get('RISE_SCORE')}점, 가중치 {w_str}")
        
        config.console.print()
        choice = Prompt.ask("기존 결과를 보시겠습니까? [dim](이전: b, 메인: q)[/dim]", choices=["y", "n", "b", "q"], default="y")
        if choice in ['b', 'q']: return False
        if choice == "y":
            buy_candidates = cached_data['data']
            params = c_params
            use_cache = True
            config.console.print(f"[dim]DB에서 {len(buy_candidates)}개의 종목 정보를 로드했습니다.[/dim]")

    # 2. 새로 분석 (캐시 미사용 시)
    if not use_cache:
        stock_list = _get_master_stock_list(market_type)
        config.console.print(f"\n[bold]{market_type} 전체 종목 수: {len(stock_list)}개[/bold]")
        
        # [추가] stock_list에 is_custom_rule 정보 주입
        for s in stock_list:
            s['is_custom_rule'] = s['code'] in rules_map
        
        c_buy = config.ANALYSIS_THRESHOLDS["BUY_SCORE"]
        c_rsi = config.ANALYSIS_THRESHOLDS["BUY_RSI_MAX"]
        c_rise = config.ANALYSIS_THRESHOLDS["RISE_SCORE"]
        c_vol = config.ANALYSIS_THRESHOLDS.get("BUY_VOL_STRENGTH", 100.0)
        
        w = config.SCORING_WEIGHTS
        w_str = f"{w.get('TREND', 4.0)}/{w.get('MOMENTUM', 2.5)}/{w.get('STRENGTH', 1.5)}/{w.get('SYNERGY', 2.0)}"

        config.console.print(f"현재 설정: 매수 {c_buy}점 / RSI {c_rsi} / 체결 {c_vol}% / 상승 {c_rise}점 / 가중치 {w_str}")

        config.console.print()
        
        # [추가] ETF 종목 포함 여부 확인 (KOSPI에서만 질문)
        if market_type == "KOSPI":
            include_etf_choice = Prompt.ask("ETF 종목을 포함하여 분석하시겠습니까? [dim](y: 포함, n: 제외, 이전: b, 메인: q)[/dim]", choices=["y", "n", "b", "q"], default="n")
            if include_etf_choice in ['b', 'q']: return False
        else:
            include_etf_choice = 'n'
        
        # [재설계 2026-08-11] 비주권(ETF/ETN/리츠/인프라펀드) 제외를 '종목명 브랜드 접두어'에서
        #  마스터의 증권그룹구분코드로 바꾼다.
        #
        #  [왜] 브랜드 목록은 운용사가 새로 생길 때마다 손으로 늘려야 하는 구조라 반드시 샌다.
        #   실제로 'DS 코스닥액티브'(0220B0)가 통과해 "지표 계산용 데이터 부족(신규상장 등)"으로
        #   실패 로그를 남겼다. 또 리츠 키워드가 후행 공백을 요구해("리츠 ") '롯데리츠'·
        #   '이리츠코크렙' 같은 실제 이름과 한 건도 매칭되지 않았다 — 필터가 사실상 무동작이었다.
        #   2026-08-11 마스터 기준 코스피에서 이 방식이 놓치던 비주권은 29건(EF 1 + RT 23 + IF 2 +
        #   MF 1 + PF 2)이고, 그 29건이 전체 분석 대상 944개에 그대로 섞여 있었다.
        #
        #  [스팩] 코스닥 71건은 증권그룹이 ST(주권)라 그룹코드로는 걸리지 않는다. 이름에 '스팩'이
        #   들어가는 것이 유일한 식별자이고(기존 키워드 "스팩 "도 후행 공백 탓에 0건 매칭),
        #   공모가 2,000원에 고정된 껍데기라 추세추종 후보가 될 수 없으므로 시장과 무관하게 뺀다.
        original_len = len(stock_list)
        if include_etf_choice == 'n' and market_type == "KOSPI":
            # grp가 빈 값이면(마스터 포맷 변경 등) 거르지 않는다 — 판정 실패로 전 종목이
            # 사라지는 것보다 섞여 들어오는 편이 낫다(분석 단계에서 걸러진다).
            stock_list = [s for s in stock_list if s.get('grp', '') in ('', 'ST')]
        stock_list = [s for s in stock_list if '스팩' not in s['name']]
        if len(stock_list) != original_len:
            config.console.print(
                f"[dim]비주권(ETF/ETN/리츠/스팩 등) 제외: {original_len}개 -> {len(stock_list)}개[/dim]\n")
            
        # [수정] 매수 체결강도 사용 여부 확인 프롬프트 간결화 및 기본값(n) 변경
        # [추가] 토스증권은 체결강도(수급)를 제공하지 않으므로 프롬프트 없이 무조건 미사용 처리
        if config.session.is_toss:
            config.console.print("[dim]토스증권은 체결강도(수급)를 제공하지 않아 해당 조건을 미사용합니다.[/dim]")
            use_vol = False
        else:
            use_vol_choice = Prompt.ask("매수 체결강도(수급) 조건을 사용하여 분석하시겠습니까? [dim](y: 사용, n: 미사용, 이전: b, 메인: q)[/dim]", choices=["y", "n", "b", "q"], default="n")
            if use_vol_choice in ['b', 'q']: return False
            use_vol = (use_vol_choice == 'y')
            
        config.console.print()
            
        # 파라미터 설정
        change_settings = Prompt.ask("분석 조건을 변경하시겠습니까? [dim](이전: b, 메인: q)[/dim]", choices=["y", "n", "b", "q"], default="n")
        if change_settings in ['b', 'q']: return False

        if change_settings == 'y':
            params = get_analysis_params(use_vol=use_vol)
            if params is None: return False
            params['INCLUDE_ETF'] = (include_etf_choice == 'y')
            params['USE_VOL'] = use_vol
        else:
            params = {
                "BUY_SCORE": config.ANALYSIS_THRESHOLDS["BUY_SCORE"],
                "BUY_RSI_MAX": config.ANALYSIS_THRESHOLDS["BUY_RSI_MAX"],
                "BUY_VOL_STRENGTH": config.ANALYSIS_THRESHOLDS.get("BUY_VOL_STRENGTH", 100.0) if use_vol else 0.0,
                "MR_VOL_STRENGTH": config.ANALYSIS_THRESHOLDS.get("MR_VOL_STRENGTH", 120.0) if use_vol else 0.0,
                "RISE_SCORE": config.ANALYSIS_THRESHOLDS["RISE_SCORE"],
                "OUTPUT_FILTER": "BUY",
                "WEIGHTS": config.SCORING_WEIGHTS,
                "INCLUDE_ETF": (include_etf_choice == 'y'),
                "USE_VOL": use_vol
            }
            config.console.print(f"[dim]기본 설정으로 진행합니다. (매수: {params['BUY_SCORE']}점, RSI: {params['BUY_RSI_MAX']}, 체결: {params['BUY_VOL_STRENGTH']}%, 상승: {params['RISE_SCORE']}점)[/dim]")
        
        # 설정 백업 및 적용
        original_thresholds = config.ANALYSIS_THRESHOLDS.copy()
        config.ANALYSIS_THRESHOLDS["BUY_SCORE"] = params["BUY_SCORE"]
        config.ANALYSIS_THRESHOLDS["BUY_RSI_MAX"] = params["BUY_RSI_MAX"]
        config.ANALYSIS_THRESHOLDS["RISE_SCORE"] = params["RISE_SCORE"]
        config.ANALYSIS_THRESHOLDS["BUY_VOL_STRENGTH"] = params["BUY_VOL_STRENGTH"]
        
        config.console.print("\n[bold cyan]━━━ 전체 종목 분석 시작 (중단: Ctrl+C) ━━━[/bold cyan]")

        try:
            with Progress(
                SpinnerColumn(),
                TextColumn("[progress.description]{task.description}"),
                BarColumn(),
                TextColumn("[progress.percentage]{task.percentage:>3.0f}%"),
                TimeRemainingColumn(),
                console=config.console
            ) as progress:
                task = progress.add_task(f"[cyan]{market_type} 분석 중...[/cyan]", total=len(stock_list))
                
                # [최적화] 실전: 4개 스레드 병렬 처리, 모의: 2개 스레드 병렬 처리
                completed_count = 0
                
                def _process_result(stock_info, res_data):
                    if res_data and 'error' not in res_data:
                        rsi_val = res_data['rsi']
                        rsi_str = f"{rsi_val:.1f}" if rsi_val is not None else "-"
                        adx_str = f"{res_data['adx']:.1f}" if res_data['adx'] is not None else "-"
                        cci_str = f"{res_data['cci']:.1f}" if res_data['cci'] is not None else "-"
                        obv_trend = res_data.get('obv_trend')
                        obv_str = "상승" if obv_trend is True else ("하락" if obv_trend is False else "-")
                        
                        sar_val = res_data.get('psar')
                        if sar_val is not None:
                            sar_str = "상승" if res_data['price'] > sar_val else "하락"
                        else:
                            sar_str = "-"
                        
                        macd_val = res_data.get('macd')
                        sig_val = res_data.get('macd_signal')
                        macd_str = "골든" if macd_val is not None and sig_val is not None and macd_val > sig_val else "데드"
                        
                        vol_str = ""
                        if res_data.get('vol_strength') is not None:
                            vol_str = f", 체결={res_data['vol_strength']:.0f}%"
                        else:
                            vol_str = ", 체결=확인생략"
                        
                        log_msg = f"[{completed_count}/{len(stock_list)}] [분석] {res_data['name']}({res_data['code']}): 현재가={int(res_data['price']):,}, 점수={res_data['score']:.2f}, 상태={res_data['state']}, RSI={rsi_str}, CCI={cci_str}, ADX={adx_str}, OBV={obv_str}, SAR={sar_str}, MACD={macd_str}{vol_str}"
                        
                        if res_data['is_target']:
                            log_style = "bold green" if res_data['state'] in ["매수", "강매수", "역매수"] else "bold orange3"
                            progress.console.print(f"[{log_style}]{log_msg}[/{log_style}]")
                            buy_candidates.append(res_data)
                        else:
                            progress.console.print(f"[dim]{log_msg}[/dim]")
                    else:
                        err_msg = res_data.get('error', '알 수 없는 오류') if res_data else "데이터 부족 또는 API 응답 없음"
                        progress.console.print(f"[dim red][{completed_count}/{len(stock_list)}] [실패] {stock_info['name']}({stock_info['code']}) - {err_msg}[/dim red]")

                # [최적화] 전체 종목 분석 시 모의투자(2) / 실전투자(4) 통합 멀티스레드 적용
                max_w = config.analysis_max_workers()
                with concurrent.futures.ThreadPoolExecutor(max_workers=max_w) as executor:
                    futures = {executor.submit(_analyze_stock_worker, stock, params, restricted_stocks, rules_map, reserved_codes, m_codes): stock for stock in stock_list}
                    for future in concurrent.futures.as_completed(futures):
                        completed_count += 1
                        stock = futures[future]
                        try:
                            result = future.result()
                            _process_result(stock, result)
                        except Exception as e:
                            progress.console.print(f"[dim red][{completed_count}/{len(stock_list)}] [오류] {stock['name']}({stock['code']}) - {e}[/dim red]")
                        
                        progress.advance(task)
                    
        except KeyboardInterrupt:
            config.console.print("\n[yellow]분석이 사용자에 의해 중단되었습니다.[/yellow]")
        finally:
            # 설정 복구
            config.ANALYSIS_THRESHOLDS = original_thresholds

    # 결과 테이블 출력
    if not buy_candidates:
        config.console.print("\n[yellow]조건을 만족하는 종목이 없습니다.[/yellow]")
        return

    # [추가] 선별된 종목에 대해 업종 정보 보강 (캐시에 없거나 새로 분석한 경우)
    need_sector_fetch = False
    if buy_candidates and 'sector' not in buy_candidates[0]:
        need_sector_fetch = True
        
    if need_sector_fetch:
        with Progress(
            SpinnerColumn(),
            TextColumn("[progress.description]{task.description}"),
            BarColumn(),
            TextColumn("[progress.percentage]{task.percentage:>3.0f}%"),
            console=config.console,
            transient=True
        ) as progress:
            task = progress.add_task("[cyan]선별된 종목의 업종 정보를 조회 중...[/cyan]", total=len(buy_candidates))
            
            # 병렬 처리로 업종 정보 조회
            def fetch_sector(item):
                delay = 0.05
                time.sleep(delay)
                try:
                    res = api.get_current_price_data(item['code'], is_overseas=False)
                    if res.get('rt_cd') == '0':
                        return res['output'].get('bstp_kor_isnm', '-')
                except Exception: pass
                return '-'

            # [최적화] 업종 정보 조회 시 모의투자(2) / 실전투자(4) 통합 멀티스레드 적용
            max_w_sec = config.analysis_max_workers()
            with concurrent.futures.ThreadPoolExecutor(max_workers=max_w_sec) as executor:
                future_to_idx = {executor.submit(fetch_sector, item): i for i, item in enumerate(buy_candidates)}
                for future in concurrent.futures.as_completed(future_to_idx):
                    buy_candidates[future_to_idx[future]]['sector'] = future.result()
                    progress.advance(task)
        
        # [추가] 업종(Sector) 정보를 기반으로 확실하게 2차 제외
        if not use_cache and not params.get('INCLUDE_ETF', True):
            original_len = len(buy_candidates)
            buy_candidates = [
                item for item in buy_candidates 
                if not any(kw in str(item.get('sector', '')).upper() for kw in ['ETF', 'ETN', '스팩', 'SPAC', '리츠', 'REIT', '인프라투용', '투자회사'])
            ]
            if len(buy_candidates) < original_len:
                config.console.print(f"[dim]업종 기반 ETF/ETN 등 2차 제외 완료: {original_len}개 -> {len(buy_candidates)}개[/dim]")

        # 새로 분석했거나 sector 정보가 추가된 경우 DB 저장
        if not use_cache:
            _save_analysis_result(market_type, buy_candidates, params)

    # 정렬 기준 개선: 1. 점수 높은 순, 2. RSI 낮은 순 (상승 여력)
    # RSI가 None인 경우 맨 뒤로 보내기 위해 999 처리
    buy_candidates.sort(key=lambda x: (-x['score'], x['rsi'] if x['rsi'] is not None else 999))
    
    filter_mode = params.get("OUTPUT_FILTER", "BUY")
    if filter_mode == "BUY": filter_str = "매수"
    elif filter_mode == "RISE": filter_str = "상승"
    else: filter_str = "매수/상승"
    config.console.print(f"\n[bold]분석 결과: {filter_str} 종목 {len(buy_candidates)}개[/bold]")
    
    # [수정] 페이징 처리 및 컬럼 포맷 변경 (한 줄 출력, 말줄임 방지)
    # 터미널 높이에 따라 페이지 크기 자동 조절
    try:
        terminal_lines = shutil.get_terminal_size().lines
        # 테이블 헤더, 타이틀, 여백, 프롬프트 공간 등을 고려하여 제외 (약 13줄)
        page_size = max(5, terminal_lines - 13)
    except Exception:
        page_size = 15

    total_items = len(buy_candidates)
    total_pages = (total_items + page_size - 1) // page_size
    
    for page in range(total_pages):
        start_idx = page * page_size
        end_idx = min((page + 1) * page_size, total_items)
        page_items = buy_candidates[start_idx:end_idx]
        any_restricted_in_page = False
        
        table = Table(title=f"{market_type} 유망 종목 ({filter_str}) - 페이지 {page+1}/{total_pages}", box=box.HORIZONTALS, header_style="dim", border_style="dim")
        table.add_column("No.", justify="right", width=4)
        table.add_column("종목명(코드)", justify="left", no_wrap=True)
        table.add_column("업종", justify="center", no_wrap=True)
        table.add_column("현재가", justify="right")
        table.add_column("52주(위치)", justify="right")
        table.add_column("점수", justify="center")
        table.add_column("상태", justify="center")
        table.add_column("추세SMO", justify="center")
        table.add_column("ADX", justify="right")
        table.add_column("RSI", justify="right")
        table.add_column("CCI", justify="right")
        table.add_column("체결강도", justify="right")
        
        for i, item in enumerate(page_items):
            rsi_val = item['rsi']
            rsi_str = f"{rsi_val:.1f}" if rsi_val is not None else "-"
            if rsi_val is not None:
                if rsi_val >= config.INDICATOR_PARAMS["RSI_UPPER"]: rsi_str = f"[magenta]{rsi_str}[/]"
                elif 55 <= rsi_val < config.INDICATOR_PARAMS["RSI_UPPER"]: rsi_str = f"[red]{rsi_str}[/]"
                elif 45 <= rsi_val < 55: rsi_str = f"[orange3]{rsi_str}[/]"
                elif config.INDICATOR_PARAMS["RSI_LOWER"] < rsi_val < 45: rsi_str = f"[yellow]{rsi_str}[/]"
                else: rsi_str = f"[blue]{rsi_str}[/]"

            # ADX 값 뒤에 DMI 우위 방향(▲/▼/●)을 함께 표기
            adx_val = item['adx']
            adx_str = format_adx_cell(adx_val, item.get('plus_di'), item.get('minus_di'))

            cci_val = item['cci']
            cci_str = f"{cci_val:.1f}" if cci_val is not None else "-"
            if cci_val is not None:
                if cci_val >= config.INDICATOR_PARAMS["CCI_UPPER"]: cci_str = f"[red]{cci_str}[/]"
                elif 0 < cci_val < config.INDICATOR_PARAMS["CCI_UPPER"]: cci_str = f"[orange3]{cci_str}[/]"
                elif config.INDICATOR_PARAMS["CCI_LOWER"] < cci_val <= 0: cci_str = f"[yellow]{cci_str}[/]"
                else: cci_str = f"[blue]{cci_str}[/]"
            
            # SAR 상태
            sar_val = item.get('psar')
            sar_icon = "[red]▲[/]" if sar_val and item['price'] > sar_val else "[blue]▼[/]"
            
            # MACD 상태
            macd_val = item.get('macd')
            sig_val = item.get('macd_signal')
            macd_icon = "-"
            if macd_val is not None and sig_val is not None:
                zero_sign = "+" if macd_val > 0 else "-"
                cross_char = "G" if macd_val > sig_val else "D"
                m_color = "red" if macd_val > sig_val else "blue"
                macd_icon = f"[{m_color}]{zero_sign}{cross_char}[/]"

            s_color = item.get('state_color', '[white]').replace('[', '').replace(']', '')
            display_state = item['state']
            
            # 52주 위치 색상
            pos = item.get('w52_pos', 0)
            w_color = "[white]"
            if pos >= 90: w_color = "[red]"
            elif pos >= 80: w_color = "[orange3]"
            elif pos <= 20: w_color = "[blue]"
            
            obv_trend = item.get('obv_trend')
            obv_icon = "-"
            if obv_trend is True: obv_icon = "[red]▲[/]"
            elif obv_trend is False: obv_icon = "[blue]▼[/]"
            
            trend_str = f"{sar_icon} {macd_icon} {obv_icon}"
            
            vol_val = item.get('vol_strength')
            vol_str = "-"
            if vol_val is not None:
                std_vol = config.ANALYSIS_THRESHOLDS.get("BUY_VOL_STRENGTH", 100.0)
                v_color = "[red]" if vol_val >= std_vol else "[blue]"
                vol_str = f"{v_color}{vol_val:.1f}%[/]"
            
            name_display = item['name']
            marks = []
            if item['code'] in restricted_stocks: marks.append("-")
            if item.get('is_custom_rule'): marks.append("+")
            if item['code'] in m_codes: marks.append("=")
            if item['code'] in reserved_codes: marks.append("[magenta]*[/magenta]")
            
            if marks:
                name_display += f"[dim]{''.join(marks)}[/dim]"

            table.add_row(
                str(start_idx + i + 1),
                f"{name_display} [dim]({item['code']})[/dim]",
                item.get('sector', '-'),
                f"{int(item['price']):,}원",
                f"{w_color}{pos:.1f}%[/]",
                f"[{s_color}]{item['score']}[/]",
                f"[{s_color}]{display_state}[/]",
                trend_str,
                adx_str,
                rsi_str,
                cci_str,
                vol_str
            )
            
            # 5개마다 실선 추가
            if (i + 1) % 5 == 0 and (i + 1) < len(page_items):
                table.add_section()
                
        config.console.print(table, crop=False)
        sys.stdout.flush()
        
        if page < total_pages - 1:
                if Prompt.ask(f"[dim]다음 페이지를 보시겠습니까? (이전: b, 메인: q)[/dim]", choices=["y", "n", "b", "q"], default="y").lower() in ['b', 'q', 'n']:
                    break

    # 상세 분석 이동 기능
    from modules import chart
    
    while True:
        config.console.print("\n[dim]개별 분석 및 상세 차트 분석을 보려면 종목 번호를 입력하세요 (Enter: 메뉴복귀, 이전: b, 메인: q)[/dim]")
        choice = Prompt.ask("선택", default="b", show_default=False)
        
        if choice.lower() in ['b', 'q']:
            return False
            
        if choice.isdigit():
            idx = int(choice) - 1
            if 0 <= idx < len(buy_candidates):
                selected = buy_candidates[idx]
                code = selected['code']
                name = selected['name']
                
                # [수정] 차트 분석 전 개별 종목 분석 결과 출력
                config.console.print(f"\n[bold green]>> {name}({code}) 개별 종목 심층 분석 실행[/bold green]")
                diagnose_stock(code, name, target_is_overseas=False)
            else:
                config.console.print("[red]잘못된 번호입니다. 리스트에 있는 번호를 입력해주세요.[/red]")
        else:
            config.console.print("[red]올바른 번호를 입력해주세요.[/red]")

def save_all_market_analysis():
    """코스피/코스닥 전 종목 분석 결과를 엑셀로 저장"""
    # [최적화] openpyxl은 이 함수에서만 사용 → 지연 임포트로 프로그램 시작 시간 단축
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter


    config.console.print("\n[bold]전체 종목 분석결과 저장 (Export to Excel)[/bold]")
    config.console.print("[dim]코스피 및 코스닥 전 종목을 분석하여 파일로 저장합니다.[/dim]")
    config.console.print("[dim]시간이 오래 걸릴 수 있습니다. (중단: Ctrl+C)[/dim]\n")
    
    if Prompt.ask("진행하시겠습니까?", choices=["y", "n"], default="n") != "y":
        return False

    # [추가] 개별 룰 로드 (전체 조회 최적화)
    custom_rules = db_manager.db.get_all_stock_strategies()
    rules_map = {}
    for r in custom_rules:
        r_dict = dict(r)
        if r_dict.get('weights') and isinstance(r_dict['weights'], str):
            try: r_dict['weights'] = json.loads(r_dict['weights'])
            except Exception: r_dict['weights'] = None
        rules_map[r_dict['code']] = r_dict
    
    # [추가] 예약 매매 및 메모 마커 조회
    reserved_codes = set()
    try:
        pending_reserves = db_manager.db.get_pending_reserved_orders()
        reserved_codes = set(o['code'] for o in pending_reserves)
    except Exception: pass
    m_codes = utils.get_memo_codes()

    # [추가] 트레이딩 제한 종목 로드
    from modules import auto_trade
    restricted_stocks = auto_trade.get_restricted_stocks()

    markets = ["KOSPI", "KOSDAQ"]
    results = {} # market -> list of dict

    try:
        with Progress(
            SpinnerColumn(),
            TextColumn("[progress.description]{task.description}"),
            BarColumn(),
            TextColumn("[progress.percentage]{task.percentage:>3.0f}%"),
            TimeRemainingColumn(),
            console=config.console
        ) as progress:
            
            for market_type in markets:
                stock_list = _get_master_stock_list(market_type)
                if not stock_list: continue
                
                results[market_type] = []
                
                # 1. 기술적 분석 (Chart Data)
                analyzed_data = []
                task = progress.add_task(f"[cyan]{market_type} 기술적 분석 중...[/cyan]", total=len(stock_list))

                max_w = config.analysis_max_workers()

                # 1. 기술적 분석 병렬 처리
                with concurrent.futures.ThreadPoolExecutor(max_workers=max_w) as executor:
                    futures = {executor.submit(_analyze_stock_worker, stock, None, restricted_stocks, rules_map, reserved_codes, m_codes): stock for stock in stock_list}
                    for future in concurrent.futures.as_completed(futures):
                        try:
                            result = future.result()
                            if result and 'error' not in result: analyzed_data.append(result)
                        except Exception: pass
                        progress.advance(task)
                
                # 2. 업종 정보 조회 (Price Data) 및 데이터 정제
                if analyzed_data:
                    task_sector = progress.add_task(f"[cyan]{market_type} 업종 정보 조회 및 정리 중...[/cyan]", total=len(analyzed_data))
                    
                    def fetch_sector_and_format(item):
                        sector = "-"
                        try:
                            res = api.get_current_price_data(item['code'], is_overseas=False)
                            if res.get('rt_cd') == '0':
                                sector = res['output'].get('bstp_kor_isnm', '-')
                        except Exception: pass
                        
                        # 데이터 포맷팅 (소수점 1자리, 정수 등)
                        rsi = round(item['rsi'], 1) if item['rsi'] is not None else None
                        adx = round(item['adx'], 1) if item['adx'] is not None else None
                        cci = round(item['cci'], 1) if item['cci'] is not None else None
                        w52 = int(item['w52_pos']) if item['w52_pos'] is not None else 0
                        vol = round(item['vol_strength'], 1) if item.get('vol_strength') else None
                        
                        # SAR/MACD 상태
                        sar_val = item.get('psar')
                        if sar_val is not None and not math.isnan(sar_val):
                            sar_state = "상승" if item['price'] > sar_val else "하락"
                        else:
                            sar_state = "-"
                        
                        macd_state = "-"
                        macd_val = item.get('macd')
                        sig_val = item.get('macd_signal')
                        if macd_val is not None and sig_val is not None and not math.isnan(macd_val) and not math.isnan(sig_val):
                            macd_state = "골든" if macd_val > sig_val else "데드"

                        name_display = item['name']
                        marks = []
                        if item['code'] in restricted_stocks: marks.append("-")
                        if item.get('is_custom_rule'): marks.append("+")
                        if item['code'] in m_codes: marks.append("=")
                        if item['code'] in reserved_codes: marks.append("*")
                        if marks: name_display += "".join(marks)

                        # [추가] 비고 (개별 룰 요약)
                        note = ""
                        if item['code'] in rules_map:
                            rule = rules_map[item['code']]
                            changes = []
                            
                            # 전역 설정값
                            def_buy_score = config.ANALYSIS_THRESHOLDS["BUY_SCORE"]
                            def_buy_rsi = config.ANALYSIS_THRESHOLDS["BUY_RSI_MAX"]
                            def_buy_vol = config.ANALYSIS_THRESHOLDS.get("BUY_VOL_STRENGTH", 100.0)
                            def_sell_score = config.SELL_STRATEGY["SELL_SCORE"]
                            def_tp = config.SELL_STRATEGY["TAKE_PROFIT_RATE"]
                            def_sl = config.SELL_STRATEGY["STOP_LOSS_RATE"]

                            # 비교
                            if rule.get('buy_score') != def_buy_score: changes.append(f"매수점수({rule['buy_score']})")
                            if rule.get('buy_rsi') != def_buy_rsi: changes.append(f"매수RSI({rule['buy_rsi']})")
                            if rule.get('buy_vol_strength') and rule['buy_vol_strength'] != def_buy_vol: changes.append(f"체결({rule['buy_vol_strength']}%)")
                            def_ask_ratio = config.ANALYSIS_THRESHOLDS.get('BUY_ASK_BID_RATIO', 1.0)
                            if rule.get('buy_ask_bid_ratio') is not None and rule['buy_ask_bid_ratio'] != def_ask_ratio: changes.append(f"매도잔량비({rule['buy_ask_bid_ratio']}배)")
                            def_auto = config.ANALYSIS_THRESHOLDS.get('AUTO_ADJUST_ASK_BID_RATIO', True)
                            if rule.get('auto_adjust_ask_bid_ratio') is not None and bool(rule['auto_adjust_ask_bid_ratio']) != def_auto: changes.append(f"자동연동({bool(rule['auto_adjust_ask_bid_ratio'])})")
                            if rule.get('sell_score') != def_sell_score: changes.append(f"매도점수({rule['sell_score']})")
                            if rule.get('take_profit') != def_tp: changes.append(f"익절({rule['take_profit']}%)")
                            if rule.get('stop_loss') != def_sl: changes.append(f"손절({rule['stop_loss']}%)")
                            if rule.get('weights'): changes.append("가중치")
                            
                            if changes:
                                note = f"개별룰: {', '.join(changes)}"
                            else:
                                note = "개별룰 적용"

                        return {
                            "종목코드": item['code'],
                            "종목명": name_display,
                            "업종": sector,
                            "현재가(원)": item['price'],
                            "52주위치(%)": w52,
                            "점수": item['score'],
                            "상태": item['state'],
                            "상태사유": item['state_reason'],
                            "RSI": rsi,
                            "CCI": cci,
                            "ADX": adx,
                            "SAR": sar_state,
                            "MACD": macd_state,
                            "OBV": "상승" if item['obv_trend'] is True else ("하락" if item['obv_trend'] is False else "-"),
                            "체결강도": vol,
                            "비고": note # [추가]
                        }

                    with concurrent.futures.ThreadPoolExecutor(max_workers=max_w) as executor:
                        futures_sector = {executor.submit(fetch_sector_and_format, item): item for item in analyzed_data}
                        for future in concurrent.futures.as_completed(futures_sector):
                            try:
                                formatted_result = future.result()
                                results[market_type].append(formatted_result)
                            except Exception: pass
                            progress.advance(task_sector)

        # 엑셀 저장
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = os.path.join(config.DATA_DIR, f"market_analysis_{timestamp}.xlsx")
        
        if not any(results.values()):
            config.console.print("\n[red]저장할 데이터가 없습니다. (마스터 파일 오류 또는 분석 실패)[/red]")
            return

        with Progress(
            SpinnerColumn(),
            TextColumn("[progress.description]{task.description}"),
            BarColumn(),
            TextColumn("[progress.percentage]{task.percentage:>3.0f}%"),
            console=config.console,
            transient=True
        ) as progress:
            task = progress.add_task(f"[cyan]엑셀 파일 저장 중... ({os.path.basename(filename)})[/cyan]", total=len(results))
            
            with pd.ExcelWriter(filename, engine='openpyxl') as writer:
                for market_type, data in results.items():
                    if data:
                        # 점수 높은 순 정렬
                        data.sort(key=lambda x: (-x['점수'], x['RSI'] if x['RSI'] is not None else 999))
                        df = pd.DataFrame(data)
                        df.to_excel(writer, sheet_name=market_type, index=False)
                        
                        # 엑셀 서식 적용 (필터, 컬럼 너비, 색상 등)
                        ws = writer.sheets[market_type]
                        ws.auto_filter.ref = ws.dimensions
                        
                        # 헤더에서 컬럼 인덱스 찾기
                        header = [c.value for c in ws[1]]
                        try:
                            col_price = header.index("현재가(원)") + 1
                            col_state = header.index("상태") + 1
                            col_score = header.index("점수") + 1
                            
                            # [수정] 모든 컬럼 너비 자동 조절
                            for i, col_name in enumerate(header):
                                col_idx = i + 1
                                col_letter = get_column_letter(col_idx)
                                
                                # 헤더 텍스트 길이 고려
                                s_header = str(col_name)
                                max_width = len(s_header) + sum(0.7 for c in s_header if ord(c) > 127)
                                
                                for row in range(2, ws.max_row + 1):
                                    val = ws.cell(row=row, column=col_idx).value
                                    if val:
                                        s_val = str(val)
                                        length = len(s_val) + sum(0.7 for c in s_val if ord(c) > 127)
                                        if length > max_width: max_width = length
                                
                                limit = 100 if col_name == "비고" else 60
                                ws.column_dimensions[col_letter].width = min(max_width * 1.2, limit)

                            for row in range(2, ws.max_row + 1):
                                # 현재가 쉼표 포맷
                                ws.cell(row=row, column=col_price).number_format = '#,##0'
                                
                                # 점수 소수점 2자리 포맷
                                ws.cell(row=row, column=col_score).number_format = '0.00'
                                
                                # 상태 컬럼 색상 적용
                                cell = ws.cell(row=row, column=col_state)
                                val = cell.value
                                if val in ["매수", "강매수"]: cell.font = Font(color="FF0000", bold=True)
                                elif val in ["상승", "대기"]: cell.font = Font(color="FF8C00", bold=True)
                                elif val == "주의": cell.font = Font(color="DAA520", bold=True)
                                elif val == "매도": cell.font = Font(color="0000FF", bold=True)
                        except ValueError: pass
                    
                    progress.advance(task)
        
        config.console.print(f"\n[bold green]저장 완료: {filename}[/bold green]")
        
    except KeyboardInterrupt:
        config.console.print("\n[yellow]작업이 중단되었습니다.[/yellow]")
    except Exception as e:
        config.console.print(f"\n[bold red]오류 발생: {e}[/bold red]")

# [진단] 일봉 수신이 이 시간을 넘긴 종목은 로그로 남긴다.
#  '데이터 수신' 단계가 마지막 한두 종목에서 오래 멈추는데, 화면에는 진행률만 있어 어느
#  종목이 원인인지 특정할 수 없었다. 프로그래스 바 표기는 그대로 두고 로그로만 알린다.
SLOW_CHART_FETCH_SEC = 5.0
# 단계별 총계 로그에서 '이 표에 느린 종목이 몇 건이었는지'를 세는 카운터(스레드에서 증가).
#  GIL 하에서 리스트 원소 증가는 원자적이지 않지만, 진단용 근사치라 락을 두지 않는다.
_SLOW_FETCH_COUNT = [0]


def _fetch_chart_data(item, is_overseas):
    """(내부함수) print_table 1단계: 과거(전체) 일봉 차트 데이터 수신.

    캐시 적중(6시간 이내·당일) 시 즉시 반환되고, 캐시 미스 때만 실제 250봉 다운로드가 발생한다.
    → '데이터 수신' 프로그래스 바는 실제 전체 데이터를 받아오는 동안에만 길어진다.
    """
    name, code = item
    started = time.monotonic()
    started_at = datetime.now()
    df = None
    try:
        df = api.get_chart_data(code, is_overseas, 'daily', False)
        return df
    except Exception as e:
        logger.error(f"[{code}] 차트 데이터 수신 오류: {e}")
        return None
    finally:
        elapsed = time.monotonic() - started
        if elapsed >= SLOW_CHART_FETCH_SEC:
            # 어떤 소스로 받았는지(KIS/TVDATAFEED/토스 등)와 봉 수를 함께 남겨야
            #  느린 구간이 어느 경로인지 한 번에 판별된다.
            try:
                src = (df.attrs.get('source') if df is not None and hasattr(df, 'attrs') else None) or '-'
                bars = 0 if df is None or df.empty else len(df)
            except Exception:      # noqa: BLE001 - 진단 로그가 본 흐름을 막지 않게 한다
                src, bars = '-', 0
            # 시작 시각을 함께 남긴다 — 완료 시각만으로는 여러 종목이 실제로 겹쳐 돌았는지
            #  (병렬) 차례로 돌았는지(직렬) 구분할 수 없다. 표 꼬리가 한 종목의 소요로
            #  수렴하는지, 소요의 합으로 늘어지는지가 여기서 갈린다.
            _SLOW_FETCH_COUNT[0] += 1
            logger.debug(f"[느린 일봉 수신] {name}({code}) {elapsed:.1f}초 "
                         f"(시작 {started_at.strftime('%H:%M:%S.%f')[:-3]}) "
                         f"| 소스={src} 봉={bars}")

def _collect_table_data(item, title, is_overseas, use_investor_data, chart_df=None, preloaded_curr=None):
    """(내부함수) print_table 2단계 전반부: 당일 실시간 데이터(현재가/체결강도/수급/상세) 수신.

    chart_df가 주어지면(1단계에서 수신) 차트는 재수신하지 않는다. 실제 지표 분석/행 포맷은
    _analyze_table_row가 담당한다. (메뉴1처럼 '데이터 수신' / '실시간 데이터 수신 및 분석' 단계 분리)
    preloaded_curr: 멀티시세(30종목/1콜) 프리페치로 확보한 현재가 응답({'rt_cd','output'}).
      주어지면 종목별 현재가 REST를 생략한다(TPS 절감). 없으면 종전대로 개별 조회.
    """
    name, code = item
    bundle = {'curr_data': None, 'chart_df': chart_df, 'inv_list': None,
              'rt_strength': None, 'ask_bid_ratio': None, 'detail': None}
    try:
        cached_ex = config.session.exchange_cache.get(code, "NAS") if is_overseas else None
        curr_data = inv_list = detail = None
        rt_strength = None
        ask_bid_ratio = None  # [토스] 체결강도 미제공 → 매도잔량비로 대체 표시

        # [최적화] 필요한 다수의 API를 병렬(Fan-out)로 일제히 호출하여 체감 속도 극대화
        for attempt in range(2):
            with concurrent.futures.ThreadPoolExecutor(max_workers=5) as ex:
                # [실시간/통합] 현재가·체결강도를 KRX/NXT 한 번의 호출로 통합 조회한다(include_nxt=True).
                # cache_ttl=0 → 예열 캐시를 재사용하지 않는 라이브 호출(25초 예열 폐지). 시스템 트레이딩과
                # 동일 캐시 키(cp_{code}_J / vol_{code})를 공유해 동시 조회 시 중복 호출이 합쳐진다.
                # (모의투자(VTS)는 NXT 미지원이라 내부에서 NX 조회를 건너뛴다 → 정규장 시세만 표시)
                # [멀티시세] 프리페치된 현재가가 있으면 종목별 REST 생략
                # [최적화] fast_info_ttl=30: 해외 현재가의 KIS 실패 시 TV 폴백이 직전 TV 일괄 예열 캐시를 재사용
                #  (종목별 TV 단건 재조회 제거. 정상 경로는 KIS last/diff/rate를 그대로 사용)
                fut_curr = ex.submit(api.get_current_price_data, code, is_overseas, True, 0, 30.0) if preloaded_curr is None else None
                # [최적화] 차트는 1단계에서 받았으면 재수신하지 않는다(미제공 시에만 캐시 경로로 조회).
                fut_chart = ex.submit(api.get_chart_data, code, is_overseas, 'daily', False) if chart_df is None else None
                fut_inv = ex.submit(api.get_investor_trend, code) if not is_overseas and use_investor_data else None
                fut_vol = ex.submit(api.get_realtime_vol_strength, code, is_overseas, cached_ex, True, 0) if not is_overseas and not use_investor_data else None
                fut_detail = ex.submit(api.fetch_overseas_detail_price, code, cached_ex) if is_overseas else None
                # [토스] 체결강도 대체 지표(매도잔량비)용 호가 조회
                fut_ab = ex.submit(api.get_ask_bid_ratio, code, False) if (config.session.is_toss and not is_overseas) else None

                curr_data = fut_curr.result() if fut_curr is not None else preloaded_curr
                if fut_chart is not None:
                    chart_df = fut_chart.result()
                inv_list = fut_inv.result() if fut_inv else None
                if fut_vol:
                    try: rt_strength = fut_vol.result()
                    except Exception: pass
                if fut_ab:
                    try: ask_bid_ratio = fut_ab.result()
                    except Exception: pass
                detail = fut_detail.result() if fut_detail else None

            if curr_data and curr_data.get('rt_cd') == '0' and chart_df is not None and not chart_df.empty:
                break
            time.sleep(0.5)

        bundle.update({'curr_data': curr_data, 'chart_df': chart_df, 'inv_list': inv_list,
                       'rt_strength': rt_strength, 'ask_bid_ratio': ask_bid_ratio, 'detail': detail})
    except Exception as e:
        logger.error(f"[{code}] 데이터 수집 오류: {e}")
    return bundle

def _prelisting_last_regular_change(chart_df, curr):
    """장전(NXT 프리마켓 08:00 개장 전) 무체결로 현재가=기준가가 되어 등락률이 0%로 굳는
    국내 종목에 대해, 일봉 차트에서 '직전 정규장 최종 등락률'(전일 종가 vs 전전일 종가)을
    산출한다. curr는 전일 정규장 종가(=현재가=기준가)로, 전전일 종가와 비교한다.

    apply_realtime_price가 장전에 당일(오늘) 봉을 새로 추가하므로 마지막 봉이 오늘이면
    전전일은 -3, (당일 봉이 없어) 마지막 봉이 전일이면 전전일은 -2다. 산출 불가 시 None.
    기준일은 달력 날짜(오늘)를 쓴다 — market_today는 휴장일에 직전 거래일을 반환해
    주말엔 마지막 봉(금요일)이 '오늘 봉'으로 오인돼 전전일이 -3(하루 더 과거)으로 밀린다.
    반환: (diff:int, rate:float)
    """
    try:
        if chart_df is None or chart_df.empty or len(chart_df) < 2:
            return None
        today = datetime.now().strftime('%Y%m%d')
        idx = -3
        if 'date' in chart_df.columns:
            last_val = chart_df.iloc[-1]['date']
            if hasattr(last_val, 'strftime'):
                last_str = last_val.strftime('%Y%m%d')
            else:
                last_str = str(last_val).replace('-', '').replace('/', '')[:8]
            idx = -3 if last_str >= today else -2
        if len(chart_df) < abs(idx):
            return None
        pp = float(chart_df.iloc[idx]['close'])  # 전전일 종가
        if pp <= 0:
            return None
        diff = int(round(curr - pp))
        rate = (curr - pp) / pp * 100
        return diff, rate
    except Exception:
        return None

_W52_MIN_BARS = 200   # 이보다 짧으면 창이 52주를 못 채운 것 → 차트 대신 벤더(API/스캐너) 값을 쓴다


def _w52_high_low(chart_df):
    """일봉에서 '최근 365일'(=52주) 구간의 (고가, 저가)를 구한다. 산출 불가 시 (None, None).

    tail(250)을 쓰면 안 된다 — 250거래일은 실측상 373일치(2025-07-14~2026-07-22)라 52주보다
    8일 넓고, 그 경계 밖 극값이 밴드를 통째로 왜곡한다(TIGER 조선TOP10: 창 밖 18,855가 잡혀
    20.2%, 실제 52주 기준은 11.0%). 날짜로 잘라 벤더의 52주 정의와 창을 맞춘다.

    봉 수가 _W52_MIN_BARS 미만이면 신규상장이거나 차트 수신이 잘린 경우다. 이때 좁아진 밴드를
    그대로 쓰면 52주 위치가 부풀려지므로 호출부가 벤더 값으로 폴백하도록 None을 준다.
    """
    try:
        if chart_df is None or chart_df.empty or 'date' not in chart_df.columns:
            return None, None
        cutoff = (datetime.now() - timedelta(days=365)).strftime('%Y%m%d')
        dates = chart_df['date'].astype(str).str.replace('-', '', regex=False).str[:8]
        win = chart_df[dates >= cutoff]
        if len(win) < _W52_MIN_BARS:
            return None, None
        h, l = float(win['high'].max()), float(win['low'].min())
        return (h, l) if h > l > 0 else (None, None)
    except Exception:
        return None, None


def _w52_band(chart_df):
    """52주 고/저 밴드 (h52, l52). 산출 불가 시 (0.0, 0.0).

    화면별로 창 정의가 달라 같은 종목의 52주 값이 어긋나던 것을 통일하기 위한 단일 진입점이다.
    (표는 365일, 지수·상세·점수 폴백은 tail(250)=실측 373일을 쓰고 있었다. 차이는 경계 밖
     8일에 극값이 걸릴 때만 생기지만, 걸리면 크다 — TIGER 조선TOP10 20.2%→11.0%.
     2026-07-24 보유 10종목 실측: 삼성SDI 46.2%→45.0%, 삼성전자·현대차 각 -0.3%p)

    365일 창을 못 채우는 경우(신규상장·차트 수신 절단)만 보유 봉 전체로 폴백한다.
    폴백 조건은 _w52_high_low가 판정한다(_W52_MIN_BARS 미만).

    주의: 창의 기준점은 '오늘'이다. 과거 시점 df(백테스트)에 쓰면 창이 어긋나지만, 그 경로는
    w52_pos를 사전계산해 인자로 넘기므로 여기에 도달하지 않는다. 도달하더라도 창이 비어
    _w52_high_low가 None을 주고 기존 tail(250) 동작으로 폴백한다.
    """
    if chart_df is None or getattr(chart_df, 'empty', True):
        return 0.0, 0.0
    h, l = _w52_high_low(chart_df)
    if h is None:
        try:
            recent = chart_df.tail(250)
            h, l = float(recent['high'].max()), float(recent['low'].min())
        except Exception:
            return 0.0, 0.0
    return h, l


def _analyze_table_row(item, title, is_overseas, use_investor_data, restricted_stocks, rules_map, market_regime_adj, reserved_codes, m_codes, bundle):
    """(내부함수) print_table 2단계: 수집된 데이터(bundle)로 지표 분석 및 행 포맷."""
    try:
        name, code = item
        w52_pos_str, per_str, pbr_str, shar_str = "[dim]-[/dim]", "[dim]-[/dim]", "[dim]-[/dim]", "[dim]-[/dim]"
        w52_pos_val = 0.0  # 52주 위치(%) — 표시 문자열과 슈퍼 모멘텀 판정이 공유
        foreign_rate_str = "[dim]-[/dim]"
        inv_str = "[dim]-[/dim]"
        cached_ex = config.session.exchange_cache.get(code, "NAS") if is_overseas else None
        strength_display = ""
        ask_bid_ratio = None  # [토스] 체결강도 미제공 → 매도잔량비로 대체 표시

        # [수정] 타이틀 기반으로 주식/ETF 컨텍스트 정확히 구분 (데이터 처리 및 컬럼 매칭용)
        # 기존: 코드 형태(숫자 여부)로만 판단하여 ETF(QQQ 등)를 주식으로 오인하는 문제 해결
        is_us_stock_context = is_overseas and ("주식" in title)
        is_us_etf_context = is_overseas and ("ETF" in title)

        # [최적화] 1단계(_collect_table_data)에서 수집한 원천 데이터를 사용한다.
        curr_data = bundle.get('curr_data')
        chart_df = bundle.get('chart_df')
        inv_list = bundle.get('inv_list')
        rt_strength = bundle.get('rt_strength')
        ask_bid_ratio = bundle.get('ask_bid_ratio')
        detail = bundle.get('detail')

        # [추가] 차트 데이터 당일 종가/고가/저가 실시간 갱신 (점수 0.5점 오차 방지)
        try:
            rt_price = 0.0
            if curr_data and curr_data.get('rt_cd') == '0':
                if is_overseas: 
                    rt_price = float(curr_data['output'].get('last', 0) or 0)
                else: 
                    nxt_price = float(curr_data['output'].get('ats_prpr', 0) or 0)
                    krx_price = float(curr_data['output'].get('stck_prpr', 0) or 0)
                    rt_price = nxt_price if nxt_price > 0 else krx_price

            indicators.apply_realtime_price(chart_df, api.chart_overlay_price(rt_price, is_overseas),
                                            market_date=utils.market_today(is_overseas))
        except Exception: pass

        # [모든 장 마감 후] 표시 현재가·등락·52주 위치를 확정된 KRX 일봉으로 맞춘다.
        #  이 행의 현재가는 curr_data(ats_prpr/stck_prpr)에서 직접 오는데, 20:00 이후·주말에는
        #  그 값이 '마지막 NXT 체결가'로 굳어 있다. USE_KRX_CLOSE_AFTER_HOURS(기본 True)면
        #  KRX 확정 종가로 고정하고, 끄면 마지막 실거래가를 그대로 노출한다.
        #  등락은 API 기준가 대신 일봉 직전 봉과 비교해 'KRX 정규장 등락'으로 통일한다.
        #  [Fix 2026-07-28] 게이트를 chart_overlay_enabled(지표용)에서 display_price_krx_fixed
        #   (표시용)로 분리한다. 지표 게이트가 정규장 전용이 되면서, 그대로 두면 NXT 거래시간
        #   (프리·애프터)에도 표시가 KRX로 굳어 '살아있는 NXT 가격을 본다'는 목적이 깨진다.
        #  [방어] 마지막 봉이 '직전 거래일'보다 오래됐으면(당일 봉 미수신) 확정 종가로 쓰지 않는다.
        #   차트가 하루 밀린 채 현재가 자리에 실리면 지난 거래일 종가·등락이 오늘 값으로 보인다
        #   (2026-07-27 22:40 실측: 삼성전자 249,500 -7.59% = 7/24 값. 원인이던 캐시는
        #    api._chart_disk_get에서 고쳤고, 여기서는 같은 증상이 다시 새지 않게 막는다).
        #   이 경우 현재가·등락은 실시간 시세(curr_data)에서 오는 종전 경로를 그대로 탄다.
        krx_close_px, krx_prev_px = 0.0, 0.0
        if not is_overseas and api.display_price_krx_fixed(False) \
                and chart_df is not None and not chart_df.empty:
            try:
                last_bar_date = str(chart_df.iloc[-1]['date']).replace('-', '')[:8]
                # [Fix 2026-07-28] 기준을 market_today → krx_last_settled_day로 바꾼다.
                #  자정~개장 전에는 market_today가 아직 열리지도 않은 '오늘'을 돌려줘
                #  이 조건이 항상 실패했고, 그 결과 새벽에 KRX 고정이 걸리지 않았다.
                if last_bar_date >= api.krx_last_settled_day():
                    krx_close_px = float(chart_df.iloc[-1]['close'])
                    if len(chart_df) >= 2:
                        krx_prev_px = float(chart_df.iloc[-2]['close'])
            except (TypeError, ValueError, KeyError, IndexError):
                krx_close_px, krx_prev_px = 0.0, 0.0

        # [추가] 토스: 현재가 API가 등락(전일대비)/52주 고저를 제공하지 않으므로 차트(캔들)에서 보강한다.
        # (이 함수는 이미 chart_df를 확보하므로 추가 API 호출 없이 out에 주입한다)
        if config.session.is_toss and curr_data and curr_data.get('rt_cd') == '0' \
           and chart_df is not None and not chart_df.empty:
            _o = curr_data['output']
            try:
                _h52, _l52 = _w52_high_low(chart_df)
                if _h52 is None:  # 52주에 못 미치는 차트(신규상장 등)는 확보한 전 구간으로 폴백
                    _h52, _l52 = float(chart_df['high'].max()), float(chart_df['low'].min())
                _cur = float(chart_df['close'].iloc[-1])
                _prev = float(chart_df['close'].iloc[-2]) if len(chart_df) >= 2 else _cur
                if is_overseas:
                    # 52주 위치는 가격 기반이므로 차트로 산출(detail 경로 유지).
                    # PER/PBR/상장주수는 fetch_overseas_detail_price가 TradingView 스캐너로 채운다.
                    if not detail:
                        detail = {}
                    detail.setdefault('h52p', _h52)
                    detail.setdefault('l52p', _l52)
                    detail.setdefault('last', _cur)
                    if _prev > 0:
                        _o['diff'] = _cur - _prev
                        _o['rate'] = (_cur - _prev) / _prev * 100
                else:
                    _o['w52_hgpr'] = str(_h52); _o['w52_lwpr'] = str(_l52)
                    if _prev > 0:
                        # 어댑터가 역산한 KRX 기준가(stck_sdpr)가 있으면 우선, 없을 때만 차트 전일종가로 폴백
                        _o.setdefault('stck_sdpr', str(int(_prev)))
            except Exception: pass

        # [멀티시세] 이 TR은 52주 고저(w52_*)를 제공하지 않으므로 차트(최근 365일)로 보강한다.
        if not is_overseas and curr_data and curr_data.get('rt_cd') == '0' \
           and curr_data.get('output', {}).get('_src') == 'multi' \
           and chart_df is not None and not chart_df.empty:
            _o = curr_data['output']
            _h52, _l52 = _w52_high_low(chart_df)
            if _h52 is not None:
                _o['w52_hgpr'] = str(_h52); _o['w52_lwpr'] = str(_l52)

        ind = indicators.calculate_indicators(chart_df)

        if not is_overseas:
            if use_investor_data and inv_list:
                p = api.safe_int(inv_list[0].get('prsn_ntby_qty'))
                f = api.safe_int(inv_list[0].get('frgn_ntby_qty'))
                i = api.safe_int(inv_list[0].get('orgn_ntby_qty'))
                def fmt_inv(val):
                    if val == 0: return "[dim]-[/dim]"
                    abs_val = abs(val)
                    if abs_val >= 1_000_000_000: s = f"{val/1_000_000_000:,.1f}B"
                    elif abs_val >= 1_000_000: s = f"{val/1_000_000:,.1f}M"
                    elif abs_val >= 1000: s = f"{val/1000:,.0f}K"
                    else: s = f"{val:,}"
                    return f"[red]{s}[/]" if val > 0 else f"[blue]{s}[/]"
                inv_str = f"{fmt_inv(p)} {fmt_inv(f)} {fmt_inv(i)}"
            if not use_investor_data:
                if config.session.is_toss:
                    # 토스: 체결강도 미제공 → 매도잔량비(매도/매수 총잔량)로 대체 표시(숫자만)
                    # 색상: 기준 1.0배 중심 5단계(체결강도 100% 밴딩과 동일 방향, 높을수록 빨강 계열)
                    # [수정] NXT 운영시간(08:00~20:00) 밖에는 매도비 표기 자체를 생략
                    #  (print_table 헤더의 ' [매도비]' 접미사 생략과 짝 — 컬럼 표기 삭제)
                    if not api.is_toss_ask_bid_window():
                        strength_display = ""
                    elif ask_bid_ratio is not None:
                        if ask_bid_ratio >= 2.0: ab_color = "[magenta]"
                        elif ask_bid_ratio >= 1.5: ab_color = "[red]"
                        elif ask_bid_ratio > 1.0: ab_color = "[orange3]"
                        elif ask_bid_ratio == 1.0: ab_color = "[white]"
                        elif ask_bid_ratio >= 0.7: ab_color = "[yellow]"
                        else: ab_color = "[blue]"
                        strength_display = f" {ab_color}[{ask_bid_ratio:.2f}][/]"
                    else:
                        strength_display = " [dim][-][/dim]"
                elif rt_strength is not None:
                    if rt_strength >= 150: s_color = "[magenta]"
                    elif rt_strength >= 120: s_color = "[red]"
                    elif rt_strength > 100: s_color = "[orange3]"
                    elif rt_strength == 100: s_color = "[white]"
                    elif rt_strength >= 80: s_color = "[yellow]"
                    else: s_color = "[blue]"
                    strength_display = f" {s_color}[{rt_strength:,.0f}%][/]"
                else: strength_display = " [dim][0%][/dim]"
            if curr_data and curr_data.get('rt_cd') == '0':
                out = curr_data.get('output', {})
                foreign_rate_str = f"{out.get('hts_frgn_ehrt', '-')}%"
                try:
                    # [수정] 52주 밴드는 차트(최근 365일)를 1차로 쓰고, 못 구할 때만 API 값으로 폴백한다.
                    #  KIS w52_hgpr/w52_lwpr는 수정주가가 반영되지 않아 차트와 어긋나는 종목이 있다
                    #  (삼성바이오 API 저가 982,000 vs 차트·토스 모두 1,228,000 — API 창이 차트보다
                    #   좁은데 더 낮은 저가를 주므로 산술적으로 성립하지 않는다).
                    #  차트 1차로 통일하면 mode 2/3의 산출 방식도 같아져 모드 간 값이 수렴한다.
                    h52, l52 = _w52_high_low(chart_df)
                    if h52 is None:
                        h52, l52 = float(out.get('w52_hgpr', 0) or 0), float(out.get('w52_lwpr', 0) or 0)
                    # [수정] 기준 현재가를 표시 현재가와 일치시킨다(NXT 우선 → KRX 폴백).
                    #  종전엔 stck_prpr(KRX)만 써서 같은 행의 현재가 컬럼(NXT 우선)과 어긋났다.
                    c = krx_close_px or float(out.get('ats_prpr', 0) or 0) or float(out.get('stck_prpr', 0) or 0)
                    if h52 > l52:
                        pos = (c - l52)/(h52 - l52)*100
                        w52_pos_val = pos
                        if pos >= 90: w_color = "[red]"
                        elif pos >= 80: w_color = "[orange3]"
                        elif pos <= 30: w_color = "[blue]"
                        elif pos <= 50: w_color = "[yellow]"
                        else: w_color = "[white]"
                        w52_pos_str = f"{w_color}{pos:.1f}%[/]"
                except Exception: pass
        else:
            if config.FILE_DEBUG_LEVEL == "DEBUG":
                logger.debug(f"[ANALYSIS_DEBUG] {code} Detail: {detail} | StockCtx:{is_us_stock_context} EtfCtx:{is_us_etf_context}")

            if detail:
                if is_us_stock_context: 
                    per_str = detail.get('perx', '[dim]-[/dim]')
                    pbr_str = detail.get('pbrx', '[dim]-[/dim]') if detail.get('pbrx') != '-' else '[dim]-[/dim]'
                if is_us_etf_context:
                    try:
                        shar_val = float(detail.get('shar', 0))
                        shar_str = f"{shar_val/1_000_000:.1f}M" if shar_val >= 1_000_000 else f"{shar_val:,.0f}"
                    except Exception: pass
                try:
                    # [수정] 국내와 동일 규칙 — 차트(최근 365일) 1차, 스캐너 값은 폴백.
                    #  상장 1년 미만(SPCX·SKHY 등)은 차트가 창을 못 채워 폴백을 탄다.
                    h52, l52 = _w52_high_low(chart_df)
                    if h52 is None:
                        h52, l52 = float(detail.get('h52p', 0) or 0), float(detail.get('l52p', 0) or 0)
                    # 기준 현재가도 표시 현재가(KIS/토스 last)를 우선한다.
                    _out = curr_data.get('output', {}) if (curr_data or {}).get('rt_cd') == '0' else {}
                    c = float(_out.get('last', 0) or 0) or float(detail.get('last', 0) or 0)
                    if h52 > l52:
                        pos = (c - l52)/(h52 - l52)*100
                        w52_pos_val = pos
                        if pos >= 90: w_color = "[red]"
                        elif pos >= 80: w_color = "[orange3]"
                        elif pos <= 30: w_color = "[blue]"
                        elif pos <= 50: w_color = "[yellow]"
                        else: w_color = "[white]"
                        w52_pos_str = f"{w_color}{pos:.1f}%[/]"
                except Exception: pass

        if curr_data and curr_data.get('rt_cd') == '0':
            out = curr_data['output']
            if is_overseas:
                curr = float(out.get('last', 0) or 0)
                rate = float(out.get('rate', 0) or 0)
                diff = float(out.get('diff', 0) or 0)
                if rate < 0 and diff > 0: diff = -diff
                curr_fmt = f"${curr:,.2f}"
                diff_str = f"{diff:+.2f}"
            else:
                nxt_curr = int(out.get('ats_prpr', 0) or 0)
                krx_curr = int(out.get('stck_prpr', 0) or 0)
                base_price = int(out.get('stck_sdpr', 0) or 0)
                curr = nxt_curr if nxt_curr > 0 else krx_curr

                # [장 종료 후] KRX 확정 종가로 대체 (등락 기준도 일봉 직전 봉으로 함께 교체)
                if krx_close_px > 0:
                    curr = int(round(krx_close_px))
                    base_price = int(round(krx_prev_px)) if krx_prev_px > 0 else base_price

                if base_price > 0:
                    diff = curr - base_price
                    rate = (diff / base_price) * 100
                else:
                    try: rate = float(out.get('prdy_ctrt', 0))
                    except Exception: rate = 0.0
                    try: diff = int(out.get('prdy_vrss', 0))
                    except Exception: diff = 0

                # [장전 폴백] 장전(NXT 개장 08:00 전)엔 무체결로 현재가=기준가 → 등락률 0%가 된다.
                #  이때 일봉 차트로 '직전 정규장 최종 등락률'(전일 vs 전전일)을 산출해 개장 전까지
                #  유지 표시한다. (토스 모드는 어댑터가 stck_sdpr을 이미 보정하므로 제외)
                if diff == 0 and not config.session.is_toss and api._before_nxt_premarket_open():
                    prev_reg = _prelisting_last_regular_change(chart_df, curr)
                    if prev_reg is not None:
                        diff, rate = prev_reg

                curr_fmt = f"{curr:,}"
                diff_str = f"{diff:+}"

            rate_color = "[red]" if rate > 0 else ("[blue]" if rate < 0 else "[white]")
            rate_str = f"{rate_color}{diff_str} ({rate:+.2f}%)[/]{strength_display}"

            # 전일 RSI — calculate_indicators가 계산한 값 재사용 (중복 계산 제거·SSOT)
            prev_rsi_val = ind.get('prev_rsi') if chart_df is not None and not chart_df.empty and len(chart_df) >= 16 else None

            # 적응형 임계값 적용
            thresholds = None
            rule = rules_map.get(code)
            if rule:
                # 개별 룰이 존재하는 경우 개별 룰의 임계값을 최우선 적용
                # [SSOT] NULL 컬럼은 전역 기본값으로, 가중치는 dict로 확정한다.
                from modules.auto_trade.engine import normalize_weights, rule_value
                thresholds = {
                    "BUY_SCORE": rule_value(rule, 'buy_score', config.ANALYSIS_THRESHOLDS["BUY_SCORE"]),
                    "BUY_RSI_MAX": rule_value(rule, 'buy_rsi', config.ANALYSIS_THRESHOLDS["BUY_RSI_MAX"]),
                    "BUY_VOL_STRENGTH": rule_value(rule, 'buy_vol_strength', config.ANALYSIS_THRESHOLDS.get("BUY_VOL_STRENGTH", 100.0)),
                    "WEIGHTS": normalize_weights(rule.get('weights')),
                    "RISE_SCORE": config.ANALYSIS_THRESHOLDS["RISE_SCORE"]
                }
            elif market_regime_adj and not is_overseas:
                # 개별 룰이 없으면 시장 국면에 따른 보정값 적용
                mrkt_name = str(curr_data['output'].get('rprs_mrkt_kor_name') or curr_data['output'].get('rprs_mrkt_eng_name') or '')
                score_adj = 0.0
                if "코스닥" in mrkt_name or "KOSDAQ" in mrkt_name.upper():
                    score_adj = market_regime_adj.get("KOSDAQ", 0.0)
                else:
                    score_adj = market_regime_adj.get("KOSPI", 0.0)
                
                if score_adj != 0:
                    thresholds = {
                        "BUY_SCORE": config.ANALYSIS_THRESHOLDS["BUY_SCORE"] + score_adj,
                        "BUY_RSI_MAX": config.ANALYSIS_THRESHOLDS["BUY_RSI_MAX"],
                        "RISE_SCORE": config.ANALYSIS_THRESHOLDS["RISE_SCORE"]
                    }
                    
            # 52주 위치는 위 표시 문자열 계산에서 w52_pos_val로 함께 산출됨 (슈퍼 모멘텀 마킹 공용)
            sm_flag, sm_reason = check_smart_money_turnaround(code, is_overseas)
            class_name, class_color, _ = classify_stock_state(df=chart_df, ind=ind, prev_rsi=prev_rsi_val, thresholds=thresholds, w52_pos=w52_pos_val, smart_money=sm_flag)

            # [추가] 분류 옆에 추세품질 — 같은 '매수'라도 검증된 추세인지가 갈린다.
            #  진입 순위에서 점수 동점을 가르는 값이라(후보 점수는 절반 넘게 동점이다)
            #  분류만 봐서는 왜 그 종목이 먼저 잡혔는지 알 수 없었다.
            #  색은 indicators.TREND_QUALITY_COLORS 단일 소스 — 도움말 표와 같은 밴드색이다.
            #  데이터가 90봉에 못 미치면 None → '(-)'. 이때는 밴드색이 아니라 dim으로 죽인다 —
            #  값이 없다는 뜻이므로 같은 행의 분류 '-'와 같은 취급이어야 하고, 개별 분석 표의
            #  추세품질 행도 이미 그렇게 찍는다(표기 규약 단일화).
            tq_val = indicators.get_trend_quality(chart_df)
            tq_color = "dim" if tq_val is None else indicators.TREND_QUALITY_COLORS.get(
                indicators.describe_trend_quality(tq_val), "white")
            #  [정렬] 한 컬럼에 두 값을 담으면 분류 문구 길이(매수 4폭 / 강매수 6폭)에 따라
            #   TQ 위치가 들쭉날쭉해 분류도 TQ도 세로로 안 읽힌다. 컬럼을 쪼개는 대신
            #   **분류는 좌측·TQ는 우측**으로 각각 고정폭을 채워 한 셀 안에서 두 열을 만든다.
            #   폭은 표시폭(cell_len) 기준이다 — 한글은 2폭이라 len()으로 세면 어긋난다.
            tq_txt = f"({'-' if tq_val is None else f'{tq_val:.0f}'})"
            tq_cell = f"{_class_tq_cell(class_name)}[{tq_color}]{tq_txt}[/]"

            # [추가] 수동 조회 결과도 상태 캐시에 남긴다 — 텔레그램 /stocks가 조회 시각과 함께
            #  보여준다. 시스템 트레이딩이 분석하지 않는 종목(ETF, NXT 시간대 비거래 종목)과
            #  시스템 정지 중의 공백을 메운다. 세션이 넘어가면 자동 만료된다.
            context.set_stock_state(code, class_name, src='manual')

            def fmt_idx(val): return f"{int(val):,}" if val is not None else "[dim]-[/dim]"

            # [통일] 지수 화면과 동일 규칙 — price_trend_color 단일 소스
            curr_price_color = price_trend_color(curr, ind.get('ema_20'), ind.get('ema_60'), ind=ind)
            curr_str = f"{curr_price_color}{curr_fmt}[/]"

            # [수정] 이평선 색상 규칙 단순화 (계층적 분석)
            ema5_color = "[white]"
            if ind.get('ema_5') is not None and ind.get('ema_20') is not None:
                ema5_color = "[red]" if ind['ema_5'] > ind['ema_20'] else "[blue]"

            ema20_color = "[white]"
            if ind.get('ema_20') is not None and ind.get('ema_60') is not None:
                ema20_color = "[red]" if ind['ema_20'] > ind['ema_60'] else "[blue]"

            ema60_color = "[white]"
            if ind.get('ema_60') is not None and ind.get('ema_120') is not None:
                ema60_color = "[red]" if ind['ema_60'] > ind['ema_120'] else "[blue]"

            ema120_color = "[white]"
            if chart_df is not None and not chart_df.empty and len(chart_df) > 121:
                try:
                    ema120_series = chart_df['close'].ewm(span=120, adjust=False).mean()
                    if ema120_series.iloc[-1] > ema120_series.iloc[-2]:
                        ema120_color = "[red]"
                    else:
                        ema120_color = "[blue]"
                except Exception: pass

            ema_5_str = f"{ema5_color}{fmt_idx(ind.get('ema_5'))}[/]"
            ema_20_str = f"{ema20_color}{fmt_idx(ind.get('ema_20'))}[/]"
            ema_60_str = f"{ema60_color}{fmt_idx(ind.get('ema_60'))}[/]"
            ema_120_str = f"{ema120_color}{fmt_idx(ind.get('ema_120'))}[/]"

            # SAR 상태
            sar_val = ind.get('psar')
            if sar_val is not None and not math.isnan(sar_val):
                sar_icon = "[red]▲[/]" if curr > sar_val else "[blue]▼[/]"
            else:
                sar_icon = "[dim]-[/dim]"
            
            # MACD 상태
            macd_val = ind.get('macd')
            sig_val = ind.get('macd_signal')
            macd_icon = "[dim]-[/dim]"
            if macd_val is not None and sig_val is not None and not math.isnan(macd_val) and not math.isnan(sig_val):
                zero_sign = "+" if macd_val > 0 else "-"
                cross_char = "G" if macd_val > sig_val else "D"
                m_color = "red" if macd_val > sig_val else "blue"
                macd_icon = f"[{m_color}]{zero_sign}{cross_char}[/]"

            # OBV 상태 및 Value
            obv_trend = ind.get('obv_trend')
            obv_val = ind.get('obv')
            vol_sum = chart_df['volume'].tail(5).sum() if chart_df is not None and 'volume' in chart_df.columns else 0
            
            if chart_df is None or len(chart_df) < config.INDICATOR_PARAMS.get("OBV_MA_PERIOD", 5):
                obv_trend = None
                obv_val = None
                
            if vol_sum == 0 or obv_trend is None or obv_val is None or math.isnan(obv_val):
                obv_icon = "[dim]-[/dim]"
                obv_disp = "[dim]-[/dim]"
            else:
                obv_icon = "[red]▲[/]" if obv_trend else "[blue]▼[/]"
                obv_c = "red" if obv_trend else "blue"
                abs_val = abs(obv_val)
                if abs_val >= 999_950_000_000: obv_str = f"{obv_val/1_000_000_000_000:,.1f}T"
                elif abs_val >= 999_950_000: obv_str = f"{obv_val/1_000_000_000:,.1f}B"
                elif abs_val >= 999_500: obv_str = f"{obv_val/1_000_000:,.1f}M"
                elif abs_val >= 999.5: obv_str = f"{obv_val/1_000:,.0f}K"
                else: obv_str = f"{obv_val:,.0f}"
                obv_disp = f"[{obv_c}]{obv_str}[/]"

            trend_str = f"{sar_icon} {macd_icon} {obv_icon}"

            rsi_str = f"{ind.get('rsi'):.1f}" if ind.get('rsi') is not None else "[dim]-[/dim]"
            if ind.get('rsi') is not None:
                if ind.get('rsi') >= config.INDICATOR_PARAMS["RSI_UPPER"]: rsi_str = f"[magenta]{rsi_str}[/]"
                elif 55 <= ind.get('rsi') < config.INDICATOR_PARAMS["RSI_UPPER"]: rsi_str = f"[red]{rsi_str}[/]"
                elif 45 <= ind.get('rsi') < 55: rsi_str = f"[orange3]{rsi_str}[/]"
                elif config.INDICATOR_PARAMS["RSI_LOWER"] < ind.get('rsi') < 45: rsi_str = f"[yellow]{rsi_str}[/]"
                else: rsi_str = f"[blue]{rsi_str}[/]"

            # ADX 값 뒤에 DMI 우위 방향(▲/▼/●)을 함께 표기
            adx_str = format_adx_cell(ind.get('adx'), ind.get('plus_di'), ind.get('minus_di'))

            cci_str = f"{ind.get('cci'):.1f}" if ind.get('cci') is not None else "[dim]-[/dim]"
            if ind.get('cci') is not None:
                if ind.get('cci') >= config.INDICATOR_PARAMS["CCI_UPPER"]: cci_str = f"[red]{cci_str}[/]"
                elif 0 < ind.get('cci') < config.INDICATOR_PARAMS["CCI_UPPER"]: cci_str = f"[orange3]{cci_str}[/]"
                elif config.INDICATOR_PARAMS["CCI_LOWER"] < ind.get('cci') <= 0: cci_str = f"[yellow]{cci_str}[/]"
                else: cci_str = f"[blue]{cci_str}[/]"

            final_name_str = name
            # [변경] 종목명 색상도 지수와 동일한 국면 룰(이중 EMA 교차 + 추종 확인)로 통일한다.
            #  기존 이평선 배열+ADX/RSI/CCI 복합 조건은 지수 색상과 의미가 달라 화면상
            #  같은 색이 다른 뜻이 되던 문제가 있었다. 판정은 classify_regime_from_df 단일 소스.
            try:
                _regime = classify_regime_from_df(chart_df)['regime']
                _, _regime_color = REGIME_DISPLAY.get(_regime, ("", "yellow"))
                final_name_str = f"[{_regime_color}]{name}[/]"
            except Exception: pass
            
            # 제한 종목 표시
            is_restricted = False
            marks = []
            if code in restricted_stocks:
                marks.append("-")
                is_restricted = True

            # 개별 룰 적용 종목 표시
            is_custom_rule = False
            if code in rules_map:
                marks.append("+")
                is_custom_rule = True
                
            is_memo = False
            if code in m_codes:
                marks.append("=")
                is_memo = True
            is_reserved = False
            if code in reserved_codes:
                marks.append("[magenta]*[/magenta]")
                is_reserved = True
            if marks:
                final_name_str += f"[dim]{''.join(marks)}[/dim]"

            row_data = [final_name_str, f"{code}", f"{class_color}{class_name}[/]{tq_cell}", curr_str, rate_str, w52_pos_str, ema_5_str, ema_20_str, ema_60_str, ema_120_str, trend_str, adx_str, rsi_str, cci_str]
            if not is_overseas:
                if use_investor_data: row_data.append(inv_str)
                else: row_data.append(obv_disp)
            else:
                if is_us_stock_context: row_data.extend([per_str, pbr_str])
                elif is_us_etf_context: row_data.append(shar_str)
            return row_data, is_restricted, is_custom_rule, is_memo, is_reserved, False
        else:
            # [Fix] 패딩 컬럼 수는 정의된 컬럼 총수와 일치해야 한다 — 국내 11(총 15칸)·
            #  미국주식 12(16칸)·미국ETF 11(15칸)·해외 무접미(trading.py title="") 10(14칸).
            #  기존 국내 14는 컬럼(15)보다 3칸 많아 rich가 빈 유령 컬럼을 추가,
            #  실패 종목이 하나라도 있으면 테이블 전체 레이아웃이 밀리던 문제.
            _pad = 11 if not is_overseas else (12 if is_us_stock_context else (11 if is_us_etf_context else 10))
            return [name, code, "[dim]-[/dim]", "실패", *["[dim]-[/dim]"] * _pad], False, False, False, False, True
    except Exception as e:
        logger.error(f"[{code}] 분석 오류: {e}")
        _pad = 11 if not is_overseas else (12 if is_us_stock_context else (11 if is_us_etf_context else 10))
        return [name, code, "[red]Error[/]", "[dim]-[/dim]", *["[dim]-[/dim]"] * _pad], False, False, False, False, True

def _realtime_and_analyze(item, title, is_overseas, use_investor_data, restricted_stocks, rules_map, market_regime_adj, reserved_codes, m_codes, chart_df, preloaded_curr=None):
    """(내부함수) print_table 2단계: 당일 실시간 데이터 수신 + 지표 분석.

    1단계에서 받은 과거 차트(chart_df)를 받아 현재가/체결강도/수급을 수신하고 지표를 계산한다.
    preloaded_curr가 있으면(멀티시세 프리페치) 종목별 현재가 REST를 생략한다.
    """
    bundle = _collect_table_data(item, title, is_overseas, use_investor_data, chart_df=chart_df, preloaded_curr=preloaded_curr)
    return _analyze_table_row(item, title, is_overseas, use_investor_data, restricted_stocks,
                              rules_map, market_regime_adj, reserved_codes, m_codes, bundle)

def _print_table_worker(item, title, is_overseas, use_investor_data, restricted_stocks, rules_map, market_regime_adj, reserved_codes, m_codes):
    """(호환용) 단일 종목 수집+분석을 한 번에 수행한다. (수집/분석 단계 분리 이전 호출부·테스트 호환)"""
    bundle = _collect_table_data(item, title, is_overseas, use_investor_data)
    return _analyze_table_row(item, title, is_overseas, use_investor_data, restricted_stocks,
                              rules_map, market_regime_adj, reserved_codes, m_codes, bundle)

_INV_PROBE_CACHE = {}      # 그룹 첫 종목 코드 -> (판정, 시각) — 반복(@) 조회마다 프로브 API 재호출 방지
_INV_PROBE_TTL_SEC = 300

def _probe_investor_data(data_list):
    """국내 그룹의 수급(개/외/기) 컬럼 사용 여부를 판정한다 (최대 3종목 샘플, 5분 캐시).

    기존에는 첫 종목 1건으로만 판정해 첫 종목이 거래정지·신규상장 등으로 수급이 0이면
    그룹 전체가 OBV로 잘못 폴백되던 문제를 샘플 확대로 보완한다.
    """
    key = data_list[0][1]
    now = time.time()
    ent = _INV_PROBE_CACHE.get(key)
    if ent and now - ent[1] < _INV_PROBE_TTL_SEC:
        return ent[0]
    use = False
    for _, code in data_list[:3]:
        try:
            test_data = api.get_investor_trend(code)
        except Exception:
            continue
        if test_data:
            sample = test_data[0]
            if any(api.safe_int(sample.get(k)) != 0 for k in ['prsn_ntby_qty', 'frgn_ntby_qty', 'orgn_ntby_qty']):
                use = True
                break
    _INV_PROBE_CACHE[key] = (use, now)
    return use

def _name_map_from(data_list):
    """{종목코드: 종목명} 추출. data_list는 (종목명, 종목코드) 튜플 리스트다.

    호출부에 따라 dict가 섞여 들어와도 죽지 않도록 방어한다 — 이 값은 경고 문구를
    꾸미는 용도일 뿐이라, 여기서 예외가 나면 표 자체가 출력되지 않는다.
    """
    out = {}
    for item in (data_list or []):
        try:
            if isinstance(item, dict):
                code, name = item.get('code'), item.get('name')
            elif isinstance(item, (tuple, list)) and len(item) >= 2:
                name, code = item[0], item[1]
            else:
                continue
            if code:
                out[str(code)] = name
        except Exception:
            continue
    return out


def print_table(title, data_list, is_overseas=False, market_regime_adj=None, is_etf=None):
    use_investor_data = False
    if not is_overseas and data_list:
        with Progress(
            SpinnerColumn(),
            TextColumn("[progress.description]{task.description}"),
            BarColumn(),
            console=config.console,
            transient=True
        ) as progress:
            progress.add_task("[cyan]수급 데이터 확인 중 (KIS API)...[/cyan]", total=None)
            use_investor_data = _probe_investor_data(data_list)

    # [이동] 적응형 임계값 준비 (테이블 생성 전으로 이동)
    use_adaptive = False
    if not is_overseas and config.MARKET_REGIME_PARAMS.get("USE_ADAPTIVE_THRESHOLD", True):
        use_adaptive = True
        if market_regime_adj is None:
            market_regime_adj = {}
            try:
                with Progress(
                    SpinnerColumn(),
                    TextColumn("[progress.description]{task.description}"),
                    BarColumn(),
                    console=config.console,
                    transient=True
                ) as progress:
                    progress.add_task("[cyan]시장 국면 분석 중 (KIS API)...[/cyan]", total=None)
                    _, kospi_adj = get_market_regime("KOSPI")
                    _, kosdaq_adj = get_market_regime("KOSDAQ")
                    market_regime_adj["KOSPI"] = kospi_adj
                    market_regime_adj["KOSDAQ"] = kosdaq_adj
            except Exception:
                use_adaptive = False
        elif not market_regime_adj:
            use_adaptive = False
    elif market_regime_adj is None:
        market_regime_adj = {}

    failed_list = []
    # [추가] 제목 옆에 현재 세션 표기 — 같은 표라도 08:30은 NXT 프리마켓 체결가,
    #  10:00은 KRX 정규장가, 22:00은 이미 마감된 KRX 종가라 값만으로는 구분되지 않는다.
    #  국내 ETF 표는 NXT 비거래라 NXT 시간대에도 값이 KRX 종가에서 멈추므로 따로 알린다.
    #  (title이 빈 문자열인 호출부[개별 주문 화면 등]는 제목 자체를 숨기므로 붙이지 않는다)
    kr_etf = (not is_overseas) and (("ETF" in title) if is_etf is None else bool(is_etf))
    display_title = f"\n{title}{api.market_session_tag(is_overseas, kr_etf) if title else ''}"
    table = Table(title=display_title, box=box.HORIZONTALS, show_header=True, header_style="dim", border_style="dim")
    table.add_column("종목명", justify="left", style="white", no_wrap=True)
    table.add_column("코드", justify="center", style="dim")
    # [표기] '분류' 셀은 분류 문구와 추세품질 값을 함께 담는다(예: 매수 (143)) — 헤더도
    #  그 사실을 말해야 한다. '추세품질'을 다 쓰면 폭 15로 이 컬럼 최대 셀(13)보다 넓어져
    #  표가 통째로 밀리므로 폭 9인 '(TQ)'로 줄인다. 밴드·산식은 도움말 표에 있다.
    table.add_column("분류" + _class_tq_cell("분류") + "(TQ)", justify="left")
    table.add_column("현재가", justify="right")
    col_header = "등락폭 (등락률)"
    if not is_overseas and not use_investor_data:
        if config.session.is_toss:
            # [수정] 토스 매도비는 NXT 운영시간(08:00~20:00)에만 표시 — 시간창 밖에는
            #  셀 표기와 함께 헤더의 컬럼 표기도 생략 (KIS 체결강도 표시 창과 동일 동작)
            if api.is_toss_ask_bid_window():
                col_header += " [매도비]"
        else:
            col_header += " [강도]"
    table.add_column(col_header, justify="right")
    # [표기] "52주"는 52주 고점(가격)으로 오해되기 쉬워 "52W%"로 바꾼다 — 값은 저점~고점
    #  밴드 내 위치(%)다. 표시폭 4로 기존 헤더와 같아 컬럼/전체 폭은 변하지 않는다.
    table.add_column("52W%", justify="right")
    table.add_column("EMA(5)", justify="right")
    table.add_column("EMA(20)", justify="right")
    table.add_column("EMA(60)", justify="right")
    table.add_column("EMA(120)", justify="right")
    table.add_column("추세SMO", justify="center")
    table.add_column("ADX", justify="right")
    table.add_column("RSI", justify="right")
    table.add_column("CCI", justify="right")

    is_us_stock = is_overseas and ("주식" in title)
    is_us_etf = is_overseas and ("ETF" in title)
    
    # [추가] 개별 룰 로드
    custom_rules = db_manager.db.get_all_stock_strategies()
    rules_map = {}
    for r in custom_rules:
        r_dict = dict(r)
        if r_dict.get('weights') and isinstance(r_dict['weights'], str):
            try: r_dict['weights'] = json.loads(r_dict['weights'])
            except Exception: r_dict['weights'] = None
        rules_map[r_dict['code']] = r_dict
    any_custom_rule = False
    
    # [추가] 트레이딩 제한 종목 로드
    from modules import auto_trade
    restricted_stocks = auto_trade.get_restricted_stocks()
    any_restricted = False
    
    # [추가] 예약 매매 및 메모 마커 조회
    reserved_codes = set()
    try:
        pending_reserves = db_manager.db.get_pending_reserved_orders()
        reserved_codes = set(o['code'] for o in pending_reserves)
    except Exception: pass
    m_codes = utils.get_memo_codes()
    
    if not is_overseas:
        if use_investor_data: table.add_column("수급(개/외/기)", justify="center")
        else: table.add_column("OBV", justify="right")
    else:
        if is_us_stock:
            table.add_column("PER", justify="right", style="dim")
            table.add_column("PBR", justify="right", style="dim") 
        elif is_us_etf:
            table.add_column("상장주수", justify="right", style="dim")

    # [최적화] 통신+연산 통합 처리 스레드 수 — 모의서버는 TPS 여유가 작아 4, 실전·해외는 5
    #  (야후/KIS 동시 호출 차단 방지. 해외·국내 분기가 동일 값이라 단일화)
    max_w = 5
    # [수정] 메뉴1처럼 진행 상태를 '데이터 수집'과 '지표 분석' 2단계로 분리하여 운영자 인지성을 높인다.
    # [Fix] 패딩 수는 실제 정의된 컬럼 총수에서 산출 — 기존 고정값(국내 14)이 컬럼(15)보다
    #  3칸 많아 rich가 빈 유령 컬럼을 추가해 테이블 레이아웃이 밀리던 문제
    _fail_cols = len(table.columns) - 4
    def _fail_row(idx):
        name, code = data_list[idx]
        return ([name, code, "[dim]-[/dim]", "실패", *["[dim]-[/dim]"] * _fail_cols], False, False, False, False, True)

    try:
        used_marks = set()
        charts = [None] * len(data_list)
        results = [None] * len(data_list)

        # [최적화] 해외 그룹: TV 일괄 예열(HTTP 1회)을 1단계(차트 수신)와 병렬로 백그라운드 수행.
        #  2단계 워커의 장외가 병합(get_yf_fast_info, ttl=30)이 대부분 캐시 적중으로 처리된다.
        #  프로그래스 바가 0%에 머무는 동기 대기를 없애기 위해 join하지 않는다: 예열이 채 끝나기
        #  전에 기동한 초기 워커(최대 max_w개)만 TV 단건 조회로 자체 폴백한다(출력 동일).
        #  백그라운드 워머(OverviewWarmer, 실전 15초 주기) 캐시가 신선하면 예열 자체를 생략한다.
        if is_overseas and data_list:
            def _warm_ovs_prices():
                try:
                    warm_fresh_sec = max(5, int(getattr(config, 'OVERVIEW_WARM_INTERVAL_SEC', 15))) + 5
                    api.prefetch_multiple_current_prices(
                        [c for _, c in data_list], is_overseas=True, skip_if_fresh_sec=warm_fresh_sec
                    )
                except Exception as e:
                    logger.debug(f"[print_table] 해외 일괄 예열 실패: {e}")
            threading.Thread(target=_warm_ovs_prices, name="OvsTablePrewarm", daemon=True).start()

        # 1단계: 데이터 수신 (과거 전체 일봉). 캐시 적중 시 즉시 통과, 캐시 미스 때만 실제 다운로드.
        with Progress(
            SpinnerColumn(),
            TextColumn("[progress.description]{task.description}"),
            BarColumn(),
            TextColumn("[progress.percentage]{task.percentage:>3.0f}%"),
            TimeRemainingColumn(),
            console=config.console,
            transient=True
        ) as progress:
            task_d = progress.add_task(f"[cyan]{title} (데이터 수신)[/cyan]", total=len(data_list))
            _stage1_started = time.monotonic()
            _slow_before = _SLOW_FETCH_COUNT[0]
            with concurrent.futures.ThreadPoolExecutor(max_workers=max_w) as executor:
                fut_map = {
                    executor.submit(_fetch_chart_data, item, is_overseas): i
                    for i, item in enumerate(data_list)
                }
                for future in concurrent.futures.as_completed(fut_map):
                    idx = fut_map[future]
                    try:
                        charts[idx] = future.result()
                    except Exception as e:
                        logger.error(f"Chart fetch error: {e}")
                        charts[idx] = None
                    progress.advance(task_d)
        # [진단] 느린 종목이 있었던 표만 단계 총계를 남긴다. 개별 소요의 '합'에 가까우면
        #  직렬, '최댓값'에 가까우면 병렬이다 — 워커를 늘려 효과가 있을지가 여기서 갈린다.
        _slow_n = _SLOW_FETCH_COUNT[0] - _slow_before
        if _slow_n:
            logger.debug(f"[일봉 수신 완료] {title} {len(data_list)}종목 "
                         f"{time.monotonic() - _stage1_started:.1f}초 "
                         f"(느린 종목 {_slow_n}건 / 워커 {max_w})")

        # [멀티시세] 국내 그룹 현재가를 30종목/1콜로 프리페치 (종목당 현재가 REST 제거 → TPS 절감)
        #  정규장(phase 'skip')은 KRX 대표가만, 장전/장후 NXT 시간(phase 'active')은 KRX+NXT를 각각
        #  배치(30종목/1콜)로 병합한다. 장후 'active'의 NX 병합은 종목별 fetch_nxt_price 팬아웃이
        #  EGW00201(초당한도)을 유발해 현재가가 전일종가로 stale 폴백되던 문제를 콜 배치화로 해소한다.
        #  야간·주말(offhours)은 _nxt_recalled_close 폴백을 위해 종목별 조회 유지. 실패 시 None → 종목별 폴백.
        multi_prices = None
        if not is_overseas and data_list and not config.session.is_toss and getattr(config, 'USE_MULTI_PRICE', True):
            _codes = [c for _, c in data_list]
            try:
                _phase = api._nxt_quote_phase()
                if _phase == 'skip':        # 정규장: KRX 대표가
                    multi_prices = api.get_multi_current_prices(_codes)
                elif _phase == 'active':     # 장전/장후 NXT 시간: KRX+NXT 배치 병합
                    multi_prices = api.get_multi_current_prices_nxt(_codes)
                    # offhours(야간·주말): 종목별 조회 유지
            except Exception:
                multi_prices = None

        def _preloaded(code):
            if multi_prices and code in multi_prices:
                return {'rt_cd': '0', 'output': multi_prices[code]}
            return None

        # 2단계: 실시간 데이터 수신 및 분석 (당일 현재가/체결강도/수급 수신 + 지표 연산)
        with Progress(
            SpinnerColumn(),
            TextColumn("[progress.description]{task.description}"),
            BarColumn(),
            TextColumn("[progress.percentage]{task.percentage:>3.0f}%"),
            TimeRemainingColumn(),
            console=config.console,
            transient=True
        ) as progress:
            task_a = progress.add_task(f"[cyan]{title} (실시간 데이터 수신 및 분석)[/cyan]", total=len(data_list))
            with concurrent.futures.ThreadPoolExecutor(max_workers=max_w) as executor:
                fut_map = {
                    executor.submit(
                        _realtime_and_analyze, item, title, is_overseas, use_investor_data,
                        restricted_stocks, rules_map, market_regime_adj, reserved_codes, m_codes, charts[i],
                        _preloaded(item[1])
                    ): i
                    for i, item in enumerate(data_list)
                }
                for future in concurrent.futures.as_completed(fut_map):
                    idx = fut_map[future]
                    try:
                        results[idx] = future.result()
                    except Exception as e:
                        logger.error(f"Analyze worker error: {e}")
                        results[idx] = _fail_row(idx)
                    progress.advance(task_a)

        # 결과 테이블 추가
        for idx, result_item in enumerate(results):
            if not result_item:
                failed_list.append(data_list[idx])
                continue

            row_data, is_res, is_cust, is_mem, is_rsv, is_failed = result_item

            if is_res: used_marks.add('-')
            if is_cust: used_marks.add('+')
            if is_mem: used_marks.add('=')
            if is_rsv: used_marks.add('*')

            # [수정] 실패 여부는 셀 문자열 매칭 대신 워커가 반환한 플래그로 판정
            if is_failed:
                failed_list.append(data_list[idx])

            table.add_row(*row_data)
            if table.row_count % 5 == 0 and table.row_count < len(data_list):
                table.add_section()

    except Exception as e:
        logger.error(f"데이터 분석 중 오류: {e}")

    try:
        # [경고] KRX 공식 일봉 실패로 토스 캔들(NXT 포함)이 쓰인 종목이 있으면 표 앞에 알린다.
        #  data_list는 (종목명, 종목코드) 튜플 리스트다. 부가 정보이므로 여기서 실패해도
        #  표 출력을 막지 않는다(아래 except가 '테이블 출력 실패'로 잡아먹지 않도록 분리).
        if not is_overseas:
            try:
                utils.print_krx_fallback_warning(_name_map_from(data_list))
            except Exception as e:
                logger.debug(f"KRX 폴백 경고 출력 실패: {e}")

        config.console.print(table, crop=False)

        mark_desc = []
        if '-' in used_marks: mark_desc.append("[dim]([/dim] - [dim]) 시스템 트레이딩 제한 종목[/dim]")
        if '+' in used_marks: mark_desc.append("[dim]([/dim] + [dim]) 개별 룰 적용 종목[/dim]")
        if '=' in used_marks: mark_desc.append("[dim]([/dim] = [dim]) 메모 설정 종목[/dim]")
        if '*' in used_marks: mark_desc.append("[dim]([/dim] * [dim]) 예약 매매 설정 종목[/dim]")
        if mark_desc:
            config.console.print(f" [dim]※[/dim] {' [dim]|[/dim] '.join(mark_desc)}")

        sys.stdout.flush()
    except Exception as e:
        logger.error(f"테이블 출력 중 오류(tmux 리사이즈 등): {e}")
        config.console.print(f"[red]테이블 출력 실패: {e}[/red]")

    return failed_list

def show_stock_analysis():
    base_breadcrumb_len = len(context.USER_ACTION_BREADCRUMB)
    last_choice = "5"
    while True:
        context.USER_ACTION_BREADCRUMB = context.USER_ACTION_BREADCRUMB[:base_breadcrumb_len]
        utils.clear_screen()
        menu_items = [
            ("1", "국내 주식", "Domestic Stock"), ("2", "국내 ETF", "Domestic ETF"),
            ("3", "미국 주식", "US Stock"), ("4", "미국 ETF", "US ETF"),
            ("5", "전체 보기", "View All"), ("6", "개별 종목 분석", "Individual Analysis"),
            ("7", "전체 종목 분석", "Market Analysis")
        ]
        choice_str = utils.show_menu("종목 시세 분석 (Stock Analysis)", menu_items, default_choice=last_choice, custom_prompt="번호 입력 [dim](예: 1,3 또는 12 / 반복: 1@ 또는 1@120)[/dim]")
        if choice_str.lower() in ['b', 'q']: return False
        if choice_str.lower() == 'h':
            if getattr(utils, 'show_help', None):
                utils.show_help()
                utils.pause()
            continue
        

        # [수정] 반복 주기 커스텀 지원: '1@'(기본 60초) 또는 '1@120'(120초, 최소 10초)
        interval = 0
        if '@' in choice_str:
            base, _, suffix = choice_str.partition('@')
            if suffix and not suffix.isdigit():
                config.console.print("[red]잘못된 반복 주기입니다. 예: 1@ 또는 1@120[/red]")
                time.sleep(1)
                continue
            interval = max(10, int(suffix)) if suffix else 60
            choice_str = base

        raw_choices = [c.strip() for c in choice_str.split(',') if c.strip()]
        choices = []
        for c in raw_choices:
            if c.isdigit() and len(c) > 1:
                choices.extend(list(c))
            else:
                choices.append(c)

        if not choices: continue

        if '6' in choices:
            last_choice = choice_str # [수정] 정상 처리된 유효한 입력만 기억
            context.USER_ACTION_BREADCRUMB.append("[6] 개별분석")
            if diagnose_stock() is not False: utils.pause()
            continue

        if '7' in choices:
            last_choice = choice_str # [수정] 정상 처리된 유효한 입력만 기억
            context.USER_ACTION_BREADCRUMB.append("[7] 전체분석")
            sub_menu = [("1", "코스피", "KOSPI"), ("2", "코스닥", "KOSDAQ"), ("3", "전체 종목 분석 결과 저장", "Save to Excel")]
            sub_choice = utils.show_menu("분석할 시장을 선택하세요", sub_menu, default_choice="1")
            
            if sub_choice.lower() in ['b', 'q']: continue
            
            if sub_choice == "3":
                sub_map = dict((k, v) for k, v, _ in sub_menu)
                context.USER_ACTION_BREADCRUMB.append(f"[{sub_choice}] {sub_map.get(sub_choice, '')}")
                if save_all_market_analysis() is not False: utils.pause()
                continue

            market_type = "KOSPI" if sub_choice == "1" else "KOSDAQ"
            context.USER_ACTION_BREADCRUMB.append(f"[시장선택] {market_type}")
            
            if analyze_market_stocks(market_type) is not False: utils.pause()
            continue

        selected_groups = set()
        group_names = []
        
        for c in choices:
            if c == '1': 
                selected_groups.add('stocks_kr')
                group_names.append("국내주식")
            elif c == '2': 
                selected_groups.add('etfs_kr')
                group_names.append("국내ETF")
            elif c == '3': 
                selected_groups.add('stocks_us')
                group_names.append("미국주식")
            elif c == '4': 
                selected_groups.add('etfs_us')
                group_names.append("미국ETF")
            elif c == '5': 
                selected_groups.update(['stocks_kr', 'etfs_kr', 'stocks_us', 'etfs_us'])
                group_names.append("전체보기")
        
        if not selected_groups:
            config.console.print("[red]잘못된 입력입니다.[/red]")
            time.sleep(1)
            continue

        last_choice = choice_str # [수정] 정상 처리된 유효한 입력만 기억
        context.USER_ACTION_BREADCRUMB.append(f"[{choice_str}] {','.join(group_names)}")

        # [추가] 최초 클론 등으로 stock.json이 없으면 기본 관심종목(삼성전자)으로 자동 생성
        if not os.path.exists(config.STOCK_DATA_FILE):
            config.session.save_stock_config({
                "stocks_kr": [{"name": "삼성전자", "code": "005930", "exchange": "KOSPI"}],
                "etfs_kr": [], "stocks_us": [], "etfs_us": []
            })
            config.session.load_stock_config()  # exchange 캐시 재구성
            config.console.print("[yellow]관심종목 파일(json/stock.json)이 없어 기본 종목(삼성전자)으로 새로 생성했습니다.[/yellow]\n")

        target_list = []
        order_map = [
            ('stocks_kr', "국내 주식 기술적 분석", False),
            ('etfs_kr', "국내 ETF 기술적 분석", False),
            ('stocks_us', "미국 주식 기술적 분석", True),
            ('etfs_us', "미국 ETF 기술적 분석", True)
        ]

        for key, title, is_ovs in order_map:
            if key in selected_groups:
                d_list = [(x['name'], x['code']) for x in config.session.stock_data.get(key, [])]
                target_list.append((title, d_list, is_ovs))

        # [추가] 조회 대상 종목이 하나도 없을 때 빈 화면 대신 안내 메시지 출력
        if not any(d_list for _, d_list, _ in target_list):
            if not any(config.session.stock_data.get(k) for k in ["stocks_kr", "etfs_kr", "stocks_us", "etfs_us"]):
                config.console.print("[yellow]관심종목에 추가된 종목이 없습니다. [7] 관심 종목 관리 메뉴에서 종목을 추가해주세요.[/yellow]")
            else:
                config.console.print("[yellow]선택한 분류에 등록된 종목이 없습니다.[/yellow]")
            utils.pause()
            continue

        logger.info(f"운영자 실행: {' - '.join(context.USER_ACTION_BREADCRUMB)}")

        try:
            while True:
                if interval > 0:
                    now_str = datetime.now().strftime("%H:%M:%S")
                    config.console.print(f"\n[dim]조회 시간: {now_str}[/dim]")
                
                # [최적화] 조회 주기마다 한 번만 시장 국면 분석 수행 (중복 API 호출 방지)
                shared_regime_adj = None
                has_domestic = any(not is_ovs for _, _, is_ovs in target_list)
                if has_domestic and config.MARKET_REGIME_PARAMS.get("USE_ADAPTIVE_THRESHOLD", True):
                    shared_regime_adj = {}
                    try:
                        with Progress(
                            SpinnerColumn(),
                            TextColumn("[progress.description]{task.description}"),
                            BarColumn(),
                            console=config.console,
                            transient=True
                        ) as progress:
                            progress.add_task("[cyan]시장 국면 분석 중 (KIS API)...[/cyan]", total=None)
                            _, k_adj = get_market_regime("KOSPI")
                            _, q_adj = get_market_regime("KOSDAQ")
                            shared_regime_adj["KOSPI"] = k_adj
                            shared_regime_adj["KOSDAQ"] = q_adj
                    except Exception:
                        pass

                failed_targets = []

                try:
                    for title, d_list, is_ovs in target_list:
                        if d_list: 
                            failed = print_table(title, d_list, is_ovs, market_regime_adj=shared_regime_adj)
                            if failed:
                                failed_targets.append((title, failed, is_ovs))
                except Exception as e:
                    logger.error(f"분석 루프 실행 중 오류: {e}")
                    config.console.print(f"[red]분석 중 오류 발생: {e}[/red]")
                
                if interval <= 0:
                    if failed_targets:
                        total_failed = sum(len(f_list) for _, f_list, _ in failed_targets)
                        if Prompt.ask(f"\n[yellow]⚠️ 조회 실패한 {total_failed}개 종목을 다시 시도하시겠습니까?[/yellow]", choices=["y", "n"], default="y") == "y":
                            config.console.print()
                            target_list = failed_targets
                            continue
                    break 

                config.console.print() 
                try:
                    for remaining in range(interval, -1, -1):
                        config.console.print(f"[bold yellow]다음 조회까지 {remaining}초 대기 중입니다. (중단: Ctrl+C)[/]   ", end="\r")
                        time.sleep(1)
                except KeyboardInterrupt:
                    config.console.print("\n[yellow]반복 조회를 중단하고 메뉴로 돌아갑니다.[/yellow]")
                    break
        except KeyboardInterrupt: config.console.print("\n[yellow]작업이 취소되었습니다.[/yellow]")
        except Exception as e:
            logger.error(f"분석 기능 실행 중 치명적 오류: {e}")
            config.console.print(f"\n[bold red]오류 발생: {e}[/bold red]")
            
        # [수정] 반복 조회 중단(Ctrl+C) 시 메인 메뉴로 이탈하지 않고 종목 시세 분석 메뉴를 유지
        if interval > 0:
            continue
        utils.pause()

def get_snapshot(code, is_overseas):
    """주문 시점의 종목 상태 스냅샷 생성 (DB 저장용)"""
    snapshot = {}
    try:
        # 1. 차트 데이터 및 지표
        df = api.get_chart_data(code, is_overseas)
        if df is not None and not df.empty:
            ind = indicators.calculate_indicators(df)
            # numpy float 등을 일반 float으로 변환하여 저장
            snapshot['indicators'] = {k: (float(v) if v is not None else None) for k, v in ind.items()}
            snapshot['price'] = float(df.iloc[-1]['close'])
        
        # 2. 환율 (해외인 경우)
        if is_overseas:
            snapshot['exchange_rate'] = utils.get_exchange_rate()
            
        snapshot['market'] = "Overseas" if is_overseas else "Domestic"
        
    except Exception as e:
        snapshot['error'] = str(e)
        
    return snapshot

def _print_period_price_common(code, is_overseas, limit=20):
    """기간별 시세 출력 공통 함수"""
    def _fmt_vol(v):
        val = float(v)
        if val == 0: return "[dim]-[/dim]"
        abs_val = abs(val)
        if abs_val >= 999_950_000_000: return f"{val/1_000_000_000_000:,.1f}T"
        elif abs_val >= 999_950_000: return f"{val/1_000_000_000:,.1f}B"
        elif abs_val >= 999_500: return f"{val/1_000_000:,.1f}M"
        elif abs_val >= 999.5: return f"{val/1_000:,.0f}K"
        return f"{val:,.0f}"

    # [수정] 단순 조회이므로 status 사용
    df = None
    investor_map = {} # [추가]
    market_flow_map = {}   # [추가] 지수 전용 — 시장 수급·공매도 (KRX)
    
    is_domestic_index = not is_overseas and code in ["KOSPI", "KOSDAQ", "KOSPI200", "KOSDAQ150", "VKOSPI"]

    from modules import market
    all_idx_codes = [c for n, c in market.ALL_INDICES]
    is_global_index = is_overseas and code in all_idx_codes
    is_index = is_domestic_index or is_global_index

    with Progress(
        SpinnerColumn(),
        TextColumn("[progress.description]{task.description}"),
        BarColumn(),
        console=config.console,
        transient=True
    ) as progress:
        progress.add_task("[cyan]기간별 시세 데이터 조회 중...[/cyan]", total=None)
        if is_domestic_index:
            df = get_domestic_index_data(code)
            # [Fix] 공유 캐시 객체 보호: 아래에서 ma/diff/OBV 등 컬럼을 추가하므로 복사본 사용
            if df is not None:
                df = df.copy()
        else:
            df = api.get_chart_data(code, is_overseas)
            
        # [추가] 수급 데이터 조회
        frgn_rates_map = {} # [추가] 실제 지분율 맵
        if not is_overseas and not is_domestic_index:
            try:
                inv_data = api.get_investor_trend(code)
                if inv_data:
                    for item in inv_data:
                        investor_map[item['stck_bsop_date']] = item
                        
                # 공매도 추이 데이터 조회
                try:
                    short_sale_data = api.get_daily_short_selling(code, limit)
                    if short_sale_data:
                        for item in short_sale_data:
                            d_key = item.get('stck_bsop_date')
                            if not d_key and 'date' in item:
                                d_key = item['date'].replace('-', '')
                            if d_key:
                                if d_key not in investor_map:
                                    investor_map[d_key] = {}
                                
                                if 'ssts_vol_rlim' in item:
                                    investor_map[d_key]['short_sale_rate'] = item['ssts_vol_rlim']
                                elif 'shortSellingVolumeRate' in item:
                                    val = item['shortSellingVolumeRate']
                                    investor_map[d_key]['short_sale_rate'] = f"{float(val) * 100:.2f}" if val else "0.00"
                except Exception as e:
                    pass

                # 외인 소진율 데이터 조회 및 병합 (최근 30일)
                frate_data = api.get_daily_foreign_rate(code)
                if frate_data:
                    for item in frate_data:
                        d_key = item['stck_bsop_date']
                        if d_key in investor_map:
                            investor_map[d_key]['hts_frgn_ehrt'] = item.get('hts_frgn_ehrt')
                        else:
                            investor_map[d_key] = {'hts_frgn_ehrt': item.get('hts_frgn_ehrt')}
                            
                # [추가] 실제 외국인 지분율 역산 로직
                cp_res = api.get_current_price_data(code, is_overseas=False)
                if cp_res.get('rt_cd') == '0':
                    out = cp_res['output']
                    lstn_stcn = api.safe_int(out.get('lstn_stcn'))
                    frgn_hldn_qty = api.safe_int(out.get('frgn_hldn_qty'))
                    
                    if lstn_stcn > 0:
                        sorted_dates = sorted(list(investor_map.keys()), reverse=True)
                        current_hldn = frgn_hldn_qty
                        for d_key in sorted_dates:
                            frgn_rates_map[d_key] = (current_hldn / lstn_stcn) * 100
                            f_net = api.safe_int(investor_map[d_key].get('frgn_ntby_qty'))
                            current_hldn -= f_net
            except Exception: pass

        # [추가] 지수는 종목과 원천이 다르다 — 시장 단위 투자자별 순매수·공매도(KRX).
        #  코스피·코스닥만 있다. 코스피200·코스닥150은 시장이 아니라 지수의 부분집합이라
        #  집계 자체가 존재하지 않는다(krx_data.get_market_flow_daily 참고).
        if is_domestic_index:
            try:
                from modules import krx_data
                flow_df = krx_data.get_market_flow_daily(code, days=(limit or 30) * 2 + 20)
                if flow_df is not None:
                    for _, r in flow_df.iterrows():
                        market_flow_map[str(r['date'])] = r
            except Exception as e:
                logger.debug(f"[Analysis] {code} 시장 수급 조회 실패(표에서 생략): {e}")

    if df is None or df.empty: return

    # 이동평균선 계산
    for w in [5, 20, 60, 120]:
        df[f'ma{w}'] = df['close'].ewm(span=w, adjust=False).mean()

    # 등락폭/등락률 계산 (get_chart_data는 기본 제공하지 않음)
    df['diff'] = df['close'].diff()
    df['rate'] = df['close'].pct_change() * 100

    # OBV 및 OBV_MA 계산
    df['OBV'] = indicators.get_obv_full_series(df)
    df['OBV_MA'] = df['OBV'].ewm(span=config.INDICATOR_PARAMS["OBV_MA_PERIOD"], adjust=False).mean()

    # 최신순 정렬 및 limit 적용
    df_sorted = df.sort_values('date', ascending=False)
    if limit:
        recent_df = df_sorted.head(limit)
    else:
        recent_df = df_sorted

    title_prefix = "[해외주식]" if is_overseas else "[국내주식]"
    if is_index:
        idx_name = code
        if is_domestic_index:
            d_map = {"KOSPI": "코스피", "KOSDAQ": "코스닥", "KOSPI200": "코스피200", "KOSDAQ150": "코스닥150", "VKOSPI": "V코스피200"}
            idx_name = d_map.get(code, code)
        else:
            idx_name = next((n for n, c in market.ALL_INDICES if c == code), code)
        title_prefix = f"[{idx_name}]"
        
    period_str = f"(최근 {limit}일)" if limit else "(전체)"
    table = Table(title=f"{title_prefix} 기간별 시세 {period_str}", box=box.HORIZONTALS, show_header=True, header_style="dim", border_style="dim")
    table.add_column("일자", justify="center")
    table.add_column("종가", justify="right")
    table.add_column("등락폭 (등락률)", justify="right")
    table.add_column("시가", justify="right")
    table.add_column("고가", justify="right")
    table.add_column("저가", justify="right")
    table.add_column("5일선", justify="right")
    table.add_column("20일선", justify="right")
    table.add_column("60일선", justify="right")
    table.add_column("120일선", justify="right")
    table.add_column("OBV", justify="right") # [이동]
    if not is_overseas and not is_domestic_index:
        table.add_column("외인률", justify="right") # [추가]
        table.add_column("공매도", justify="right") # [추가]
        table.add_column("수급(개/외/기)", justify="center") # [수정]
    elif market_flow_map:
        # [추가] 지수는 외인률이 없다 — 상장주식수가 없어 '외국인 보유비율'이 정의되지 않는다.
        #  **단위는 종목 표와 다르다**: 공매도=거래대금 비중, 수급=순매수 대금(원).
        #  지수엔 주식 수 개념이 없어 주식 수 기준 집계가 애초에 존재하지 않는다.
        #  (헤더에 단위를 적었다가 걷어냈다 — 표가 시끄러워진다. 값의 자릿수가 B·T 라
        #   금액임이 드러나고, 종목 표는 K·M 이라 눈으로도 갈린다.)
        table.add_column("공매도", justify="right")
        table.add_column("수급(개/외/기)", justify="center")

    for i, (idx, row) in enumerate(recent_df.iterrows()):
        date_str = str(row['date'])
        if len(date_str) == 8: date_str = f"{date_str[:4]}/{date_str[4:6]}/{date_str[6:]}"
        
        close = row['close']
        diff = row['diff'] if not pd.isna(row['diff']) else 0
        rate = row['rate'] if not pd.isna(row['rate']) else 0
        
        def fmt_p(val): 
            if is_index:
                return f"{val:,.0f}" if val >= 1000 else f"{val:,.2f}"
            return f"{val:,.2f}" if is_overseas else f"{int(val):,}"
            
        def fmt_diff(val): 
            if is_index:
                return f"{val:+.0f}" if abs(val) >= 1000 else f"{val:+.2f}"
            return f"{val:+.2f}" if is_overseas else f"{int(val):+}"
        
        c_color = "[red]" if diff > 0 else ("[blue]" if diff < 0 else "[white]")
        diff_str = f"{c_color}{fmt_diff(diff)} ({rate:+.2f}%)[/]"
        
        # [수정] 이동평균선 색상 규칙 변경 (이평선 간 배열 기준)
        ma5_val, ma20_val = row['ma5'], row['ma20']
        ma60_val, ma120_val = row['ma60'], row['ma120']

        def get_ma_color(val, ma_type):
            if pd.isna(val): return "dim"
            
            if ma_type == 5:
                if pd.isna(ma20_val): return "dim"
                return "red" if val > ma20_val else "blue"
            elif ma_type == 20:
                if pd.isna(ma60_val): return "white"
                return "red" if val > ma60_val else "blue"
            elif ma_type == 60:
                if pd.isna(ma120_val): return "white"
                return "red" if val > ma120_val else "blue"
            elif ma_type == 120:
                if i + 1 < len(recent_df):
                    prev_ma120 = recent_df.iloc[i+1]['ma120']
                    if not pd.isna(prev_ma120):
                        return "red" if val > prev_ma120 else "blue"
                return "white"
            return "dim"

        def fmt_ma(val, color):
            if pd.isna(val): return "[dim]-[/dim]"
            return f"[{color}]{fmt_p(val)}[/]"

        # OBV 포맷팅
        obv_val = row['OBV']
        obv_ma_val = row['OBV_MA']
        if pd.isna(obv_val):
            obv_disp = "[dim]-[/dim]"
        else:
            obv_trend = obv_val > obv_ma_val if not pd.isna(obv_ma_val) else None
            if obv_trend is None:
                obv_c = "white"
            else:
                obv_c = "red" if obv_trend else "blue"
            
            abs_val = abs(obv_val)
            if abs_val >= 999_950_000_000: obv_str = f"{obv_val/1_000_000_000_000:,.1f}T"
            elif abs_val >= 999_950_000: obv_str = f"{obv_val/1_000_000_000:,.1f}B"
            elif abs_val >= 999_500: obv_str = f"{obv_val/1_000_000:,.1f}M"
            elif abs_val >= 999.5: obv_str = f"{obv_val/1_000:,.0f}K"
            else: obv_str = f"{obv_val:,.0f}"
            obv_disp = f"[{obv_c}]{obv_str}[/]"

        # [추가] 수급 데이터 포맷팅
        inv_str = "[dim]-[/dim]"
        foreign_rate_str = "[dim]-[/dim]"
        short_sale_str = "[dim]-[/dim]"
        if not is_overseas and not is_domestic_index:
            d_key = str(row['date']).replace('-', '')[:8]
            if d_key in investor_map:
                item = investor_map[d_key]
                
                if 'short_sale_rate' in item:
                    val = item['short_sale_rate']
                    try: short_sale_str = f"{float(val):.2f}%"
                    except Exception: pass

                p = api.safe_int(item.get('prsn_ntby_qty'))
                f = api.safe_int(item.get('frgn_ntby_qty'))
                o = api.safe_int(item.get('orgn_ntby_qty'))
                
                # [수정] 역산된 실제 외국인 지분율 우선 적용
                if d_key in frgn_rates_map:
                    foreign_rate_str = f"{frgn_rates_map[d_key]:.2f}%"
                else:
                    f_rate = item.get('hts_frgn_ehrt')
                    if f_rate is not None and str(f_rate).strip():
                        try: foreign_rate_str = f"{float(f_rate):.2f}%"
                        except Exception: pass

                def _fmt_i(val):
                    if val == 0: return "[dim]-[/dim]"
                    abs_val = abs(val)
                    # elif 여야 한다 — if 를 두 번 쓰면 10억 이상이 곧바로 M 분기에
                    #  덮여 B 표기가 **한 번도 나오지 않는다**(억 단위 수급이 'M' 으로 뜬다).
                    if abs_val >= 1_000_000_000: s = f"{val/1_000_000_000:,.1f}B"
                    elif abs_val >= 1_000_000: s = f"{val/1_000_000:,.1f}M"
                    elif abs_val >= 1000: s = f"{val/1000:,.0f}K"
                    else: s = f"{val:,}"
                    return f"[red]{s}[/]" if val > 0 else f"[blue]{s}[/]"
                
                inv_str = f"{_fmt_i(p)} {_fmt_i(f)} {_fmt_i(o)}"

        # [추가] 지수 — 시장 수급·공매도 (금액 단위)
        flow_inv_str = "[dim]-[/dim]"
        flow_short_str = "[dim]-[/dim]"
        if market_flow_map:
            fr = market_flow_map.get(str(row['date']).replace('-', '')[:8])
            if fr is not None:
                ratio = fr.get('short_ratio')
                if ratio is not None and not pd.isna(ratio):
                    flow_short_str = f"{float(ratio):.2f}%"

                def _fmt_money(val):
                    if val is None or pd.isna(val) or float(val) == 0:
                        return "[dim]-[/dim]"
                    val = float(val)
                    s_abs = abs(val)
                    if s_abs >= 999_950_000_000: txt = f"{val/1_000_000_000_000:+,.1f}T"
                    elif s_abs >= 999_950_000: txt = f"{val/1_000_000_000:+,.1f}B"
                    elif s_abs >= 999_500: txt = f"{val/1_000_000:+,.1f}M"
                    else: txt = f"{val:+,.0f}"
                    return f"[red]{txt}[/]" if val > 0 else f"[blue]{txt}[/]"

                flow_inv_str = (f"{_fmt_money(fr.get('indi'))} "
                                f"{_fmt_money(fr.get('frgn'))} "
                                f"{_fmt_money(fr.get('inst'))}")

        row_data = [
            date_str, 
            fmt_p(close), 
            diff_str, 
            fmt_p(row['open']), 
            fmt_p(row['high']), 
            fmt_p(row['low']), 
            fmt_ma(ma5_val, get_ma_color(ma5_val, 5)),
            fmt_ma(ma20_val, get_ma_color(ma20_val, 20)),
            fmt_ma(ma60_val, get_ma_color(ma60_val, 60)),
            fmt_ma(ma120_val, get_ma_color(ma120_val, 120)),
            obv_disp
        ]
        if not is_overseas and not is_domestic_index:
            row_data.append(foreign_rate_str)
            row_data.append(short_sale_str)
            row_data.append(inv_str)
        elif market_flow_map:
            row_data.append(flow_short_str)
            row_data.append(flow_inv_str)

        table.add_row(*row_data)
        
        if (i + 1) % 5 == 0 and (i + 1) < len(recent_df):
            table.add_section()
    
    config.console.print(table)

def _print_period_price_30(code, is_overseas):
    """기간별 시세 30일치 출력"""
    _print_period_price_common(code, is_overseas, limit=30)
    config.console.print()
